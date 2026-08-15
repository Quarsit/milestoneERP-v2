"""
Jenerik dışa aktarma (export) modülü — Milestone ERP
Herhangi bir liste verisini xlsx veya pdf olarak üretir.

Kullanım:
    from export_utils import liste_xlsx, liste_pdf
    headers = ['Sipariş No', 'Müşteri', 'Tutar']
    rows = [['SIP-1', 'Anka', '1000'], ['SIP-2', 'Beta', '2000']]
    return liste_xlsx('Siparişler', headers, rows, dosya_adi='siparisler')
    return liste_pdf('Siparişler', headers, rows, dosya_adi='siparisler')
"""
import io
from datetime import date


# Icerik tipi → dosya uzantisi. Uzanti burada, TEK YERDE ekleniyor.
# Onceden her cagiran kendisi eklemek zorundaydi; liste_pdf'in
# WeasyPrint yolu bunu unutmustu ve PDF'ler uzantisiz iniyordu.
# Windows dosyayi taniyamayip Word'e veriyor, icerik gecerli PDF
# oldugu halde "bozuk" gorunuyordu.
_UZANTILAR = {
    'application/pdf': '.pdf',
    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet': '.xlsx',
    'text/csv': '.csv',
}


def _make_response(data_bytes, dosya_adi, content_type, inline=False):
    from flask import make_response
    resp = make_response(data_bytes)
    resp.headers['Content-Type'] = content_type

    dosya_adi = (dosya_adi or 'dosya').strip() or 'dosya'
    uzanti = _UZANTILAR.get((content_type or '').split(';')[0].strip().lower())
    if uzanti and not dosya_adi.lower().endswith(uzanti):
        dosya_adi += uzanti

    # inline=True → tarayıcıda aç (PDF için); inline=False → indir (Excel için)
    yerlesim = 'inline' if inline else 'attachment'

    # Dosya adinda ASCII disi karakter varsa (Turkce ad, firma unvani)
    # duz filename= basligi bazi tarayicilarda bozulur. RFC 6266/5987
    # geregi ASCII yedek + filename* birlikte veriliyor.
    ascii_ad = dosya_adi.encode('ascii', 'replace').decode('ascii').replace('?', '_')
    from urllib.parse import quote
    resp.headers['Content-Disposition'] = (
        f'{yerlesim}; filename="{ascii_ad}"; '
        f"filename*=UTF-8''{quote(dosya_adi)}")
    return resp


def _firma_bilgisi():
    """Kayıtlı firma logosu (base64) ve firma adını döner.

    Veriler tablosu: kategori='firma_logo', deger='logo' → uzun_deger
    Erişilemezse (bağlam yoksa vb.) sessizce boş döner; logo yüzünden
    dışa aktarma başarısız olmamalı.
    """
    logo, ad = None, ''
    try:
        from models import Veriler
        k = Veriler.query.filter_by(kategori='firma_logo', deger='logo').first()
        logo = (k.uzun_deger if k else None) or None
        f = Veriler.query.filter_by(kategori='firma').first()
        ad = (f.deger if f else '') or ''
    except Exception:
        pass
    return logo, ad


def _logo_bytes(logo_veri):
    """'data:image/png;base64,...' → ham bayt. Çözülemezse None."""
    if not logo_veri:
        return None
    try:
        import base64
        t = str(logo_veri)
        if ',' in t:
            t = t.split(',', 1)[1]
        return base64.b64decode(t)
    except Exception:
        return None


def liste_xlsx(baslik, headers, rows, dosya_adi='liste', sayisal_sutunlar=None):
    """
    headers: ['Sütun1', 'Sütun2', ...]
    rows: [[deger1, deger2, ...], ...]
    sayisal_sutunlar: sağa yaslanacak/sayı formatı uygulanacak sütun indeksleri (0-tabanlı)
    """
    sayisal_sutunlar = sayisal_sutunlar or []
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError('Excel (.xlsx) çıktısı için "openpyxl" paketi kurulu olmalı. '
                           'Kurulum: pip install openpyxl')
    wb = Workbook()
    ws = wb.active
    ws.title = (baslik or 'Liste')[:31]

    n_col = max(1, len(headers))

    # ── FİRMA LOGOSU VE ADI ──
    _logo_veri, _firma_adi = _firma_bilgisi()
    _ham = _logo_bytes(_logo_veri)
    _kaydir = 0
    if _ham:
        try:
            from openpyxl.drawing.image import Image as XLImage
            _img = XLImage(io.BytesIO(_ham))
            _oran = (_img.height / _img.width) if _img.width else 1
            _img.width = 130
            _img.height = max(20, int(130 * _oran))
            ws.add_image(_img, 'A1')
            _kaydir = max(1, int(_img.height / 19) + 1)   # ~19 px = 1 satır
            for _i in range(1, _kaydir + 1):
                ws.row_dimensions[_i].height = 19
        except Exception:
            _kaydir = 0            # bozuk logo dosyayı bozmasın

    if _firma_adi:
        _kaydir += 1
        ws.merge_cells(start_row=_kaydir, start_column=1, end_row=_kaydir, end_column=n_col)
        f = ws.cell(row=_kaydir, column=1, value=_firma_adi)
        f.font = Font(bold=True, size=11, color='1E3A5F')
        f.alignment = Alignment(horizontal='left')

    # Başlık satırı
    _b = _kaydir + 1
    ws.merge_cells(start_row=_b, start_column=1, end_row=_b, end_column=n_col)
    c = ws.cell(row=_b, column=1, value=baslik)
    c.font = Font(bold=True, size=14, color='1E3A5F')
    c.alignment = Alignment(horizontal='left')
    # Tarih satırı
    ws.merge_cells(start_row=_b + 1, start_column=1, end_row=_b + 1, end_column=n_col)
    t = ws.cell(row=_b + 1, column=1, value=f'Oluşturma: {date.today().strftime("%d.%m.%Y")}')
    t.font = Font(size=9, color='888888')

    # Sütun başlıkları
    header_row = _b + 3
    thin = Side(style='thin', color='DDDDDD')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ci, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=ci, value=h)
        cell.font = Font(bold=True, color='FFFFFF', size=10)
        cell.fill = PatternFill('solid', start_color='1E3A5F')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Veri satırları
    r = header_row + 1
    for row in rows:
        for ci, val in enumerate(row, start=1):
            # Güvenli değer: None → boş, diğer her şey string (openpyxl uyumsuz tipleri önle)
            guvenli = '' if val is None else (val if isinstance(val, (int, float, str)) else str(val))
            cell = ws.cell(row=r, column=ci, value=guvenli)
            cell.border = border
            cell.font = Font(size=10)
            if (ci - 1) in sayisal_sutunlar:
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.alignment = Alignment(horizontal='left')
        # Zebra
        if (r - header_row) % 2 == 0:
            for ci in range(1, n_col + 1):
                ws.cell(row=r, column=ci).fill = PatternFill('solid', start_color='F7F5F0')
        r += 1

    # Sütun genişlikleri (içeriğe göre)
    for ci in range(1, n_col + 1):
        maxlen = len(str(headers[ci - 1])) if ci - 1 < len(headers) else 10
        for row in rows:
            if ci - 1 < len(row):
                maxlen = max(maxlen, len(str(row[ci - 1] if row[ci - 1] is not None else '')))
        ws.column_dimensions[get_column_letter(ci)].width = min(max(maxlen + 3, 10), 50)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    if not dosya_adi.endswith('.xlsx'):
        dosya_adi += '.xlsx'
    return _make_response(
        bio.getvalue(), dosya_adi,
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


def _turkce_font_kaydet():
    """Türkçe karakterleri destekleyen fontu kaydeder. Döner: (normal_font, bold_font)."""
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import os
    # Olası font yolları (Linux/Pardus, Windows)
    normal_yollar = [
        '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
        '/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf',
        '/usr/local/lib/python3.12/dist-packages/matplotlib/mpl-data/fonts/ttf/DejaVuSans.ttf',
        'C:\\Windows\\Fonts\\arial.ttf',
        'C:\\Windows\\Fonts\\tahoma.ttf',
    ]
    bold_yollar = [
        '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf',
        '/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf',
        'C:\\Windows\\Fonts\\arialbd.ttf',
        'C:\\Windows\\Fonts\\tahomabd.ttf',
    ]
    try:
        normal = next((p for p in normal_yollar if os.path.exists(p)), None)
        bold = next((p for p in bold_yollar if os.path.exists(p)), None)
        if normal:
            if 'TR' not in pdfmetrics.getRegisteredFontNames():
                pdfmetrics.registerFont(TTFont('TR', normal))
            if bold and 'TR-Bold' not in pdfmetrics.getRegisteredFontNames():
                pdfmetrics.registerFont(TTFont('TR-Bold', bold))
            return 'TR', ('TR-Bold' if bold else 'TR')
    except Exception:
        pass
    return 'Helvetica', 'Helvetica-Bold'


def liste_pdf(baslik, headers, rows, dosya_adi='liste', sayisal_sutunlar=None,
              ozet=None):
    """Liste PDF'i — HTML şablonu üzerinden (templates/liste_print.html).

    NEDEN HTML?
        Önceki sürüm reportlab ile düz tablo çiziyordu: firma logosu yoktu,
        tipografi ve renkler belge kimliğiyle uyuşmuyordu. Ekstre, proforma
        ve packing list zaten HTML şablonuyla üretiliyor ve çok daha iyi
        görünüyor. Liste çıktıları da aynı görsel dile taşındı; logo
        {{ firma_logo() }} bağlam işlevinden otomatik gelir.

    İMZA DEĞİŞMEDİ — flask_app.py tarafında düzenleme gerekmez.

    weasyprint kurulu değilse veya PDF üretimi başarısız olursa eski
    reportlab yolu devreye girer; çıktı alınamaması yerine sade bir PDF
    üretilir.
    """
    from datetime import date as _date
    sayisal = set(sayisal_sutunlar or [])

    try:
        from flask import render_template
        from weasyprint import HTML

        html = render_template(
            'liste_print.html',
            baslik=baslik or 'Liste',
            headers=headers or [],
            rows=rows or [],
            sayisal=sayisal,
            ozet=ozet or [],
            bugun=_date.today().strftime('%d.%m.%Y'),
        )
        pdf = HTML(string=html, base_url='.').write_pdf()
        # inline=True: reportlab yedegiyle AYNI davranis. Onceden
        # birincil yol dosya olarak indiriyor, yedek yol tarayicida
        # aciyordu — ayni istek, iki farkli sonuc.
        return _make_response(pdf, dosya_adi, 'application/pdf', inline=True)

    except Exception as hata:                                    # noqa: BLE001
        # Şablon yoksa, weasyprint kurulu değilse veya render hata verirse:
        # çıktı hiç alınamamaktansa sade PDF üretilsin.
        try:
            import logging
            logging.getLogger(__name__).warning(
                'liste_pdf: HTML yolu basarisiz (%s) — reportlab yedegine dusuldu', hata)
        except Exception:
            pass
        return _liste_pdf_reportlab(baslik, headers, rows, dosya_adi, sayisal_sutunlar,
                                    ozet=ozet)


def _liste_pdf_reportlab(baslik, headers, rows, dosya_adi='liste', sayisal_sutunlar=None,
                         ozet=None):
    """Yedek PDF üretici (eski reportlab yolu). Yalnızca HTML yolu
    çalışmadığında kullanılır."""
    """PDF tablo üretir (reportlab). Türkçe karakter destekli."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_LEFT
    except ImportError:
        raise RuntimeError('PDF çıktısı için "reportlab" paketi kurulu olmalı. '
                           'Kurulum: pip install reportlab')

    font_normal, font_bold = _turkce_font_kaydet()
    sayisal_sutunlar = sayisal_sutunlar or []
    bio = io.BytesIO()
    # Çok sütun varsa yatay sayfa
    yatay = len(headers) > 5
    sayfa = landscape(A4) if yatay else A4
    doc = SimpleDocTemplate(bio, pagesize=sayfa,
                            topMargin=15 * mm, bottomMargin=15 * mm,
                            leftMargin=12 * mm, rightMargin=12 * mm)
    styles = getSampleStyleSheet()
    baslik_stil = ParagraphStyle('Baslik', parent=styles['Title'], fontSize=15, fontName=font_bold,
                                 textColor=colors.HexColor('#1E3A5F'), alignment=TA_LEFT, spaceAfter=4)
    tarih_stil = ParagraphStyle('Tarih', parent=styles['Normal'], fontSize=8, fontName=font_normal,
                                textColor=colors.HexColor('#888888'), spaceAfter=10)

    elemanlar = [
        Paragraph(baslik or 'Liste', baslik_stil),
        Paragraph(f'Oluşturma: {date.today().strftime("%d.%m.%Y")}', tarih_stil),
        Spacer(1, 4),
    ]

    # ÖZET — HTML yolundaki blokla ayni bilgi. Yedege dusuldugunde
    # ozetin sessizce kaybolmamasi icin burada da basiliyor.
    if ozet:
        ozet_stil = ParagraphStyle('Ozet', parent=styles['Normal'], fontSize=8.5,
                                   leading=12, fontName=font_normal,
                                   textColor=colors.HexColor('#333333'))
        for _etiket, _deger in ozet:
            elemanlar.append(Paragraph(f'<b>{_etiket}:</b> {_deger}', ozet_stil))
        elemanlar.append(Spacer(1, 8))

    # Hücre içeriklerini Paragraph yap (uzun metin kayması için)
    hucre_stil = ParagraphStyle('Hucre', parent=styles['Normal'], fontSize=8, leading=10, fontName=font_normal)
    hucre_sag = ParagraphStyle('HucreSag', parent=hucre_stil, alignment=2)  # sağ
    head_stil = ParagraphStyle('Head', parent=styles['Normal'], fontSize=8, fontName=font_bold,
                               textColor=colors.white, leading=10)

    data = [[Paragraph(str(h), head_stil) for h in headers]]
    for row in rows:
        satir = []
        for ci, val in enumerate(row):
            stil = hucre_sag if ci in sayisal_sutunlar else hucre_stil
            satir.append(Paragraph(str(val if val is not None else ''), stil))
        data.append(satir)

    tablo = Table(data, repeatRows=1)
    tablo.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A5F')),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#DDDDDD')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING', (0, 0), (-1, -1), 5),
        ('RIGHTPADDING', (0, 0), (-1, -1), 5),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F7F5F0')]),
    ]))
    elemanlar.append(tablo)
    doc.build(elemanlar)
    bio.seek(0)
    if not dosya_adi.endswith('.pdf'):
        dosya_adi += '.pdf'
    return _make_response(bio.getvalue(), dosya_adi, 'application/pdf', inline=True)
