"""
/api/v2/excel/* — Excel dosyası ayrıştırma (toplu stok içe aktarma için).

NEDEN VAR:
    Tarayıcı .xlsx dosyasını kendi başına okuyamaz. Bu uç nokta dosyayı
    openpyxl ile açar; sayfa adlarını, sütun başlıklarını ve satırları
    JSON olarak döner.

NE YAPMAZ:
    Veritabanına HİÇBİR ŞEY yazmaz — salt okuma, yan etkisiz.

    Stok kayıtlarını tarayıcı, mevcut ``POST /api/stok`` uç noktası
    üzerinden satır satır oluşturur. Bu bilinçli bir tercihtir: toplu
    giriş ile tekli giriş böylece BİREBİR aynı muhasebe mantığını
    kullanır — KDV ayrıştırma (kdv dahil ise fiyat / (1+oran/100)),
    'Devreden KDV' maliyet kalemi ve tedarikçi cariye 'Alış Faturası'
    borcu. Bu mantık flask_app.create_app() içinde kapalı olduğundan
    buraya kopyalansaydı, iki yol zamanla birbirinden ayrılır ve toplu
    girişten gelen stokların maliyeti/borcu yanlış olurdu.
"""
from __future__ import annotations

import io

from flask import Blueprint, jsonify, request, session

bp = Blueprint("excel_import", __name__, url_prefix="/api/v2/excel")

MAKS_SATIR = 5000                 # bir sayfadan okunacak azami veri satırı
MAKS_BOYUT = 15 * 1024 * 1024     # 15 MB


def _auth_required() -> bool:
    return session.get("kullanici") is None


def _hucre(v):
    """openpyxl hücre değerini JSON'a uygun sade bir değere çevir."""
    if v is None:
        return ""
    if hasattr(v, "isoformat"):          # datetime / date
        return v.isoformat()[:10]
    if isinstance(v, bool):
        return "EVET" if v else "HAYIR"
    if isinstance(v, str):
        return v.strip()
    return v


@bp.post("/analiz")
def analiz():
    """Yüklenen çalışma kitabını ayrıştır.

    Dönen yapı::

        {ok: true, dosya_adi: "...", sayfalar: [
            {ad: "Sayfa1", basliklar: [...], satirlar: [[...], ...],
             satir_sayisi: 42}, ...
        ]}

    Tüm sayfalar tek seferde döner; kullanıcı sayfalar arasında geçiş
    yaparken dosyayı yeniden yüklemek zorunda kalmaz.
    """
    if _auth_required():
        return jsonify(error="Unauthorized"), 401

    dosya = request.files.get("dosya")
    if dosya is None or not dosya.filename:
        return jsonify(ok=False, mesaj="Dosya seçilmedi."), 400

    ad = dosya.filename.lower()
    if not (ad.endswith(".xlsx") or ad.endswith(".xlsm")):
        return jsonify(
            ok=False,
            mesaj="Yalnızca .xlsx / .xlsm desteklenir. Dosyanız .xls veya .csv "
                  "ise Excel'de 'Farklı Kaydet' ile .xlsx biçimine çevirin.",
        ), 400

    ham = dosya.read()
    if len(ham) > MAKS_BOYUT:
        return jsonify(ok=False, mesaj="Dosya çok büyük (en fazla 15 MB)."), 400

    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(ham), data_only=True, read_only=True)
    except Exception as e:                                   # noqa: BLE001
        return jsonify(ok=False, mesaj=f"Dosya okunamadı: {e}"), 400

    sayfalar = []
    try:
        for ws in wb.worksheets:
            basliklar, satirlar, kesildi = [], [], False

            for i, satir in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    basliklar = [_hucre(h) for h in satir]
                    continue
                if len(satirlar) >= MAKS_SATIR:
                    kesildi = True
                    break
                deger = [_hucre(h) for h in satir]
                if not any(str(d).strip() for d in deger):   # tamamen boş satır
                    continue
                satirlar.append(deger)

            # Başlığı boş olan sütunlara ad ver ki eşleştirmede seçilebilsinler
            basliklar = [
                b if str(b).strip() else f"Sütun {i + 1}"
                for i, b in enumerate(basliklar)
            ]

            sayfalar.append({
                "ad": ws.title,
                "basliklar": basliklar,
                "satirlar": satirlar,
                "satir_sayisi": len(satirlar),
                "kesildi": kesildi,
            })
    finally:
        try:
            wb.close()
        except Exception:                                     # noqa: BLE001
            pass

    if not sayfalar:
        return jsonify(ok=False, mesaj="Dosyada okunabilir sayfa bulunamadı."), 400

    return jsonify(ok=True, dosya_adi=dosya.filename, sayfalar=sayfalar)
