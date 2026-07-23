from flask import Flask, render_template, request, jsonify, session, redirect, url_for, flash, send_file
from werkzeug.security import generate_password_hash, check_password_hash
from sqlalchemy import func
from flask_migrate import Migrate
import logging, uuid, os, json
from datetime import datetime, date, timedelta
from decimal import Decimal, ROUND_HALF_UP
import requests
import xml.etree.ElementTree as ET

from models import db, Kullanici, BlokStok, PlakaStok, EbatliStok, StokCikis, Proforma, ProformaKalem, AuditLog
from models import Siparis, SiparisKalem, Rezervasyon, Cari, CariHareket, Maliyet, Sevkiyat, DovizKur, Veriler, SatisKaydi, Fatura, Banka, Kasa, KasaHareket, Kesim, KesimDetay
from models import Cek, CekHareket
import yedek as yedek_modul
import threading

def create_app():
    app = Flask(__name__)
    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
    # ── OTURUM GİZLİ ANAHTARI ──
    # Bu anahtar oturum çerezlerini imzalar. Herkesçe bilinen sabit bir değer
    # kalırsa, saldırgan kendi oturum çerezini üretip admin gibi girebilir.
    # Üretimde SECRET_KEY ortam değişkeni MUTLAKA ayarlanmalıdır.
    _gizli = os.environ.get('SECRET_KEY')
    if not _gizli:
        _gizli = 'milestone-erp-gelistirme-ANAHTARI-DEGISTIRIN'
        logging.warning(
            'SECRET_KEY ortam degiskeni ayarlanmamis — gelistirme anahtari kullaniliyor. '
            'URETIMDE MUTLAKA AYARLAYIN: export SECRET_KEY="rastgele-uzun-bir-deger"')
    app.secret_key = _gizli

    # ═══ BELGE ÇEVİRİ SİSTEMİ (_ceviri) ═══
    # Print şablonları (proforma, commercial invoice, packing list, ekstre)
    # bu fonksiyonla EN/TR başlık üretir. Anahtar sözlükte yoksa çökmez;
    # anahtarın son parçasını okunur metne çevirir (load_port → Load Port).
    _CEVIRI = {
        'belge.proforma_invoice': ('PROFORMA INVOICE', 'PROFORMA FATURA'),
        'belge.commercial_invoice': ('COMMERCIAL INVOICE', 'TİCARİ FATURA'),
        'belge.packing_list': ('PACKING LIST', 'ÇEKİ LİSTESİ'),
        'belge.date_label': ('Date:', 'Tarih:'),
        'belge.pi_no': ('PI No:', 'PI No:'),
        'belge.ci_no': ('CI No:', 'CI No:'),
        'belge.pl_no': ('PL No:', 'PL No:'),
        'belge.consignee': ('CONSIGNEE', 'ALICI'),
        'belge.order_ref': ('Order Ref', 'Sipariş Ref'),
        'belge.country_origin': ('Country of Origin', 'Menşe Ülke'),
        'belge.cont_no': ('Cont.', 'Sıra'),
        'belge.no': ('No', 'No'),
        'belge.description_of_goods': ('Description of Goods', 'Mal Açıklaması'),
        'belge.block_no': ('Block No', 'Blok No'),
        'belge.slab_no': ('Slab No', 'Plaka No'),
        'belge.crate_no': ('Crate No', 'Kasa No'),
        'belge.bundle': ('Bundle', 'Bundle'),
        'belge.spec': ('Spec', 'Özellik'),
        'belge.special': ('Special', 'Özel'),
        'belge.surface': ('Surface', 'Yüzey'),
        'belge.length': ('Length', 'Boy'),
        'belge.height': ('Height', 'Yükseklik'),
        'belge.thk': ('Thk', 'Kal.'),
        'belge.width': ('Width', 'En'),
        'belge.ton': ('TON', 'TON'),
        'belge.m3': ('M³', 'm³'),
        'belge.qnty': ('Qnty', 'Miktar'),
        'belge.pcs': ('Pcs', 'Adet'),
        'belge.sqm': ('Sqm', 'm²'),
        'belge.sqft': ('Sqft', 'Sqft'),
        'belge.unit': ('Unit', 'Birim'),
        'belge.price': ('Price', 'Fiyat'),
        'belge.amount': ('Amount', 'Tutar'),
        'belge.weight': ('Weight', 'Ağırlık'),
        'belge.total_weight': ('Total Weight', 'Toplam Ağırlık'),
        'belge.total_pcs': ('Total Pcs', 'Toplam Adet'),
        'belge.total_label': ('TOTAL:', 'TOPLAM:'),
        'belge.totally': ('TOTALLY', 'GENEL TOPLAM'),
        'belge.subtotal': ('Subtotal', 'Ara Toplam'),
        'belge.subtotal_gross': ('Gross Subtotal', 'Brüt Ara Toplam'),
        'belge.subtotal_ex_vat': ('Subtotal (Ex. VAT)', 'Ara Toplam (KDV Hariç)'),
        'belge.discount': ('Discount', 'İskonto'),
        'belge.disc': ('Disc.', 'İsk.'),
        'belge.discount_fixed': ('Discount', 'İskonto'),
        'belge.item_discounts': ('Item Discounts', 'Kalem İskontoları'),
        'belge.advance_fixed': ('Advance', 'Avans'),
        'belge.balance_due': ('Balance Due', 'Kalan Bakiye'),
        'belge.bank': ('Bank', 'Banka'),
        'belge.bank_details': ('Bank Details', 'Banka Bilgileri'),
        'belge.payment': ('Payment', 'Ödeme'),
        'belge.delivery': ('Delivery', 'Teslimat'),
        'belge.incoterms': ('Incoterms', 'Teslim Şekli'),
        'belge.shipment_details': ('Shipment Details', 'Sevkiyat Bilgileri'),
        'belge.shipment_terms': ('Shipment Terms', 'Sevkiyat Şartları'),
        'belge.load_port': ('Port of Loading', 'Yükleme Limanı'),
        'belge.discharge': ('Port of Discharge', 'Varış Limanı'),
        'belge.container': ('Container', 'Konteyner'),
        'belge.terms_conditions': ('Terms & Conditions', 'Şartlar ve Koşullar'),
        'belge.declaration': ('We declare that this invoice shows the actual price of the goods described and that all particulars are true and correct.',
                              'Bu faturanın, tanımlanan malların gerçek fiyatını gösterdiğini ve tüm bilgilerin doğru olduğunu beyan ederiz.'),
        'belge.buyer_acceptance': ("Buyer's Acceptance", 'Alıcı Onayı'),
        'belge.customer_approval': ('Customer Approval', 'Müşteri Onayı'),
        'belge.authorized_signature': ('Authorized Signature', 'Yetkili İmza'),
        # Ekstre başlıkları
        'tarih': ('Date', 'Tarih'),
        'islem_tipi': ('Transaction', 'İşlem Tipi'),
        'evrak_no': ('Doc No', 'Evrak No'),
        'aciklama': ('Description', 'Açıklama'),
        'borc': ('Debit', 'Borç'),
        'alacak': ('Credit', 'Alacak'),
        'doviz': ('Currency', 'Döviz'),
        'kumulatif_bakiye': ('Balance', 'Bakiye'),
        'toplam_borc': ('Total Debit', 'Toplam Borç'),
        'toplam_alacak': ('Total Credit', 'Toplam Alacak'),
        'net_ozet': ('Net Summary', 'Net Özet'),
        'musteri_bilgileri': ('Customer Details', 'Müşteri Bilgileri'),
        'musteri_onayi': ('Customer Approval', 'Müşteri Onayı'),
        'kase_imza': ('Stamp & Signature', 'Kaşe & İmza'),
        'yetkili_imza': ('Authorized Signature', 'Yetkili İmza'),
    }

    def _ceviri(anahtar, dil='en'):
        try:
            d = str(dil).lower() if dil else 'en'
        except Exception:
            d = 'en'
        cift = _CEVIRI.get(anahtar)
        if cift:
            return cift[1] if d.startswith('tr') else cift[0]
        # Bilinmeyen anahtar: son parçayı okunur yap (belge.load_port → Load Port)
        son = str(anahtar).split('.')[-1].replace('_', ' ')
        return son.title()

    app.jinja_env.globals['_ceviri'] = _ceviri
    app.jinja_env.globals['dil'] = 'en'  # şablonlar dil değişkeni beklerse varsayılan

    # Kategori bazlı değer çevirileri (yüzey işlemi, cari işlem tipi)
    _YUZEY_EN = {'Cilalı': 'Polished', 'Cilali': 'Polished', 'Honlu': 'Honed', 'Honed': 'Honed',
                 'Polished': 'Polished', 'Fırçalı': 'Brushed', 'Fircali': 'Brushed', 'Brushed': 'Brushed',
                 'Eskitme': 'Tumbled', 'Ham': 'Raw', 'Kumlama': 'Sandblasted', 'Patinato': 'Patinato',
                 'Mat': 'Matte', 'Dolgulu': 'Filled', 'Epoksi': 'Epoxy'}
    _ISLEM_EN = {'Tahsilat': 'Collection', 'Odeme': 'Payment', 'Ödeme': 'Payment',
                 'Avans Tahsilati': 'Advance Received', 'Avans Odemesi': 'Advance Paid',
                 'Avans Devri (Giriş)': 'Advance Transfer (In)', 'Avans Devri (Çıkış)': 'Advance Transfer (Out)',
                 'Fatura (Satis)': 'Sales Invoice', 'Fatura (Alis)': 'Purchase Invoice',
                 'Iade (Satis)': 'Sales Return', 'Iade (Alis)': 'Purchase Return',
                 'Vade Farki (Borc)': 'Late Fee (Debit)', 'Vade Farki (Alacak)': 'Late Fee (Credit)',
                 'Kur Farki (Borc)': 'FX Diff (Debit)', 'Kur Farki (Alacak)': 'FX Diff (Credit)',
                 'Çek (Borc)': 'Cheque (Issued)', 'Çek (Alacak)': 'Cheque (Received)',
                 'Çek Tahsilatı': 'Cheque Collection', 'Mahsup': 'Offset', 'Mahsup (Karşılıklı)': 'Offset (Mutual)',
                 'Mahsup (Fatura)': 'Offset (Invoice)', 'Bakiye Transfer (Giriş)': 'Balance Transfer (In)',
                 'Bakiye Transfer (Çıkış)': 'Balance Transfer (Out)', 'Bakiye Transfer (Kasaya)': 'Balance Transfer (Cash)',
                 'Acilis Bakiyesi': 'Opening Balance'}

    def _deger_cevir(kategori, deger, dil='en'):
        if deger is None:
            return ''
        deger = str(deger)
        try:
            d = str(dil).lower() if dil else 'en'
        except Exception:
            d = 'en'
        if d.startswith('tr'):
            return deger
        if kategori == 'yuzey':
            return _YUZEY_EN.get(deger, deger)
        if kategori == 'islem':
            return _ISLEM_EN.get(deger, deger)
        return deger

    _ODEME_EN = {'Pesin': 'Cash in Advance', 'Peşin': 'Cash in Advance',
                 '%30 Avans %70 Yukleme Oncesi': '30% Advance, 70% Before Loading',
                 '%50 Avans %50 Yukleme Oncesi': '50% Advance, 50% Before Loading',
                 'Akreditif (L/C)': 'Letter of Credit (L/C)', 'Mal Mukabili': 'Cash Against Goods',
                 'Vesaik Mukabili': 'Cash Against Documents', 'Vadeli': 'Deferred Payment'}

    def _odeme_cevir(deger, dil='en'):
        if not deger:
            return ''
        try:
            d = str(dil).lower() if dil else 'en'
        except Exception:
            d = 'en'
        if d.startswith('tr'):
            return str(deger)
        return _ODEME_EN.get(str(deger), str(deger))

    def _aciklama_cevir(metin, dil='en'):
        # Serbest metin — güvenli davranış: olduğu gibi döndür
        return '' if metin is None else str(metin)

    app.jinja_env.globals['_deger_cevir'] = _deger_cevir
    app.jinja_env.globals['_odeme_cevir'] = _odeme_cevir
    app.jinja_env.globals['_aciklama_cevir'] = _aciklama_cevir

    # Heartbeat gate icin: login sonrasi ilk sayfa yuklemesini isaretler.
    # session.pop ile bir kez okunur ve silinir — sonraki yuklemeler normal.
    app.jinja_env.globals['taze_giris_al'] = lambda: session.pop('taze_giris', False)

    db_url = os.environ.get('DATABASE_URL', 'sqlite:///milestone.db')
    if db_url.startswith('postgres://'):
        db_url = db_url.replace('postgres://', 'postgresql://', 1)
    app.config['SQLALCHEMY_DATABASE_URI'] = db_url
    app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

    db.init_app(app)
    migrate = Migrate(app, db)

    # ---------- YARDIMCI FONKSİYONLAR ----------
    def _auth_required():
        if 'kullanici' not in session:
            return redirect(url_for('giris'))
        return None

    # ─── ESNEK YETKİ SİSTEMİ ──────────────────────────────────────────
    # Modül listesi (yetki atanabilir)
    YETKI_MODULLERI = ['dashboard', 'stok', 'siparis', 'rezervasyon', 'proforma',
                       'fatura', 'cari', 'maliyet', 'sevkiyat', 'karlilik',
                       'satislar', 'raporlar', 'kasa', 'kesim', 'ayarlar', 'denetim']

    def _kullanici_yetkileri():
        """Aktif kullanıcının modül yetkilerini döner: {modul: 'gizli|okuma|yazma'}.
        admin -> hepsine 'yazma'."""
        rol = session.get('rol', '')
        if rol == 'admin' or rol == 'ADMIN':
            return {m: 'yazma' for m in YETKI_MODULLERI}
        ad = session.get('kullanici')
        if not ad:
            return {m: 'gizli' for m in YETKI_MODULLERI}
        k = Kullanici.query.filter_by(ad=ad).first()
        if not k:
            return {m: 'gizli' for m in YETKI_MODULLERI}
        try:
            kayitli = json.loads(k.yetkiler or '{}')
        except Exception:
            kayitli = {}
        # Kayitli yetki JSON'i hic yoksa (eski kullanici): tum modullere 'yazma' (geriye uyumluluk)
        # Kayitli yetki JSON'i varsa: tanimsiz modul -> 'gizli' (yeni eklenen modullere otomatik erisim yok)
        if not kayitli:
            return {m: 'yazma' for m in YETKI_MODULLERI}
        return {m: kayitli.get(m, 'gizli') for m in YETKI_MODULLERI}

    def _yetki_var_mi(modul, seviye='okuma'):
        """Aktif kullanıcının modüle erişimi var mı?
        seviye='okuma' -> okuma veya yazma yeterli
        seviye='yazma' -> sadece yazma yeterli"""
        yetkiler = _kullanici_yetkileri()
        mevcut = yetkiler.get(modul, 'gizli')
        if seviye == 'yazma':
            return mevcut == 'yazma'
        return mevcut in ('okuma', 'yazma')

    def _yetki_kontrol(modul, seviye='okuma'):
        """Endpoint başında çağrılır. Yetki yoksa JSON 403 döner, varsa None."""
        if not _yetki_var_mi(modul, seviye):
            return jsonify({'ok': False, 'error': 'yetki_yok',
                'mesaj': f'Bu işlem için yetkiniz yok ({modul}/{seviye}). Yöneticinizle görüşün.'}), 403
        return None

    def _proforma_onay_yetkisi_var_mi():
        """
        Proforma İÇ ONAYINI verme yetkisi. Modül yazma yetkisinden AYRI, ince taneli:
        yetkiler JSON'inde "proforma_onay": true olan (veya admin) onaylayabilir.
        Bu sayede proforma hazırlayabilen herkes onaylayamaz — çift kontrol sağlanır.
        Eski kullanıcılar (yetki JSON'i boş) geriye uyumluluk için onaylayabilir.
        """
        rol = session.get('rol', '')
        if rol in ('admin', 'ADMIN'):
            return True
        ad = session.get('kullanici')
        if not ad:
            return False
        k = Kullanici.query.filter_by(ad=ad).first()
        if not k:
            return False
        try:
            kayitli = json.loads(k.yetkiler or '{}')
        except Exception:
            kayitli = {}
        if not kayitli:
            return True  # eski kullanıcı: geriye uyumluluk
        return bool(kayitli.get('proforma_onay', False))
    # ─── YAZMA YETKİ GUARD (POST/PUT/DELETE) ──────────────────────────
    # URL -> Modül eşleşmesi
    URL_MODUL_MAP = [
        ('/api/stok', 'stok'),
        ('/api/siparis', 'siparis'),
        ('/api/rezervasyon', 'rezervasyon'),
        ('/api/proforma', 'proforma'),
        ('/api/fatura', 'fatura'),
        ('/api/tahsilat', 'fatura'),  # Tahsilat fatura modülüne dahil
        ('/api/cari', 'cari'),
        ('/api/maliyet', 'maliyet'),
        ('/api/sevkiyat', 'sevkiyat'),
        ('/api/satislar', 'satislar'),
        ('/api/satis', 'satislar'),
        ('/api/kasa', 'kasa'),
        ('/api/kesim', 'kesim'),
        ('/api/lookup', 'ayarlar'),
        ('/api/ayarlar', 'ayarlar'),
        ('/api/yedek', 'ayarlar'),
        ('/api/banka', 'ayarlar'),
        ('/api/denetim', 'denetim'),
    ]

    # Yazma kontrolünden MUAF endpoint'ler (genel veya kişisel)
    YAZMA_MUAF_PATHS = [
        '/api/yetkilerim',
        '/api/diag/',  # Diagnostic endpoint'leri
    ]

    @app.before_request
    def _yazma_yetki_guard():
        """POST/PUT/DELETE isteklerinde, kullanicinin ilgili modul icin
        YAZMA yetkisi olup olmadigini kontrol eder."""
        # Sadece API yazma istekleri
        if request.method not in ('POST', 'PUT', 'DELETE', 'PATCH'):
            return None
        if not request.path.startswith('/api/'):
            return None
        # Login ve cikis muaf
        if 'kullanici' not in session:
            return None
        # Admin her seye yetkili
        if session.get('rol') in ('admin', 'ADMIN'):
            return None
        # Muaf path'ler
        for muaf in YAZMA_MUAF_PATHS:
            if request.path.startswith(muaf):
                return None
        # Modul tespit et
        modul = None
        for prefix, m in URL_MODUL_MAP:
            if request.path.startswith(prefix):
                modul = m
                break
        if not modul:
            # Tanimlanmamis API - sessizce gec
            return None
        # Yazma yetkisi var mi?
        try:
            yetkiler = _kullanici_yetkileri()
            mevcut = yetkiler.get(modul, 'gizli')
            if mevcut != 'yazma':
                return jsonify({
                    'ok': False,
                    'error': 'yetki_yok',
                    'mesaj': f'⛔ Bu işlem için YAZMA yetkiniz yok. Modül: {modul} (mevcut: {mevcut})'
                }), 403
        except Exception as e:
            app.logger.error(f'Yazma yetki kontrol hatasi: {e}')
        return None
    # ─── /YAZMA YETKİ GUARD ───────────────────────────────────────────

    # ─── OKUMA YETKİ GUARD (GET) ──────────────────────────────────────
    # Bir modülü OKUMA yetkisi olmayan kullanıcı, o modülün GET verilerini
    # de çekemesin. Çapraz-modül yardımcı GET'ler muaftır (dropdown'lar,
    # kur, ayar okuma vb.) — aksi halde yetkili kullanıcı bile kendi
    # ekranını kullanamaz hale gelir.
    OKUMA_MUAF_PATHS = [
        '/api/yetkilerim',        # kullanıcının kendi yetkileri
        '/api/diag/',             # diagnostic
        '/api/doviz_kur',         # kur — her ekranda lazım
        '/api/dashboard',         # dashboard özetleri
        '/api/ayarlar/lookup',    # cins/ülke/banka dropdown'ları (çapraz)
        '/api/ayarlar/kdv_oran',  # KDV oranı (çapraz)
        '/api/ayarlar/firma',     # firma bilgisi (belge/proforma çapraz)
    ]

    @app.before_request
    def _okuma_yetki_guard():
        """GET isteklerinde, kullanicinin ilgili modul icin en az OKUMA
        yetkisi olup olmadigini kontrol eder."""
        if request.method != 'GET':
            return None
        if not request.path.startswith('/api/'):
            return None
        if 'kullanici' not in session:
            # Oturum yoksa _auth_required zaten 401 döndürecek; burada karışma
            return None
        if session.get('rol') in ('admin', 'ADMIN'):
            return None
        for muaf in OKUMA_MUAF_PATHS:
            if request.path.startswith(muaf):
                return None
        modul = None
        for prefix, m in URL_MODUL_MAP:
            if request.path.startswith(prefix):
                modul = m
                break
        if not modul:
            return None  # tanimlanmamis API — sessizce gec
        try:
            yetkiler = _kullanici_yetkileri()
            mevcut = yetkiler.get(modul, 'gizli')
            # okuma veya yazma yeterli; sadece 'gizli' engellenir
            if mevcut not in ('okuma', 'yazma'):
                return jsonify({
                    'ok': False,
                    'error': 'yetki_yok',
                    'mesaj': f'⛔ Bu modülü görüntüleme yetkiniz yok. Modül: {modul}'
                }), 403
        except Exception as e:
            app.logger.error(f'Okuma yetki kontrol hatasi: {e}')
        return None
    # ─── /OKUMA YETKİ GUARD ───────────────────────────────────────────

    @app.context_processor
    def _yetki_context():
        """Tum template'lere menu yetkilerini otomatik gecirir."""
        try:
            if 'kullanici' not in session:
                return {'menu_yetki': {}, 'menu_admin': False, 'menu_proforma_onay': False}
            return {
                'menu_yetki': _kullanici_yetkileri(),
                'menu_admin': session.get('rol') in ('admin', 'ADMIN'),
                'menu_proforma_onay': _proforma_onay_yetkisi_var_mi()
            }
        except Exception as e:
            app.logger.error(f'context_processor hatasi: {e}')
            # Hata olsa bile TUM modulleri yazma olarak ver (menü kaybolmasın)
            return {
                'menu_yetki': {m: 'yazma' for m in YETKI_MODULLERI},
                'menu_admin': False,
                'menu_proforma_onay': False
            }
    # ─── /ESNEK YETKİ SİSTEMİ ─────────────────────────────────────────

    @app.context_processor
    def _firma_kimlik_context():
        """
        Firma logosunu ve adını TÜM şablonlara geçirir (belgeler dahil).
        Logo CALLABLE olarak geçer — yalnızca şablon {{ firma_logo() }} çağırırsa
        veritabanından okunur. Böylece her sayfa yüklemesinde büyük base64
        veriyi boşuna çekmeyiz; sadece belge yazdırılırken okunur.
        """
        def _logo():
            try:
                k = Veriler.query.filter_by(kategori='firma_logo', deger='logo').first()
                return (k.uzun_deger if k else None) or None
            except Exception:
                return None
        def _firma_adi():
            try:
                k = Veriler.query.filter_by(kategori='firma').first()
                return (k.deger if k else None) or ''
            except Exception:
                return ''
        return {'firma_logo': _logo, 'firma_adi_ayar': _firma_adi}


    # ════════════════════════════════════════════════════════
    # STOK-REZERVASYON YARDIMCI FONKSİYONLARI
    # ════════════════════════════════════════════════════════
    def _stok_getir(stok_id, stok_tip):
        """Tip ve ID'ye göre doğru tablo'dan stok kaydını döner."""
        if stok_tip == 'BLOK':   return BlokStok.query.get(stok_id)
        elif stok_tip == 'PLAKA': return PlakaStok.query.get(stok_id)
        elif stok_tip == 'EBATLI': return EbatliStok.query.get(stok_id)
        return None

    # ════════════════════════════════════════════════════════
    # BİRİM DÖNÜŞTÜRME MOTORU
    # ════════════════════════════════════════════════════════
    # Temel birimler:  BLOK -> ton (+ m3),  PLAKA/EBATLI -> m2
    # Türetilmiş:      sqft = m2 * 10.7639
    # m3 ve ton BLOK'ta bağımsız ölçülerdir (ikisi de gerçek değer).
    # ════════════════════════════════════════════════════════
    # MERKEZI YUVARLAMA HELPER'LARI (Karma yaklaşım — Madde 15)
    # ════════════════════════════════════════════════════════
    # q3()     -> para ve miktar: 3 basamak, HALF_UP (banker's rounding değil)
    # q_kur()  -> kur değeri: 6 basamak (TCMB hassasiyetiyle uyumlu)
    # q_oran() -> yüzde/marj: 4 basamak (görsel olarak 2 basamak gösterilir)
    #
    # Kural: 4. basamak >=5 ise yukarı yuvarlanır (23.9768 -> 23.977).
    # Toplam hesaplamadan ÖNCE kalemleri q3() ile yuvarla -> yuvarlama farkı birikmesin.

    def _q(value, ndigits):
        if value is None or value == '':
            return 0.0
        try:
            d = Decimal(str(value)).quantize(
                Decimal('1.' + '0' * ndigits), rounding=ROUND_HALF_UP)
            return float(d)
        except Exception:
            try:
                return float(value)
            except Exception:
                return 0.0

    def q3(value):
        """Para ve miktar için 3 basamak yuvarlama."""
        return _q(value, 3)

    def q_kur(value):
        """Döviz kuru için 6 basamak yuvarlama."""
        return _q(value, 6)

    def q_oran(value):
        """Yüzde/oran (KDV, marj, iskonto) için 4 basamak yuvarlama."""
        return _q(value, 4)

    # ════════════════════════════════════════════════════════
    # KDV HELPER'LARI (Faz 6 - Maddeler 5, 6, 7)
    # ════════════════════════════════════════════════════════
    def _kdv_varsayilan_oran():
        """Ayarlardan varsayilan KDV oranini al (yoksa %20)."""
        try:
            kayit = Veriler.query.filter_by(
                kategori='kdv_ayar', deger='varsayilan_oran').first()
            if kayit and kayit.kisaltma:
                return float(kayit.kisaltma)
        except Exception:
            pass
        return 20.0

    def _uretim_avans_esigi():
        """
        Üretime geçmek için gereken MİNİMUM tahsilat yüzdesi (sipariş toplamına oranla).
        Ayarlardan okunur (kategori='siparis_ayar', deger='uretim_avans_yuzdesi').
        0 = kapı kapalı (ödeme kontrolü yapılmaz). Varsayılan 0 (geriye uyumlu).
        """
        try:
            kayit = Veriler.query.filter_by(
                kategori='siparis_ayar', deger='uretim_avans_yuzdesi').first()
            if kayit and kayit.kisaltma is not None:
                return float(kayit.kisaltma)
        except Exception:
            pass
        return 0.0

    def _siparis_tahsilat_durumu(siparis_id):
        """
        Bir siparişe yapılan toplam tahsilatı sipariş dövizi cinsinden döner.
        Kaynak: siparişe bağlı faturaların tahsilatları (cari hareketler, alacak).
        Döner: {'toplam': sipariş toplamı, 'tahsil': tahsil edilen, 'yuzde': %,
                'doviz': sipariş dövizi, 'fatura_sayisi': n}
        """
        sip = Siparis.query.get(siparis_id)
        if not sip:
            return None
        s_doviz = sip.doviz or 'USD'
        toplam = float(sip.toplam_tutar or 0)
        # Siparişe bağlı (iptal olmayan) faturalar
        faturalar = Fatura.query.filter_by(siparis_id=siparis_id).filter(
            Fatura.durum != 'Iptal').all()
        tahsil = 0.0
        for f in faturalar:
            # Her faturanın tahsilatını fatura dövizi eşdeğeriyle topla (çapraz döviz destekli)
            try:
                tahsil += _fatura_odenen_esdeger(f)
            except Exception:
                # Yardımcı yoksa/başarısızsa: doğrudan alacak toplamı
                dt = db.session.query(db.func.sum(CariHareket.alacak)).filter(
                    CariHareket.baglanti_tip == 'fatura', CariHareket.baglanti_id == f.id,
                    CariHareket.alacak > 0).scalar() or 0
                tahsil += float(dt)
        yuzde = (tahsil / toplam * 100) if toplam > 0 else 0
        return {'toplam': q3(toplam), 'tahsil': q3(tahsil), 'yuzde': round(yuzde, 1),
                'doviz': s_doviz, 'fatura_sayisi': len(faturalar)}

    def _kdv_hesapla(toplam_tutar, kdv_oran, kdv_dahil_mi):
        """
        Tutar ve KDV oranindan matrah + KDV tutarini hesaplar.
        - kdv_dahil_mi=True  -> tutar KDV dahildir, matrah ve KDV ayristirilir
        - kdv_dahil_mi=False -> tutar matrahdir, KDV ustune eklenir
        Donus: (matrah, kdv_tutar, toplam_kdv_dahil)
        """
        tutar = float(toplam_tutar or 0)
        oran = float(kdv_oran or 0) / 100.0
        if oran <= 0:
            return q3(tutar), 0.0, q3(tutar)
        if kdv_dahil_mi:
            matrah = tutar / (1 + oran)
            kdv = tutar - matrah
            return q3(matrah), q3(kdv), q3(tutar)
        else:
            kdv = tutar * oran
            return q3(tutar), q3(kdv), q3(tutar + kdv)

    # ════════════════════════════════════════════════════════
    # ÇOK DÖVİZLİ MUHASEBE (Faz 7 - Madde 11)
    # ════════════════════════════════════════════════════════
    # Ana para birimi: TRY (ayarlardan değiştirilebilir)
    # Her parasal işlem kayıt: doviz + tutar + kur + try_karsilik
    # Kur farkı: tahsilat/ödeme anında otomatik hesaplanır

    def _ana_para_birimi():
        """Ayarlardan ana para birimini al (default TRY)."""
        try:
            k = Veriler.query.filter_by(kategori='muhasebe', deger='ana_para_birimi').first()
            if k and k.kisaltma:
                return k.kisaltma
        except Exception:
            pass
        return 'TRY'

    def _kur_getir(doviz, tarih=None):
        """
        Belirtilen para birimi için TRY karşılığı kuru (TCMB'den) döner.
        doviz='USD' -> 1 USD kaç TRY?
        doviz='TRY' -> 1.0 (kendisi)
        doviz='EUR' -> 1 EUR kaç TRY?
        Tarih verilmezse en yakın günün kuru.
        """
        if not doviz or doviz == 'TRY':
            return 1.0
        try:
            q = DovizKur.query.filter_by(doviz=doviz)
            if tarih:
                q = q.filter(DovizKur.tarih <= tarih)
            k = q.order_by(DovizKur.tarih.desc()).first()
            if k:
                # Oncelik: alis -> efektif -> satis (hangi alanda veri varsa)
                val = k.alis or k.efektif or k.satis
                if val:
                    return q_kur(val)
        except Exception as e:
            app.logger.warning(f'Kur getirme hatasi ({doviz}): {e}')
        return 0  # bulunamadı

    def _try_karsilik(tutar, doviz, kur=None, tarih=None):
        """
        Verilen tutarı ana para birimine (TRY) çevirir.
        - kur verilmezse otomatik TCMB'den çekilir
        - doviz='TRY' ise kur=1 (kendisi)
        Donus: (try_karsilik, kullanilan_kur)
        """
        if not tutar:
            return 0.0, 1.0
        if doviz == 'TRY' or not doviz:
            return q3(tutar), 1.0
        if kur is None or kur <= 0:
            kur = _kur_getir(doviz, tarih)
        if not kur or kur <= 0:
            return 0.0, 0.0
        return q3(float(tutar) * float(kur)), q_kur(kur)

    def _kur_farki_hesapla_ve_olustur(yeni_hareket):
        """
        Tahsilat/odeme hareketi olusturulduktan sonra cagrilir.
        Ayni cari + ayni doviz icin kapanmamis (en eski) borc/alacak hareketini bulur,
        kur farkini hesaplar, otomatik 'Kur Farki' hareketi acar.
        """
        if not yeni_hareket.cari_id or not yeni_hareket.doviz:
            return None

        is_tahsilat = (yeni_hareket.alacak or 0) > 0
        is_odeme = (yeni_hareket.borc or 0) > 0
        if not (is_tahsilat or is_odeme):
            return None

        # ═══ ÇAPRAZ DÖVİZ: hareket bir FATURAYA bağlıysa önce o faturanın
        # açılış hareketini hedefle — döviz FARKLI olsa bile (TRY köprüsü ile).
        # Örn: TRY tahsilat, USD faturayı ödeme günündeki kurdan kapatır. ═══
        karsi = None
        karsi_tutar_attr = 'borc' if is_tahsilat else 'alacak'
        if yeni_hareket.baglanti_tip == 'fatura' and yeni_hareket.baglanti_id:
            fq = CariHareket.query.filter(
                CariHareket.baglanti_tip == 'fatura',
                CariHareket.baglanti_id == yeni_hareket.baglanti_id,
                CariHareket.cari_id == yeni_hareket.cari_id,
                CariHareket.kapatildi == False,
                CariHareket.id != yeni_hareket.id)
            if is_tahsilat:
                fq = fq.filter(CariHareket.borc > 0)
            else:
                fq = fq.filter(CariHareket.alacak > 0)
            karsi = fq.order_by(CariHareket.hareket_tarihi.asc(),
                                CariHareket.guncelleme.asc(),
                                CariHareket.id.asc()).first()

        if karsi is None and yeni_hareket.doviz == 'TRY':
            # Fatura bağlantısı olmayan TRY hareketi kör FIFO'ya girmez —
            # hangi dövizli borcu kapattığı belirsiz olur (muhasebesel risk).
            return None

        # Eslesecek karsi taraf (FIFO: en eski kapanmamis, AYNI dövizde) —
        # yalnızca fatura hedefli eşleşme bulunamadıysa
        if karsi is None:
            if is_tahsilat:
                karsi_q = CariHareket.query.filter(
                    CariHareket.cari_id == yeni_hareket.cari_id,
                    CariHareket.doviz == yeni_hareket.doviz,
                    CariHareket.borc > 0,
                    CariHareket.kapatildi == False,
                    CariHareket.id != yeni_hareket.id
                ).order_by(CariHareket.hareket_tarihi.asc(), CariHareket.guncelleme.asc(), CariHareket.id.asc())
            else:
                karsi_q = CariHareket.query.filter(
                    CariHareket.cari_id == yeni_hareket.cari_id,
                    CariHareket.doviz == yeni_hareket.doviz,
                    CariHareket.alacak > 0,
                    CariHareket.kapatildi == False,
                    CariHareket.id != yeni_hareket.id
                ).order_by(CariHareket.hareket_tarihi.asc(), CariHareket.guncelleme.asc(), CariHareket.id.asc())
            karsi = karsi_q.first()

        if not karsi:
            return None

        karsi_tutar = float(getattr(karsi, karsi_tutar_attr) or 0)
        yeni_tutar = float((yeni_hareket.alacak or 0) if is_tahsilat else (yeni_hareket.borc or 0))
        if karsi_tutar <= 0 or yeni_tutar <= 0:
            return None

        eski_kur = karsi.kur_uygulanan or _kur_getir(karsi.doviz, karsi.hareket_tarihi)
        yeni_kur = yeni_hareket.kur_uygulanan or _kur_getir(yeni_hareket.doviz, yeni_hareket.hareket_tarihi)
        if not eski_kur or not yeni_kur:
            return None

        capraz = (yeni_hareket.doviz or 'TRY') != (karsi.doviz or 'TRY')
        if not capraz:
            # AYNI DÖVİZ (mevcut mantık): eşleşen tutar üzerinden iki kur karşılaştırılır
            eslesen = min(karsi_tutar, yeni_tutar)
            if eslesen <= 0:
                return None
            eski_try = q3(eslesen * float(eski_kur))
            yeni_try = q3(eslesen * float(yeni_kur))
            karsi.kapatildi = (karsi_tutar <= yeni_tutar)
            yeni_hareket.kapatildi = (yeni_tutar <= karsi_tutar)
        else:
            # ═══ ÇAPRAZ DÖVİZ (TRY köprüsü) ═══
            # Örn: 47.000 TRY tahsilat → 1.000 USD fatura borcu.
            # Ödemenin TRY değeri, ödeme günündeki karşı-döviz kuruyla karşı
            # dövize çevrilir; fark = (ödeme günü kuru − fatura günü kuru) × eşleşen.
            guncel_karsi_kur = 1.0 if (karsi.doviz or 'TRY') == 'TRY' else \
                _kur_getir(karsi.doviz, yeni_hareket.hareket_tarihi)
            if not guncel_karsi_kur or guncel_karsi_kur <= 0:
                return None
            yeni_toplam_try = yeni_tutar * float(yeni_kur)
            # Ödemenin karşı döviz cinsinden bugünkü değeri
            yeni_karsi_esdeger = yeni_toplam_try / float(guncel_karsi_kur)
            eslesen = min(karsi_tutar, yeni_karsi_esdeger)  # karşı döviz cinsinden
            if eslesen <= 0:
                return None
            eski_try = q3(eslesen * float(eski_kur))          # fatura günü kuruyla
            yeni_try = q3(eslesen * float(guncel_karsi_kur))  # ödeme günü kuruyla
            _eps = 0.01
            karsi.kapatildi = (karsi_tutar <= yeni_karsi_esdeger + _eps)
            yeni_hareket.kapatildi = (yeni_karsi_esdeger <= karsi_tutar + _eps)

        fark = q3(yeni_try - eski_try)

        # Kapanis isaretle
        if karsi.kapatildi:
            karsi.kapanis_hareket_id = yeni_hareket.id

        if abs(fark) < 0.01:
            return None  # onemsiz fark

        # KUR FARKI MODU: Bagli faturadan oku
        # 'gider' = ihracat/ithalat default: cari kapansin, fark gider/gelire yazilir
        # 'cari'  = yurt ici default: kur farki cari'ye yansir, musteri TRY borclu/alacakli kalir
        kur_farki_modu = 'gider'  # varsayilan
        if karsi.baglanti_tip == 'fatura' and karsi.baglanti_id:
            bagli_fatura = Fatura.query.get(karsi.baglanti_id)
            if bagli_fatura:
                kur_farki_modu = getattr(bagli_fatura, 'kur_farki_modu', None) or 'gider'

        # Yön mantığı modlara göre:
        # fark > 0: yeni_try > eski_try (tahsilatta kur yukseldi/odemede kur yukseldi)
        # fark < 0: yeni_try < eski_try (tahsilatta kur dustu/odemede kur dustu)
        if kur_farki_modu == 'cari':
            # CARI MODU (yurt ici): musteri TRY borcu/alacagi kalir
            if is_tahsilat:
                if fark > 0:
                    # Musteri USD bazinda fazla geldi -> bize fazla TRY ödedi -> Alacak (cariden bize odeme)
                    kf_islem = 'Kur Farki (Alacak)'
                    kf_borc, kf_alacak = 0, abs(fark)
                else:
                    # Musteri USD bazinda eksik geldi -> musteri hala TRY borclu -> Borc (cariden alacak)
                    kf_islem = 'Kur Farki (Borc)'
                    kf_borc, kf_alacak = abs(fark), 0
            else:
                # Odeme (alis faturasi)
                if fark > 0:
                    # Fazla TRY odedik -> tedarikciden alacaklisin -> Alacak
                    kf_islem = 'Kur Farki (Alacak)'
                    kf_borc, kf_alacak = 0, abs(fark)
                else:
                    # Az TRY odedik -> tedarikciye hala TRY borclusun -> Borc
                    kf_islem = 'Kur Farki (Borc)'
                    kf_borc, kf_alacak = abs(fark), 0
        else:
            # GIDER MODU (ihracat/ithalat default): cari kapansin
            if is_tahsilat:
                if fark > 0:
                    # Kazanc: cariye Borc yazilir (kazanc gelire)
                    kf_islem = 'Kur Farki (Borc)'
                    kf_borc, kf_alacak = abs(fark), 0
                else:
                    # Kayip: cariden Alacak yazilir (cari kapansin, kayip gidere)
                    kf_islem = 'Kur Farki (Alacak)'
                    kf_borc, kf_alacak = 0, abs(fark)
            else:
                # Odeme (alis faturasi)
                if fark > 0:
                    # Fazla TRY odedik -> Borc (kayip)
                    kf_islem = 'Kur Farki (Borc)'
                    kf_borc, kf_alacak = abs(fark), 0
                else:
                    # Az TRY odedik -> Alacak (kazanc)
                    kf_islem = 'Kur Farki (Alacak)'
                    kf_borc, kf_alacak = 0, abs(fark)

        _capraz_not = f' (çapraz döviz {yeni_hareket.doviz}→{karsi.doviz})' if capraz else ''
        kf_hareket = CariHareket(
            id=_yeni_id('HR'),
            hareket_tarihi=yeni_hareket.hareket_tarihi or date.today(),
            cari_id=yeni_hareket.cari_id,
            cari_unvan=yeni_hareket.cari_unvan,
            islem_tip=kf_islem,
            aciklama=f'Otomatik kur farki - {karsi.id} ile {yeni_hareket.id} eslesmesi{_capraz_not}',
            borc=q3(kf_borc),
            alacak=q3(kf_alacak),
            doviz='TRY',
            kur_uygulanan=1.0,
            kur_kaynak='TCMB',
            borc_try=q3(kf_borc),
            alacak_try=q3(kf_alacak),
            kaynak='otomatik_kur_farki',
            baglanti_tip='hareket',
            baglanti_id=karsi.id,
            kullanici=session.get('kullanici', 'sistem')
        )
        db.session.add(kf_hareket)
        return kf_hareket

    def _kur_farki_geri_al(hareket_id):
        """
        Bir tahsilat/ödeme hareketi silinirken kur farkı eşleşmesini geri alır:
        - Bu hareketin kapattığı karşı hareketler yeniden açılır (kapatildi=False)
        - Bu hareketin eşleşmesinden doğan otomatik kur farkı kayıtları silinir
        Döner: (yeniden_acilan_sayisi, silinen_kf_sayisi). Commit ETMEZ.
        """
        acilan = 0
        for k in CariHareket.query.filter_by(kapanis_hareket_id=hareket_id).all():
            k.kapatildi = False
            k.kapanis_hareket_id = None
            acilan += 1
        silinen = 0
        for kf in CariHareket.query.filter(
                CariHareket.kaynak == 'otomatik_kur_farki',
                CariHareket.aciklama.like(f'%{hareket_id}%')).all():
            db.session.delete(kf)
            silinen += 1
        return acilan, silinen

    def _baglanti_okunabilir(baglanti_tip, baglanti_id):
        """
        Bir bağlantı referansını KULLANICIYA GÖSTERİLECEK okunabilir hale çevirir.
        Sistem ID'leri (BLK-87CBE9 gibi) rastgele kombinasyonlardır; kullanıcı
        bunları takip edemez. Bunun yerine blok no / kasa no / slab no gösterilir.
        Anlamlı alternatifi olmayan referanslar (sipariş, proforma, fatura no)
        olduğu gibi döner — onlar zaten takip edilebilir belge numaralarıdır.
        """
        if not baglanti_id:
            return ''
        tip = (baglanti_tip or '').lower()
        if tip == 'stok':
            s = BlokStok.query.get(baglanti_id)
            if s:
                return s.blok_no or baglanti_id
            s = PlakaStok.query.get(baglanti_id)
            if s:
                if getattr(s, 'slab_no', None):
                    return f"{s.blok_no or ''}#{s.slab_no}".lstrip('#')
                return s.blok_no or baglanti_id
            s = EbatliStok.query.get(baglanti_id)
            if s:
                return s.kasa_no or baglanti_id
        return baglanti_id

    def _devreden_kdv_kalemi_olustur(stok_id, kdv_tutar, doviz='TRY', aciklama='', fatura_no=None):
        """
        KDV'li stok girisi olunca otomatik 'Devreden KDV' maliyet kalemi olusturur.
        Bu kalem stoga bagli ama urunun kar hesabini etkilemez
        (sadece KDV mahsubu icin izlenir).
        """
        if not kdv_tutar or kdv_tutar <= 0:
            return None
        try:
            # USD karsiligi hesapla (kur)
            if doviz == 'USD':
                usd_k = kdv_tutar
            else:
                k = DovizKur.query.filter_by(doviz='USD').order_by(
                    DovizKur.tarih.desc()).first()
                kur = (k.alis if k else 0) or 1
                if doviz == 'TRY':
                    usd_k = kdv_tutar / kur if kur else 0
                elif doviz == 'EUR':
                    eur_k = DovizKur.query.filter_by(doviz='EUR').order_by(
                        DovizKur.tarih.desc()).first()
                    eur_kur = (eur_k.alis if eur_k else 0) or 0
                    usd_k = (kdv_tutar * eur_kur / kur) if kur else 0
                else:
                    usd_k = kdv_tutar

            m = Maliyet(
                id=_yeni_id('MYT'),
                maliyet_tip='Devreden KDV',
                baglanti_tip='stok',
                baglanti_id=stok_id,
                tutar=q3(kdv_tutar),
                doviz=doviz,
                usd_karsilik=q3(usd_k),
                fatura_no=(fatura_no or '').strip() or None,
                aciklama=aciklama or None,
                kullanici=session.get('kullanici') if 'kullanici' in session else 'sistem'
            )
            db.session.add(m)
            return m
        except Exception as e:
            app.logger.warning(f'Devreden KDV kalem hatasi: {e}')
            return None

    def _stok_cari_hareket_olustur(stok_id, uretici_unvan, toplam_tutar, doviz, fatura_no='', aciklama='',
                                   fatura_durumu='faturali', alis_tarihi=None):
        """
        Stok girişi yapılınca tedarikçi/üretici cariye ALACAK kaydı oluşturur.
        uretici_unvan: Cari.unvan ile eşleşir (urun_tedarikcisi=True)
        toplam_tutar: matrah + KDV (stoğun toplam alış bedeli)
        fatura_durumu: 'faturali' → borç oluştur | 'faturasiz' → borç OLUŞTURMA (mal geldi, fatura bekliyor)
                       'mal_bekliyor' → borç oluştur (fatura kesildi, mal henüz gelmedi)
        alis_tarihi: cari borcun (fatura) tarihi. Verilmezse bugün.
        Döner: oluşturulan CariHareket veya None
        """
        # FATURASIZ giriş: mal depoda ama henüz fatura yok → cari borç OLUŞMAZ.
        # Borç, fatura sonradan eşleştirilince oluşturulur.
        if fatura_durumu == 'faturasiz':
            return None
        if not uretici_unvan or not toplam_tutar or toplam_tutar <= 0:
            return None
        cari = Cari.query.filter_by(unvan=uretici_unvan, urun_tedarikcisi=True).first()
        if not cari:
            cari = Cari.query.filter_by(unvan=uretici_unvan).first()
        if not cari:
            return None
        try:
            # Borç tarihi = fatura/alış tarihi (mal giriş tarihinden farklı olabilir)
            borc_tarihi = alis_tarihi or date.today()
            # KUR: fatura tarihindeki kuru kullan (kayıt günündeki değil!)
            alacak_try_val, kullanilan_kur = _try_karsilik(toplam_tutar, doviz, tarih=borc_tarihi)
            ch = CariHareket(
                id=_yeni_id('HR'),
                hareket_tarihi=borc_tarihi,
                cari_id=cari.id, cari_unvan=cari.unvan,
                islem_tip='Alış Faturası',
                aciklama=aciklama or f'Stok girişi — {fatura_no or stok_id}',
                borc=0, alacak=toplam_tutar,
                alacak_try=alacak_try_val,
                doviz=doviz, kur_uygulanan=kullanilan_kur,
                evrak_no=fatura_no, baglanti_tip='stok', baglanti_id=stok_id,
                kaynak='stok', kullanici=session.get('kullanici', 'sistem'))
            db.session.add(ch)
            return ch
        except Exception as e:
            app.logger.warning(f'Stok cari hareket hatasi: {e}')
            return None

    M2_TO_SQFT = 10.7639  # 1 m² = 10.7639 sqft (sabit)

    # ════════════════════════════════════════════════════════
    # EBATLI STOK REFERANS KODU SISTEMI (Madde 2, 3, 4)
    # ════════════════════════════════════════════════════════
    # Kod yapisi: URT-CCCBYYKKS-N
    #   URT: 3 harfli uretici kisaltmasi (Cari kayitlarindan)
    #   CCC: 3 harfli cins kisaltmasi
    #   BY:  boy (cm, tam sayi)
    #   YY:  yukseklik (cm, tam sayi)
    #   KK:  kalinlik x 10 (2.0 -> "20", 1.5 -> "15")
    #   S:   seri harfi (A, B, C ... 99 doluyunca sonrakine geÃ§)
    #   N:   sira numarasi (1-99)
    # Ornek: MRM-IVT122612A-1  =  MARMADOS, IVORY TRAVERTEN, 122x61x2.0, A serisi #1

    UNLU_HARFLER = set('AEIOU')  # Y sessiz sayilir

    def _cins_kisaltma(cins):
        """Cins isminden 3 harfli kisaltma uretir.
        Tek kelime: 1. harf + 1. sessiz + 2. sessiz
        Cok kelime: 1. harf + 1. sessiz + 2. kelimenin 1. harfi
        Sessiz yetmezse sesli dahil edilir."""
        if not cins:
            return 'XXX'
        s = cins.upper().strip()
        kelimeler = [k for k in s.split() if k]
        if not kelimeler:
            return 'XXX'

        def _is_sessiz(h):
            return h.isalpha() and h not in UNLU_HARFLER

        k1 = kelimeler[0]
        ilk_harf = k1[0] if k1 else 'X'

        # 1. kelimenin ilk harften SONRAKI ilk sessiz harf
        ilk_sessiz_idx = None
        for i, h in enumerate(k1):
            if i == 0:
                continue
            if _is_sessiz(h):
                ilk_sessiz_idx = i
                break
        ilk_sessiz = k1[ilk_sessiz_idx] if ilk_sessiz_idx is not None else (k1[1] if len(k1) > 1 else '')

        if len(kelimeler) >= 2:
            ucuncu = kelimeler[1][0] if kelimeler[1] else 'X'
        else:
            # Tek kelime: 2. sessiz
            ikinci_sessiz_idx = None
            for i in range((ilk_sessiz_idx or 0) + 1, len(k1)):
                if _is_sessiz(k1[i]):
                    ikinci_sessiz_idx = i
                    break
            if ikinci_sessiz_idx is not None:
                ucuncu = k1[ikinci_sessiz_idx]
            else:
                # Sessiz kalmadi, kalan herhangi harf
                kalan = k1[(ilk_sessiz_idx or 0) + 1:]
                ucuncu = next((h for h in kalan), 'X')

        sonuc = (ilk_harf + ilk_sessiz + ucuncu).upper()
        while len(sonuc) < 3:
            sonuc += 'X'
        return sonuc[:3]

    def _uretici_kisaltma(uretici_unvan):
        """Carideki uretici kisaltma alanindan oku.
        Bulunamazsa cins kisaltma kuralini uretici unvanina uygula."""
        if not uretici_unvan:
            return 'XXX'
        cari = Cari.query.filter_by(unvan=uretici_unvan).first()
        if cari and getattr(cari, 'uretici_kisaltma', None):
            k = cari.uretici_kisaltma.strip().upper()
            if len(k) >= 1:
                return (k + 'XXX')[:3]
        # Fallback: cins kisaltma mantigi
        return _cins_kisaltma(uretici_unvan)

    def _referans_kodu_uret(uretici_unvan, cins, boy, yukseklik, kalinlik, adet=1):
        """
        Yeni N adet kasa icin referans kodlari uretir.
        Ayni boy+yukseklik+kalinlik+cins+uretici icin en son N'den devam eder.
        99 dolarsa B serisine gecer.
        Donus: liste, ornegin ['MRM-IVT122612A-6', 'MRM-IVT122612A-7', ...]
        """
        urt = _uretici_kisaltma(uretici_unvan)
        cks = _cins_kisaltma(cins)
        by = int(boy or 0)
        yk = int(yukseklik or 0)
        kk = int(round((kalinlik or 0) * 10))
        # Prefix kismi (seri ve sira disinda)
        prefix = f"{urt}-{cks}{by}{yk}{kk}"

        # Mevcut stoklara bak: ayni boy+yk+kalinlik+cins+uretici
        # Hem aktif (Serbest/Rezerve/Satildi vb.) tum durumlari kapsayalim
        from sqlalchemy import or_
        mevcut = EbatliStok.query.filter(
            EbatliStok.uretici == uretici_unvan,
            EbatliStok.cins == cins,
            EbatliStok.boy == boy,
            EbatliStok.yukseklik == yukseklik,
            EbatliStok.kalinlik == kalinlik,
            EbatliStok.kasa_no.like(prefix + '%')
        ).all()

        # En son seri harfi ve numarayi bul
        son_seri = 'A'
        son_sira = 0
        import re
        kod_re = re.compile(rf'^{re.escape(prefix)}([A-Z])-(\d+)$')
        for s in mevcut:
            kn = s.kasa_no or ''
            m = kod_re.match(kn)
            if not m:
                continue
            seri_h, sira_str = m.group(1), m.group(2)
            try:
                sira = int(sira_str)
            except Exception:
                continue
            # En buyuk serinin en buyuk sirasi
            if seri_h > son_seri or (seri_h == son_seri and sira > son_sira):
                son_seri = seri_h
                son_sira = sira

        # Bir sonraki sira/seri
        sonuclar = []
        seri = son_seri
        sira = son_sira
        for _ in range(int(adet or 1)):
            sira += 1
            if sira > 99:
                # Bir sonraki seri
                seri = chr(ord(seri) + 1)
                if seri > 'Z':
                    raise ValueError(f"Seri sinirina ulasildi (Z-99): {prefix}")
                sira = 1
            sonuclar.append(f"{prefix}{seri}-{sira}")
        return sonuclar



    def _stok_olcu(stok, birim):
        """
        Bir stok kaydının verilen birimdeki ölçü değerini döner.
        birim: 'ton' | 'm3' | 'm2' | 'sqft'
        - BLOK:   ton -> tonaj,   m3 -> hacim_m3
        - PLAKA/EBATLI: m2 -> metraj_m2,  sqft -> metraj_m2 * 10.7639
        Stok kaydı yoksa veya ölçü yoksa 0 döner.
        """
        if not stok:
            return 0
        if birim == 'ton':
            return getattr(stok, 'tonaj', 0) or 0
        elif birim == 'm3':
            return getattr(stok, 'hacim_m3', 0) or 0
        elif birim == 'm2':
            return getattr(stok, 'metraj_m2', 0) or 0
        elif birim == 'sqft':
            m2 = getattr(stok, 'metraj_m2', 0) or 0
            return m2 * M2_TO_SQFT
        return 0

    def _stok_temel_olcu(stok, stok_tip):
        """
        Stoğun TEMEL birimindeki ölçüsü (hesap motorunun kullandığı).
        BLOK -> tonaj,  PLAKA/EBATLI -> metraj_m2
        """
        if stok_tip == 'BLOK':
            return getattr(stok, 'tonaj', 0) or 0
        else:
            return getattr(stok, 'metraj_m2', 0) or 0

    def _birim_cevir(deger, kaynak_birim, hedef_birim, stok=None):
        """
        Bir ölçü değerini kaynak birimden hedef birime çevirir.
        m2 <-> sqft sabit katsayıyla çevrilir.
        ton <-> m3 BLOK'ta bağımsız ölçülerdir; çevrim için stok kaydı gerekir
        (stoğun kendi tonaj ve hacim_m3 oranı kullanılır).
        """
        if kaynak_birim == hedef_birim:
            return deger
        # m2 <-> sqft
        if kaynak_birim == 'm2' and hedef_birim == 'sqft':
            return deger * M2_TO_SQFT
        if kaynak_birim == 'sqft' and hedef_birim == 'm2':
            return deger / M2_TO_SQFT
        # ton <-> m3 : stoğun kendi oranıyla
        if stok and {kaynak_birim, hedef_birim} == {'ton', 'm3'}:
            tonaj = getattr(stok, 'tonaj', 0) or 0
            hacim = getattr(stok, 'hacim_m3', 0) or 0
            if kaynak_birim == 'ton' and hedef_birim == 'm3':
                return (deger / tonaj * hacim) if tonaj else 0
            if kaynak_birim == 'm3' and hedef_birim == 'ton':
                return (deger / hacim * tonaj) if hacim else 0
        return deger  # çevrilemezse aynen döner

    def _tutar_hesapla(stok, stok_tip, fiyat, fiyat_birim):
        """
        Bir stok için tutar = fiyat × (fiyat biriminin karşılığı olan ölçü).
        Örn: PLAKA, fiyat=90, fiyat_birim='m2' -> 90 × metraj_m2
             BLOK, fiyat=500, fiyat_birim='ton' -> 500 × tonaj
        """
        if not stok or not fiyat:
            return 0
        olcu = _stok_olcu(stok, fiyat_birim)
        return fiyat * olcu

    def _birim_etiket(birim):
        """Birim kodunu okunabilir etikete çevirir."""
        return {'ton': 'ton', 'm3': 'm³', 'm2': 'm²', 'sqft': 'sqft'}.get(birim, birim)

    def _stok_durum_degistir(stok_id, stok_tip, yeni_durum):
        """Stok durumunu güvenli şekilde değiştir. Geçerli durumlar: Serbest, Rezerve, Satildi, Sevkedildi."""
        stok = _stok_getir(stok_id, stok_tip)
        if stok:
            stok.durum = yeni_durum
            return True
        return False

    def _proforma_stoklarini_rezerve_et(proforma):
        """
        Proforma kalemlerindeki stokları REZERVE et.
        Sadece proforma siparişe bağlı DEĞİLSE çalışır (bağlıysa zaten Satıldı durumunda).
        Stoklar Serbest ise Rezerve yapılır, Rezervasyon kaydı oluşturulur.
        """
        if proforma.siparis_id:
            return 0  # Siparişe bağlı proforma -> zaten Satildi

        kalemler = ProformaKalem.query.filter_by(proforma_id=proforma.id).all()
        rezerve_sayisi = 0
        for k in kalemler:
            # Slab_no genelde stok_id'yi tutar
            sid = k.slab_no or ''
            if not sid:
                continue
            stip = k.urun_tip or 'PLAKA'
            stok = _stok_getir(sid, stip)
            if not stok or stok.durum != 'Serbest':
                continue

            # Aynı stok için aktif rezervasyon var mı?
            mevcut = Rezervasyon.query.filter_by(
                stok_id=sid, stok_tip=stip, iptal_nedeni=None
            ).first()
            if mevcut:
                continue  # zaten rezerve

            rez = Rezervasyon(
                id=_yeni_id('REZ'),
                musteri=proforma.musteri,
                proforma_id=proforma.id,
                siparis_id=None,
                stok_tip=stip,
                cins=stok.cins,
                ozellik=getattr(stok, 'ozellik', None),
                stok_id=sid,
                miktar=getattr(stok, 'metraj_m2', None) or getattr(stok, 'hacim_m3', None),
                rez_tip='Proforma',
                kullanici=session.get('kullanici')
            )
            db.session.add(rez)
            eski_d = stok.durum
            stok.durum = 'Rezerve'
            _log_audit('DURUM', 'stok', sid,
                       eski={'durum': eski_d}, yeni={'durum': 'Rezerve'},
                       aciklama=f'Proforma {proforma.id} icin rezerve edildi')
            rezerve_sayisi += 1
        return rezerve_sayisi

    def _proforma_rezervasyonlarini_iptal_et(proforma_id, neden='Proforma iptal'):
        """
        Bu proformaya bağlı aktif rezervasyonları iptal et.
        Stokları Serbest yap (hem Rezerve hem Satildi durumunda olanlari).
        Not: Proforma onaylanip siparise donusmusse stoklar Satildi durumunda olur,
        bu durumda da serbest birakilmali.
        """
        rezler = Rezervasyon.query.filter_by(
            proforma_id=proforma_id, iptal_nedeni=None
        ).all()
        iptal_sayisi = 0
        for r in rezler:
            r.iptal_nedeni = neden
            stok = _stok_getir(r.stok_id, r.stok_tip)
            if stok and stok.durum in ('Rezerve', 'Satildi'):
                eski_d = stok.durum
                stok.durum = 'Serbest'
                _log_audit('DURUM', 'stok', r.stok_id,
                           eski={'durum': eski_d}, yeni={'durum': 'Serbest'},
                           aciklama=f'Proforma {proforma_id} iptali ({neden}) - stok serbest birakildi')
                iptal_sayisi += 1
        return iptal_sayisi

    def _proformayi_siparise_baglava(proforma, siparis_id):
        """
        Proforma onaylandığında çağrılır. Mevcut Rezerve stokları Satildi'ya çevir,
        Rezervasyonların siparis_id'sini set et, rez_tip='Siparis' yap.
        # FAZ 16.5.2: PROFORMA-SIPARIS KALEM-AWARE
        FAZ 16: siparis_kalem_id'yi de stok_id eslesmesi ile set eder.
        """
        import json as _json
        rezler = Rezervasyon.query.filter_by(
            proforma_id=proforma.id, iptal_nedeni=None
        ).all()
        # Siparise bagli kalemleri onceden cek
        kalemler = SiparisKalem.query.filter_by(siparis_id=siparis_id).all()
        guncellenen = 0
        for r in rezler:
            r.siparis_id = siparis_id
            r.rez_tip = 'Siparis'
            # FAZ 16: Stok ID'ye gore eslesen kalemi bul ve baglan
            if not r.siparis_kalem_id:
                for k in kalemler:
                    if not k.stok_ids_json:
                        continue
                    try:
                        stok_ids = _json.loads(k.stok_ids_json)
                        if r.stok_id in stok_ids:
                            r.siparis_kalem_id = k.id
                            break
                    except Exception:
                        pass
            stok = _stok_getir(r.stok_id, r.stok_tip)
            if stok and stok.durum == 'Rezerve':
                stok.durum = 'Satildi'
                _log_audit('DURUM', 'stok', r.stok_id,
                           eski={'durum': 'Rezerve'}, yeni={'durum': 'Satildi'},
                           aciklama=f'Siparis {siparis_id} olusturuldu (proforma {proforma.id}) - stok satildi')
                guncellenen += 1
        return guncellenen

    def _siparis_stoklarini_serbest_birak(siparis_id, neden='Siparis iptal'):
        """Sipariş iptal/silme: bağlı rezervasyonları iptal, Satildi/Rezerve -> Serbest."""
        rezler = Rezervasyon.query.filter_by(
            siparis_id=siparis_id, iptal_nedeni=None
        ).all()
        serbest_sayisi = 0
        for r in rezler:
            r.iptal_nedeni = neden
            stok = _stok_getir(r.stok_id, r.stok_tip)
            if stok and stok.durum in ('Satildi', 'Rezerve'):
                eski_d = stok.durum
                stok.durum = 'Serbest'
                _log_audit('DURUM', 'stok', r.stok_id,
                           eski={'durum': eski_d}, yeni={'durum': 'Serbest'},
                           aciklama=f'Siparis {siparis_id} iptali ({neden}) - stok serbest birakildi')
                serbest_sayisi += 1
        return serbest_sayisi

    def _siparis_stoklarini_durum_guncelle(siparis_id, yeni_stok_durum):
        """
        Sipariş durumu değişince bağlı stokların durumunu güncelle.
        Sadece AKTİF rezervasyonların stoklarına dokunur.
        Sevkedildi/Serbest durumundaki stokları değiştirmez (güvenlik).
        """
        rezler = Rezervasyon.query.filter_by(
            siparis_id=siparis_id, iptal_nedeni=None
        ).all()
        guncellenen = 0
        for r in rezler:
            stok = _stok_getir(r.stok_id, r.stok_tip)
            # Sadece Rezerve/Satildi durumundaki stoklari guncelle (Sevkedildi/Serbest'e dokunma)
            if stok and stok.durum in ('Rezerve', 'Satildi') and stok.durum != yeni_stok_durum:
                stok.durum = yeni_stok_durum
                guncellenen += 1
        return guncellenen

    def _siparis_teslim_edildi(siparis_id, teslim_tarihi=None):
        """
        Sipariş 'Teslim Edildi' durumuna geçtiğinde her bağlı stok için
        SatisKaydi oluşturur ve stok durumunu 'Teslim Edildi' yapar.

        - Maliyet: Maliyet tablosundan baglanti_id=siparis_id ile toplanır
        - Kur: o anki TCMB kuru (DovizKur tablosundan)
        - Bağlı en son aktif Proforma'dan proforma_id alınır
        - Bağlı en son Sevkiyat varsa sevkiyat_id alınır
        - Eğer bu sipariş için zaten SatisKaydi varsa atlar (idempotent)
        """
        sip = Siparis.query.get(siparis_id)
        if not sip:
            return 0

        # Mevcut satış kayıtlarını kontrol et (idempotent)
        mevcut_stoklar = {sk.stok_id for sk in SatisKaydi.query.filter_by(siparis_id=siparis_id).all()}

        # Aktif rezervasyonlar
        rezler = Rezervasyon.query.filter_by(
            siparis_id=siparis_id, iptal_nedeni=None
        ).all()
        if not rezler:
            # ── STOKSUZ TESLİM FALLBACK ──
            # Rezervasyon yok. Kâr kaybını önlemek için:
            # 1) Bu sipariş için zaten SatisKaydi varsa (fatura yolundan) dokunma.
            # 2) Siparişe bağlı KESİLMİŞ fatura varsa kaydı fatura motoruna yaptır
            #    (maliyet fatura üzerinden doğru hesaplanır).
            # 3) Hiçbiri yoksa SiparisKalem'den satış kaydı üret (maliyet 0 —
            #    sonradan Maliyet modülünden girilirse otomatik güncellenir).
            if mevcut_stoklar:
                return 0
            _fat = Fatura.query.filter_by(siparis_id=siparis_id).filter(
                Fatura.durum != 'Iptal').order_by(Fatura.olusturma.desc()).first() \
                if hasattr(Fatura, 'olusturma') else \
                Fatura.query.filter_by(siparis_id=siparis_id).filter(Fatura.durum != 'Iptal').first()
            if _fat and _fat.durum in ('Kesildi', 'Tahsil Edildi', 'Kismi Tahsil'):
                try:
                    olusan, _ = _fatura_satis_kaydi_olustur(_fat.id)
                    return olusan or 0
                except Exception as _e:
                    app.logger.warning(f'Stoksuz teslim fatura kaydı: {_e}')
            # 3) SiparisKalem'den üret
            _kalemler = SiparisKalem.query.filter_by(siparis_id=siparis_id).order_by(SiparisKalem.sira).all()
            if not _kalemler:
                return 0
            _u = DovizKur.query.filter_by(doviz='USD').order_by(DovizKur.tarih.desc()).first()
            _e = DovizKur.query.filter_by(doviz='EUR').order_by(DovizKur.tarih.desc()).first()
            _ku = (_u.alis if _u else 0) or 0
            _ke = (_e.alis if _e else 0) or 0

            def _sk_usd(t, d):
                if not t: return 0
                if d in ('USD', None, ''): return t
                if d == 'EUR': return (t * _ke / _ku) if _ku else 0
                if d == 'TRY': return (t / _ku) if _ku else 0
                return t

            olusan = 0
            for _k in _kalemler:
                _tut = q3(_k.toplam_fiyat or ((_k.miktar or 0) * (_k.birim_fiyat or 0)))
                _tut_usd = q3(_sk_usd(_tut, _k.doviz or sip.doviz))
                sk = SatisKaydi(
                    id=_yeni_id('SAT'),
                    stok_id=f'STOKSUZ-{siparis_id}-{_k.id}',
                    stok_tip=(_k.urun_tip or 'STOKSUZ'),
                    cins=_k.cins,
                    siparis_id=siparis_id,
                    siparis_kalem_id=_k.id,
                    kaynak='siparis_teslim',
                    musteri=sip.musteri,
                    satis_tarihi=teslim_tarihi or date.today(),
                    miktar=_k.miktar, birim=_k.birim,
                    doviz=_k.doviz or sip.doviz or 'USD',
                    tutar=_tut_usd,
                    kur_usd=q_kur(_ku), kur_eur=q_kur(_ke),
                    tutar_usd=_tut_usd,
                    tutar_try=q3(_tut_usd * _ku if _ku else 0),
                    maliyet_usd=0, maliyet_try=0,
                    kar_usd=_tut_usd,
                    marj_yuzde=q_oran(100 if _tut_usd else 0),
                    kullanici=session.get('kullanici'))
                db.session.add(sk)
                olusan += 1
            if olusan:
                app.logger.info(f'Stoksuz teslim: {siparis_id} için {olusan} satış kaydı '
                                f'SiparisKalem üzerinden oluşturuldu (maliyet 0).')
            return olusan

        # Güncel kurlar (her döviz ayrı satır) - alım fiyatı çevrimi için önce lazım
        usd_kur_kayit = DovizKur.query.filter_by(doviz='USD').order_by(DovizKur.tarih.desc()).first()
        eur_kur_kayit = DovizKur.query.filter_by(doviz='EUR').order_by(DovizKur.tarih.desc()).first()
        kur_usd = (usd_kur_kayit.alis if usd_kur_kayit else 0) or 0
        kur_eur = (eur_kur_kayit.alis if eur_kur_kayit else 0) or 0

        def _usd_cevir(tutar, dvz):
            """Verilen tutarı USD'ye çevir."""
            if not tutar:
                return 0
            if dvz == 'USD' or not dvz:
                return tutar
            elif dvz == 'EUR':
                return (tutar * kur_eur / kur_usd) if kur_usd else 0
            elif dvz == 'TRY':
                return (tutar / kur_usd) if kur_usd else 0
            return tutar

        # ── MALİYET HESABI ──
        # 1) EK MALİYETLER: Maliyet tablosundan (nakliye, işçilik, gümrük vs.)
        #    Sipariş bazlı + stok bazlı kayıtlar toplanır
        ek_maliyet_usd = db.session.query(
            db.func.sum(Maliyet.usd_karsilik)
        ).filter(Maliyet.baglanti_id == siparis_id).scalar() or 0
        ek_maliyet_try = db.session.query(
            db.func.sum(Maliyet.try_karsilik)
        ).filter(Maliyet.baglanti_id == siparis_id).scalar() or 0
        stok_idler = [r.stok_id for r in rezler]
        if stok_idler:
            # KDV kalemleri HARIC (Devreden/Iade KDV maliyet hesabina girmez)
            ek_maliyet_usd += db.session.query(db.func.sum(Maliyet.usd_karsilik)).filter(
                Maliyet.baglanti_id.in_(stok_idler),
                ~Maliyet.maliyet_tip.in_(['Devreden KDV', 'Iade KDV'])
            ).scalar() or 0
            ek_maliyet_try += db.session.query(db.func.sum(Maliyet.try_karsilik)).filter(
                Maliyet.baglanti_id.in_(stok_idler),
                ~Maliyet.maliyet_tip.in_(['Devreden KDV', 'Iade KDV'])
            ).scalar() or 0

        # Ek maliyetler stoklara eşit dağıtılır
        stok_sayi = len(rezler)
        per_stok_ek_maliyet_usd = ek_maliyet_usd / stok_sayi if stok_sayi > 0 else 0
        per_stok_ek_maliyet_try = ek_maliyet_try / stok_sayi if stok_sayi > 0 else 0
        # Not: her stoğun ALIM FİYATI kendi kaydından alınır (asağıdaki döngüde)

        # Bağlı proforma (en son aktif olan)
        proforma_id = None
        baglı_pf = Proforma.query.filter_by(siparis_id=siparis_id).filter(
            ~Proforma.durum.in_(['Iptal'])
        ).order_by(Proforma.olusturma.desc()).first()
        if baglı_pf:
            proforma_id = baglı_pf.id

        # Bağlı sevkiyat (varsa)
        sevkiyat_id = None
        sev = Sevkiyat.query.filter_by(siparis_id=siparis_id).order_by(Sevkiyat.sevk_tarihi.desc()).first()
        if sev:
            sevkiyat_id = sev.id

        # Müşteri ülke (Cari'den)
        cari = Cari.query.filter_by(unvan=sip.musteri).first()
        musteri_ulke = cari.ulke if cari else None

        olusturulan = 0
        for r in rezler:
            if r.stok_id in mevcut_stoklar:
                continue  # Zaten satış kaydı var

            stok = _stok_getir(r.stok_id, r.stok_tip)
            if not stok:
                continue

            doviz = sip.doviz or 'USD'
            # FAZ 16.5.1: KALEM-AWARE TESLIM
            # ── SATIŞ TUTARI (KALEM-AWARE) ──
            # FAZ 16: Siparis artık tek bir birim/fiyat tutmuyor.
            # Bilgiler rezervasyonun bagli oldugu kalemden alinir.
            kalem = SiparisKalem.query.get(r.siparis_kalem_id) if r.siparis_kalem_id else None
            if not kalem:
                # Rezervasyon bir kaleme bagli degil — bu satis kaydi olusturulamaz.
                # Boyle bir durum normalde olmamali; FAZ 16 sonrası tüm rezervasyonlar
                # bir kaleme baglidir. Eski kayitlar icin atlanir.
                logging.warning(f"_siparis_teslim_edildi: Rezervasyon {r.id} bir kaleme bagli degil, atlandi.")
                continue

            # Birim ve fiyat kalemden gelir
            satis_birim = (kalem.birim or '').lower()
            satis_birim = {'ton': 'ton', 'm3': 'm3', 'm2': 'm2', 'sqft': 'sqft'}.get(
                satis_birim, 'ton' if r.stok_tip == 'BLOK' else 'm2')
            birim_fiyat = kalem.birim_fiyat or 0

            # Stoğun satış birimindeki ölçüsü (birim motoru)
            satis_olcu = _stok_olcu(stok, satis_birim)
            tutar = birim_fiyat * satis_olcu

            # SatisKaydi'na yazılacak miktar/birim = satış birimi cinsinden
            miktar = satis_olcu
            birim = satis_birim

            # USD/TRY karşılık
            tutar_usd = tutar
            tutar_try = 0
            if doviz == 'USD':
                tutar_usd = tutar
                tutar_try = tutar * kur_usd if kur_usd else 0
            elif doviz == 'EUR':
                tutar_usd = tutar * kur_eur / kur_usd if kur_usd else 0
                tutar_try = tutar * kur_eur if kur_eur else 0
            elif doviz == 'TRY':
                tutar_try = tutar
                tutar_usd = tutar / kur_usd if kur_usd else 0

            # ── STOĞUN KENDİ ALIM MALİYETİ (birim sistemine göre) ──
            # alis_fiyati, alis_fiyat_birim biriminde birim fiyat
            # Alım maliyeti = alış_fiyatı × (alış biriminin stoktaki ölçüsü)
            stok_alis_birim_fiyat = getattr(stok, 'alis_fiyati', None) or 0
            stok_alis_birim = getattr(stok, 'alis_fiyat_birim', None) or (
                'ton' if r.stok_tip == 'BLOK' else 'm2')
            stok_doviz = getattr(stok, 'doviz', None) or 'USD'
            alis_olcu = _stok_olcu(stok, stok_alis_birim)
            alim_maliyet_usd = _usd_cevir(stok_alis_birim_fiyat * alis_olcu, stok_doviz)

            # TOPLAM MALİYET = stoğun alım maliyeti + paya düşen ek maliyet
            stok_maliyet_usd = alim_maliyet_usd + per_stok_ek_maliyet_usd
            stok_maliyet_try = (alim_maliyet_usd * kur_usd if kur_usd else 0) + per_stok_ek_maliyet_try

            # Kar/marj
            kar_usd = tutar_usd - stok_maliyet_usd
            marj = (kar_usd / tutar_usd * 100) if tutar_usd > 0 else 0

            sk = SatisKaydi(
                id=_yeni_id('SAT'),
                stok_id=r.stok_id,
                stok_tip=r.stok_tip,
                cins=stok.cins,
                ozellik=getattr(stok, 'ozellik', None),
                blok_no=getattr(stok, 'blok_no', None) or getattr(stok, 'kasa_no', None),
                boy=getattr(stok, 'boy', None),
                yukseklik=getattr(stok, 'yukseklik', None),
                kalinlik=getattr(stok, 'kalinlik', None),
                en=getattr(stok, 'en', None),
                metraj_m2=getattr(stok, 'metraj_m2', None),
                metraj_sqft=getattr(stok, 'metraj_sqft', None),
                hacim_m3=getattr(stok, 'hacim_m3', None),
                tonaj=getattr(stok, 'tonaj', None),
                agirlik_kg=getattr(stok, 'm2_kg', None),
                siparis_id=siparis_id,
                siparis_kalem_id=kalem.id,
                proforma_id=proforma_id,
                sevkiyat_id=sevkiyat_id,
                musteri=sip.musteri,
                musteri_ulke=musteri_ulke,
                satis_tarihi=sip.siparis_tarihi or date.today(),
                teslim_tarihi=teslim_tarihi or date.today(),
                birim_fiyat=q3(birim_fiyat),
                miktar=q3(miktar),
                birim=birim,
                doviz=doviz,
                tutar=q3(tutar),
                kur_usd=q_kur(kur_usd),
                kur_eur=q_kur(kur_eur),
                tutar_usd=q3(tutar_usd),
                tutar_try=q3(tutar_try),
                maliyet_usd=q3(stok_maliyet_usd),
                maliyet_try=q3(stok_maliyet_try),
                kar_usd=q3(kar_usd),
                marj_yuzde=q_oran(marj),
                kullanici=session.get('kullanici')
            )
            db.session.add(sk)

            # Stok durumunu Teslim Edildi yap
            stok.durum = 'Teslim Edildi'
            olusturulan += 1

        # ── OTOMATİK BORÇ (çift kayıt koruması) ──
        # Bu siparişten/proformasından fatura kesilmişse otomatik borç OLUŞMAZ
        # (fatura zaten borcu işledi). Faturası yoksa teslim borcu işlenir.
        if olusturulan > 0:
            # Siparişe bağlı fatura var mı?
            fatura_var = Fatura.query.filter_by(siparis_id=siparis_id).filter(
                Fatura.durum != 'Iptal').first()
            # Bu sipariş için zaten otomatik teslim borcu işlenmiş mi?
            mevcut_borc = CariHareket.query.filter_by(
                baglanti_tip='siparis', baglanti_id=siparis_id, kaynak='otomatik-teslim'
            ).first()

            if not fatura_var and not mevcut_borc:
                # Sipariş toplam tutarı (USD bazlı kayıtlı tutar)
                toplam_satis = sum((s.tutar or 0) for s in
                    SatisKaydi.query.filter_by(siparis_id=siparis_id).all())
                if toplam_satis > 0:
                    # Cari bulunamazsa teslim akisi cokmesin - borc atlanir
                    try:
                        _cari_hareket_ekle(
                            cari_unvan=sip.musteri,
                            islem_tip='Teslimat',
                            borc=toplam_satis,
                            doviz=sip.doviz or 'USD',
                            aciklama=f'Siparis teslimi {siparis_id}',
                            kaynak='otomatik-teslim',
                            baglanti_tip='siparis',
                            baglanti_id=siparis_id
                        )
                    except ValueError as e:
                        # Cari yok - borc islenemedi ama teslim devam eder
                        app.logger.warning(f'Teslim borc atlandi ({siparis_id}): {e}')

        # ════════════════════════════════════════════════════
        # KDV İADE DÖNÜŞÜMÜ (Madde 6)
        # Satis tipi 'ihracat' | 'ihrac_kayitli' | 'yurtici_kdvsiz' ise
        # bu siparis kapsamindaki stoklarin 'Devreden KDV' kalemleri
        # 'Iade KDV' olarak yeniden etiketlenir (devletten geri alinacak).
        # 'yurtici_kdvli' ise Devreden KDV olarak kalir (mahsub).
        # ════════════════════════════════════════════════════
        iade_kapsami = ('ihracat', 'ihrac_kayitli', 'yurtici_kdvsiz')
        if getattr(sip, 'satis_tipi', 'ihracat') in iade_kapsami:
            stok_idler = [r.stok_id for r in rezler if r.stok_id]
            if stok_idler:
                kdv_kalemleri = Maliyet.query.filter(
                    func.lower(Maliyet.baglanti_tip) == 'stok',
                    Maliyet.baglanti_id.in_(stok_idler),
                    Maliyet.maliyet_tip == 'Devreden KDV',
                    Maliyet.aktif == True  # Sadece aktif kayitlari donustur
                ).all()
                bugun_donusum = date.today()
                for m in kdv_kalemleri:
                    # C SIKKI: Eski Devreden kaydini PASIF yap, yeni Iade kaydi olustur
                    m.aktif = False
                    m.donusum_tarihi = bugun_donusum
                    # Yeni Iade KDV kaydi (aynı stok, aynı tutar, donusum_id=eski.id)
                    yeni_id = _yeni_id('MYT')
                    iade_kayit = Maliyet(
                        id=yeni_id,
                        maliyet_tarihi=bugun_donusum,
                        maliyet_tip='Iade KDV',
                        baglanti_tip=m.baglanti_tip,
                        baglanti_id=m.baglanti_id,
                        tutar=m.tutar,
                        doviz=m.doviz,
                        kur=m.kur,
                        try_karsilik=m.try_karsilik,
                        usd_karsilik=m.usd_karsilik,
                        eur_karsilik=m.eur_karsilik,
                        fatura_no=m.fatura_no,
                        aciklama=f'KDV iade donusumu — {_baglanti_okunabilir(m.baglanti_tip, m.baglanti_id)}'
                                 + (f' ({m.fatura_no})' if m.fatura_no else ''),
                        kullanici=session.get('kullanici'),
                        toplam_miktar=m.toplam_miktar,
                        birim_maliyet=m.birim_maliyet,
                        aktif=True,
                        donusum_id=m.id  # Es kaydin ID'si
                    )
                    db.session.add(iade_kayit)
                    # Eski kayda yeni donusum_id'yi yazalim (cift yon iz)
                    m.donusum_id = yeni_id
                if kdv_kalemleri:
                    app.logger.info(
                        f'KDV iade donusumu: {len(kdv_kalemleri)} Devreden -> Iade '
                        f'(eski pasif, yeni Iade aktif | siparis={siparis_id}, tip={sip.satis_tipi})')

        return olusturulan

    def _siparis_teslimi_iptal_et(siparis_id):
        """
        Sipariş Teslim Edildi'den geri alındığında SatisKaydi'ları sil,
        stokları tekrar Satildi yap. (Geri alma için kullanılır.)
        Ayrica KDV donusumu tersine cevirir: aktif Iade KDV'leri pasif yapar,
        eski Devreden'leri yeniden aktif yapar.
        """
        skler = SatisKaydi.query.filter_by(siparis_id=siparis_id).all()
        silinen = 0
        stok_idler = []
        for sk in skler:
            stok = _stok_getir(sk.stok_id, sk.stok_tip)
            if stok and stok.durum == 'Teslim Edildi':
                stok.durum = 'Satildi'
            stok_idler.append(sk.stok_id)
            db.session.delete(sk)
            silinen += 1

        # KDV donusumunu geri al: aktif Iade'leri pasif, eski Devreden'leri aktif yap
        if stok_idler:
            aktif_iadeler = Maliyet.query.filter(
                func.lower(Maliyet.baglanti_tip) == 'stok',
                Maliyet.baglanti_id.in_(stok_idler),
                Maliyet.maliyet_tip == 'Iade KDV',
                Maliyet.aktif == True
            ).all()
            geri_alinan = 0
            for iade in aktif_iadeler:
                # Iade'yi pasif yap
                iade.aktif = False
                iade.donusum_tarihi = date.today()
                # Es Devreden kaydini bul (donusum_id alani)
                if iade.donusum_id:
                    eski_devreden = Maliyet.query.get(iade.donusum_id)
                    if eski_devreden:
                        eski_devreden.aktif = True
                        eski_devreden.donusum_tarihi = None
                        geri_alinan += 1
            if aktif_iadeler:
                app.logger.info(
                    f'KDV donusumu geri alindi: {len(aktif_iadeler)} Iade pasif, '
                    f'{geri_alinan} Devreden tekrar aktif (siparis={siparis_id})')

        return silinen

    def _teslimde_varis_takibi(teslim_sekli):
        """
        Incoterm'e göre teslim akışını belirler.
        Grup A (yüklemede teslim): EXW, FCA, FOB, CFR, CIF, CPT, CIP
          → Satıcının sorumluluğu yüklemede biter. Sevkiyat 'Sevk Edildi' olunca
            her şey biter (stok Teslim Edildi + satış kaydı). Ara durum YOK.
        Grup B (varışta teslim): DAP, DDP, DPU
          → Satıcı malı varış noktasına kadar teslim etmiş sayılır. İki aşamalı:
            Yolda (Sevkedildi) → Teslim Edildi. Varış takibi anlamlı.
        Incoterm boş/tanımsız (eski kayıt): güvenli tarafta B (iki aşamalı) sayılır.
        Döner: True = varış takibi VAR (Grup B), False = yok (Grup A).
        """
        ts = (teslim_sekli or '').strip().upper()
        # İlk kelimeyi al (örn "FOB Izmir" → "FOB", "CIF - Long Beach" → "CIF")
        kod = ts.replace('-', ' ').split()[0] if ts else ''
        GRUP_A = {'EXW', 'FCA', 'FOB', 'CFR', 'CIF', 'CPT', 'CIP', 'FAS'}
        GRUP_B = {'DAP', 'DDP', 'DPU', 'DAT'}
        if kod in GRUP_A:
            return False   # yüklemede teslim
        if kod in GRUP_B:
            return True    # varışta teslim
        return True        # bilinmeyen/boş → güvenli: iki aşamalı

    def _sevkiyat_stoklarini_guncelle(sevkiyat_id, yeni_stok_durum):
        """
        Sevkiyata bağlı siparişin stoklarını yeni duruma getirir.
        Sevkiyat 'Yolda' -> stoklar Sevkedildi
        Sevkiyat geri alınırsa -> stoklar Satildi
        Sadece Satildi/Sevkedildi durumundaki stoklara dokunur.
        """
        sev = Sevkiyat.query.get(sevkiyat_id)
        if not sev or not sev.siparis_id:
            return 0
        rezler = Rezervasyon.query.filter_by(
            siparis_id=sev.siparis_id, iptal_nedeni=None
        ).all()
        guncellenen = 0
        for r in rezler:
            stok = _stok_getir(r.stok_id, r.stok_tip)
            if stok and stok.durum in ('Satildi', 'Sevkedildi') and stok.durum != yeni_stok_durum:
                stok.durum = yeni_stok_durum
                guncellenen += 1
        return guncellenen

    def _parse_date(s):
        if not s:
            return None
        try:
            return datetime.strptime(s, '%Y-%m-%d').date()
        except:
            return None

    def _yeni_id(prefix):
        return f"{prefix}-{uuid.uuid4().hex[:6].upper()}"

    def _log_audit(islem, tablo, kayit_id, eski=None, yeni=None, aciklama=None):
        """
        Denetim kaydı ekler. db.session'a EKLER ama commit ETMEZ —
        çağıran fonksiyon kendi commit'inde birlikte yazar (atomik).
        islem: 'EKLE' / 'GUNCELLE' / 'SIL' / 'DURUM' / 'GIRIS'
        """
        try:
            log = AuditLog(
                kullanici=session.get('kullanici', 'sistem'),
                islem_tipi=islem,
                tablo_adi=tablo,
                kayit_id=str(kayit_id) if kayit_id else None,
                eski_veri=json.dumps(eski, ensure_ascii=False, default=str) if eski else None,
                yeni_veri=json.dumps(yeni, ensure_ascii=False, default=str) if yeni else (
                    json.dumps({'aciklama': aciklama}, ensure_ascii=False) if aciklama else None
                ),
                ip_adresi=request.remote_addr if request else None
            )
            db.session.add(log)
        except Exception as e:
            logging.error(f"Audit log hatası: {e}")

    def _cari_bul(unvan):
        """
        Unvandan cari kaydini bulur. Once birebir, olmazsa
        bosluk/buyuk-kucuk harf duyarsiz eslesme dener.
        """
        if not unvan:
            return None
        # 1) Birebir
        cari = Cari.query.filter_by(unvan=unvan).first()
        if cari:
            return cari
        # 2) Bosluk + harf duyarsiz (Turkce karakterler dahil)
        def _norm(s):
            return ' '.join((s or '').split()).strip().lower() \
                .replace('i̇', 'i').replace('İ', 'i')
        hedef = _norm(unvan)
        for c2 in Cari.query.all():
            if _norm(c2.unvan) == hedef:
                return c2
        return None

    def _cari_risk_durumu(cari, ek_tutar_doviz=0, ek_doviz=None):
        """
        Bir carinin risk durumunu carinin KENDİ PARA BİRİMİNDE hesaplar.
        - risk_limiti: cari kartında tanımlı üst sınır (carinin para_birimi cinsinden)
        - acik_risk: müşterinin ödenmemiş net borcu (cari para birimine çevrilmiş)
        - kullanilabilir: risk_limiti - acik_risk
        - ek_tutar_doviz / ek_doviz: onaylanmak istenen yeni proformanın tutarı ve dövizi
        Döner dict: limit tanımlı değilse limit_var=False (uyarı yapılmaz).
        """
        if not cari:
            return {'limit_var': False}
        limit = cari.risk_limiti
        if not limit or limit <= 0:
            return {'limit_var': False, 'unvan': cari.unvan}

        risk_doviz = (cari.para_birimi or 'USD').upper()
        # risk dövizinin bugünkü kuru (TRY karşılığı) — çevrim köprüsü
        risk_kur = 1.0 if risk_doviz == 'TRY' else (_kur_getir(risk_doviz) or 0)
        if not risk_kur:
            # Kur yoksa TRY bazına düş (güvenli), ama işaretle
            risk_kur = 1.0
            risk_doviz_gecerli = False
        else:
            risk_doviz_gecerli = True

        # Açık risk: cariye ait tüm hareketlerin TRY karşılığı (borc - alacak), sonra risk dövizine çevir
        hareketler = CariHareket.query.filter_by(cari_id=cari.id).all()
        acik_try = 0.0
        for h in hareketler:
            borc_try = h.borc_try if h.borc_try else (
                (h.borc or 0) * (h.kur_uygulanan or _kur_getir(h.doviz, h.hareket_tarihi) or 1))
            alacak_try = h.alacak_try if h.alacak_try else (
                (h.alacak or 0) * (h.kur_uygulanan or _kur_getir(h.doviz, h.hareket_tarihi) or 1))
            acik_try += (borc_try or 0) - (alacak_try or 0)
        # TRY → risk dövizi
        acik = q3(acik_try / risk_kur) if risk_kur else q3(acik_try)

        # Yeni proforma tutarını risk dövizine çevir
        ek_risk_doviz = 0.0
        if ek_tutar_doviz and ek_doviz:
            ed = (ek_doviz or 'USD').upper()
            if ed == risk_doviz:
                ek_risk_doviz = ek_tutar_doviz
            else:
                ek_kur = 1.0 if ed == 'TRY' else (_kur_getir(ed) or 0)
                # ek dövizi → TRY → risk dövizi
                ek_try = ek_tutar_doviz * ek_kur if ek_kur else 0
                ek_risk_doviz = (ek_try / risk_kur) if risk_kur else ek_try

        kullanilabilir = q3(limit - acik)
        sonrasi = q3(kullanilabilir - (ek_risk_doviz or 0))
        return {
            'limit_var': True, 'unvan': cari.unvan, 'risk_doviz': risk_doviz,
            'risk_doviz_gecerli': risk_doviz_gecerli,
            'risk_limiti': q3(limit), 'acik_risk': acik,
            'kullanilabilir': kullanilabilir,
            'ek_tutar': q3(ek_risk_doviz or 0),
            'kalan_sonrasi': sonrasi,
            'asiliyor': sonrasi < 0,
            'asim_tutari': q3(-sonrasi) if sonrasi < 0 else 0
        }

    def _cari_hareket_ekle(cari_unvan, islem_tip, borc=0, alacak=0, doviz='USD',
                            aciklama=None, kaynak='manuel', baglanti_tip=None,
                            baglanti_id=None, vade_tarihi=None, evrak_no=None,
                            kur=None):
        """
        Cari hareket kaydı oluşturur (commit ETMEZ - çağıran commit eder).
        Müşteri unvanından cari_id bulur. Otomatik borç/alacak için kullanılır.
        Cari bulunamazsa ValueError firlatir (sessizce kaybetmemek icin).
        kur: verilirse (manuel/çapraz döviz senaryosu) TCMB yerine bu kur kullanılır.
        """
        cari = _cari_bul(cari_unvan)
        if not cari:
            raise ValueError(
                f'"{cari_unvan}" cari hesabi sistemde bulunamadi. '
                f'Once Cari modulunden bu musteriyi/tedarikciyi ekleyin.')
        # ÇOK DÖVİZLİ: kur ve TRY karşılığı
        kullanilan_kur = kur if (kur and kur > 0) else _kur_getir(doviz or 'USD')
        borc_try, _ = _try_karsilik(borc or 0, doviz or 'USD', kullanilan_kur)
        alacak_try, _ = _try_karsilik(alacak or 0, doviz or 'USD', kullanilan_kur)

        hareket = CariHareket(
            id=_yeni_id('HR'),
            hareket_tarihi=date.today(),
            cari_id=cari.id,
            cari_unvan=cari.unvan,
            islem_tip=islem_tip,
            evrak_no=evrak_no,
            aciklama=aciklama,
            borc=borc or 0,
            alacak=alacak or 0,
            doviz=doviz or 'USD',
            kur_uygulanan=q_kur(kullanilan_kur),
            kur_kaynak='TCMB',
            borc_try=q3(borc_try),
            alacak_try=q3(alacak_try),
            vade_tarihi=vade_tarihi,
            kaynak=kaynak,
            baglanti_tip=baglanti_tip,
            baglanti_id=baglanti_id,
            siparis_id=baglanti_id if baglanti_tip == 'siparis' else None,
            kullanici=session.get('kullanici', 'sistem')
        )
        db.session.add(hareket)
        return hareket

    def _kesilmis_blok_maliyet_yeniden_dagit(blok_id):
        """
        Bir bloğa SONRADAN maliyet eklendiğinde, o bloktan kesilmiş plakaların
        birim maliyetini yeniden hesaplar.
        Yeni birim maliyet = (bloğun matrahı + bloğa bağlı tüm ek maliyetler) / üretilen toplam m².
        Döner: güncellenen plaka sayısı.
        """
        blok = BlokStok.query.get(blok_id)
        if not blok:
            return 0
        kesimler = Kesim.query.filter_by(kaynak_id=blok_id).all()
        if not kesimler:
            return 0  # blok kesilmemiş

        matrah = blok.matrah or ((blok.alis_fiyati or 0) * (blok.tonaj or 0)
                                 if blok.alis_fiyat_birim == 'ton' else
                                 (blok.alis_fiyati or 0) * (blok.hacim_m3 or 0))
        ek_top = db.session.query(db.func.sum(Maliyet.usd_karsilik)).filter(
            Maliyet.baglanti_id == blok_id,
            ~Maliyet.maliyet_tip.in_(['Devreden KDV', 'Iade KDV'])
        ).scalar() or 0
        yeni_toplam = (matrah or 0) + ek_top

        guncellenen = 0
        for kesim in kesimler:
            detaylar = KesimDetay.query.filter_by(kesim_id=kesim.id, hedef_tip='PLAKA').all()
            plakalar, toplam_m2 = [], 0
            for d in detaylar:
                p = PlakaStok.query.get(d.hedef_stok_id) if d.hedef_stok_id else None
                if p:
                    plakalar.append(p)
                    toplam_m2 += (p.metraj_m2 or 0)
            if toplam_m2 <= 0:
                continue
            yeni_birim = q3(yeni_toplam / toplam_m2)
            for p in plakalar:
                p.alis_fiyati = yeni_birim
                guncellenen += 1
        return guncellenen

    def _satis_kaydi_maliyet_guncelle(stok_idler):
        """
        Verilen stok ID'lerine ait SatisKaydi'larin maliyet/kar/marj
        degerlerini Maliyet tablosundan yeniden hesaplar.
        Satilmis plakaya gecmise donuk maliyet eklendiginde cagrilir.
        Donus: guncellenen kayit sayisi
        """
        if not stok_idler:
            return 0
        guncellenen = 0
        for sk in SatisKaydi.query.filter(SatisKaydi.stok_id.in_(list(stok_idler))).all():
            # Bu stoga ait toplam maliyet (Maliyet tablosu)
            stok_maliyet = db.session.query(db.func.sum(Maliyet.usd_karsilik)).filter(
                Maliyet.baglanti_id == sk.stok_id,
                ~Maliyet.maliyet_tip.in_(['Devreden KDV', 'Iade KDV'])
            ).scalar() or 0
            # Stogun kendi alis maliyeti (stok kaydindan)
            alim_usd = 0
            stok = _stok_getir(sk.stok_id, sk.stok_tip)
            if stok:
                alis_fiyat = getattr(stok, 'alis_fiyati', 0) or 0
                alis_birim = getattr(stok, 'alis_fiyat_birim', None) or (
                    'ton' if sk.stok_tip == 'BLOK' else 'm2')
                try:
                    alis_olcu = _stok_olcu(stok, alis_birim)
                except Exception:
                    alis_olcu = 0
                stok_doviz = getattr(stok, 'doviz', 'USD') or 'USD'
                # USD'ye cevir
                if stok_doviz == 'USD':
                    alim_usd = alis_fiyat * alis_olcu
                else:
                    usd_k = DovizKur.query.filter_by(doviz='USD').order_by(
                        DovizKur.tarih.desc()).first()
                    eur_k = DovizKur.query.filter_by(doviz='EUR').order_by(
                        DovizKur.tarih.desc()).first()
                    ku = (usd_k.alis if usd_k else 0) or 0
                    ke = (eur_k.alis if eur_k else 0) or 0
                    ham = alis_fiyat * alis_olcu
                    if stok_doviz == 'EUR':
                        alim_usd = (ham * ke / ku) if ku else 0
                    elif stok_doviz == 'TRY':
                        alim_usd = (ham / ku) if ku else 0
                    else:
                        alim_usd = ham

            yeni_maliyet = alim_usd + stok_maliyet
            satis = sk.tutar_usd or sk.tutar or 0
            yeni_kar = satis - yeni_maliyet
            yeni_marj = (yeni_kar / satis * 100) if satis else 0

            sk.maliyet_usd = q3(yeni_maliyet)
            sk.maliyet_try = q3(yeni_maliyet * (sk.kur_usd or 0))
            sk.kar_usd = q3(yeni_kar)
            sk.marj_yuzde = q_oran(yeni_marj)
            guncellenen += 1
        return guncellenen

    def _fatura_satis_kaydi_olustur(fatura_id):
        """
        Fatura 'Kesildi' olunca SatisKaydi olusturur (kar hesabi icin).
        - Stoklu fatura (siparise bagli + rezervasyon var): maliyet stoklardan
        - Transit fatura (rezervasyon yok): maliyet faturadaki alis_maliyeti'nden
        Cift kayit korumasi: bu fatura veya bagli siparisi icin zaten
        SatisKaydi varsa atlar.
        Donus: (olusan_adet, hata_mesaji)
        """
        f = Fatura.query.get(fatura_id)
        if not f:
            return 0, 'Fatura bulunamadi'

        # ── ÇİFT KAYIT KORUMASI ──
        # Bu fatura icin zaten SatisKaydi olustu mu?
        mevcut = SatisKaydi.query.filter_by(fatura_id=fatura_id).first()
        if mevcut:
            return 0, None  # zaten var, sessizce atla
        # Bagli siparis teslim edilip SatisKaydi olusmus mu?
        if f.siparis_id:
            sip_sk = SatisKaydi.query.filter_by(siparis_id=f.siparis_id).first()
            if sip_sk:
                return 0, None  # siparis teslimi zaten kayit olusturmus

        # Kurlar
        usd_k = DovizKur.query.filter_by(doviz='USD').order_by(DovizKur.tarih.desc()).first()
        eur_k = DovizKur.query.filter_by(doviz='EUR').order_by(DovizKur.tarih.desc()).first()
        kur_usd = (usd_k.alis if usd_k else 0) or 0
        kur_eur = (eur_k.alis if eur_k else 0) or 0

        def _usd(tutar, dvz):
            if not tutar: return 0
            if dvz == 'USD' or not dvz: return tutar
            if dvz == 'EUR': return (tutar * kur_eur / kur_usd) if kur_usd else 0
            if dvz == 'TRY': return (tutar / kur_usd) if kur_usd else 0
            return tutar

        # Fatura kalemleri
        try:
            kalemler = json.loads(f.kalemler_json) if f.kalemler_json else []
        except Exception:
            kalemler = []

        # Stoklu/transit kararini faturanin KAYITLI tipinden ver.
        # Rezervasyonlar teslim/sevk sonrasi kapanmis olabilir -> iptal filtresi koyma.
        stoklu = (getattr(f, 'fatura_tipi', 'stoklu') != 'transit')
        rezler = []
        if stoklu and f.siparis_id:
            rezler = Rezervasyon.query.filter_by(siparis_id=f.siparis_id).all()

        # Ek maliyetler (Maliyet tablosu - siparise bagli)
        ek_maliyet_usd = 0
        if f.siparis_id:
            ek_maliyet_usd = db.session.query(db.func.sum(Maliyet.usd_karsilik)).filter(
                Maliyet.baglanti_id == f.siparis_id, ~Maliyet.maliyet_tip.in_(['Devreden KDV', 'Iade KDV'])).scalar() or 0

        # Karlilik matrah uzerinden: ara_toplam (KDV haric). Eger 0 ise f.toplam fallback.
        satis_matrah = f.ara_toplam or f.toplam or 0
        satis_toplam_usd = _usd(satis_matrah, f.doviz or 'USD')

        if rezler:
            # ── STOKLU FATURA: her stok icin ayri SatisKaydi ──
            olusan = 0
            n = len(rezler)
            for rz in rezler:
                stok = _stok_getir(rz.stok_id, rz.stok_tip)
                if not stok:
                    continue
                # Stok alim maliyeti (alis birimi x olcu)
                alis_fiyat = getattr(stok, 'alis_fiyati', 0) or 0
                alis_birim = getattr(stok, 'alis_fiyat_birim', None) or (
                    'ton' if rz.stok_tip == 'BLOK' else 'm2')
                alis_olcu = _stok_olcu(stok, alis_birim)
                stok_doviz = getattr(stok, 'doviz', 'USD') or 'USD'
                alim_usd = _usd(alis_fiyat * alis_olcu, stok_doviz)
                # Ek maliyet payi (esit bolustur)
                pay_ek = ek_maliyet_usd / n if n else 0
                maliyet_usd = alim_usd + pay_ek
                # Bu stogun satis payi (faturayi stok sayisina bol)
                satis_pay_usd = satis_toplam_usd / n if n else 0
                kar_usd = satis_pay_usd - maliyet_usd
                marj = (kar_usd / satis_pay_usd * 100) if satis_pay_usd else 0

                # MIKTAR ve BIRIM - sip.birim ve _stok_olcu
                # FAZ 16: Siparis.birim kaldirildi; kalemden al (yoksa stok tipine gore varsayilan)
                _sb = None
                try:
                    _s_id = getattr(rz, 'siparis_id', None)
                    if _s_id:
                        _k = (SiparisKalem.query.filter_by(siparis_id=_s_id)
                              .order_by(SiparisKalem.sira).first())
                        _sb = _k.birim if _k else None
                except Exception:
                    _sb = None
                sat_birim = _sb or ('ton' if rz.stok_tip == 'BLOK' else 'm2')
                # sip referansı yoksa fatura'dan al
                if not sat_birim:
                    sat_birim = 'm2'
                stok_miktar = _stok_olcu(stok, sat_birim)

                # Orijinal döviz hesabı (USD/TRY/EUR)
                fdoviz = (f.doviz or 'USD').upper()
                # Pay tutarı orijinal döviz cinsinden
                fatura_toplam_orj = f.ara_toplam or f.toplam or 0  # KDV haric (matrah)
                satis_pay_orj = (fatura_toplam_orj / n) if n else 0
                # Maliyet orijinal dövize çevir (USD -> hedef)
                if fdoviz == 'USD':
                    maliyet_orj = maliyet_usd
                elif fdoviz == 'TRY':
                    maliyet_orj = maliyet_usd * (kur_usd or 1)
                elif fdoviz == 'EUR':
                    maliyet_orj = maliyet_usd * (kur_usd or 1) / (kur_eur or 1) if (kur_eur and kur_usd) else maliyet_usd
                else:
                    maliyet_orj = maliyet_usd
                kar_orj = satis_pay_orj - maliyet_orj

                sk = SatisKaydi(
                    id=_yeni_id('SAT'),
                    stok_id=stok.id,
                    stok_tip=rz.stok_tip,
                    cins=getattr(stok, 'cins', None),
                    ozellik=getattr(stok, 'ozellik', None),
                    blok_no=getattr(stok, 'blok_no', None) or getattr(stok, 'kasa_no', None),
                    boy=getattr(stok, 'boy', None),
                    yukseklik=getattr(stok, 'yukseklik', None),
                    kalinlik=getattr(stok, 'kalinlik', None),
                    en=getattr(stok, 'en', None),
                    metraj_m2=getattr(stok, 'metraj_m2', None),
                    metraj_sqft=getattr(stok, 'metraj_sqft', None),
                    hacim_m3=getattr(stok, 'hacim_m3', None),
                    tonaj=getattr(stok, 'tonaj', None),
                    siparis_id=f.siparis_id,
                    proforma_id=f.proforma_id,
                    fatura_id=fatura_id,
                    kaynak='fatura',
                    musteri=f.musteri,
                    musteri_ulke=f.musteri_ulke,
                    satis_tarihi=f.fatura_tarihi or date.today(),
                    miktar=q3(stok_miktar),
                    birim=sat_birim,
                    birim_fiyat=q3(satis_pay_orj / stok_miktar) if stok_miktar else 0,
                    doviz=fdoviz,
                    tutar=q3(satis_pay_orj),
                    kur_usd=q_kur(kur_usd), kur_eur=q_kur(kur_eur),
                    tutar_usd=q3(satis_pay_usd),
                    tutar_try=q3(satis_pay_usd * kur_usd if kur_usd else 0),
                    maliyet_usd=q3(maliyet_usd),
                    maliyet_try=q3(maliyet_usd * kur_usd if kur_usd else 0),
                    kar_usd=q3(kar_usd),
                    marj_yuzde=q_oran(marj),
                    fatura_no=f.fatura_no,
                    fatura_tarihi=f.fatura_tarihi,
                    kullanici=session.get('kullanici')
                )
                db.session.add(sk)
                stok.durum = 'Teslim Edildi'
                olusan += 1
            return olusan, None
        else:
            # ── TRANSIT FATURA: tek SatisKaydi, maliyet faturadan ──
            alis_mal_usd = _usd(f.alis_maliyeti or 0, f.maliyet_doviz or 'USD')
            maliyet_usd = alis_mal_usd + ek_maliyet_usd
            kar_usd = satis_toplam_usd - maliyet_usd
            marj = (kar_usd / satis_toplam_usd * 100) if satis_toplam_usd else 0

            # Kalemlerden ozet bilgi
            ilk_cins = kalemler[0].get('cins') if kalemler else None
            top_miktar = sum((k.get('miktar') or 0) for k in kalemler)
            ilk_birim = kalemler[0].get('birim') if kalemler else None

            sk = SatisKaydi(
                id=_yeni_id('SAT'),
                stok_id=f'TRANSIT-{fatura_id}',
                stok_tip='TRANSIT',
                cins=ilk_cins,
                siparis_id=f.siparis_id,
                proforma_id=f.proforma_id,
                fatura_id=fatura_id,
                kaynak='fatura',
                musteri=f.musteri,
                musteri_ulke=f.musteri_ulke,
                satis_tarihi=f.fatura_tarihi or date.today(),
                miktar=top_miktar,
                birim=ilk_birim,
                doviz=f.doviz or 'USD',
                tutar=q3(satis_toplam_usd),
                kur_usd=q_kur(kur_usd), kur_eur=q_kur(kur_eur),
                tutar_usd=q3(satis_toplam_usd),
                tutar_try=q3(satis_toplam_usd * kur_usd if kur_usd else 0),
                maliyet_usd=q3(maliyet_usd),
                maliyet_try=q3(maliyet_usd * kur_usd if kur_usd else 0),
                kar_usd=q3(kar_usd),
                marj_yuzde=q_oran(marj),
                fatura_no=f.fatura_no,
                fatura_tarihi=f.fatura_tarihi,
                kullanici=session.get('kullanici')
            )
            db.session.add(sk)
            return 1, None

        # ─── TRANSACTION HELPER (Veri butunlugu icin) ────────────────────────
    def _safe_commit(islem_adi='işlem'):
        """db.session.commit() yapar, hata olursa rollback yapar.
        return: (success, hata_mesaji_veya_None)"""
        try:
            db.session.commit()
            return True, None
        except Exception as e:
            db.session.rollback()
            app.logger.exception(f'[{islem_adi}] commit hatasi - rollback yapildi')
            return False, str(e)

    def _safe_transaction(func, islem_adi='işlem'):
        """Bir fonksiyonu cagiri, hata olursa otomatik rollback yapar.
        Decorator olarak da kullanilabilir.
        return: (success, sonuc_veya_hata)"""
        try:
            sonuc = func()
            db.session.commit()
            return True, sonuc
        except Exception as e:
            db.session.rollback()
            app.logger.exception(f'[{islem_adi}] islem hatasi - rollback yapildi')
            return False, str(e)
    # ─── /TRANSACTION HELPER ──────────────────────────────────────────────

    def _otomatik_migrasyon():
        """db.create_all() yeni SUTUN eklemez (sadece yeni TABLO). Modele sonradan
        eklenen sutunlari mevcut tablolara ALTER TABLE ile ekler. Idempotent."""
        from sqlalchemy import inspect as _inspect, text as _text
        try:
            mufettis = _inspect(db.engine)
            # Lehçeye göre boolean literali (PostgreSQL: TRUE, SQLite: 1)
            _dialect = db.engine.dialect.name
            _bool_true = 'TRUE' if _dialect in ('postgresql', 'postgres') else '1'
            _bool_default = 'BOOLEAN DEFAULT TRUE' if _dialect in ('postgresql', 'postgres') else 'BOOLEAN DEFAULT 1'
            eklenecek = [
                # (tablo, sutun, SQL tipi)
                ('kasa', 'banka_id', 'INTEGER'),
                ('kasa_hareket', 'siparis_id', 'VARCHAR(20)'),
                ('kasa_hareket', 'evrak_no', 'VARCHAR(50)'),
                ('proforma', 'ana_pi_id', 'VARCHAR(20)'),
                ('proforma', 'revizyon_no', 'INTEGER DEFAULT 0'),
                ('proforma', 'aktif_surum', _bool_default),
                ('proforma', 'revizyon_notu', 'TEXT'),
                ('proforma', 'onaya_gonderen', 'VARCHAR(50)'),
                ('proforma', 'onaya_gonderme_tarihi', 'TIMESTAMP'),
                ('proforma', 'onaylayan', 'VARCHAR(50)'),
                ('proforma', 'onay_tarihi', 'TIMESTAMP'),
                ('proforma', 'onay_reddeden', 'VARCHAR(50)'),
                ('proforma', 'onay_red_notu', 'TEXT'),
                ('kesim', 'kaynak_onceki_durum', 'TEXT'),
                ('maliyet', 'olusturma', 'TIMESTAMP'),
                ('cari_hareket', 'kdv_dahil_mi', 'BOOLEAN'),
                ('cari_hareket', 'kdv_oran', 'FLOAT'),
                ('cari_hareket', 'kdv_tutar', 'FLOAT'),
                ('cari_hareket', 'matrah', 'FLOAT'),
                ('veriler', 'uzun_deger', 'TEXT'),
            ]
            for tablo, sutun, tip in eklenecek:
                if tablo not in mufettis.get_table_names():
                    continue
                mevcut = {s['name'] for s in mufettis.get_columns(tablo)}
                if sutun in mevcut:
                    continue
                with db.engine.begin() as baglanti:
                    baglanti.execute(_text(f'ALTER TABLE {tablo} ADD COLUMN {sutun} {tip}'))
                print(f'Migrasyon: {tablo}.{sutun} sutunu eklendi.')

            # Geriye donuk: revizyon alanlari bos olan proformalari kok (Rev.0) yap
            if 'proforma' in mufettis.get_table_names():
                sutunlar = {s['name'] for s in mufettis.get_columns('proforma')}
                if {'ana_pi_id', 'aktif_surum', 'revizyon_no'} <= sutunlar:
                    with db.engine.begin() as baglanti:
                        baglanti.execute(_text(
                            "UPDATE proforma SET ana_pi_id = id "
                            "WHERE ana_pi_id IS NULL OR ana_pi_id = ''"))
                        baglanti.execute(_text(
                            "UPDATE proforma SET revizyon_no = 0 WHERE revizyon_no IS NULL"))
                        baglanti.execute(_text(
                            f"UPDATE proforma SET aktif_surum = {_bool_true} WHERE aktif_surum IS NULL"))
        except Exception as hata:
            print(f'Migrasyon uyarisi: {hata}')

    def _seed_data():
        if Kullanici.query.count() == 0:
            admin = Kullanici(ad='admin', sifre=generate_password_hash('admin123'), rol='ADMIN')
            db.session.add(admin)
        # Varsayilan lookup listeleri — hepsi BUYUK HARF (belge/etiket tutarliligi icin).
        # NOT: 'uretici' kategorisi ARTIK SEED EDILMEZ; ureticiler Cari modulunden
        # (cari_tip='Uretici' / urun_tedarikcisi) yonetilir.
        _cinsler = [
            'IVORY TRAVERTINE', 'STRATA TRAVERTINE', 'SILVER ROOT', 'FRENCH VANILIA',
            'BEIGE MARBLE', 'ROSSO LEVANTO', 'ROSSO LAGUNA', 'MYRA BEIGE',
            'THUNDRA GREY', 'GALAXY GREY', 'PIETRA GREY', 'SILVER TRAVERTINE',
            'TITANIUM TRAVERTINE', 'FILDIŞI', 'EMPERADOR', 'NERO', 'TRAVERTEN',
        ]
        # Ulkeler: (TURKCE_AD, ISO3_KOD). Cari'de KOD saklanir (USA, TUR...);
        # formda tam ad secilir, listelerde/belgelerde kod gorunur.
        _ulkeler = [
            ('AFGANİSTAN','AFG'),('ARNAVUTLUK','ALB'),('CEZAYİR','DZA'),('ANDORRA','AND'),
            ('ANGOLA','AGO'),('ANTİGUA VE BARBUDA','ATG'),('ARJANTİN','ARG'),('ERMENİSTAN','ARM'),
            ('AVUSTRALYA','AUS'),('AVUSTURYA','AUT'),('AZERBAYCAN','AZE'),('BAHAMALAR','BHS'),
            ('BAHREYN','BHR'),('BANGLADEŞ','BGD'),('BARBADOS','BRB'),('BELARUS','BLR'),
            ('BELÇİKA','BEL'),('BELİZE','BLZ'),('BENİN','BEN'),('BHUTAN','BTN'),('BOLİVYA','BOL'),
            ('BOSNA-HERSEK','BIH'),('BOTSVANA','BWA'),('BREZİLYA','BRA'),('BRUNEİ','BRN'),
            ('BULGARİSTAN','BGR'),('BURKİNA FASO','BFA'),('BURUNDİ','BDI'),
            ('YEŞİL BURUN ADALARI','CPV'),('KAMBOÇYA','KHM'),('KAMERUN','CMR'),('KANADA','CAN'),
            ('ORTA AFRİKA CUMHURİYETİ','CAF'),('ÇAD','TCD'),('ŞİLİ','CHL'),('ÇİN','CHN'),
            ('KOLOMBİYA','COL'),('KOMORLAR','COM'),('KONGO','COG'),('KOSTA RİKA','CRI'),
            ('FİLDİŞİ SAHİLİ','CIV'),('HIRVATİSTAN','HRV'),('KÜBA','CUB'),('KIBRIS','CYP'),
            ('ÇEKYA','CZE'),('KUZEY KORE','PRK'),('KONGO DEMOKRATİK CUMHURİYETİ','COD'),
            ('DANİMARKA','DNK'),('CİBUTİ','DJI'),('DOMİNİKA','DMA'),('DOMİNİK CUMHURİYETİ','DOM'),
            ('EKVADOR','ECU'),('MISIR','EGY'),('EL SALVADOR','SLV'),('EKVATOR GİNESİ','GNQ'),
            ('ERİTRE','ERI'),('ESTONYA','EST'),('ESVATİNİ','SWZ'),('ETİYOPYA','ETH'),('FİJİ','FJI'),
            ('FİNLANDİYA','FIN'),('FRANSA','FRA'),('GABON','GAB'),('GAMBİYA','GMB'),
            ('GÜRCİSTAN','GEO'),('ALMANYA','DEU'),('GANA','GHA'),('YUNANİSTAN','GRC'),
            ('GRENADA','GRD'),('GUATEMALA','GTM'),('GİNE','GIN'),('GİNE-BİSSAU','GNB'),
            ('GUYANA','GUY'),('HAİTİ','HTI'),('HONDURAS','HND'),('MACARİSTAN','HUN'),
            ('İZLANDA','ISL'),('HİNDİSTAN','IND'),('ENDONEZYA','IDN'),('İRAN','IRN'),('IRAK','IRQ'),
            ('İRLANDA','IRL'),('İSRAİL','ISR'),('İTALYA','ITA'),('JAMAİKA','JAM'),('JAPONYA','JPN'),
            ('ÜRDÜN','JOR'),('KAZAKİSTAN','KAZ'),('KENYA','KEN'),('KİRİBATİ','KIR'),
            ('KUVEYT','KWT'),('KIRGIZİSTAN','KGZ'),('LAOS','LAO'),('LETONYA','LVA'),
            ('LÜBNAN','LBN'),('LESOTHO','LSO'),('LİBERYA','LBR'),('LİBYA','LBY'),
            ('LİHTENŞTAYN','LIE'),('LİTVANYA','LTU'),('LÜKSEMBURG','LUX'),('MADAGASKAR','MDG'),
            ('MALAVİ','MWI'),('MALEZYA','MYS'),('MALDİVLER','MDV'),('MALİ','MLI'),('MALTA','MLT'),
            ('MARSHALL ADALARI','MHL'),('MORİTANYA','MRT'),('MAURİTİUS','MUS'),('MEKSİKA','MEX'),
            ('MİKRONEZYA','FSM'),('MONAKO','MCO'),('MOĞOLİSTAN','MNG'),('KARADAĞ','MNE'),
            ('FAS','MAR'),('MOZAMBİK','MOZ'),('MYANMAR','MMR'),('NAMİBYA','NAM'),('NAURU','NRU'),
            ('NEPAL','NPL'),('HOLLANDA','NLD'),('YENİ ZELANDA','NZL'),('NİKARAGUA','NIC'),
            ('NİJER','NER'),('NİJERYA','NGA'),('KUZEY MAKEDONYA','MKD'),('NORVEÇ','NOR'),
            ('UMMAN','OMN'),('PAKİSTAN','PAK'),('PALAU','PLW'),('PANAMA','PAN'),
            ('PAPUA YENİ GİNE','PNG'),('PARAGUAY','PRY'),('PERU','PER'),('FİLİPİNLER','PHL'),
            ('POLONYA','POL'),('PORTEKİZ','PRT'),('KATAR','QAT'),('GÜNEY KORE','KOR'),
            ('MOLDOVA','MDA'),('ROMANYA','ROU'),('RUSYA','RUS'),('RUANDA','RWA'),
            ('SAİNT KİTTS VE NEVİS','KNA'),('SAİNT LUCİA','LCA'),
            ('SAİNT VİNCENT VE GRENADİNLER','VCT'),('SAMOA','WSM'),('SAN MARİNO','SMR'),
            ('SAO TOME VE PRİNCİPE','STP'),('SUUDİ ARABİSTAN','SAU'),('SENEGAL','SEN'),
            ('SIRBİSTAN','SRB'),('SEYŞELLER','SYC'),('SİERRA LEONE','SLE'),('SİNGAPUR','SGP'),
            ('SLOVAKYA','SVK'),('SLOVENYA','SVN'),('SOLOMON ADALARI','SLB'),('SOMALİ','SOM'),
            ('GÜNEY AFRİKA','ZAF'),('GÜNEY SUDAN','SSD'),('İSPANYA','ESP'),('SRİ LANKA','LKA'),
            ('SUDAN','SDN'),('SURİNAM','SUR'),('İSVEÇ','SWE'),('İSVİÇRE','CHE'),('SURİYE','SYR'),
            ('TACİKİSTAN','TJK'),('TAYLAND','THA'),('DOĞU TİMOR','TLS'),('TOGO','TGO'),
            ('TONGA','TON'),('TRİNİDAD VE TOBAGO','TTO'),('TUNUS','TUN'),('TÜRKİYE','TUR'),
            ('TÜRKMENİSTAN','TKM'),('TUVALU','TUV'),('UGANDA','UGA'),('UKRAYNA','UKR'),
            ('BİRLEŞİK ARAP EMİRLİKLERİ','ARE'),('BİRLEŞİK KRALLIK','GBR'),('TANZANYA','TZA'),
            ('AMERİKA BİRLEŞİK DEVLETLERİ','USA'),('URUGUAY','URY'),('ÖZBEKİSTAN','UZB'),
            ('VANUATU','VUT'),('VENEZUELA','VEN'),('VİETNAM','VNM'),('YEMEN','YEM'),
            ('ZAMBİYA','ZMB'),('ZİMBABVE','ZWE'),
        ]
        _yuzeyler = [
            'POLISHED', 'HONED', 'BRUSHED', 'LEADHERED', 'SAND BLASTED',
            'VC / UF / H', 'VC / F / H', 'CC / UF / H', 'CC / F / H',
            'POLISHED / BM', 'HONED / BM', 'LEADHERED / BM', 'BRUSHED / BM',
        ]
        _lookup = ([('siparis_durum','Teklif Asam.'),('siparis_durum','Onaylandi'),('siparis_durum','Uretimde'),
                    ('siparis_durum','Hazir'),('siparis_durum','Teslim Edildi'),('siparis_durum','Iptal Edildi'),
                    ('durum','Serbest'),('durum','Rezerve'),('durum','Satildi'),
                    ('durum','Sevkedildi'),('durum','Teslim Edildi'),('durum','Tukendi'),
                    ('durum','Hasarlı')]
                   + [('cins', c) for c in _cinsler]
                   + [('ozellik', y) for y in _yuzeyler]
                   + [('odeme','Pesin'),('odeme','%30 Avans %70 Yukleme Oncesi'),
                      ('odeme','%50 Avans %50 Yukleme Oncesi'),('odeme','Akreditif (L/C)'),
                      ('odeme','Mal Mukabili'),('odeme','Vesaik Mukabili'),('odeme','Banka Havalesi (T/T)'),
                      ('teslim','EXW'),('teslim','FOB'),('teslim','CIF'),('teslim','CFR'),
                      ('teslim','FCA'),('teslim','DAP'),('teslim','DDP'),('teslim','CPT')])
        # Ulkeler: kisaltma = ISO3 kodu
        for _ad, _kod in _ulkeler:
            if not Veriler.query.filter_by(kategori='ulke', deger=_ad).first():
                db.session.add(Veriler(kategori='ulke', deger=_ad, kisaltma=_kod))

        for kat, deger in _lookup:
            if not Veriler.query.filter_by(kategori=kat, deger=deger).first():
                db.session.add(Veriler(kategori=kat, deger=deger))
        db.session.commit()

    # ---------- DÖVİZ KURU GÜNCELLEME FONKSİYONU ----------
    def guncel_kurlari_cek():
        """TCMB'den güncel kurları çeker ve veritabanına kaydeder"""
        varsayilan_usd = 45.07
        varsayilan_eur = 48.50
        usd_kur = None
        eur_kur = None
        
        try:
            url = "https://www.tcmb.gov.tr/kurlar/today.xml"
            response = requests.get(url, timeout=10)
            
            if response.status_code == 200:
                root = ET.fromstring(response.content)
                
                for currency in root.findall('Currency'):
                    kod = currency.get('CurrencyCode')
                    if kod == 'USD':
                        alis = currency.find('ForexBuying').text
                        satis = currency.find('ForexSelling').text
                        if satis and satis != '':
                            usd_kur = float(satis)
                        elif alis and alis != '':
                            usd_kur = float(alis)
                    elif kod == 'EUR':
                        alis = currency.find('ForexBuying').text
                        satis = currency.find('ForexSelling').text
                        if satis and satis != '':
                            eur_kur = float(satis)
                        elif alis and alis != '':
                            eur_kur = float(alis)
                
                print(f"TCMB'den kur çekildi - USD: {usd_kur}, EUR: {eur_kur}")
        except Exception as e:
            print(f"TCMB'den kur çekilemedi: {e}")
        
        if usd_kur:
            yeni_usd = DovizKur(doviz='USD', alis=usd_kur, satis=usd_kur, efektif=usd_kur, tarih=date.today())
            db.session.add(yeni_usd)
        else:
            son_usd = DovizKur.query.filter_by(doviz='USD').order_by(DovizKur.tarih.desc()).first()
            if son_usd:
                usd_kur = son_usd.efektif
            else:
                usd_kur = varsayilan_usd
                yeni_usd = DovizKur(doviz='USD', alis=usd_kur, satis=usd_kur, efektif=usd_kur, tarih=date.today())
                db.session.add(yeni_usd)
        
        if eur_kur:
            yeni_eur = DovizKur(doviz='EUR', alis=eur_kur, satis=eur_kur, efektif=eur_kur, tarih=date.today())
            db.session.add(yeni_eur)
        else:
            son_eur = DovizKur.query.filter_by(doviz='EUR').order_by(DovizKur.tarih.desc()).first()
            if son_eur:
                eur_kur = son_eur.efektif
            else:
                eur_kur = varsayilan_eur
                yeni_eur = DovizKur(doviz='EUR', alis=eur_kur, satis=eur_kur, efektif=eur_kur, tarih=date.today())
                db.session.add(yeni_eur)
        
        db.session.commit()
        return usd_kur, eur_kur

    def _tcmb_gun_kuru_cek(gun):
        """Belirli bir günün TCMB kurunu çeker. Döner: (usd, eur) veya (None, None).
        Hafta sonu/tatil günleri TCMB kur yayınlamaz → 404 → (None, None)."""
        try:
            url = f"https://www.tcmb.gov.tr/kurlar/{gun.strftime('%Y%m')}/{gun.strftime('%d%m%Y')}.xml"
            r = requests.get(url, timeout=10)
            if r.status_code != 200:
                return None, None
            root = ET.fromstring(r.content)
            usd = eur = None
            for c in root.findall('Currency'):
                kod = c.get('CurrencyCode')
                if kod in ('USD', 'EUR'):
                    satis_el = c.find('ForexSelling')
                    alis_el = c.find('ForexBuying')
                    val = None
                    if satis_el is not None and satis_el.text:
                        val = float(satis_el.text)
                    elif alis_el is not None and alis_el.text:
                        val = float(alis_el.text)
                    if kod == 'USD':
                        usd = val
                    else:
                        eur = val
            return usd, eur
        except Exception:
            return None, None

    def gecmis_kurlari_arsivle(baslangic=None):
        """
        Kur arşivini doldurur. baslangic'tan (varsayılan 1 Ocak 2026) bugüne kadar
        eksik günlerin kurlarını TCMB'den çeker. Zaten kayıtlı günleri atlar.
        Açılışta bir kez çalışır; arşiv doluysa hızlıca geçer (sadece eksikleri çeker).
        Döner: eklenen gün sayısı.
        """
        if baslangic is None:
            baslangic = date(2026, 1, 1)
        bugun = date.today()
        # Hangi günler zaten kayıtlı? (USD bazında — USD/EUR aynı gün çekilir)
        kayitli = set(
            t[0] for t in db.session.query(DovizKur.tarih).filter_by(doviz='USD').all()
        )
        eklenen = 0
        gun = baslangic
        denenecek = []
        while gun <= bugun:
            if gun not in kayitli and gun.weekday() < 5:  # hafta içi + kayıtsız
                denenecek.append(gun)
            gun += timedelta(days=1)
        if not denenecek:
            return 0  # arşiv tam, çekilecek bir şey yok
        app.logger.info(f'Kur arşivi: {len(denenecek)} eksik gün çekiliyor...')
        for g in denenecek:
            usd, eur = _tcmb_gun_kuru_cek(g)
            if usd:
                db.session.add(DovizKur(doviz='USD', alis=usd, satis=usd, efektif=usd, tarih=g, kaynak='TCMB'))
                eklenen += 1
            if eur:
                db.session.add(DovizKur(doviz='EUR', alis=eur, satis=eur, efektif=eur, tarih=g, kaynak='TCMB'))
            if eklenen % 20 == 0 and eklenen > 0:
                db.session.commit()  # ara ara kaydet (uzun liste için)
        db.session.commit()
        app.logger.info(f'Kur arşivi: {eklenen} günlük kur eklendi.')
        return eklenen

    with app.app_context():
        db.create_all()
        _otomatik_migrasyon()
        _seed_data()
        # Geçmiş kur arşivi: 1 Ocak 2026'dan bugüne eksik günleri doldur
        # (ilk açılışta yavaş, sonraki açılışlarda sadece eksik günler — hızlı)
        try:
            print("Kur arşivi kontrol ediliyor (1 Ocak 2026'dan itibaren)...")
            eklenen = gecmis_kurlari_arsivle()
            if eklenen:
                print(f"Kur arşivi: {eklenen} günlük geçmiş kur eklendi.")
            else:
                print("Kur arşivi güncel.")
        except Exception as e:
            print(f"Kur arşivi doldurulamadı (atlandı): {e}")
        print("Güncel döviz kurları çekiliyor...")
        guncel_kurlari_cek()
        print("Kur güncelleme tamamlandı.")

    # ---------- SAYFA ROTALARI ----------
    @app.route('/')
    def index():
        return redirect(url_for('dashboard') if 'kullanici' in session else url_for('giris'))

    @app.route('/giris', methods=['GET', 'POST'])
    def giris():
        if request.method == 'POST':
            ad = request.form.get('kullanici', '').strip()
            sifre = request.form.get('sifre', '')
            k = Kullanici.query.filter_by(ad=ad, aktif=True).first()
            if k and check_password_hash(k.sifre, sifre):
                session['kullanici'] = k.ad
                session['rol'] = k.rol
                session['taze_giris'] = True  # heartbeat gate'i login sonrasi ilk yuklemede atlatir
                _log_audit('GIRIS', 'kullanici', k.id, aciklama=f'{k.ad} giris yapti')
                db.session.commit()
                # Arka planda otomatik yedek al (giriş bekletmesin)
                threading.Thread(target=yedek_modul.yedek_al, args=('auto',), daemon=True).start()
                return redirect(url_for('dashboard'))
            flash('Kullanıcı adı veya şifre hatalı', 'error')
        return render_template('giris.html')

    @app.route('/cikis')
    def cikis():
        session.clear()
        return redirect(url_for('giris'))

    @app.route('/dashboard')
    def dashboard():
        if _auth_required():
            return _auth_required()
        bugun = date.today()
        kur_usd = DovizKur.query.filter_by(doviz='USD').order_by(DovizKur.tarih.desc()).first()
        kur_eur = DovizKur.query.filter_by(doviz='EUR').order_by(DovizKur.tarih.desc()).first()
        k_usd = kur_usd.efektif if kur_usd else 1
        k_eur = kur_eur.efektif if kur_eur else 1

        # COKLU DOVIZ TOPLAMI: TRY uzerinden topla, ana dovize cevir
        # SUM(borc) ile 1000 USD + 500 EUR birbirine eklenirdi - yanlis.
        # borc_try / alacak_try alanlari hareket olusurken TRY karsiligi olarak saklanir.
        ana_pb = _ana_para_birimi() or 'TRY'

        # TRY uzerinden topla (TRY hareketler 1.0 kur ile zaten dogru)
        toplam_alacak_try = db.session.query(func.sum(CariHareket.borc_try)).scalar() or 0
        toplam_borc_try = db.session.query(func.sum(CariHareket.alacak_try)).scalar() or 0

        # borc_try / alacak_try dolmamis eski kayitlar icin fallback: TRY hareketleri direkt borc/alacak'i kullan
        eski_borc = db.session.query(func.sum(CariHareket.borc)).filter(
            CariHareket.doviz == 'TRY',
            (CariHareket.borc_try.is_(None)) | (CariHareket.borc_try == 0),
            CariHareket.borc > 0
        ).scalar() or 0
        eski_alacak = db.session.query(func.sum(CariHareket.alacak)).filter(
            CariHareket.doviz == 'TRY',
            (CariHareket.alacak_try.is_(None)) | (CariHareket.alacak_try == 0),
            CariHareket.alacak > 0
        ).scalar() or 0
        toplam_alacak_try += eski_borc
        toplam_borc_try += eski_alacak

        # Frontend toggle ile çevireceği için TRY karşılığı gönder
        toplam_alacak = toplam_alacak_try
        toplam_borc = toplam_borc_try
        net_bakiye = toplam_alacak_try - toplam_borc_try

        otuz_gun_sonra = bugun + timedelta(days=30)
        odemeler = CariHareket.query.filter(
            CariHareket.vade_tarihi >= bugun,
            CariHareket.vade_tarihi <= otuz_gun_sonra,
            CariHareket.alacak > 0
        ).order_by(CariHareket.vade_tarihi).all()
        yaklasan_odemeler = [{
            'tarih': h.vade_tarihi,
            'cari': Cari.query.get(h.cari_id).unvan if h.cari_id else 'Bilinmeyen',
            'tutar': h.alacak,
            'doviz': h.doviz,
            'aciklama': h.aciklama
        } for h in odemeler]

        tahsilatlar = CariHareket.query.filter(
            CariHareket.vade_tarihi >= bugun,
            CariHareket.vade_tarihi <= otuz_gun_sonra,
            CariHareket.borc > 0
        ).order_by(CariHareket.vade_tarihi).all()
        yaklasan_tahsilatlar = [{
            'tarih': h.vade_tarihi,
            'cari': Cari.query.get(h.cari_id).unvan if h.cari_id else 'Bilinmeyen',
            'tutar': h.borc,
            'doviz': h.doviz,
            'aciklama': h.aciklama
        } for h in tahsilatlar]

        maliyet_dict = {}
        maliyetler = db.session.query(Maliyet.baglanti_id, func.sum(Maliyet.usd_karsilik)).filter(func.lower(Maliyet.baglanti_tip) == 'stok').group_by(Maliyet.baglanti_id).all()
        for m in maliyetler:
            maliyet_dict[m[0]] = m[1] if m[1] is not None else 0

        def stok_deger(rows, tip):
            usd = 0.0
            for r in rows:
                f = r.alis_fiyati or 0
                dov = r.doviz or 'USD'
                if tip == 'BLOK':
                    m3 = r.hacim_m3 or 0
                    if dov == 'USD':
                        usd += f * m3
                    elif dov == 'EUR':
                        usd += f * m3 * (k_eur / k_usd) if k_usd else f * m3
                    else:
                        usd += (f * m3) / k_usd if k_usd else f * m3
                else:
                    m2 = r.metraj_m2 or 0
                    if dov == 'USD':
                        usd += f * m2
                    elif dov == 'EUR':
                        usd += f * m2 * (k_eur / k_usd) if k_usd else f * m2
                    else:
                        usd += (f * m2) / k_usd if k_usd else f * m2
                maliyet_deger = maliyet_dict.get(r.id) or 0
                usd += maliyet_deger
            return q3(usd)

        blok_rows = BlokStok.query.filter_by(durum='Serbest').all()
        plaka_rows = PlakaStok.query.filter_by(durum='Serbest').all()
        ebatli_rows = EbatliStok.query.filter_by(durum='Serbest').all()
        aktif_sip = Siparis.query.filter(Siparis.durum.notin_(['Teslim Edildi','Iptal Edildi'])).all()
        # Aggregate active orders' total via Siparis.toplam_tutar (per-line fields live on SiparisKalem).
        sip_usd = 0
        for _s in aktif_sip:
            _t = _s.toplam_tutar or 0
            _d = _s.doviz or 'USD'
            if _d == 'USD':
                sip_usd += _t
            elif _d == 'EUR':
                sip_usd += _t * (k_eur / k_usd) if k_usd else _t
            else:
                sip_usd += _t / k_usd if k_usd else _t
        yaklasan_siparis_raw = Siparis.query.filter(
            Siparis.termin >= bugun,
            Siparis.termin <= bugun + timedelta(days=14),
            Siparis.durum.notin_(['Teslim Edildi','Iptal Edildi'])
        ).order_by(Siparis.termin).all()
        # Kalem bilgisini ekleyerek dict listesine çevir (dashboard tablosu için)
        yaklasan_siparis = []
        for s in yaklasan_siparis_raw:
            ik = s.kalemler[0] if s.kalemler else None
            yaklasan_siparis.append({
                'id': s.id, 'musteri': s.musteri, 'durum': s.durum,
                'termin': s.termin, 'toplam_tutar': q3(s.toplam_tutar or 0),
                'doviz': s.doviz or 'USD',
                'urun_tip': ik.urun_tip if ik else '—',
                'cins': (ik.cins if ik else None) or '—',
                'miktar': q3(ik.miktar or 0) if (ik and ik.miktar) else '',
                'birim': (ik.birim if ik else '') or '',
            })

        stats = {
            'blok': len(blok_rows), 'blok_usd': stok_deger(blok_rows, 'BLOK'), 'blok_try': q3(stok_deger(blok_rows, 'BLOK')*k_usd),
            'blok_ton': q3(sum(r.tonaj or 0 for r in blok_rows)), 'blok_m3': q3(sum(r.hacim_m3 or 0 for r in blok_rows)),
            'plaka': len(plaka_rows), 'plaka_usd': stok_deger(plaka_rows, 'PLAKA'), 'plaka_try': q3(stok_deger(plaka_rows, 'PLAKA')*k_usd),
            'plaka_m2': q3(sum(r.metraj_m2 or 0 for r in plaka_rows)), 'plaka_sqft': q3(sum(r.metraj_sqft or 0 for r in plaka_rows)),
            'ebatli': len(ebatli_rows), 'ebatli_usd': stok_deger(ebatli_rows, 'EBATLI'), 'ebatli_try': q3(stok_deger(ebatli_rows, 'EBATLI')*k_usd),
            'ebatli_m2': q3(sum(r.metraj_m2 or 0 for r in ebatli_rows)),
            'ebatli_sqft': q3(sum(r.metraj_sqft or 0 for r in ebatli_rows)),
            'siparis': len(aktif_sip), 'siparis_usd': q3(sip_usd), 'siparis_try': q3(sip_usd * k_usd),
            'rezervasyon': Rezervasyon.query.filter_by(iptal_nedeni=None).count(),
            'sevkiyat': Sevkiyat.query.filter(Sevkiyat.durum.notin_(['Teslim Edildi','Iptal'])).count(),
        }

        # ═══ İŞ AKIŞI PANOSU — Bekleyen işler sayıları ═══
        # Faturalanmayı bekleyen proformalar (faturaya dönüşmemiş, iptal olmayan)
        try:
            fatura_bekleyen = Proforma.query.filter(
                Proforma.durum.notin_(['Faturalandi', 'Iptal'])
            ).count()
        except Exception:
            fatura_bekleyen = 0
        # Sevkiyat bekleyen siparişler (aktif durumda)
        try:
            sevk_bekleyen = Siparis.query.filter(
                Siparis.durum.in_(['Onaylandi', 'Uretimde', 'Hazir'])
            ).count()
        except Exception:
            sevk_bekleyen = 0
        # Tahsilat bekleyen satış faturaları (açık)
        try:
            tahsilat_bekleyen = Fatura.query.filter(
                Fatura.yon == 'satis', Fatura.durum.in_(['Kesildi', 'Kismi Tahsil'])
            ).count()
        except Exception:
            tahsilat_bekleyen = 0
        # Ödeme bekleyen alış faturaları (açık)
        try:
            odeme_bekleyen = Fatura.query.filter(
                Fatura.yon == 'alis', Fatura.durum.in_(['Kesildi', 'Kismi Tahsil'])
            ).count()
        except Exception:
            odeme_bekleyen = 0
        stats['akis'] = {
            'fatura_bekleyen': fatura_bekleyen,
            'sevk_bekleyen': sevk_bekleyen,
            'tahsilat_bekleyen': tahsilat_bekleyen,
            'odeme_bekleyen': odeme_bekleyen,
        }

        # ═══ AKTİF SİPARİŞLER — Akışı başlat tablosu (en yeni 15) ═══
        akis_siparisler = []
        try:
            aktif_sip_list = Siparis.query.filter(
                Siparis.durum.notin_(['Teslim Edildi', 'Iptal Edildi'])
            ).order_by(Siparis.siparis_tarihi.desc()).limit(15).all()
            for s in aktif_sip_list:
                # İlk kalemden ürün bilgisi
                ilk_kalem = s.kalemler[0] if s.kalemler else None
                cins = ilk_kalem.cins if ilk_kalem else None
                urun_tip = ilk_kalem.urun_tip if ilk_kalem else None
                miktar = ilk_kalem.miktar if ilk_kalem else None
                birim = ilk_kalem.birim if ilk_kalem else None
                # Bu siparişin proforması/sevkiyatı var mı?
                proforma_var = Proforma.query.filter_by(siparis_id=s.id).count() > 0
                sevkiyat_var = Sevkiyat.query.filter_by(siparis_id=s.id).count() > 0
                sevke_uygun = s.durum in ('Onaylandi', 'Uretimde', 'Hazir')
                akis_siparisler.append({
                    'id': s.id, 'musteri': s.musteri, 'durum': s.durum,
                    'termin': s.termin.strftime('%d.%m.%Y') if s.termin else '—',
                    'toplam_tutar': q3(s.toplam_tutar or 0),
                    'tutar': q3(s.toplam_tutar or 0),
                    'doviz': s.doviz or 'USD', 'urun_tip': urun_tip, 'cins': cins or '—',
                    'miktar': q3(miktar or 0) if miktar else None, 'birim': birim,
                    'proforma_var': proforma_var, 'sevkiyat_var': sevkiyat_var,
                    'sevke_uygun': sevke_uygun,
                })
        except Exception:
            akis_siparisler = []

        return render_template('dashboard.html', stats=stats, yaklasan=yaklasan_siparis, kur_usd=kur_usd, kur_eur=kur_eur,
                               today=bugun, toplam_alacak=q3(toplam_alacak), toplam_borc=q3(toplam_borc),
                               net_bakiye=q3(net_bakiye), yaklasan_odemeler=yaklasan_odemeler, yaklasan_tahsilatlar=yaklasan_tahsilatlar,
                               akis_siparisler=akis_siparisler,
                               ana_pb=ana_pb)

    # ---------- DİĞER SAYFALAR ----------
    @app.route('/stok')
    def stok_sayfa():
        if _auth_required(): return _auth_required()
        if not _yetki_var_mi('stok', 'okuma'):
            return redirect(url_for('dashboard'))
        return render_template('stok.html')

    @app.route('/siparis')
    def siparis_sayfa():
        if _auth_required(): return _auth_required()
        if not _yetki_var_mi('siparis', 'okuma'):
            return redirect(url_for('dashboard'))
        return render_template('siparis.html')

    @app.route('/rezervasyon')
    def rezervasyon_sayfa():
        if _auth_required(): return _auth_required()
        if not _yetki_var_mi('rezervasyon', 'okuma'):
            return redirect(url_for('dashboard'))
        return render_template('rezervasyon.html')

    @app.route('/cari')
    def cari_sayfa():
        if _auth_required(): return _auth_required()
        if not _yetki_var_mi('cari', 'okuma'):
            return redirect(url_for('dashboard'))
        return render_template('cari.html')

    @app.route('/maliyet')
    def maliyet_sayfa():
        if _auth_required(): return _auth_required()
        if not _yetki_var_mi('maliyet', 'okuma'):
            return redirect(url_for('dashboard'))
        return render_template('maliyet.html')

    @app.route('/sevkiyat')
    def sevkiyat_sayfa():
        if _auth_required(): return _auth_required()
        if not _yetki_var_mi('sevkiyat', 'okuma'):
            return redirect(url_for('dashboard'))
        return render_template('sevkiyat.html')

    @app.route('/kesim')
    def kesim_sayfa():
        if _auth_required(): return _auth_required()
        if not _yetki_var_mi('kesim', 'okuma'):
            return redirect(url_for('dashboard'))
        return render_template('kesim.html')

    @app.route('/kasa')
    def kasa_sayfa():
        if _auth_required(): return _auth_required()
        if not _yetki_var_mi('kasa', 'okuma'):
            return redirect(url_for('dashboard'))
        return render_template('kasa.html')

    @app.route('/cek')
    def cek_sayfa():
        if _auth_required(): return _auth_required()
        if not _yetki_var_mi('kasa', 'okuma'):
            return redirect(url_for('dashboard'))
        return render_template('cek.html')

    @app.route('/karlilik')
    def karlilik_sayfa():
        if _auth_required(): return _auth_required()
        if not _yetki_var_mi('karlilik', 'okuma'):
            return redirect(url_for('dashboard'))
        return render_template('karlilik.html')

    @app.route('/raporlar')
    def raporlar_sayfa():
        if _auth_required(): return _auth_required()
        if not _yetki_var_mi('raporlar', 'okuma'):
            return redirect(url_for('dashboard'))
        return render_template('raporlar.html')

    @app.route('/ayarlar')
    def ayarlar_sayfa():
        if _auth_required(): return _auth_required()
        if not _yetki_var_mi('ayarlar', 'okuma'):
            return redirect(url_for('dashboard'))
        return render_template('ayarlar.html')

    @app.route('/proforma')
    def proforma_sayfa():
        if _auth_required(): return _auth_required()
        if not _yetki_var_mi('proforma', 'okuma'):
            return redirect(url_for('dashboard'))
        return render_template('proforma.html')

    @app.route('/fatura')
    def fatura_sayfa():
        if _auth_required(): return _auth_required()
        if not _yetki_var_mi('fatura', 'okuma'):
            return redirect(url_for('dashboard'))
        return render_template('fatura.html')

    # ---------- API: DÖVİZ KURU ----------
    @app.route('/api/doviz_kur/guncelle', methods=['POST'])
    def api_doviz_kur_guncelle():
        """TCMB'den USD ve EUR kurlarini ceker (manuel tetiklenir)."""
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        try:
            guncel_kurlari_cek()
            usd = DovizKur.query.filter_by(doviz='USD').order_by(DovizKur.tarih.desc()).first()
            eur = DovizKur.query.filter_by(doviz='EUR').order_by(DovizKur.tarih.desc()).first()
            return jsonify({
                'ok': True,
                'usd': q_kur(usd.alis or usd.efektif or 0) if usd else 0,
                'eur': q_kur(eur.alis or eur.efektif or 0) if eur else 0,
                'mesaj': 'Kurlar guncellendi (TCMB)'
            })
        except Exception as e:
            return jsonify({'ok': False, 'mesaj': f'Kur cekme hatasi: {e}'}), 500

    @app.route('/api/doviz_kur', methods=['GET'])
    def api_doviz_kur():
        if _auth_required():
            return jsonify({'error': 'Unauthorized'}), 401
        # Tek doviz icin (cari hareket formu icin)
        single = request.args.get('doviz')
        tarih_param = request.args.get('tarih')  # YYYY-MM-DD: o tarihteki (veya öncesi) kur
        if single:
            single = single.upper()
            if single == 'TRY':
                return jsonify({'doviz': 'TRY', 'alis': 1.0, 'satis': 1.0, 'efektif': 1.0})
            q = DovizKur.query.filter_by(doviz=single)
            if tarih_param:
                t = _parse_date(tarih_param)
                if t:
                    # O tarih veya öncesi en yakın kur (tatil/hafta sonu → bir önceki iş günü)
                    q = q.filter(DovizKur.tarih <= t)
            k = q.order_by(DovizKur.tarih.desc()).first()
            if not k:
                return jsonify({'doviz': single, 'alis': 0, 'satis': 0, 'efektif': 0, 'mesaj': 'Kur bulunamadi'})
            return jsonify({'doviz': single, 'alis': q_kur(k.alis or 0),
                            'satis': q_kur(k.satis or 0), 'efektif': q_kur(k.efektif or 0),
                            'tarih': k.tarih.isoformat() if k.tarih else None})

        # Tum dovizler (geri uyumluluk)
        usd = DovizKur.query.filter_by(doviz='USD').order_by(DovizKur.tarih.desc()).first()
        eur = DovizKur.query.filter_by(doviz='EUR').order_by(DovizKur.tarih.desc()).first()
        usd_efektif = usd.efektif if usd else 45.07
        eur_efektif = eur.efektif if eur else 48.50
        return jsonify({'USD': {'efektif': usd_efektif}, 'EUR': {'efektif': eur_efektif}, 'TRY': {'efektif': 1.0}})

    # ---------- API: STOK ----------
    @app.route('/api/stok', methods=['GET'])
    def api_stok_liste():
        """Stok listesi.

        Query params:
            tip      : 'BLOK' | 'PLAKA' | 'EBATLI'  (default: 'PLAKA')
                       NOTE: API tek seferde TEK TİP döndürür. UI'da
                       her tip için ayrı sekme olduğundan default
                       'PLAKA' bilinçli bir tercihtir; tüm stokları
                       almak için üç ayrı çağrı yapın.
            durum    : 'Serbest' | 'Rezerve' | 'Teslim Edildi' | 'aktif'
            cins, uretici, ozellik, blok_no  : Eşitlik filtreleri
            page, per_page                   : Sayfalama

        Yanıt: ``{ data: [...], meta: {page, per_page, total, tip} }``
        """
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        tip = request.args.get('tip', 'PLAKA')
        durum = request.args.get('durum', '')
        cins = request.args.get('cins', '')
        uretici = request.args.get('uretici', '')
        ozellik = request.args.get('ozellik', '')
        blok_no = request.args.get('blok_no', '')
        per_page = request.args.get('per_page', type=int, default=50)
        page = request.args.get('page', type=int, default=1)

        if tip == 'BLOK':
            query = BlokStok.query
            sort_col = BlokStok.giris_tarihi.desc()
        elif tip == 'PLAKA':
            query = PlakaStok.query
            sort_col = PlakaStok.giris_tarihi.desc()
        else:
            query = EbatliStok.query
            sort_col = EbatliStok.giris_tarihi.desc()

        if durum:
            if durum == 'aktif':
                # Teslim Edildi hariç tüm durumlar
                if tip == 'BLOK':
                    query = query.filter(BlokStok.durum != 'Teslim Edildi')
                elif tip == 'PLAKA':
                    query = query.filter(PlakaStok.durum != 'Teslim Edildi')
                else:
                    query = query.filter(EbatliStok.durum != 'Teslim Edildi')
            else:
                query = query.filter_by(durum=durum)
        if cins: query = query.filter_by(cins=cins)
        if uretici: query = query.filter_by(uretici=uretici)
        if ozellik: query = query.filter_by(ozellik=ozellik)
        # Fatura durumu filtresi (faturasiz / mal_bekliyor / faturali)
        fatura_durumu = request.args.get('fatura_durumu', '')
        if fatura_durumu:
            if tip == 'BLOK':
                query = query.filter(BlokStok.fatura_durumu == fatura_durumu)
            elif tip == 'PLAKA':
                query = query.filter(PlakaStok.fatura_durumu == fatura_durumu)
            else:
                query = query.filter(EbatliStok.fatura_durumu == fatura_durumu)
        if blok_no:
            # BLOK -> blok_no, EBATLI -> kasa_no
            if tip == 'EBATLI':
                query = query.filter_by(kasa_no=blok_no)
            else:
                query = query.filter_by(blok_no=blok_no)

        paginated = query.order_by(sort_col).paginate(page=page, per_page=per_page, error_out=False)

        # Bu sayfadaki stoklarin ek maliyetlerini tek sorguda topla
        sayfa_idler = [s.id for s in paginated.items]
        ek_maliyet_map = {}
        kdv_devreden_map = {}  # Devreden KDV (mahsup beklenen)
        kdv_iade_map = {}      # Iade KDV (devletten geri alinacak)
        if sayfa_idler:
            # Ek maliyet: KDV kalemleri HARIC (Devreden/Iade KDV maliyet degil)
            maliyet_sonuc = db.session.query(
                Maliyet.baglanti_id,
                db.func.sum(Maliyet.usd_karsilik)
            ).filter(
                Maliyet.baglanti_id.in_(sayfa_idler),
                ~Maliyet.maliyet_tip.in_(['Devreden KDV', 'Iade KDV'])
            ).group_by(Maliyet.baglanti_id).all()
            for bid, tutar in maliyet_sonuc:
                ek_maliyet_map[bid] = tutar or 0

            # KDV kalemleri ayri toplam (bilgi amacli, gosterim icin)
            for tip_adi, hedef_map in [('Devreden KDV', kdv_devreden_map), ('Iade KDV', kdv_iade_map)]:
                kdv_sonuc = db.session.query(
                    Maliyet.baglanti_id,
                    db.func.sum(Maliyet.usd_karsilik)
                ).filter(
                    Maliyet.baglanti_id.in_(sayfa_idler),
                    Maliyet.maliyet_tip == tip_adi
                ).group_by(Maliyet.baglanti_id).all()
                for bid, tutar in kdv_sonuc:
                    hedef_map[bid] = tutar or 0

        # USD kuru (alim fiyati USD'ye cevrimi icin)
        _uk = DovizKur.query.filter_by(doviz='USD').order_by(DovizKur.tarih.desc()).first()
        _ek = DovizKur.query.filter_by(doviz='EUR').order_by(DovizKur.tarih.desc()).first()
        kur_usd_v = (_uk.alis if _uk else 0) or 0
        kur_eur_v = (_ek.alis if _ek else 0) or 0

        def _alim_usd(fiyat, dvz):
            """Alim fiyatini USD'ye cevir."""
            if not fiyat: return 0
            if dvz == 'USD' or not dvz: return fiyat
            if dvz == 'EUR': return (fiyat * kur_eur_v / kur_usd_v) if kur_usd_v else 0
            if dvz == 'TRY': return (fiyat / kur_usd_v) if kur_usd_v else 0
            return fiyat

        items = []
        for s in paginated.items:
            item = {
                'id': s.id, 'tip': tip, 'cins': s.cins, 'durum': s.durum, 'konum': s.konum,
                'fatura_durumu': getattr(s, 'fatura_durumu', None) or 'faturali',
                'uretici': getattr(s, 'uretici', None),
                'aciklama': getattr(s, 'aciklama', None) or '',
                'fiyat': s.alis_fiyati, 'alis_fiyati': s.alis_fiyati, 'doviz': s.doviz,
                'fiyat_birim': getattr(s, 'alis_fiyat_birim', None) or ('ton' if tip == 'BLOK' else 'm2'),
                'kdv_dahil_mi': bool(getattr(s, 'kdv_dahil_mi', False)),
                'kdv_oran': getattr(s, 'kdv_oran', 0) or 0,
                'kdv_tutar': getattr(s, 'kdv_tutar', 0) or 0,
                'matrah': getattr(s, 'matrah', 0) or 0
            }

            # TEMEL birim: BLOK -> ton, PLAKA/EBATLI -> m2
            if tip == 'BLOK':
                temel_birim = 'ton'
                temel_olcu = s.tonaj or 0
                birim_ad = 'ton'
            else:
                temel_birim = 'm2'
                temel_olcu = s.metraj_m2 or 0
                birim_ad = 'm²'

            alis_birim = item['fiyat_birim']
            alis_fiyat = s.alis_fiyati or 0

            # TOPLAM ALIM MALİYETİ (USD):
            #   alim fiyati kendi biriminde -> o birimin ölçüsüyle çarp -> USD'ye çevir
            alim_olcu = _stok_olcu(s, alis_birim)  # alış biriminin karşılığı ölçü
            alim_toplam_usd = _alim_usd(alis_fiyat * alim_olcu, s.doviz or 'USD')

            # EK MALİYET (Maliyet tablosundan, zaten USD toplam)
            ek_maliyet_toplam = ek_maliyet_map.get(s.id, 0)

            # TOPLAM MALİYET (USD)
            toplam_maliyet_usd = alim_toplam_usd + ek_maliyet_toplam

            # BİRİM MALİYET = toplam maliyet / temel ölçü (ton veya m2 başına)
            birim_maliyet_usd = q3(toplam_maliyet_usd / temel_olcu) if temel_olcu > 0 else 0

            item['birim_maliyet_usd'] = birim_maliyet_usd
            item['toplam_maliyet_usd'] = q3(toplam_maliyet_usd)
            item['birim_ad'] = birim_ad
            item['ek_maliyet_toplam'] = q3(ek_maliyet_toplam)
            item['kdv_devreden_usd'] = q3(kdv_devreden_map.get(s.id, 0))
            item['kdv_iade_usd'] = q3(kdv_iade_map.get(s.id, 0))
            item['alis_tipi'] = getattr(s, 'alis_tipi', None) or 'yurtici_kdvli'
            item['fatura_no'] = getattr(s, 'fatura_no', None) or ''
            item['fatura_durumu'] = getattr(s, 'fatura_durumu', None) or 'faturali'
            item['alis_tarihi'] = s.alis_tarihi.isoformat() if getattr(s, 'alis_tarihi', None) else None
            item['giris_tarihi'] = s.giris_tarihi.isoformat() if getattr(s, 'giris_tarihi', None) else None

            if tip == 'BLOK':
                item.update({'boy': s.boy, 'yukseklik': s.yukseklik, 'en': s.en, 'hacim': s.hacim_m3, 'tonaj': s.tonaj, 'blok_no': s.blok_no})
            elif tip == 'PLAKA':
                item.update({'boy': s.boy, 'yukseklik': s.yukseklik, 'kalinlik': s.kalinlik, 'm2': s.metraj_m2, 'sqft': s.metraj_sqft, 'blok_no': s.blok_no, 'ozellik': s.ozellik, 'slab_no': s.slab_no})
            else:
                item.update({'boy': s.boy, 'yukseklik': s.yukseklik, 'kalinlik': s.kalinlik, 'm2': s.metraj_m2, 'kasa_no': s.kasa_no, 'blok_no': s.kasa_no, 'ozellik': s.ozellik, 'kasa_ici_adet': s.kasa_ici_adet})
            items.append(item)
        return jsonify({'data': items, 'meta': {'page': page, 'per_page': per_page, 'total': paginated.total, 'tip': tip, 'note': 'Endpoint returns a single stok type per call; use ?tip=BLOK|PLAKA|EBATLI for the others.'}})

    @app.route('/api/blok/<blok_no>/izleme', methods=['GET'])
    def api_blok_izleme(blok_no):
        """Bir blok numarası üzerinden tüm yaşam döngüsü izlemesi.

        ÖZELLİK: blok_no hem orijinal blok numarası hem de **üretim blok
        numarası** (kesim sonrası verilen yeni numara) olarak aranır.
        Böylece hem orijinal blok hem de yeni üretim numarası üzerinden
        tüm maliyet/karlılık zinciri sorgulanabilir.

        Yanıt:
            {
              ok: true,
              blok_no: "A-2026-001",
              kaynak_bloklar: [...]   # bu numaraya sahip blok stoklar
              kesimler: [...]         # bu numarayı kaynak_no veya
                                      # uretim_blok_no olarak içeren kesimler
              uretilen_stoklar: [...] # bu numaraya sahip plakalar/ebatlılar
              maliyetler: [...]       # ilgili tüm maliyet kayıtları
            }
        """
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        bn = (blok_no or '').strip()

        # 1) Kaynak blok stoklar (orijinal blok numarası)
        kaynak_bloklar = []
        for b in BlokStok.query.filter_by(blok_no=bn).all():
            kaynak_bloklar.append({
                'id': b.id, 'cins': b.cins, 'uretici': b.uretici,
                'boy': b.boy, 'yukseklik': b.yukseklik, 'en': b.en,
                'durum': b.durum, 'doviz': b.doviz,
                'matrah': b.matrah, 'tonaj': b.tonaj,
                'giris_tarihi': b.giris_tarihi.isoformat() if b.giris_tarihi else None,
                'fatura_no': getattr(b, 'fatura_no', None) or ''
            })

        # 2) Kesimler: kaynak_no veya uretim_blok_no eşleşen
        kesimler = []
        for k in Kesim.query.filter(
            db.or_(Kesim.kaynak_no == bn, Kesim.uretim_blok_no == bn)
        ).all():
            kesimler.append({
                'id': k.id,
                'tarih': k.kesim_tarihi.isoformat() if k.kesim_tarihi else None,
                'kaynak_no': k.kaynak_no, 'uretim_blok_no': k.uretim_blok_no,
                'kaynak_cins': k.kaynak_cins,
                'fire_orani': k.fire_orani or 0,
                'toplam_maliyet': k.kaynak_toplam_maliyet or 0,
                'doviz': k.kaynak_doviz or 'USD',
                'kullanici': k.kullanici
            })

        # 3) Üretilen stoklar: blok_no=bn olan tüm plaka ve ebatlılar
        uretilen = []
        for p in PlakaStok.query.filter_by(blok_no=bn).all():
            uretilen.append({
                'tip': 'PLAKA', 'id': p.id, 'cins': p.cins,
                'boy': p.boy, 'yukseklik': p.yukseklik, 'kalinlik': p.kalinlik,
                'metraj_m2': p.metraj_m2, 'durum': p.durum,
                'alis_fiyati': p.alis_fiyati, 'matrah': p.matrah, 'doviz': p.doviz,
                'aciklama': p.aciklama
            })
        # EbatliStok'ta blok_no yok; aciklama içinde aramaya bakalım
        for e in EbatliStok.query.filter(
            db.or_(EbatliStok.aciklama.contains(bn))
        ).all():
            uretilen.append({
                'tip': 'EBATLI', 'id': e.id, 'cins': e.cins,
                'boy': e.boy, 'yukseklik': e.yukseklik, 'kalinlik': e.kalinlik,
                'kasa_no': e.kasa_no, 'metraj_m2': e.metraj_m2, 'durum': e.durum,
                'alis_fiyati': e.alis_fiyati, 'matrah': e.matrah, 'doviz': e.doviz,
                'aciklama': e.aciklama
            })

        # 4) Maliyet kayıtları (ilgili kesim'lerin tüm maliyetleri)
        maliyetler = []
        kesim_ids = [k['id'] for k in kesimler]
        blk_ids = [b['id'] for b in kaynak_bloklar]
        if kesim_ids or blk_ids:
            q = Maliyet.query.filter(
                db.or_(
                    db.and_(Maliyet.baglanti_tip == 'kesim', Maliyet.baglanti_id.in_(kesim_ids)) if kesim_ids else False,
                    db.and_(func.lower(Maliyet.baglanti_tip) == 'stok',  Maliyet.baglanti_id.in_(blk_ids))    if blk_ids   else False,
                )
            )
            for m in q.all():
                maliyetler.append({
                    'id': m.id, 'tarih': m.maliyet_tarihi.isoformat() if m.maliyet_tarihi else None,
                    'tip': m.maliyet_tip, 'tutar': m.tutar, 'doviz': m.doviz,
                    'kaynak': f'{m.baglanti_tip}:{m.baglanti_id}',
                    'aciklama': m.aciklama
                })

        return jsonify({
            'ok': True, 'blok_no': bn,
            'kaynak_bloklar': kaynak_bloklar,
            'kesimler': kesimler,
            'uretilen_stoklar': uretilen,
            'maliyetler': maliyetler,
            'ozet': {
                'orijinal_blok_sayisi': len(kaynak_bloklar),
                'kesim_sayisi': len(kesimler),
                'uretilen_stok_sayisi': len(uretilen),
                'maliyet_kayit_sayisi': len(maliyetler)
            }
        })

    @app.route('/api/stok/<stok_id>/gecmis', methods=['GET'])
    def api_stok_gecmis(stok_id):
        """Bir stok kaydının tüm yaşam döngüsünü timeline olarak döndürür.

        Olay tipleri (renk + tarih + açıklama):
            • giris           - Stoğa girdi
            • rezervasyon     - Sipariş kalemine bağlandı (Rezerve)
            • rezervasyon_iptal - Rezervasyon iptal edildi
            • cikis           - Stoktan çıktı (satıldı / sevkiyat)
            • kesim_kaynak    - Bu stoktan başka stok üretildi
            • kesim_hedef     - Bu stok bir kesimden üretildi
            • cari_hareket    - İlgili cari hareket (alış maliyeti, vs.)
        """
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401

        # Stoğu üç tablodan birinde bul
        stok = (BlokStok.query.get(stok_id) or
                PlakaStok.query.get(stok_id) or
                EbatliStok.query.get(stok_id))
        if not stok:
            return jsonify({'ok': False, 'mesaj': 'Stok bulunamadı'}), 404

        olaylar = []
        # 1) Giriş
        olaylar.append({
            'tip': 'giris', 'renk': 'success',
            'tarih': stok.giris_tarihi.isoformat() if stok.giris_tarihi else None,
            'baslik': 'Stok Girişi',
            'aciklama': f"Üretici: {stok.uretici or '-'} | Fatura: {getattr(stok, 'fatura_no', None) or '-'} | Doviz: {stok.doviz}",
            'kullanici': stok.kullanici or '-',
        })

        # 2) Rezervasyonlar
        for r in Rezervasyon.query.filter_by(stok_id=stok_id).order_by(Rezervasyon.id.asc()).all():
            olaylar.append({
                'tip': 'rezervasyon' if not r.iptal_nedeni else 'rezervasyon_iptal',
                'renk': 'info' if not r.iptal_nedeni else 'danger',
                'tarih': r.olusturma.isoformat() if r.olusturma else None,
                'baslik': 'Rezerve Edildi' if not r.iptal_nedeni else 'Rezervasyon İptal',
                'aciklama': f"Sipariş: {r.siparis_id or '-'} | Müşteri: {r.musteri or '-'}"
                            + (f" | İptal: {r.iptal_nedeni}" if r.iptal_nedeni else ''),
                'kullanici': '-',
            })

        # 3) Çıkışlar
        for c in StokCikis.query.filter_by(stok_id=stok_id).order_by(StokCikis.cikis_tarihi.asc()).all():
            olaylar.append({
                'tip': 'cikis', 'renk': 'danger',
                'tarih': c.cikis_tarihi.isoformat() if c.cikis_tarihi else None,
                'baslik': f"Stok Çıkış ({c.cikis_nedeni or 'Satış'})",
                'aciklama': f"Müşteri: {c.musteri or '-'} | Tutar: {c.satis_fiyati or 0} {c.doviz or 'USD'}",
                'kullanici': c.kullanici or '-',
            })

        # 4) Kesim - bu stok bir kesimde kaynak mı?
        kesim_kaynak = Kesim.query.filter_by(kaynak_id=stok_id).all()
        for k in kesim_kaynak:
            olaylar.append({
                'tip': 'kesim_kaynak', 'renk': 'warning',
                'tarih': k.kesim_tarihi.isoformat() if k.kesim_tarihi else None,
                'baslik': 'Kesime Verildi',
                'aciklama': f"Kesim: {k.id} | Fire %: {k.fire_orani or 0} | Hedef stoklar üretildi",
                'kullanici': k.kullanici or '-',
            })
        # ya da bir kesimin hedefi mi?
        kdetay = KesimDetay.query.filter_by(hedef_stok_id=stok_id).first()
        if kdetay:
            k = Kesim.query.get(kdetay.kesim_id) if kdetay.kesim_id else None
            if k:
                olaylar.append({
                    'tip': 'kesim_hedef', 'renk': 'success',
                    'tarih': k.kesim_tarihi.isoformat() if k.kesim_tarihi else None,
                    'baslik': 'Kesimden Üretildi',
                    'aciklama': f"Kaynak: {k.kaynak_tip} {k.kaynak_id or k.kaynak_no or '-'}",
                    'kullanici': k.kullanici or '-',
                })

        # 5) İlgili cari hareketler
        for ch in CariHareket.query.filter_by(baglanti_tip='stok', baglanti_id=stok_id).all():
            olaylar.append({
                'tip': 'cari_hareket', 'renk': 'info',
                'tarih': ch.hareket_tarihi.isoformat() if ch.hareket_tarihi else None,
                'baslik': f"Cari Hareket: {ch.islem_tip or '-'}",
                'aciklama': f"{ch.cari_unvan or '-'} | Borç: {ch.borc or 0} / Alacak: {ch.alacak or 0} {ch.doviz}",
                'kullanici': ch.kullanici or '-',
            })

        # Tarihe göre sırala
        olaylar.sort(key=lambda x: (x.get('tarih') or '', x.get('tip') or ''))
        return jsonify({'ok': True, 'olaylar': olaylar, 'stok_id': stok_id})

    @app.route('/api/stok/ekle', methods=['POST'])
    def api_stok_ekle():
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        data = request.json
        tip = data.get('tip')

        # ── ZORUNLU ALAN DOĞRULAMASI ──
        # UI bu kontrolleri yapıyor ama API doğrudan da çağrılabilir; sunucu
        # kendi doğrulamasını yapmalı, yoksa cinssiz/tipsiz kayıt oluşabiliyor.
        if tip not in ('BLOK', 'PLAKA', 'EBATLI'):
            return jsonify({'ok': False, 'error': 'gecersiz_tip',
                'mesaj': 'Gecerli bir stok tipi secin (BLOK / PLAKA / EBATLI).'}), 400
        if not (data.get('cins') or '').strip():
            return jsonify({'ok': False, 'error': 'cins_zorunlu',
                'mesaj': 'Cins zorunludur — urun tanimi icin gereklidir.'}), 400

        # Form degerleri JSON'da string gelebilir -> guvenli sayi cevirici
        def _sayi(anahtar, varsayilan=0):
            v = data.get(anahtar, varsayilan)
            if v is None or v == '':
                return varsayilan
            try:
                return float(v)
            except (ValueError, TypeError):
                return varsayilan

        # ─── FATURA DURUMU + TARİHLER (mal/fatura zaman farkı yönetimi) ───
        # fatura_durumu: 'faturali' (mal+fatura birlikte) | 'faturasiz' (mal geldi, fatura bekliyor)
        #                | 'mal_bekliyor' (fatura kesildi, mal henüz gelmedi)
        _fatura_durumu = (data.get('fatura_durumu') or 'faturali').strip()
        if _fatura_durumu not in ('faturali', 'faturasiz', 'mal_bekliyor'):
            _fatura_durumu = 'faturali'
        # giris_tarihi: malın depoya girdiği gün (verilmezse bugün)
        _giris_tarihi = _parse_date(data.get('giris_tarihi')) or date.today()
        # alis_tarihi: fatura/borç tarihi (verilmezse giriş tarihi ile aynı)
        _alis_tarihi = _parse_date(data.get('alis_tarihi')) or _giris_tarihi

        # ─── VERİ DOĞRULAMA ───────────────────────────────────────────
        # 1) Negatif/sıfır fiyat kontrolü
        _fiyat_kontrol = _sayi('fiyat')
        if _fiyat_kontrol < 0:
            return jsonify({'ok': False, 'mesaj': '❌ Fiyat negatif olamaz!'}), 400

        # 2) Tekrar (duplicate) kontrolü — aynı blok_no/kasa_no varsa uyar
        _blok_no = (data.get('blok_no') or '').strip()
        _kasa_no = (data.get('bas_kasa_no') or data.get('kasa_no') or '').strip()
        if tip == 'BLOK' and _blok_no:
            mevcut = BlokStok.query.filter(
                BlokStok.blok_no == _blok_no,
                BlokStok.durum != 'Teslim Edildi'
            ).first()
            if mevcut:
                return jsonify({'ok': False,
                    'mesaj': f'⚠️ Bu blok no zaten kayıtlı: "{_blok_no}" (ID: {mevcut.id}, durum: {mevcut.durum}). Farklı bir blok no girin veya mevcut kaydı düzenleyin.',
                    'duplicate': True, 'mevcut_id': mevcut.id}), 409
        elif tip == 'PLAKA' and _blok_no:
            # PLAKA'da aynı blok_no birden fazla olabilir (aynı bloktan kesilmiş plakalar)
            # ama UYARI verelim - kullanıcı bilsin
            mevcut_sayi = PlakaStok.query.filter(
                PlakaStok.blok_no == _blok_no,
                PlakaStok.durum != 'Teslim Edildi'
            ).count()
            # PLAKA için engelleme YOK, sadece bilgi (frontend'de gösterilebilir)
        # EBATLI: referans kodu sistemi otomatik benzersiz üretir, duplicate kontrolü gereksiz
        # ─── /VERİ DOĞRULAMA ──────────────────────────────────────────

        try:
            if tip == 'BLOK':
                # BLOK tek parca - adet yok
                boy, yuk, en = _sayi('boy'), _sayi('yukseklik'), _sayi('en')
                m3 = (boy * yuk * en) / 1000000
                fiyat_girilen = _sayi('fiyat')
                fiyat_birim = data.get('fiyat_birim', 'ton')
                doviz = data.get('doviz','USD')
                # KDV hesabi
                kdv_dahil_mi = bool(data.get('kdv_dahil_mi', False))
                kdv_oran = q_oran(data.get('kdv_oran', 0) or 0)
                alis_tipi = (data.get('alis_tipi') or 'yurtici_kdvli').strip()
                # KDV'siz alis tipleri icin oranni zorla 0 yap
                if alis_tipi != 'yurtici_kdvli':
                    kdv_oran = 0
                    kdv_dahil_mi = False
                # Birim fiyat KDV hariç (matrah) saklanır
                if kdv_dahil_mi and kdv_oran > 0:
                    fiyat = q3(fiyat_girilen / (1 + kdv_oran/100))
                else:
                    fiyat = fiyat_girilen
                # Toplam tutarlar
                miktar = _sayi('tonaj') if fiyat_birim == 'ton' else q3(m3)
                matrah = q3(fiyat * miktar)
                kdv_tutar = q3(matrah * kdv_oran / 100) if kdv_oran > 0 else 0
                stok = BlokStok(id=_yeni_id('BLK'), uretici=data.get('uretici'), cins=data.get('cins'), blok_no=data.get('blok_no'),
                                boy=boy, yukseklik=yuk, en=en, hacim_m3=q3(m3),
                                tonaj=_sayi('tonaj'), alis_fiyati=fiyat,
                                alis_fiyat_birim=fiyat_birim,
                                doviz=doviz, durum='Serbest',
                                kdv_dahil_mi=kdv_dahil_mi, kdv_oran=kdv_oran,
                                kdv_tutar=q3(kdv_tutar), matrah=q3(matrah),
                                fatura_no=(data.get('fatura_no') or '').strip() or None,
                                giris_tarihi=_giris_tarihi, alis_tarihi=_alis_tarihi,
                                fatura_durumu=_fatura_durumu,
                                aciklama=data.get('aciklama', ''), alis_tipi=alis_tipi)
                db.session.add(stok)
                db.session.flush()
                # KDV varsa otomatik 'Devreden KDV' maliyet kalemi (faturalıysa)
                if kdv_tutar > 0 and _fatura_durumu != 'faturasiz':
                    _devreden_kdv_kalemi_olustur(stok.id, kdv_tutar, doviz,
                        aciklama=f'BLOK stok girisi - {stok.cins} {stok.blok_no}',
                        fatura_no=stok.fatura_no)
                # Tedarikçi cariye alış faturası (matrah + KDV) — faturasız ise borç oluşmaz
                _stok_cari_hareket_olustur(stok.id, stok.uretici, q3(matrah + kdv_tutar), doviz,
                    fatura_no=stok.fatura_no or '',
                    aciklama=f'BLOK alış — {stok.cins} {stok.blok_no} — {stok.fatura_no or ""}',
                    fatura_durumu=_fatura_durumu, alis_tarihi=_alis_tarihi)
                _log_audit('EKLE', 'stok_blok', stok.id, yeni={'cins': stok.cins, 'uretici': stok.uretici, 'kdv': kdv_tutar})
                db.session.commit()
                return jsonify({'ok': True, 'id': stok.id, 'adet': 1,
                                'kdv_tutar': kdv_tutar, 'matrah': matrah})

            elif tip == 'PLAKA':
                # PLAKA: 'adet' kadar ayri plaka kaydi olustur
                adet = int(_sayi('adet', 1)) or 1
                bas_no = int(_sayi('baslangic_no', 1)) or 1
                boy, yuk = _sayi('boy'), _sayi('yukseklik')
                m2 = (boy * yuk) / 10000
                fiyat_girilen = _sayi('fiyat')
                fiyat_birim = data.get('fiyat_birim', 'm2')
                doviz = data.get('doviz','USD')
                # KDV ayarlari
                kdv_dahil_mi = bool(data.get('kdv_dahil_mi', False))
                kdv_oran = q_oran(data.get('kdv_oran', 0) or 0)
                alis_tipi = (data.get('alis_tipi') or 'yurtici_kdvli').strip()
                # KDV'siz alis tipleri icin oranni zorla 0 yap
                if alis_tipi != 'yurtici_kdvli':
                    kdv_oran = 0
                    kdv_dahil_mi = False
                # Birim fiyat KDV hariç (matrah) saklanır
                if kdv_dahil_mi and kdv_oran > 0:
                    # KDV dahil girilmis: ayristir
                    fiyat = q3(fiyat_girilen / (1 + kdv_oran/100))
                else:
                    # KDV haric girilmis: fiyat zaten matrah
                    fiyat = fiyat_girilen
                # Plaka basina KDV hariç matrah ve KDV tutarı
                miktar_birim = m2 if fiyat_birim == 'm2' else q3(m2 * M2_TO_SQFT)
                matrah = q3(fiyat * miktar_birim)
                # KDV tutari: kdv_oran > 0 ise her zaman hesaplanir (KDV haric girilmise de)
                kdv_tutar = q3(matrah * kdv_oran / 100) if kdv_oran > 0 else 0
                olusan_idler = []
                for i in range(adet):
                    slab = bas_no + i
                    stok = PlakaStok(id=_yeni_id('PLK'), uretici=data.get('uretici'), cins=data.get('cins'), blok_no=data.get('blok_no'),
                                     boy=boy, yukseklik=yuk, kalinlik=_sayi('kalinlik'), ozellik=data.get('ozellik'),
                                     metraj_m2=q3(m2), metraj_sqft=q3(m2*10.764),
                                     slab_no=slab, alis_fiyati=fiyat,
                                     alis_fiyat_birim=fiyat_birim,
                                     doviz=doviz,
                                     kdv_dahil_mi=kdv_dahil_mi, kdv_oran=kdv_oran,
                                     kdv_tutar=q3(kdv_tutar), matrah=q3(matrah),
                                     fatura_no=(data.get('fatura_no') or '').strip() or None,
                                     giris_tarihi=_giris_tarihi, alis_tarihi=_alis_tarihi,
                                     fatura_durumu=_fatura_durumu,
                                     durum='Serbest', aciklama=data.get('aciklama', ''), alis_tipi=alis_tipi)
                    db.session.add(stok)
                    db.session.flush()
                    if kdv_tutar > 0 and _fatura_durumu != 'faturasiz':
                        _devreden_kdv_kalemi_olustur(stok.id, kdv_tutar, doviz,
                            aciklama=f'PLAKA stok girisi - {stok.cins} {stok.blok_no} #{slab}',
                            fatura_no=stok.fatura_no)
                    # Tedarikçi cariye alış faturası (her plaka için matrah + KDV) — faturasız ise borç oluşmaz
                    _stok_cari_hareket_olustur(stok.id, stok.uretici, q3(matrah + kdv_tutar), doviz,
                        fatura_no=stok.fatura_no or '',
                        aciklama=f'PLAKA alış — {stok.cins} {stok.blok_no} #{slab} — {stok.fatura_no or ""}',
                        fatura_durumu=_fatura_durumu, alis_tarihi=_alis_tarihi)
                    olusan_idler.append(stok.id)
                son_no = bas_no + adet - 1
                _log_audit('EKLE', 'stok_plaka', f'{adet} adet',
                           yeni={'cins': data.get('cins'), 'adet': adet,
                                 'slab_araligi': f'{bas_no}-{son_no}', 'blok_no': data.get('blok_no')})
                db.session.commit()
                return jsonify({'ok': True, 'adet': adet, 'idler': olusan_idler,
                                'mesaj': f'{adet} plaka stoğa eklendi (slab no: {bas_no}-{son_no})'})

            else:
                # EBATLI: her kasa AYRI kayit, her birine otomatik referans kodu
                boy = _sayi('boy')
                yuk = _sayi('yukseklik')
                kal = _sayi('kalinlik')
                kasa_adet = int(_sayi('kasa_adedi', 1) or 1)
                if kasa_adet < 1: kasa_adet = 1
                kasa_ici_adet = int(_sayi('kasa_ici_adet', 1) or 1)
                if kasa_ici_adet < 1: kasa_ici_adet = 1
                m2_per_kasa = (boy * yuk) / 10000 * kasa_ici_adet
                uretici = data.get('uretici') or ''
                cins = data.get('cins') or ''

                # Kasa adedi kadar referans kodu uret (en son N'den devam eder)
                try:
                    referans_kodlari = _referans_kodu_uret(
                        uretici, cins, boy, yuk, kal, adet=kasa_adet)
                except ValueError as ve:
                    return jsonify({'ok': False, 'mesaj': str(ve)}), 400

                # Manuel referans kodu girilmisse ilk kasa icin onu kullan
                manuel_ref = (data.get('blok_no') or '').strip()
                if manuel_ref:
                    referans_kodlari[0] = manuel_ref

                # KDV ayarlari
                kdv_dahil_mi = bool(data.get('kdv_dahil_mi', False))
                kdv_oran = q_oran(data.get('kdv_oran', 0) or 0)
                alis_tipi = (data.get('alis_tipi') or 'yurtici_kdvli').strip()
                # KDV'siz alis tipleri icin oranni zorla 0 yap
                if alis_tipi != 'yurtici_kdvli':
                    kdv_oran = 0
                    kdv_dahil_mi = False
                fiyat_girilen = _sayi('fiyat')
                fiyat_birim = data.get('fiyat_birim', 'm2')
                doviz = data.get('doviz','USD')
                # Birim fiyat KDV hariç (matrah) saklanır
                if kdv_dahil_mi and kdv_oran > 0:
                    fiyat = q3(fiyat_girilen / (1 + kdv_oran/100))
                else:
                    fiyat = fiyat_girilen
                # Kasa basina KDV hariç matrah ve KDV tutarı
                miktar_birim = m2_per_kasa if fiyat_birim == 'm2' else q3(m2_per_kasa * M2_TO_SQFT)
                matrah_k = q3(fiyat * miktar_birim)
                kdv_t_k = q3(matrah_k * kdv_oran / 100) if kdv_oran > 0 else 0

                olusturulan_ids = []
                for ref_kod in referans_kodlari:
                    stok = EbatliStok(
                        id=_yeni_id('EBT'),
                        uretici=uretici, cins=cins, kasa_no=ref_kod,
                        boy=boy, yukseklik=yuk, kalinlik=kal,
                        ozellik=data.get('ozellik'),
                        kasa_adedi=1,
                        kasa_ici_adet=kasa_ici_adet,
                        m2_kg=_sayi('m2_kg'),
                        metraj_m2=q3(m2_per_kasa),
                        metraj_sqft=q3(m2_per_kasa * M2_TO_SQFT),
                        alis_fiyati=fiyat,
                        alis_fiyat_birim=fiyat_birim,
                        doviz=doviz,
                        kdv_dahil_mi=kdv_dahil_mi, kdv_oran=kdv_oran,
                        alis_tipi=alis_tipi,
                        kdv_tutar=q3(kdv_t_k), matrah=q3(matrah_k),
                        durum='Serbest',
                        aciklama=data.get('aciklama', ''),
                        fatura_no=(data.get('fatura_no') or '').strip() or None,
                        giris_tarihi=_giris_tarihi, alis_tarihi=_alis_tarihi,
                        fatura_durumu=_fatura_durumu,
                        kullanici=session.get('kullanici')
                    )
                    db.session.add(stok)
                    db.session.flush()
                    olusturulan_ids.append(stok.id)
                    if kdv_t_k > 0 and _fatura_durumu != 'faturasiz':
                        _devreden_kdv_kalemi_olustur(stok.id, kdv_t_k, doviz,
                            aciklama=f'EBATLI stok girisi - {cins} {ref_kod}',
                            fatura_no=stok.fatura_no)
                    # Tedarikçi cariye alış faturası (her kasa için matrah + KDV) — faturasız ise borç oluşmaz
                    _stok_cari_hareket_olustur(stok.id, uretici, q3(matrah_k + kdv_t_k), doviz,
                        fatura_no=stok.fatura_no or '',
                        aciklama=f'EBATLI alış — {cins} {ref_kod} — {stok.fatura_no or ""}',
                        fatura_durumu=_fatura_durumu, alis_tarihi=_alis_tarihi)
                    _log_audit('EKLE', 'stok_ebatli', stok.id,
                               yeni={'cins': cins, 'uretici': uretici, 'ref': ref_kod, 'kdv': kdv_t_k})
                db.session.commit()
                return jsonify({
                    'ok': True, 'adet': len(referans_kodlari),
                    'referans_kodlari': referans_kodlari,
                    'mesaj': f'{len(referans_kodlari)} kasa eklendi: {referans_kodlari[0]} - {referans_kodlari[-1]}'
                        if len(referans_kodlari) > 1 else f'1 kasa eklendi: {referans_kodlari[0]}'
                })
        except Exception as e:
            db.session.rollback()
            return jsonify({'ok': False, 'mesaj': str(e)}), 500

    @app.route('/api/stok/<tip>/<stok_id>/faturalandir', methods=['POST'])
    def api_stok_faturalandir(tip, stok_id):
        """
        FATURASIZ stoğu (mal geldi, fatura bekliyordu) faturalandırır.
        Fatura no + tarih girilince tedarikçi cariye borç (alış faturası) oluşturur.
        """
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        data = request.json or {}
        stok = _stok_getir(stok_id, tip.upper())
        if not stok:
            return jsonify({'ok': False, 'mesaj': 'Stok bulunamadı'}), 404
        if getattr(stok, 'fatura_durumu', 'faturali') != 'faturasiz':
            return jsonify({'ok': False, 'mesaj': 'Bu stok zaten faturalı (veya mal bekliyor durumunda).'}), 400

        fatura_no = (data.get('fatura_no') or '').strip()
        alis_tarihi = _parse_date(data.get('alis_tarihi')) or date.today()
        # Stoğu güncelle
        stok.fatura_no = fatura_no or stok.fatura_no
        stok.alis_tarihi = alis_tarihi
        stok.fatura_durumu = 'faturali'
        # KDV maliyet kalemi (varsa şimdi oluştur)
        toplam = q3((stok.matrah or 0) + (stok.kdv_tutar or 0))
        if (stok.kdv_tutar or 0) > 0:
            _devreden_kdv_kalemi_olustur(stok.id, stok.kdv_tutar, stok.doviz,
                aciklama=f'{tip.upper()} faturalandırma - {stok.cins} {getattr(stok,"blok_no","") or getattr(stok,"kasa_no","")}',
                fatura_no=stok.fatura_no)
        # Tedarikçi cariye borç (fatura tarihinde)
        ch = _stok_cari_hareket_olustur(stok.id, stok.uretici, toplam, stok.doviz,
            fatura_no=fatura_no,
            aciklama=f'{tip.upper()} faturalandırma — {stok.cins} — {fatura_no}',
            fatura_durumu='faturali', alis_tarihi=alis_tarihi)
        _log_audit('GUNCELLE', f'stok_{tip.lower()}_faturalandir', stok.id,
                   yeni={'fatura_no': fatura_no, 'tarih': alis_tarihi.isoformat()})
        ok, hata = _safe_commit(f'Stok faturalandırma: {stok_id}')
        if not ok:
            return jsonify({'ok': False, 'mesaj': f'Hata: {hata}'}), 500
        return jsonify({'ok': True,
            'mesaj': f'Stok faturalandırıldı. Tedarikçi cariye {toplam:,.2f} {stok.doviz} borç işlendi.',
            'cari_hareket': ch.id if ch else None})

    @app.route('/api/stok/<tip>/<stok_id>/mal_teslim', methods=['POST'])
    def api_stok_mal_teslim(tip, stok_id):
        """
        MAL BEKLEYEN stoğu (fatura kesildi, borç vardı, mal henüz gelmemişti)
        teslim alındı olarak işaretler. Borç zaten oluşmuştu; sadece mal giriş
        tarihini ve durumu günceller.
        """
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        data = request.json or {}
        stok = _stok_getir(stok_id, tip.upper())
        if not stok:
            return jsonify({'ok': False, 'mesaj': 'Stok bulunamadı'}), 404
        if getattr(stok, 'fatura_durumu', 'faturali') != 'mal_bekliyor':
            return jsonify({'ok': False, 'mesaj': 'Bu stok zaten teslim alınmış.'}), 400
        # Mal giriş tarihini güncelle (fiziksel teslim günü)
        stok.giris_tarihi = _parse_date(data.get('giris_tarihi')) or date.today()
        stok.fatura_durumu = 'faturali'
        _log_audit('GUNCELLE', f'stok_{tip.lower()}_mal_teslim', stok.id,
                   yeni={'giris_tarihi': stok.giris_tarihi.isoformat()})
        ok, hata = _safe_commit(f'Mal teslim: {stok_id}')
        if not ok:
            return jsonify({'ok': False, 'mesaj': f'Hata: {hata}'}), 500
        return jsonify({'ok': True, 'mesaj': 'Mal teslim alındı olarak işaretlendi.'})

    @app.route('/api/stok/<stok_id>', methods=['PUT'])
    def api_stok_guncelle(stok_id):
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        data = request.json
        for model in [BlokStok, PlakaStok, EbatliStok]:
            stok = model.query.get(stok_id)
            if stok:
                eski_durum = stok.durum
                if 'durum' in data:
                    yeni_durum = data['durum']
                    stok.durum = yeni_durum
                    # VERI BUTUNLUGU: Stok "Serbest" yapildiysa aktif rezervasyonlarini iptal et
                    if yeni_durum == 'Serbest' and eski_durum in ('Rezerve', 'Teslim Edildi'):
                        aktif_rez = Rezervasyon.query.filter_by(stok_id=stok_id).filter(
                            (Rezervasyon.iptal_nedeni.is_(None)) | (Rezervasyon.iptal_nedeni == '')
                        ).all()
                        for r in aktif_rez:
                            r.iptal_nedeni = f'Stok manuel Serbest yapildi (eski durum: {eski_durum})'
                        if aktif_rez:
                            app.logger.warning(f'Stok {stok_id} Serbest -> {len(aktif_rez)} rezervasyon iptal edildi')
                if 'konum' in data: stok.konum = data['konum']
                if 'uretici' in data and data['uretici']:
                    stok.uretici = data['uretici']
                if 'alis_fiyati' in data:
                    try: stok.alis_fiyati = float(data['alis_fiyati']) if data['alis_fiyati'] not in (None,'') else None
                    except (ValueError, TypeError): pass
                if 'fiyat_birim' in data and data['fiyat_birim']:
                    stok.alis_fiyat_birim = data['fiyat_birim']
                if 'aciklama' in data: stok.aciklama = data['aciklama']
                db.session.commit()
                return jsonify({'ok': True})
        return jsonify({'ok': False, 'mesaj': 'Stok bulunamadı'}), 404

    @app.route('/api/stok/<tip>/<stok_id>', methods=['DELETE'])
    def api_stok_sil(tip, stok_id):
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        if tip == 'BLOK': stok = BlokStok.query.get(stok_id)
        elif tip == 'PLAKA': stok = PlakaStok.query.get(stok_id)
        else: stok = EbatliStok.query.get(stok_id)
        if not stok: return jsonify({'ok': False, 'mesaj': 'Bulunamadi'}), 404
        # Okunabilir numarayı silmeden ÖNCE yakala (mesajda kullanılacak)
        stok_no = _baglanti_okunabilir('stok', stok_id)

        # Rezerve veya Satildi durumundaki stoklar silinemez
        if stok.durum == 'Rezerve':
            aktif_rez = Rezervasyon.query.filter_by(stok_id=stok_id, stok_tip=tip, iptal_nedeni=None).first()
            if aktif_rez:
                return jsonify({'ok': False, 'mesaj': f'Bu stok {aktif_rez.musteri} musterisinin {aktif_rez.siparis_id or "rezervasyonuna"} bagli, once rezervasyonu iptal edin'}), 400
        if stok.durum == 'Satildi':
            return jsonify({'ok': False, 'mesaj': 'Satilmis stoklar silinemez (gecmis kayit korunmali)'}), 400
        if stok.durum in ('Sevkedildi', 'Teslim Edildi'):
            return jsonify({'ok': False, 'mesaj': f'"{stok.durum}" durumundaki stok silinemez (gecmis kayit korunmali)'}), 400

        # KESİM BAĞLANTISI: kaynak veya hedef olarak bir kesime girmişse silinemez
        _kesim_kaynak = Kesim.query.filter_by(kaynak_id=stok_id).first()
        if not _kesim_kaynak:
            for _k in Kesim.query.filter(Kesim.kaynak_ids_json.isnot(None)).all():
                try:
                    if stok_id in (json.loads(_k.kaynak_ids_json) or []):
                        _kesim_kaynak = _k
                        break
                except Exception:
                    pass
        _kesim_hedef = KesimDetay.query.filter_by(hedef_stok_id=stok_id).first()
        if _kesim_kaynak or _kesim_hedef:
            _kid = _kesim_kaynak.id if _kesim_kaynak else _kesim_hedef.kesim_id
            return jsonify({'ok': False, 'error': 'kesim_bagli',
                'mesaj': f'Bu stok {_kid} kesimine bagli. Once kesimi geri alin.'}), 400

        # ÖDENMİŞ BORÇ: alış borcu kapatılmışsa silme (muhasebe izi bozulur)
        _kapali = CariHareket.query.filter_by(
            baglanti_tip='stok', baglanti_id=stok_id, kapatildi=True).first()
        if _kapali:
            return jsonify({'ok': False, 'error': 'borc_kapali',
                'mesaj': 'Bu stogun alis borcu odenmis/kapatilmis. Silmek icin once '
                         'ilgili odemeyi geri alin.'}), 400

        # ── ZİNCİRLEME TEMİZLİK ──
        # Stok silinince ona bağlı mali kayıtlar öksüz kalmamalı:
        #   1) Maliyet kayıtları + her birinin cari hareketi
        #   2) Stok alışının tedarikçi cari hareketi
        #   3) İptal edilmiş rezervasyon kayıtları
        silinen_maliyet = 0
        silinen_ch = 0
        for _m in Maliyet.query.filter_by(baglanti_tip='stok', baglanti_id=stok_id).all():
            silinen_ch += CariHareket.query.filter_by(
                baglanti_tip='maliyet', baglanti_id=_m.id).delete()
            db.session.delete(_m)
            silinen_maliyet += 1
        # Alış (tedarikçi borcu) hareketi
        silinen_ch += CariHareket.query.filter_by(
            baglanti_tip='stok', baglanti_id=stok_id).delete()
        # İptal edilmiş rezervasyonlar (aktif olan zaten yukarıda engellendi)
        silinen_rez = Rezervasyon.query.filter_by(stok_id=stok_id, stok_tip=tip).delete()

        _log_audit('SIL', 'stok_' + tip.lower(), stok_id,
                   eski={'cins': stok.cins, 'durum': stok.durum,
                         'silinen_maliyet': silinen_maliyet,
                         'silinen_cari_hareket': silinen_ch})
        db.session.delete(stok)
        db.session.commit()
        logging.info(f"Stok silindi: {tip}/{stok_id} — {silinen_maliyet} maliyet, "
                     f"{silinen_ch} cari hareket, {silinen_rez} rezervasyon temizlendi "
                     f"- by {session.get('kullanici')}")
        _ek = []
        if silinen_maliyet: _ek.append(f'{silinen_maliyet} maliyet')
        if silinen_ch: _ek.append(f'{silinen_ch} cari hareket')
        if silinen_rez: _ek.append(f'{silinen_rez} rezervasyon kaydi')
        return jsonify({'ok': True,
            'silinen_maliyet': silinen_maliyet,
            'silinen_cari_hareket': silinen_ch,
            'mesaj': f'{stok_no or stok_id} silindi'
                     + (' · ' + ', '.join(_ek) + ' de temizlendi' if _ek else '')})

    @app.route('/api/stok/toplu_import', methods=['POST'])
    def api_stok_toplu_import():
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        data = request.json
        plakalar = data.get('plakalar', [])
        if not plakalar: return jsonify({'ok': False, 'mesaj': 'Plaka listesi boş'}), 400
        eklenen = 0
        for p in plakalar:
            if not p.get('boy') or not p.get('yukseklik'): continue
            boy, yuk = float(p['boy']), float(p['yukseklik'])
            if boy <= 0 or yuk <= 0: continue
            metraj_m2 = (boy * yuk) / 10000
            metraj_sqft = metraj_m2 * 10.764
            aciklama = p.get('aciklama', '') or f"Toplu import - {datetime.now().strftime('%Y-%m-%d %H:%M')}"
            plaka = PlakaStok(id=_yeni_id('PLK'), uretici=p.get('uretici', ''), cins=p.get('cins', ''), blok_no=p.get('blok_no', ''),
                              boy=boy, yukseklik=yuk, kalinlik=p.get('kalinlik'), ozellik=p.get('ozellik', ''),
                              metraj_m2=q3(metraj_m2), metraj_sqft=q3(metraj_sqft), alis_fiyati=p.get('alis_fiyati',0),
                              doviz=p.get('doviz','USD'), konum=p.get('konum',''), durum='Serbest', aciklama=aciklama, kullanici=session['kullanici'])
            db.session.add(plaka)
            eklenen += 1
        db.session.commit()
        return jsonify({'ok': True, 'eklenen': eklenen})

    # ---------- API: CARİ VE HAREKETLER ----------
    @app.route('/api/cari', methods=['GET'])
    def api_cari_liste():
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        per_page = request.args.get('per_page', type=int, default=100)
        page = request.args.get('page', type=int, default=1)
        query = Cari.query.order_by(Cari.unvan)
        paginated = query.paginate(page=page, per_page=per_page, error_out=False)
        return jsonify({'data': [{'id': c.id, 'unvan': c.unvan, 'cari_tip': c.cari_tip, 'telefon': c.telefon,
                                  'risk_limiti': c.risk_limiti, 'urun_tedarikcisi': c.urun_tedarikcisi,
                                  'uretici_kisaltma': c.uretici_kisaltma, 'email': c.email, 'adres': c.adres,
                                  # ulke: ISO3 kodu (USA, TUR...). Serializer'da YOKTU -> arayuzde hic gorunmuyordu.
                                  'ulke': c.ulke, 'para_birimi': c.para_birimi,
                                  'vergi_dairesi': c.vergi_dairesi, 'vergi_no': c.vergi_no,
                                  'yetkili': c.yetkili, 'iban': c.iban,
                                  'aciklama': c.aciklama} for c in paginated.items],
                        'meta': {'page': page, 'per_page': per_page, 'total': paginated.total}})

    @app.route('/api/cari', methods=['POST'])
    def api_cari_ekle():
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        data = request.json
        if not data.get('unvan'): return jsonify({'ok': False, 'mesaj': 'Unvan zorunlu'}), 400

        risk = data.get('risk_limiti')
        try:
            risk = float(risk) if risk not in (None, '') else None
        except (ValueError, TypeError):
            risk = None

        # cari_tip: liste (checkbox) veya string olabilir, normalize et
        cari_tip_in = data.get('cari_tip')
        if isinstance(cari_tip_in, list):
            cari_tip_str = ','.join([t for t in cari_tip_in if t])
        else:
            cari_tip_str = cari_tip_in or 'Musteri'
        if not cari_tip_str:
            cari_tip_str = 'Musteri'

        # Üretici / Tedarikçi rolü varsa kısaltma otomatik üret
        # NOT: cari_tip "Üretici", "Tedarikçi", "Tedarikci", "Uretici" - hepsini destekle
        _tip_lower = (cari_tip_str or '').lower()
        is_uretici = ('uretici' in _tip_lower) or ('üretici' in _tip_lower) or \
                     ('tedarikci' in _tip_lower) or ('tedarikçi' in _tip_lower)
        uretici_kis = data.get('uretici_kisaltma')
        if is_uretici and not uretici_kis:
            try:
                uretici_kis = _uretici_kisaltma(data['unvan'])
            except Exception:
                uretici_kis = (data['unvan'][:3] or 'XXX').upper()

        try:
            cari = Cari(id=_yeni_id('CR'), unvan=data['unvan'], email=data.get('email'), risk_limiti=risk,
                        cari_tip=cari_tip_str, telefon=data.get('telefon'), adres=data.get('adres'),
                        ulke=data.get('ulke'),
            vergi_dairesi=data.get('vergi_dairesi'),
            vergi_no=data.get('vergi_no'),
            yetkili=data.get('yetkili'),
            iban=data.get('iban'),
            aciklama=data.get('aciklama'),
                        para_birimi=data.get('para_birimi', 'USD'),
                        urun_tedarikcisi=is_uretici, uretici_kisaltma=uretici_kis)
            db.session.add(cari)
            _log_audit('EKLE', 'cari', cari.id, yeni={'unvan': cari.unvan, 'tip': cari.cari_tip})
            db.session.commit()
            return jsonify({'ok': True, 'id': cari.id, 'mesaj': 'Cari kaydedildi',
                            'uretici_kisaltma': uretici_kis if is_uretici else None})
        except Exception as e:
            db.session.rollback()
            return jsonify({'ok': False, 'mesaj': f'Kayit hatasi: {str(e)}'}), 500

    @app.route('/api/cari/<cari_id>', methods=['PUT'])
    def api_cari_guncelle(cari_id):
        """Cari guncelleme. ONCEDEN sadece unvan + uretici_kisaltma aliniyordu;
        ulke, telefon, email, adres gibi alanlar duzenlemede KAYBOLUYORDU."""
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        if _yazma_yetki_guard(): return _yazma_yetki_guard()
        c = Cari.query.get(cari_id)
        if not c:
            return jsonify({'ok': False, 'mesaj': 'Cari bulunamadi'}), 404
        data = request.json or {}
        _eski = {'unvan': c.unvan, 'cari_tip': c.cari_tip, 'ulke': c.ulke}

        for alan in ('unvan', 'cari_tip', 'ulke', 'telefon', 'email', 'adres',
                     'para_birimi', 'vergi_dairesi', 'vergi_no', 'yetkili', 'iban',
                     'uretici_kisaltma', 'aciklama'):
            if alan in data:
                d = data.get(alan)
                setattr(c, alan, (d.strip() if isinstance(d, str) else d) or None)

        if 'risk_limiti' in data:
            try:
                c.risk_limiti = float(data['risk_limiti']) if data['risk_limiti'] not in (None, '') else None
            except (TypeError, ValueError):
                pass
        if 'urun_tedarikcisi' in data:
            c.urun_tedarikcisi = bool(data['urun_tedarikcisi'])

        # Denetim izi commit'ten ÖNCE — _log_audit kendi commit'ini yapmaz
        _log_audit('GUNCELLE', 'cari', c.id, eski=_eski,
                   yeni={'unvan': c.unvan, 'cari_tip': c.cari_tip, 'ulke': c.ulke})
        db.session.commit()
        return jsonify({'ok': True, 'mesaj': 'Cari guncellendi'})

    @app.route('/api/cari/<cari_id>/tahsilat_kontrol', methods=['GET'])
    def api_cari_tahsilat_kontrol(cari_id):
        """Cari için tahsilat fazla mı kontrol.
        Query: ?tutar=X&doviz=USD (girilmek istenen tutar)
        Aynı dövizdeki açık fatura toplamı vs girilmek istenen tutar.
        """
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        cari = Cari.query.get(cari_id)
        if not cari:
            return jsonify({'ok': False, 'mesaj': 'Cari bulunamadi'}), 404

        try:
            yeni_tutar = float(request.args.get('tutar') or 0)
        except Exception:
            yeni_tutar = 0
        yeni_doviz = (request.args.get('doviz') or 'USD').upper()

        # Bu cariye ait, kesilmiş ve iptal olmayan satış faturaları (aynı dövizde)
        # Toplam fatura tutarı - toplam tahsilat = açık bakiye
        aktif_faturalar = Fatura.query.filter(
            Fatura.musteri == cari.unvan,
            Fatura.durum == 'Kesildi',
            Fatura.yon == 'satis',
            Fatura.doviz == yeni_doviz
        ).all()
        toplam_fatura = sum(f.toplam or 0 for f in aktif_faturalar)

        # Önceki tahsilatlar
        onceki_tahsilat = db.session.query(db.func.sum(CariHareket.alacak)).filter(
            CariHareket.cari_id == cari_id,
            CariHareket.doviz == yeni_doviz,
            CariHareket.alacak > 0
        ).scalar() or 0

        acik_bakiye = toplam_fatura - onceki_tahsilat

        # Fazla mı?
        fazla = max(0, yeni_tutar - acik_bakiye)
        uyari = fazla > 0.01

        return jsonify({
            'ok': True,
            'cari': cari.unvan,
            'doviz': yeni_doviz,
            'toplam_fatura': q3(toplam_fatura),
            'onceki_tahsilat': q3(onceki_tahsilat),
            'acik_bakiye': q3(acik_bakiye),
            'yeni_tahsilat': q3(yeni_tutar),
            'fazla_tutar': q3(fazla),
            'uyari': uyari,
            'mesaj': f'Açık fatura bakiyesi {acik_bakiye:,.2f} {yeni_doviz}. Girilen {yeni_tutar:,.2f} {yeni_doviz}. Fazla: {fazla:,.2f} {yeni_doviz} (avans olarak kaydedilecek).' if uyari else None
        })

    @app.route('/api/cari/<cari_id>/hareketler', methods=['GET'])
    def api_cari_hareketler(cari_id):
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        # Cari gerçekten var mı — yoksa boş liste yerine net 404 dön
        # (aksi halde silinmiş/yanlış ID sessizce boş ekstre gösterir)
        if not Cari.query.get(cari_id):
            return jsonify({'ok': False, 'mesaj': 'Cari bulunamadi'}), 404
        # Ekstre döviz parametresi: ?ekstre_doviz=TRY|USD|EUR (varsayilan ana para birimi)
        ekstre_doviz = (request.args.get('ekstre_doviz') or _ana_para_birimi()).upper()
        # Kur modu:
        #   orijinal (default): hareketin kendi tarihindeki kur (orijinal, kur_uygulanan)
        #   guncel: tüm hareketler bugünün TCMB kuru üzerinden ("kur farkı işlet")
        kur_modu = (request.args.get('kur_modu') or 'orijinal').lower()
        # Bugünün kurları (kur_modu='guncel' için)
        guncel_usd_try = _kur_getir('USD') or 1.0
        guncel_eur_try = _kur_getir('EUR') or 1.0
        hareketler = CariHareket.query.filter_by(cari_id=cari_id).order_by(
            CariHareket.hareket_tarihi.asc(), CariHareket.guncelleme.asc(), CariHareket.id.asc()).all()

        # ISLEM_TIP siralama oncelik atamasi: aynı timestamp icin mantiksal sira
        # Fatura/Acilis 1, Tahsilat/Odeme 2, Kur Farki 3 (hep sonra)
        def _islem_sirasi(h):
            t = (h.islem_tip or '').lower()
            if 'kur farki' in t or 'kur farkı' in t:
                return 3  # her zaman sonra
            if 'tahsilat' in t or 'odeme' in t or 'ödeme' in t:
                return 2  # ortada
            return 1  # fatura/acilis vs hep once
        # Stable sort - mevcut DB sirasini koruyup, sadece esit timestamp icin onceligi uygula
        from datetime import datetime as _dt
        hareketler.sort(key=lambda h: (
            h.hareket_tarihi or _dt.min.date(),
            _islem_sirasi(h),
            h.guncelleme or _dt.min,
            h.id or ''
        ))

        # ÇOK DÖVİZLİ MUHASEBE:
        # Hareketler kendi dovizlerinde + TRY karsiligi saklanir.
        # Ekstre istenen dovize gore donusturulur:
        #   1) TRY karsiligi al
        #   2) Istenen dovize cevir (guncel kur)
        # Bu yontem 1000 USD + 500 EUR = 1500 TL hatasini engeller.
        kur_hedef = _kur_getir(ekstre_doviz) if ekstre_doviz != 'TRY' else 1.0
        if not kur_hedef or kur_hedef <= 0:
            kur_hedef = 1.0  # fallback

        # Yuruyen (kumulatif) bakiye - ekstre dovizinde
        yuruyen_try = 0
        kayitlar = []
        for h in hareketler:
            # Kullanılan kur seçimi:
            #   orijinal: kayıtlı kur_uygulanan (geçmiş tarih)
            #   guncel:   bugünün TCMB kuru
            if kur_modu == 'guncel':
                if h.doviz == 'TRY':
                    h_kur = 1.0
                elif h.doviz == 'USD':
                    h_kur = guncel_usd_try
                elif h.doviz == 'EUR':
                    h_kur = guncel_eur_try
                else:
                    h_kur = _kur_getir(h.doviz) or 0
            else:
                h_kur = h.kur_uygulanan or _kur_getir(h.doviz, h.hareket_tarihi) or 0
            # Kayitli TRY karsiligi (yoksa hesapla)
            if kur_modu == 'guncel':
                # 'kur farkı işlet' → her hareket bugünün kuruyla yeniden TRY'ye çevrilir
                borc_try = q3((h.borc or 0) * h_kur) if (h.borc or 0) > 0 else 0
                alacak_try = q3((h.alacak or 0) * h_kur) if (h.alacak or 0) > 0 else 0
            else:
                borc_try = h.borc_try or 0
                alacak_try = h.alacak_try or 0
                if not borc_try and (h.borc or 0) > 0:
                    # Eski kayit - kayitli kur yoksa o tarihten cek
                    if h_kur and h_kur > 0:
                        borc_try = q3((h.borc or 0) * h_kur)
                if not alacak_try and (h.alacak or 0) > 0:
                    if h_kur and h_kur > 0:
                        alacak_try = q3((h.alacak or 0) * h_kur)

            yuruyen_try += borc_try - alacak_try

            # Ekstre dovizinde gosterim
            if ekstre_doviz == 'TRY':
                borc_ekstre = borc_try
                alacak_ekstre = alacak_try
                bakiye_ekstre = yuruyen_try
            else:
                borc_ekstre = q3(borc_try / kur_hedef)
                alacak_ekstre = q3(alacak_try / kur_hedef)
                bakiye_ekstre = q3(yuruyen_try / kur_hedef)

            kayitlar.append({
                'id': h.id,
                'hareket_tarihi': h.hareket_tarihi.isoformat() if h.hareket_tarihi else None,
                'islem_tip': h.islem_tip, 'evrak_no': h.evrak_no, 'aciklama': h.aciklama,
                # Kendi dovizinde tutar
                'borc': h.borc, 'alacak': h.alacak, 'doviz': h.doviz,
                'kur_uygulanan': h.kur_uygulanan or 0,
                # Aktif kur (kur_modu'na göre - UI'da gösterilecek)
                'kur': q_kur(h_kur),
                'kur_modu': kur_modu,
                # TRY karsiligi (ana para birimi)
                'borc_try': q3(borc_try), 'alacak_try': q3(alacak_try),
                # Ekstre dovizinde
                'borc_ekstre': borc_ekstre, 'alacak_ekstre': alacak_ekstre,
                'bakiye_ekstre': bakiye_ekstre,
                'vade_tarihi': h.vade_tarihi.isoformat() if h.vade_tarihi else None,
                'siparis_id': h.siparis_id,
                'kaynak': h.kaynak,
                'kapatildi': bool(h.kapatildi),
                # KDV ayrımı (fatura nitelikli hareketlerde dolu)
                'kdv_oran': h.kdv_oran or 0,
                'kdv_tutar': q3(h.kdv_tutar or 0),
                'matrah': q3(h.matrah or 0),
                'kdv_dahil_mi': bool(h.kdv_dahil_mi),
                # Geri uyumluluk
                'bakiye': bakiye_ekstre
            })
        kayitlar.reverse()

        # Ozet - hem TRY hem ekstre dovizinde
        def _hareket_try(h, kind='borc'):
            stored = (h.borc_try if kind == 'borc' else h.alacak_try) or 0
            if stored > 0:
                return stored
            raw = (h.borc if kind == 'borc' else h.alacak) or 0
            if raw <= 0:
                return 0
            kur = h.kur_uygulanan or _kur_getir(h.doviz, h.hareket_tarihi)
            if not kur or kur <= 0:
                return 0
            return raw * kur

        toplam_borc_try = sum(_hareket_try(h, 'borc') for h in hareketler)
        toplam_alacak_try = sum(_hareket_try(h, 'alacak') for h in hareketler)

        return jsonify({
            'data': kayitlar,
            'ekstre_doviz': ekstre_doviz,
            'kur_modu': kur_modu,
            'kur_hedef': q_kur(kur_hedef),
            'guncel_kurlar': {'USD': q_kur(guncel_usd_try), 'EUR': q_kur(guncel_eur_try)},
            'ana_para_birimi': _ana_para_birimi(),
            'ozet': {
                'borc': q3(toplam_borc_try / kur_hedef) if ekstre_doviz != 'TRY' else q3(toplam_borc_try),
                'alacak': q3(toplam_alacak_try / kur_hedef) if ekstre_doviz != 'TRY' else q3(toplam_alacak_try),
                'net': q3((toplam_borc_try - toplam_alacak_try) / kur_hedef) if ekstre_doviz != 'TRY' else q3(toplam_borc_try - toplam_alacak_try),
                'borc_try': q3(toplam_borc_try),
                'alacak_try': q3(toplam_alacak_try),
                'net_try': q3(toplam_borc_try - toplam_alacak_try),
                'hareket_sayisi': len(hareketler)
            }
        })

    @app.route('/api/cari/<cari_id>/bakiye', methods=['GET'])
    def api_cari_bakiye(cari_id):
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        hareketler = CariHareket.query.filter_by(cari_id=cari_id).all()
        borc = sum(h.borc or 0 for h in hareketler)
        alacak = sum(h.alacak or 0 for h in hareketler)
        net = borc - alacak
        return jsonify({'borc': borc, 'alacak': alacak, 'net': net})

    @app.route('/api/cari/<cari_id>', methods=['DELETE'])
    def api_cari_sil(cari_id):
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        if session.get('rol') != 'ADMIN':
            return jsonify({'ok': False, 'mesaj': 'Sadece admin cari silebilir'}), 403
        cari = Cari.query.get(cari_id)
        if not cari:
            return jsonify({'ok': False, 'mesaj': 'Cari bulunamadi'}), 404
        # Bagli kayit kontrolu
        hareket_sayisi = CariHareket.query.filter_by(cari_id=cari_id).count()
        siparis_sayisi = Siparis.query.filter_by(musteri=cari.unvan).count()
        proforma_sayisi = Proforma.query.filter_by(musteri=cari.unvan).count()

        force = request.args.get('force', '0') == '1'
        if (hareket_sayisi or siparis_sayisi or proforma_sayisi) and not force:
            return jsonify({
                'ok': False,
                'baglantili': True,
                'mesaj': f'Bu cariye bagli kayitlar var: {hareket_sayisi} hareket, {siparis_sayisi} siparis, {proforma_sayisi} proforma. Yine de silmek icin force=1 parametresi gerekli.',
                'detay': {'hareket': hareket_sayisi, 'siparis': siparis_sayisi, 'proforma': proforma_sayisi}
            }), 400

        # Cascade ile hareketler otomatik silinir (models.py)
        # Siparis/proforma musteri ismine baglanmis, onlar etkilenmez (string ref)
        _log_audit('SIL', 'cari', cari_id, eski={'unvan': cari.unvan, 'tip': cari.cari_tip})
        db.session.delete(cari)
        db.session.commit()
        logging.info(f"Cari silindi: {cari_id} - {cari.unvan} ({hareket_sayisi} hareket) - by {session.get('kullanici')}")
        return jsonify({'ok': True, 'mesaj': f'Cari silindi ({hareket_sayisi} hareket de temizlendi)'})



    @app.route('/api/diag/siparis/<siparis_id>/sil_zorla', methods=['POST', 'GET'])
    def api_diag_sip_sil_zorla(siparis_id):
        """Diagnostic: bir siparisi zorla sil (bagli kayitlar varsa uyari)."""
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        sip = Siparis.query.get(siparis_id)
        if not sip:
            return jsonify({'ok': False, 'mesaj': 'Siparis bulunamadi'}), 404

        # Bagli kayitlari say
        rez_sayi = Rezervasyon.query.filter_by(siparis_id=siparis_id).count()
        prf_sayi = Proforma.query.filter_by(siparis_id=siparis_id).count()
        ftr_sayi = Fatura.query.filter_by(siparis_id=siparis_id).count()

        if rez_sayi > 0 or prf_sayi > 0 or ftr_sayi > 0:
            return jsonify({
                'ok': False,
                'mesaj': f'Siparise bagli kayitlar var: {rez_sayi} rezervasyon, {prf_sayi} proforma, {ftr_sayi} fatura. Once bunlari farkli siparise tasiyin veya silin.',
                'rezervasyon': rez_sayi, 'proforma': prf_sayi, 'fatura': ftr_sayi
            }), 400

        db.session.delete(sip)
        db.session.commit()
        return jsonify({'ok': True, 'mesaj': f'Siparis {siparis_id} silindi'})

    @app.route('/api/diag/fatura/<fatura_id>/kdv_duzelt', methods=['POST', 'GET'])
    def api_diag_fatura_kdv_duzelt(fatura_id):
        """Fatura ara_toplam/kdv_oran/kdv_tutar/toplam degerlerini duzelt.
        ?kdv_oran=0 -> KDV sifir, ara_toplam = toplam
        ?kdv_oran=20 -> matrah ayristir
        Bagli SatisKaydi'lari ve cari hareketleri de gunceller.
        """
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        f = Fatura.query.get(fatura_id)
        if not f:
            return jsonify({'ok': False, 'mesaj': 'Fatura bulunamadi'}), 404

        # GET veya POST destekle, JSON Content-Type zorunlu degil
        body = {}
        try:
            body = request.get_json(silent=True) or {}
        except Exception:
            body = {}
        yeni_kdv_oran = float(request.args.get('kdv_oran', body.get('kdv_oran', 0)) or 0)
        yeni_satis_tipi = request.args.get('satis_tipi') or body.get('satis_tipi')

        # Mevcut toplam (KDV dahil) korunur, matrah ve kdv_tutar yeniden hesaplanir
        mevcut_toplam = f.toplam or 0
        if yeni_kdv_oran > 0:
            yeni_matrah = mevcut_toplam / (1 + yeni_kdv_oran/100)
            yeni_kdv_tutar = mevcut_toplam - yeni_matrah
        else:
            # KDV YOK: matrah = toplam, kdv_tutar = 0
            yeni_matrah = mevcut_toplam
            yeni_kdv_tutar = 0

        eski = {
            'ara_toplam': f.ara_toplam, 'kdv_oran': f.kdv_oran,
            'kdv_tutar': f.kdv_tutar, 'toplam': f.toplam, 'satis_tipi': f.satis_tipi
        }
        f.ara_toplam = q3(yeni_matrah)
        f.kdv_oran = yeni_kdv_oran
        f.kdv_tutar = q3(yeni_kdv_tutar)
        if yeni_satis_tipi:
            f.satis_tipi = yeni_satis_tipi

        # Bagli SatisKaydi'lari yeniden hesapla
        n = SatisKaydi.query.filter_by(fatura_id=fatura_id).count() or 1
        satislar = SatisKaydi.query.filter_by(fatura_id=fatura_id).all()
        matrah_pay = yeni_matrah / n
        for s in satislar:
            s.tutar = q3(matrah_pay)
            if s.miktar:
                s.birim_fiyat = q3(matrah_pay / s.miktar)
            if (f.doviz or 'USD') == 'TRY':
                s.tutar_try = q3(matrah_pay)
                kur = s.kur_usd or 1
                s.tutar_usd = q3(matrah_pay / kur) if kur else 0
            elif (f.doviz or 'USD') == 'USD':
                s.tutar_usd = q3(matrah_pay)
                kur = s.kur_usd or 1
                s.tutar_try = q3(matrah_pay * kur)
            if s.maliyet_usd is not None:
                s.kar_usd = q3((s.tutar_usd or 0) - (s.maliyet_usd or 0))
                if s.tutar_usd:
                    s.marj_yuzde = q_oran((s.kar_usd / s.tutar_usd) * 100)

        db.session.commit()
        return jsonify({
            'ok': True, 'fatura_id': fatura_id, 'eski': eski,
            'yeni': {
                'ara_toplam': f.ara_toplam, 'kdv_oran': f.kdv_oran,
                'kdv_tutar': f.kdv_tutar, 'toplam': f.toplam, 'satis_tipi': f.satis_tipi
            },
            'guncellenen_satis_kaydi': len(satislar)
        })

    @app.route('/api/diag/faturalar/tumu')
    def api_diag_faturalar_tumu():
        """Diagnostic: tum faturalarin matrah/kdv/toplam degerleri."""
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        faturalar = Fatura.query.order_by(Fatura.fatura_tarihi.desc()).all()
        return jsonify({
            'toplam': len(faturalar),
            'kayitlar': [{
                'id': f.id, 'fatura_no': f.fatura_no, 'musteri': f.musteri,
                'doviz': f.doviz, 'satis_tipi': f.satis_tipi,
                'ara_toplam': f.ara_toplam, 'kdv_oran': f.kdv_oran,
                'kdv_tutar': f.kdv_tutar, 'toplam': f.toplam,
                'durum': f.durum, 'fatura_tipi': f.fatura_tipi
            } for f in faturalar]
        })

    @app.route('/api/diag/satislar/duzelt_eksik_alanlar', methods=['POST', 'GET'])
    def api_diag_satislar_duzelt():
        """Mevcut SatisKaydi'larin eksik alanlarini + KDV dahil tutarlarini matrah'a cevir."""
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        # TUM kayitlari kontrol et (sadece NULL olanlari degil, KDV dahil olanlari da duzelt)
        kayitlar = SatisKaydi.query.all()
        duzeltilen = 0
        for s in kayitlar:
            try:
                stok = _stok_getir(s.stok_id, s.stok_tip)
                # Faturayi bul - matrah uzerinden hesap icin
                fat = None
                if s.fatura_id:
                    fat = Fatura.query.get(s.fatura_id)
                elif s.fatura_no:
                    fat = Fatura.query.filter_by(fatura_no=s.fatura_no).first()

                sip = Siparis.query.get(s.siparis_id) if s.siparis_id else None
                _k = (SiparisKalem.query.filter_by(siparis_id=s.siparis_id)
                      .order_by(SiparisKalem.sira).first()) if s.siparis_id else None
                sat_birim = (_k.birim if _k else None) or s.birim or ('ton' if s.stok_tip == 'BLOK' else 'm2')

                if stok:
                    stok_miktar = _stok_olcu(stok, sat_birim)
                    s.miktar = q3(stok_miktar)
                    s.birim = sat_birim
                    s.metraj_m2 = getattr(stok, 'metraj_m2', None)
                    s.metraj_sqft = getattr(stok, 'metraj_sqft', None)
                    s.hacim_m3 = getattr(stok, 'hacim_m3', None)
                    s.tonaj = getattr(stok, 'tonaj', None)
                    s.boy = getattr(stok, 'boy', None)
                    s.yukseklik = getattr(stok, 'yukseklik', None)
                    s.kalinlik = getattr(stok, 'kalinlik', None)

                # MATRAH uzerinden hesap: KDV'yi cikar
                if fat:
                    # n = bu faturanin satis kaydi sayisi
                    n_satis = SatisKaydi.query.filter_by(fatura_id=fat.id).count() or 1
                    matrah_toplam = fat.ara_toplam or fat.toplam or 0
                    matrah_pay = matrah_toplam / n_satis
                    s.tutar = q3(matrah_pay)
                    if stok and s.miktar:
                        s.birim_fiyat = q3(matrah_pay / s.miktar)
                    # Tutar USD ve TRY karsiliklar
                    if (fat.doviz or 'USD') == 'TRY':
                        s.tutar_try = q3(matrah_pay)
                        kur = s.kur_usd or 1
                        s.tutar_usd = q3(matrah_pay / kur) if kur else 0
                    elif (fat.doviz or 'USD') == 'USD':
                        s.tutar_usd = q3(matrah_pay)
                        kur = s.kur_usd or 1
                        s.tutar_try = q3(matrah_pay * kur)
                    # Kar yeniden hesapla
                    if s.maliyet_usd is not None:
                        s.kar_usd = q3((s.tutar_usd or 0) - (s.maliyet_usd or 0))
                        if s.tutar_usd:
                            s.marj_yuzde = q_oran((s.kar_usd / s.tutar_usd) * 100)

                duzeltilen += 1
            except Exception as e:
                app.logger.warning(f'Satis {s.id} duzeltme hatasi: {e}')
        db.session.commit()
        return jsonify({
            'ok': True,
            'kontrol_edilen': len(kayitlar),
            'duzeltilen': duzeltilen,
            'not': 'Tutarlar MATRAH (KDV haric) uzerinden duzeltildi'
        })

    @app.route('/api/diag/satislar/tumu')
    def api_diag_satislar_tumu():
        """Diagnostic: tum satis kayitlari ve miktar/tutar."""
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        kayitlar = SatisKaydi.query.order_by(SatisKaydi.satis_tarihi.desc()).limit(50).all()
        return jsonify({
            'toplam': len(kayitlar),
            'kayitlar': [{
                'id': s.id, 'stok_id': s.stok_id, 'stok_tip': s.stok_tip,
                'cins': s.cins, 'musteri': s.musteri,
                'metraj_m2': s.metraj_m2, 'hacim_m3': s.hacim_m3, 'tonaj': s.tonaj,
                'miktar': s.miktar, 'birim': s.birim,
                'birim_fiyat': s.birim_fiyat,
                'doviz': s.doviz, 'tutar': s.tutar,
                'tutar_usd': s.tutar_usd, 'tutar_try': s.tutar_try,
                'maliyet_usd': s.maliyet_usd, 'kar_usd': s.kar_usd,
                'fatura_no': s.fatura_no
            } for s in kayitlar]
        })

    @app.route('/api/diag/siparisler/olcu_temizle', methods=['POST', 'GET'])
    def api_diag_siparis_olcu_temizle():
        """Mevcut Siparis kayitlarindaki olcu metnini integer formatla.
        Or: 60.0x30.0x2.0 -> 60x30x2"""
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        import re as _re
        siparisler = Siparis.query.all()
        guncellenen = []
        for s in siparisler:
            if not s.olcu:
                continue
            # 60.0 gibi numaralari temizle
            def _temizle(m):
                v = m.group(0)
                try:
                    f = float(v)
                    return str(int(f)) if f == int(f) else str(f)
                except ValueError:
                    return v
            yeni_olcu = _re.sub(r'\d+\.\d+', _temizle, s.olcu)
            if yeni_olcu != s.olcu:
                guncellenen.append({'id': s.id, 'eski': s.olcu, 'yeni': yeni_olcu})
                s.olcu = yeni_olcu
        db.session.commit()
        return jsonify({
            'ok': True,
            'toplam_siparis': len(siparisler),
            'guncellenen': len(guncellenen),
            'ornek': guncellenen[:10]
        })

    @app.route('/api/diag/durum_otomatik_duzelt', methods=['POST', 'GET'])
    def api_diag_durum_otomatik_duzelt():
        """Sistemin durum tutarsizliklarini otomatik tespit edip duzeltir.
        Kurallar:
        1) Fatura Kesildi -> bagli siparis "Hazir" yap
        2) Fatura Iptal -> bagli siparis geri "Onaylandi", proforma geri "Onaylandi"
        3) Siparis Teslim Edildi -> bagli stoklar "Teslim Edildi"
        4) Proforma Onaylandi -> bagli siparis "Onaylandi"
        """
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401

        duzeltilen = []

        # 1) Kesilmis faturalar, bagli siparis geride
        for f in Fatura.query.filter(Fatura.durum == 'Kesildi').all():
            if f.siparis_id:
                sip = Siparis.query.get(f.siparis_id)
                if sip and sip.durum not in ('Hazir', 'Teslim Edildi', 'Iptal Edildi'):
                    eski = sip.durum
                    sip.durum = 'Hazir'
                    duzeltilen.append({
                        'tip': 'KURAL_2_SIP_HAZIR',
                        'detay': f'Siparis {sip.id}: {eski} -> Hazir (fatura {f.fatura_no} Kesildi)'
                    })
            if f.proforma_id:
                pf = Proforma.query.get(f.proforma_id)
                if pf and pf.durum not in ('Faturalandi', 'Iptal'):
                    eski = pf.durum
                    pf.durum = 'Faturalandi'
                    duzeltilen.append({
                        'tip': 'KURAL_2_PRF_FAT',
                        'detay': f'Proforma {pf.id}: {eski} -> Faturalandi (fatura {f.fatura_no} Kesildi)'
                    })

        # 2) Siparis Teslim Edildi, stoklar geride
        for sip in Siparis.query.filter(Siparis.durum == 'Teslim Edildi').all():
            rezler = Rezervasyon.query.filter_by(siparis_id=sip.id).filter(
                (Rezervasyon.iptal_nedeni.is_(None)) | (Rezervasyon.iptal_nedeni == '')
            ).all()
            for r in rezler:
                for cls in (BlokStok, PlakaStok, EbatliStok):
                    s = cls.query.get(r.stok_id)
                    if s and s.durum not in ('Teslim Edildi', 'Satildi'):
                        eski = s.durum
                        s.durum = 'Teslim Edildi'
                        duzeltilen.append({
                            'tip': 'KURAL_3_STOK_TESLIM',
                            'detay': f'Stok {r.stok_id}: {eski} -> Teslim Edildi (siparis {sip.id} teslim edildi)'
                        })
                        break

        # 3) Proforma Onaylandi, bagli siparis Teklif Asam.
        for pf in Proforma.query.filter(Proforma.durum == 'Onaylandi').all():
            if pf.siparis_id:
                sip = Siparis.query.get(pf.siparis_id)
                if sip and sip.durum == 'Teklif Asam.':
                    sip.durum = 'Onaylandi'
                    duzeltilen.append({
                        'tip': 'KURAL_4_SIP_ONAY',
                        'detay': f'Siparis {sip.id}: Teklif Asam. -> Onaylandi (proforma {pf.proforma_no} Onaylandi)'
                    })

        # 4) Iptal fatura, bagli siparis hala Hazir/Teslim Edildi
        for f in Fatura.query.filter(Fatura.durum == 'Iptal').all():
            if f.siparis_id:
                sip = Siparis.query.get(f.siparis_id)
                if sip and sip.durum in ('Hazir', 'Teslim Edildi'):
                    eski = sip.durum
                    sip.durum = 'Onaylandi'
                    duzeltilen.append({
                        'tip': 'KURAL_5_SIP_GERI',
                        'detay': f'Siparis {sip.id}: {eski} -> Onaylandi (fatura {f.fatura_no} Iptal)'
                    })

        # Kaydet
        if duzeltilen:
            ok, hata = _safe_commit('Otomatik durum duzeltme')
            if not ok:
                return jsonify({'ok': False, 'mesaj': f'Hata: {hata}'}), 500

        return jsonify({
            'ok': True,
            'duzeltilen_sayisi': len(duzeltilen),
            'duzeltilenler': duzeltilen
        })

    @app.route('/api/rapor/acente_performans', methods=['GET'])
    def api_rapor_acente_performans():
        """Acente bazlı komisyon raporu. Hangi acente ne kadar sipariş getirdi, komisyon tutarı ne."""
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401

        # Komisyonlu siparişler
        siparisler = Siparis.query.filter(
            Siparis.acente_cari_id.isnot(None),
            Siparis.acente_cari_id != ''
        ).all()

        # Acente bazında grupla
        acente_map = {}
        for s in siparisler:
            aid = s.acente_cari_id
            cari = Cari.query.get(aid)
            unvan = cari.unvan if cari else f'(Silinmiş #{aid})'
            satis_tutari = (s.satis_fiyati or 0) * (s.miktar or 0)
            kom_tutari = s.komisyon_tutar or 0
            if aid not in acente_map:
                acente_map[aid] = {
                    'acente_id': aid, 'acente_unvan': unvan,
                    'siparis_adedi': 0,
                    'toplam_satis_usd': 0,
                    'toplam_komisyon_usd': 0,
                    'siparisler': []
                }
            # USD çevirimi (kaba)
            satis_usd = satis_tutari if (s.doviz or 'USD') == 'USD' else satis_tutari  # basit, ileride kur çevirimi
            kom_usd = kom_tutari if (s.komisyon_doviz or s.doviz or 'USD') == 'USD' else kom_tutari
            acente_map[aid]['siparis_adedi'] += 1
            acente_map[aid]['toplam_satis_usd'] += satis_usd
            acente_map[aid]['toplam_komisyon_usd'] += kom_usd
            acente_map[aid]['siparisler'].append({
                'siparis_id': s.id,
                'tarih': s.siparis_tarihi.isoformat() if s.siparis_tarihi else None,
                'musteri': s.musteri,
                'cins': s.cins,
                'durum': s.durum,
                'satis_tutari': q3(satis_tutari), 'doviz': s.doviz,
                'komisyon_yontem': s.komisyon_yontem,
                'komisyon_deger': s.komisyon_deger,
                'komisyon_tutar': q3(kom_tutari),
                'komisyon_doviz': s.komisyon_doviz or s.doviz,
                'komisyon_aciklama': s.komisyon_aciklama or ''
            })

        acente_listesi = []
        for a in acente_map.values():
            a['toplam_satis_usd'] = q3(a['toplam_satis_usd'])
            a['toplam_komisyon_usd'] = q3(a['toplam_komisyon_usd'])
            a['oran'] = q_oran((a['toplam_komisyon_usd'] / a['toplam_satis_usd'] * 100) if a['toplam_satis_usd'] else 0)
            acente_listesi.append(a)
        acente_listesi.sort(key=lambda x: -x['toplam_komisyon_usd'])

        return jsonify({
            'ok': True,
            'acente_listesi': acente_listesi,
            'toplam_komisyon_usd': q3(sum(a['toplam_komisyon_usd'] for a in acente_listesi)),
            'toplam_siparis': sum(a['siparis_adedi'] for a in acente_listesi)
        })

    @app.route('/api/rapor/konteyner_takip', methods=['GET'])
    def api_rapor_konteyner_takip():
        """Konteyner bazlı sevkiyat takip raporu.
        Her konteyner için: içindeki sevkiyatlar, durum, müşteri, ürün özetleri.
        """
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401

        durum_filter = request.args.get('durum', '')  # boş = hepsi
        konteyner_no_filter = (request.args.get('konteyner_no') or '').strip().upper()

        q = Sevkiyat.query.filter(
            Sevkiyat.konteyner_no.isnot(None),
            Sevkiyat.konteyner_no != ''
        )
        if durum_filter:
            q = q.filter_by(durum=durum_filter)
        if konteyner_no_filter:
            q = q.filter(Sevkiyat.konteyner_no.ilike(f'%{konteyner_no_filter}%'))

        sevkiyatlar = q.order_by(Sevkiyat.sevk_tarihi.desc(), Sevkiyat.konteyner_no).all()

        # Konteyner bazında grupla
        konteyner_map = {}
        for s in sevkiyatlar:
            k_no = (s.konteyner_no or '').strip().upper()
            if not k_no:
                continue
            if k_no not in konteyner_map:
                konteyner_map[k_no] = {
                    'konteyner_no': k_no,
                    'sevkiyat_sayisi': 0,
                    'son_durum': s.durum or '',
                    'son_tarih': s.sevk_tarihi.isoformat() if s.sevk_tarihi else None,
                    'nakliye_firma': s.nakliye_firma or '',
                    'varis_noktasi': s.varis_noktasi or '',
                    'tah_teslim': s.tah_teslim.isoformat() if s.tah_teslim else None,
                    'gercek_teslim': s.gercek_teslim.isoformat() if s.gercek_teslim else None,
                    'musteri_listesi': set(),
                    'sevkiyatlar': []
                }
            grup = konteyner_map[k_no]
            grup['sevkiyat_sayisi'] += 1
            if s.musteri:
                grup['musteri_listesi'].add(s.musteri)

            # Sipariş bilgisi (varsa)
            siparis_info = None
            if s.siparis_id:
                sip = Siparis.query.get(s.siparis_id)
                if sip:
                    # FAZ 16: cins/urun_tip/miktar/birim/fiyat artik SiparisKalem'de.
                    # Cok kalemli siparisde ilk kalem temsili, birden fazlaysa "+N kalem".
                    _kl = SiparisKalem.query.filter_by(siparis_id=sip.id).order_by(SiparisKalem.sira).all()
                    _ilk = _kl[0] if _kl else None
                    siparis_info = {
                        'id': sip.id,
                        'cins': (_ilk.cins if _ilk else None) or '-',
                        'urun_tip': (_ilk.urun_tip if _ilk else None) or '-',
                        'miktar': sum((k.miktar or 0) for k in _kl) if _kl else 0,
                        'birim': (_ilk.birim if _ilk else None) or '',
                        'fiyat': (_ilk.birim_fiyat if _ilk else None) or 0,
                        'kalem_sayisi': len(_kl),
                        'doviz': sip.doviz
                    }

            grup['sevkiyatlar'].append({
                'id': s.id,
                'sevk_tarihi': s.sevk_tarihi.isoformat() if s.sevk_tarihi else None,
                'musteri': s.musteri,
                'siparis_id': s.siparis_id,
                'siparis_info': siparis_info,
                'cikis_noktasi': s.cikis_noktasi,
                'varis_noktasi': s.varis_noktasi,
                'durum': s.durum,
                'tah_teslim': s.tah_teslim.isoformat() if s.tah_teslim else None,
                'gercek_teslim': s.gercek_teslim.isoformat() if s.gercek_teslim else None,
                'belge_no': s.belge_no,
                'belge_tip': s.belge_tip,
                'aciklama': s.aciklama
            })

        # Set'leri liste'ye çevir
        for k in konteyner_map.values():
            k['musteri_listesi'] = sorted(list(k['musteri_listesi']))

        # Durum bazında özet
        durum_ozet = {}
        for k in konteyner_map.values():
            d = k['son_durum'] or 'Belirsiz'
            durum_ozet[d] = durum_ozet.get(d, 0) + 1

        konteyner_listesi = sorted(konteyner_map.values(),
                                    key=lambda x: x['son_tarih'] or '', reverse=True)

        return jsonify({
            'ok': True,
            'konteyner_listesi': konteyner_listesi,
            'toplam_konteyner': len(konteyner_listesi),
            'toplam_sevkiyat': sum(k['sevkiyat_sayisi'] for k in konteyner_listesi),
            'durum_ozet': durum_ozet
        })

    @app.route('/api/acente_liste', methods=['GET'])
    def api_acente_liste():
        """Acente olarak kullanılabilir cari listesi (Acente, Tedarikci veya Her ikisi tipi)."""
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        # Acente listesi: cari_tip'te 'Acente' veya 'Tedarikci' veya 'Her ikisi' geçenler
        cariler = Cari.query.filter(
            db.or_(
                Cari.cari_tip.like('%Acente%'),
                Cari.cari_tip.like('%Tedarikci%'),
                Cari.cari_tip == 'Her ikisi'
            )
        ).order_by(Cari.unvan).all()
        return jsonify({
            'ok': True,
            'acenteler': [{'id': c.id, 'unvan': c.unvan, 'ulke': c.ulke} for c in cariler]
        })

    @app.route('/api/rapor/maliyet_analiz', methods=['GET'])
    def api_rapor_maliyet_analiz():
        """Maliyet Analiz Raporu - sipariş bazlı maliyet kırılımı.
        Her teslim edilmiş sipariş için: satış, alış maliyeti, ek maliyetler,
        KDV durumu, net kâr.
        Query: ?siparis_id=X (tek sipariş detayı) veya hepsini listele
        """
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401

        tek_siparis = request.args.get('siparis_id')

        # SatisKaydi olan siparişleri al (yani teslim edilmiş/faturalanmış)
        sk_query = db.session.query(SatisKaydi.siparis_id).distinct()
        if tek_siparis:
            sk_query = sk_query.filter(SatisKaydi.siparis_id == tek_siparis)
        siparis_idler = [row[0] for row in sk_query.all() if row[0]]

        sonuc = []
        for sip_id in siparis_idler:
            sip = Siparis.query.get(sip_id)
            if not sip:
                continue

            # Bu siparise bagli satis kayitlari
            satislar = SatisKaydi.query.filter_by(siparis_id=sip_id).all()
            if not satislar:
                continue

            toplam_satis_usd = sum(s.tutar_usd or 0 for s in satislar)
            toplam_maliyet_usd = sum(s.maliyet_usd or 0 for s in satislar)
            toplam_kar_usd = sum(s.kar_usd or 0 for s in satislar)

            # Stok ID'leri (maliyet kalemlerini bulmak için)
            stok_idler = [s.stok_id for s in satislar if s.stok_id]

            # Maliyet kalemlerini topla (stok bazli + siparis bazli)
            maliyet_kalemleri = []
            if stok_idler:
                mlist = Maliyet.query.filter(
                    func.lower(Maliyet.baglanti_tip) == 'stok',
                    Maliyet.baglanti_id.in_(stok_idler),
                    Maliyet.aktif == True
                ).all()
                # maliyet_tip bazinda grupla
                tip_gruplari = {}
                for m in mlist:
                    t = m.maliyet_tip or 'Diğer'
                    if t not in tip_gruplari:
                        tip_gruplari[t] = {'tip': t, 'tutar_usd': 0, 'adet': 0}
                    tip_gruplari[t]['tutar_usd'] += m.usd_karsilik or 0
                    tip_gruplari[t]['adet'] += 1
                maliyet_kalemleri = list(tip_gruplari.values())

            # Siparis bazli maliyetler (nakliye, ek gider vb.)
            sip_maliyetleri = Maliyet.query.filter(
                Maliyet.baglanti_tip == 'siparis',
                Maliyet.baglanti_id == sip_id,
                Maliyet.aktif == True
            ).all()
            for m in sip_maliyetleri:
                t = m.maliyet_tip or 'Sipariş Gideri'
                maliyet_kalemleri.append({
                    'tip': t, 'tutar_usd': q3(m.usd_karsilik or 0), 'adet': 1
                })

            # Yuvarlama
            for mk in maliyet_kalemleri:
                mk['tutar_usd'] = q3(mk['tutar_usd'])

            marj = q_oran((toplam_kar_usd / toplam_satis_usd * 100) if toplam_satis_usd else 0)

            ilk_satis = satislar[0]
            sonuc.append({
                'siparis_id': sip_id,
                'musteri': ilk_satis.musteri or sip.musteri or '?',
                'musteri_ulke': ilk_satis.musteri_ulke or '-',
                'cins': ilk_satis.cins or '?',
                'stok_tip': ilk_satis.stok_tip or '?',
                'satis_tarihi': ilk_satis.satis_tarihi.isoformat() if ilk_satis.satis_tarihi else None,
                'stok_adedi': len(satislar),
                'satis_usd': q3(toplam_satis_usd),
                'maliyet_usd': q3(toplam_maliyet_usd),
                'kar_usd': q3(toplam_kar_usd),
                'marj': marj,
                'maliyet_kalemleri': maliyet_kalemleri
            })

        # Kâra göre sırala
        sonuc.sort(key=lambda x: -x['kar_usd'])

        # Genel toplam
        genel = {
            'siparis_adedi': len(sonuc),
            'toplam_satis_usd': q3(sum(s['satis_usd'] for s in sonuc)),
            'toplam_maliyet_usd': q3(sum(s['maliyet_usd'] for s in sonuc)),
            'toplam_kar_usd': q3(sum(s['kar_usd'] for s in sonuc))
        }
        genel['ort_marj'] = q_oran((genel['toplam_kar_usd'] / genel['toplam_satis_usd'] * 100) if genel['toplam_satis_usd'] else 0)

        return jsonify({
            'ok': True,
            'genel': genel,
            'siparisler': sonuc
        })

    @app.route('/api/rapor/satis_performans', methods=['GET'])
    def api_rapor_satis_performans():
        """Satış Performans Raporu - müşteri/ürün/ülke bazlı analiz.
        Query: ?bas_tarih=YYYY-MM-DD&bit_tarih=YYYY-MM-DD (opsiyonel tarih araligi)
        """
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401

        from datetime import date as _date_class
        bas_str = request.args.get('bas_tarih')
        bit_str = request.args.get('bit_tarih')

        q = SatisKaydi.query
        if bas_str:
            try:
                bas = _date_class.fromisoformat(bas_str)
                q = q.filter(SatisKaydi.satis_tarihi >= bas)
            except Exception:
                pass
        if bit_str:
            try:
                bit = _date_class.fromisoformat(bit_str)
                q = q.filter(SatisKaydi.satis_tarihi <= bit)
            except Exception:
                pass

        satislar = q.all()

        # Genel toplam
        genel = {
            'adet': len(satislar),
            'satis_usd': sum(s.tutar_usd or 0 for s in satislar),
            'maliyet_usd': sum(s.maliyet_usd or 0 for s in satislar),
            'kar_usd': sum(s.kar_usd or 0 for s in satislar)
        }
        genel['marj'] = q_oran((genel['kar_usd'] / genel['satis_usd'] * 100) if genel['satis_usd'] else 0)

        # Müşteri bazlı
        musteri_map = {}
        for s in satislar:
            m = s.musteri or '?'
            if m not in musteri_map:
                musteri_map[m] = {'musteri': m, 'ulke': s.musteri_ulke or '-', 'adet': 0,
                                  'satis_usd': 0, 'maliyet_usd': 0, 'kar_usd': 0}
            mm = musteri_map[m]
            mm['adet'] += 1
            mm['satis_usd'] += s.tutar_usd or 0
            mm['maliyet_usd'] += s.maliyet_usd or 0
            mm['kar_usd'] += s.kar_usd or 0

        # Ürün (cins) bazlı
        cins_map = {}
        for s in satislar:
            cn = s.cins or '?'
            if cn not in cins_map:
                cins_map[cn] = {'cins': cn, 'stok_tip': s.stok_tip or '-', 'adet': 0,
                                'satis_usd': 0, 'maliyet_usd': 0, 'kar_usd': 0}
            cm = cins_map[cn]
            cm['adet'] += 1
            cm['satis_usd'] += s.tutar_usd or 0
            cm['maliyet_usd'] += s.maliyet_usd or 0
            cm['kar_usd'] += s.kar_usd or 0

        # Ülke bazlı
        ulke_map = {}
        for s in satislar:
            u = s.musteri_ulke or 'Bilinmeyen'
            if u not in ulke_map:
                ulke_map[u] = {'ulke': u, 'adet': 0, 'satis_usd': 0, 'maliyet_usd': 0, 'kar_usd': 0}
            um = ulke_map[u]
            um['adet'] += 1
            um['satis_usd'] += s.tutar_usd or 0
            um['maliyet_usd'] += s.maliyet_usd or 0
            um['kar_usd'] += s.kar_usd or 0

        def _finalize(d):
            """Marj hesapla ve yuvarla"""
            d['marj'] = q_oran((d['kar_usd'] / d['satis_usd'] * 100) if d['satis_usd'] else 0)
            d['satis_usd'] = q3(d['satis_usd'])
            d['maliyet_usd'] = q3(d['maliyet_usd'])
            d['kar_usd'] = q3(d['kar_usd'])
            return d

        musteri_listesi = sorted([_finalize(m) for m in musteri_map.values()], key=lambda x: -x['kar_usd'])
        cins_listesi = sorted([_finalize(cm) for cm in cins_map.values()], key=lambda x: -x['kar_usd'])
        ulke_listesi = sorted([_finalize(u) for u in ulke_map.values()], key=lambda x: -x['kar_usd'])

        genel['satis_usd'] = q3(genel['satis_usd'])
        genel['maliyet_usd'] = q3(genel['maliyet_usd'])
        genel['kar_usd'] = q3(genel['kar_usd'])

        return jsonify({
            'ok': True,
            'genel': genel,
            'musteri_listesi': musteri_listesi,
            'cins_listesi': cins_listesi,
            'ulke_listesi': ulke_listesi
        })

    @app.route('/api/rapor/stok_durum', methods=['GET'])
    def api_rapor_stok_durum():
        """Stok Durum Raporu - mevcut envanter durumu.
        Cins, üretici, durum bazında gruplama + toplam değer.
        Query: ?tip=BLOK|PLAKA|EBATLI (opsiyonel), ?durum=Serbest|... (opsiyonel)
        """
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401

        tip_filtre = request.args.get('tip')
        durum_filtre = request.args.get('durum')

        # Kur bilgisi (USD karşılığı için)
        usd_kur = DovizKur.query.filter_by(doviz='USD').order_by(DovizKur.tarih.desc()).first()
        eur_kur = DovizKur.query.filter_by(doviz='EUR').order_by(DovizKur.tarih.desc()).first()
        kur_usd = (usd_kur.efektif if usd_kur and usd_kur.efektif else 45.07) or 45.07
        kur_eur = (eur_kur.efektif if eur_kur and eur_kur.efektif else 48.50) or 48.50

        def _usd(tutar, doviz):
            if not tutar:
                return 0
            d = (doviz or 'USD').upper()
            if d == 'USD':
                return tutar
            if d == 'TRY':
                return tutar / kur_usd if kur_usd else 0
            if d == 'EUR':
                return tutar * kur_eur / kur_usd if kur_usd else 0
            return tutar

        # Tüm stokları topla
        kayitlar = []  # her stok: tip, cins, uretici, durum, miktar, birim, deger_usd

        # BLOK
        if not tip_filtre or tip_filtre == 'BLOK':
            q = BlokStok.query
            if durum_filtre:
                q = q.filter(BlokStok.durum == durum_filtre)
            for s in q.all():
                miktar = s.tonaj or 0
                alis = s.alis_fiyati or 0
                # Birim fiyat * miktar
                if s.alis_fiyat_birim == 'm3':
                    deger = alis * (s.hacim_m3 or 0)
                else:
                    deger = alis * miktar
                kayitlar.append({
                    'tip': 'BLOK', 'cins': s.cins or '?', 'uretici': s.uretici or '?',
                    'durum': s.durum, 'miktar': miktar, 'birim': 'ton',
                    'deger_usd': _usd(deger, s.doviz)
                })

        # PLAKA
        if not tip_filtre or tip_filtre == 'PLAKA':
            q = PlakaStok.query
            if durum_filtre:
                q = q.filter(PlakaStok.durum == durum_filtre)
            for s in q.all():
                miktar = s.metraj_m2 or 0
                deger = (s.alis_fiyati or 0) * miktar
                kayitlar.append({
                    'tip': 'PLAKA', 'cins': s.cins or '?', 'uretici': s.uretici or '?',
                    'durum': s.durum, 'miktar': miktar, 'birim': 'm²',
                    'deger_usd': _usd(deger, s.doviz)
                })

        # EBATLI
        if not tip_filtre or tip_filtre == 'EBATLI':
            q = EbatliStok.query
            if durum_filtre:
                q = q.filter(EbatliStok.durum == durum_filtre)
            for s in q.all():
                # m2 = boy*yuk/10000 * kasa_ici_adet * adet
                m2_birim = ((s.boy or 0) * (s.yukseklik or 0) / 10000)
                kasa_ici = s.kasa_ici_adet or 1
                miktar = m2_birim * kasa_ici
                deger = (s.alis_fiyati or 0) * miktar
                kayitlar.append({
                    'tip': 'EBATLI', 'cins': s.cins or '?', 'uretici': s.uretici or '?',
                    'durum': s.durum, 'miktar': miktar, 'birim': 'm²',
                    'deger_usd': _usd(deger, s.doviz)
                })

        # Durum bazında özet
        durum_ozet = {}
        for k in kayitlar:
            d = k['durum']
            if d not in durum_ozet:
                durum_ozet[d] = {'adet': 0, 'deger_usd': 0}
            durum_ozet[d]['adet'] += 1
            durum_ozet[d]['deger_usd'] += k['deger_usd']

        # Tip bazında özet
        tip_ozet = {}
        for k in kayitlar:
            t = k['tip']
            if t not in tip_ozet:
                tip_ozet[t] = {'adet': 0, 'miktar': 0, 'deger_usd': 0, 'birim': k['birim']}
            tip_ozet[t]['adet'] += 1
            tip_ozet[t]['miktar'] += k['miktar']
            tip_ozet[t]['deger_usd'] += k['deger_usd']

        # Cins bazında özet (en detaylı)
        cins_ozet = {}
        for k in kayitlar:
            anahtar = f"{k['tip']}|{k['cins']}"
            if anahtar not in cins_ozet:
                cins_ozet[anahtar] = {
                    'tip': k['tip'], 'cins': k['cins'], 'birim': k['birim'],
                    'serbest_adet': 0, 'rezerve_adet': 0, 'satildi_adet': 0, 'teslim_adet': 0,
                    'toplam_adet': 0, 'toplam_miktar': 0, 'deger_usd': 0
                }
            co = cins_ozet[anahtar]
            co['toplam_adet'] += 1
            co['toplam_miktar'] += k['miktar']
            co['deger_usd'] += k['deger_usd']
            d = k['durum']
            if d == 'Serbest':
                co['serbest_adet'] += 1
            elif d == 'Rezerve':
                co['rezerve_adet'] += 1
            elif d == 'Satildi':
                co['satildi_adet'] += 1
            elif d == 'Teslim Edildi':
                co['teslim_adet'] += 1

        # Yuvarlama
        for d in durum_ozet.values():
            d['deger_usd'] = q3(d['deger_usd'])
        for t in tip_ozet.values():
            t['miktar'] = q3(t['miktar'])
            t['deger_usd'] = q3(t['deger_usd'])
        cins_listesi = []
        for co in cins_ozet.values():
            co['toplam_miktar'] = q3(co['toplam_miktar'])
            co['deger_usd'] = q3(co['deger_usd'])
            cins_listesi.append(co)
        cins_listesi.sort(key=lambda x: -x['deger_usd'])

        return jsonify({
            'ok': True,
            'kur_usd': q_kur(kur_usd),
            'kur_eur': q_kur(kur_eur),
            'toplam_stok_adedi': len(kayitlar),
            'toplam_deger_usd': q3(sum(k['deger_usd'] for k in kayitlar)),
            'durum_ozet': durum_ozet,
            'tip_ozet': tip_ozet,
            'cins_listesi': cins_listesi
        })

    # ─── KESİM YÖNETİMİ (BLOK→PLAKA, PLAKA→EBATLI) ─────────────────────
    # ─── KESİM SİSTEMİ ─────────────────────────────────────────────────
    @app.route('/api/diag/kesim/kaynaklar', methods=['GET'])
    def api_diag_kesim_kaynaklar():
        """DEBUG: Tum stoklarin durumlarini listele."""
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        bloklar = BlokStok.query.all()
        plakalar = PlakaStok.query.all()
        return jsonify({
            'ok': True,
            'bloklar_toplam': len(bloklar),
            'bloklar': [{'id': b.id, 'blok_no': b.blok_no, 'durum': repr(b.durum), 'hacim': b.hacim_m3, 'fiyat': b.alis_fiyati} for b in bloklar],
            'plakalar_toplam': len(plakalar),
            'plakalar': [{'id': p.id, 'durum': repr(p.durum), 'blok_no': p.blok_no, 'm2': p.metraj_m2} for p in plakalar[:50]]
        })

    @app.route('/api/kesim/kaynak_listesi', methods=['GET'])
    def api_kesim_kaynak_listesi():
        """Kesilebilir kaynakları listele (Serbest BLOK ve PLAKA)."""
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        kaynak_tip = (request.args.get('tip') or 'BLOK').upper()
        sonuc = []
        if kaynak_tip == 'BLOK':
            # Kesilebilir durumlar: Serbest (boş veya None de Serbest sayilir, geriye uyumluluk)
            bloklar = BlokStok.query.filter(
                db.or_(BlokStok.durum == 'Serbest', BlokStok.durum.is_(None), BlokStok.durum == '')
            ).all()
            app.logger.info(f'Kesim kaynak BLOK: {len(bloklar)} adet bulundu')
            for b in bloklar:
                sonuc.append({
                    'id': b.id, 'no': b.blok_no, 'cins': b.cins, 'uretici': b.uretici,
                    'mevcut': b.hacim_m3 or 0, 'birim': 'm³',
                    'birim_maliyet': b.alis_fiyati or 0, 'doviz': b.doviz,
                    'durum': b.durum, 'konum': b.konum,
                    'ozellik': '',  # BlokStok'ta ozellik alani yok
                    'tonaj': b.tonaj or 0,
                    # 3-boyutlu ölçü bilgisi (kart üzerinde gösterilir)
                    'boy': b.boy, 'yukseklik': b.yukseklik, 'en': b.en,
                    'fatura_no': getattr(b, 'fatura_no', '') or ''
                })
        elif kaynak_tip == 'PLAKA':
            plakalar = PlakaStok.query.filter(
                db.or_(PlakaStok.durum == 'Serbest', PlakaStok.durum.is_(None), PlakaStok.durum == '')
            ).all()
            app.logger.info(f'Kesim kaynak PLAKA: {len(plakalar)} adet bulundu')
            for p in plakalar:
                sonuc.append({
                    'id': p.id, 'no': p.blok_no, 'cins': p.cins, 'uretici': p.uretici,
                    'mevcut': p.metraj_m2 or 0, 'birim': 'm²',
                    'birim_maliyet': p.alis_fiyati or 0, 'doviz': p.doviz,
                    'durum': p.durum, 'konum': p.konum,
                    'boy': p.boy, 'yukseklik': p.yukseklik, 'kalinlik': p.kalinlik,
                    'slab_no': p.slab_no or '', 'blok_no': p.blok_no or '',
                    'ozellik': p.ozellik or ''
                })
        return jsonify({'ok': True, 'kaynaklar': sonuc})

    @app.route('/api/kesim', methods=['GET'])
    def api_kesim_liste():
        """Tüm kesimleri listele (detaylar dahil)."""
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        per_page = request.args.get('per_page', type=int, default=100)
        page = request.args.get('page', type=int, default=1)
        q = Kesim.query.order_by(Kesim.kesim_tarihi.desc(), Kesim.id.desc())
        paginated = q.paginate(page=page, per_page=per_page, error_out=False)
        result = []
        for k in paginated.items:
            detaylar_list = []
            try:
                for d in (k.detaylar or []):
                    detaylar_list.append({
                        'id': d.id, 'hedef_tip': d.hedef_tip, 'hedef_stok_id': d.hedef_stok_id,
                        'cins': d.cins, 'boy': d.boy, 'yukseklik': d.yukseklik,
                        'kalinlik': d.kalinlik, 'miktar_m2': d.miktar_m2, 'adet': d.adet,
                        'kasa_no': d.kasa_no, 'slab_no': d.slab_no, 'ozellik': d.ozellik,
                        'birim_maliyet': d.birim_maliyet, 'toplam_maliyet': d.toplam_maliyet
                    })
            except Exception:
                pass
            result.append({
                'id': k.id,
                'kesim_tarihi': k.kesim_tarihi.isoformat() if k.kesim_tarihi else None,
                'kaynak_tip': k.kaynak_tip, 'kaynak_id': k.kaynak_id,
                'kaynak_no': k.kaynak_no, 'kaynak_cins': k.kaynak_cins,
                'kaynak_miktar_once': k.kaynak_miktar_once or 0,
                'kaynak_miktar_sonra': k.kaynak_miktar_sonra or 0,
                'kaynak_durum': k.kaynak_durum,
                'kaynak_birim_maliyet': k.kaynak_birim_maliyet or 0,
                'kaynak_toplam_maliyet': k.kaynak_toplam_maliyet or 0,
                'kaynak_doviz': k.kaynak_doviz or 'USD',
                'uretim_blok_no': getattr(k, 'uretim_blok_no', None) or '',
                'fire_orani': k.fire_orani or 0,
                'fire_miktar': k.fire_miktar or 0,
                'detay_sayisi': len(detaylar_list),
                'detaylar': detaylar_list,
                'aciklama': k.aciklama, 'kullanici': k.kullanici
            })
        return jsonify({
            'ok': True,
            'kesimler': result,
            'meta': {'page': page, 'per_page': per_page, 'total': paginated.total}
        })

    @app.route('/api/kesim/<kesim_id>', methods=['GET'])
    def api_kesim_detay(kesim_id):
        """Kesim detayı + tüm hedef ürünler."""
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        k = Kesim.query.get(kesim_id)
        if not k:
            return jsonify({'ok': False, 'mesaj': 'Kesim bulunamadi'}), 404
        detaylar_list = [{
            'id': d.id, 'hedef_tip': d.hedef_tip, 'hedef_stok_id': d.hedef_stok_id,
            'cins': d.cins, 'boy': d.boy, 'yukseklik': d.yukseklik, 'kalinlik': d.kalinlik,
            'en': d.en, 'miktar_m2': d.miktar_m2 or 0, 'adet': d.adet,
            'kasa_no': d.kasa_no, 'slab_no': d.slab_no, 'ozellik': d.ozellik,
            'birim_maliyet': d.birim_maliyet or 0,
            'toplam_maliyet': d.toplam_maliyet or 0,
            'aciklama': d.aciklama
        } for d in (k.detaylar or [])]
        return jsonify({
            'ok': True,
            'kesim': {
                'id': k.id,
                'kesim_tarihi': k.kesim_tarihi.isoformat() if k.kesim_tarihi else None,
                'tarih': k.kesim_tarihi.isoformat() if k.kesim_tarihi else None,  # eski uyum
                'kaynak_tip': k.kaynak_tip, 'kaynak_id': k.kaynak_id,
                'kaynak_no': k.kaynak_no, 'kaynak_cins': k.kaynak_cins,
                'kaynak_miktar_once': k.kaynak_miktar_once or 0,
                'kaynak_miktar_sonra': k.kaynak_miktar_sonra or 0,
                'kaynak_durum': k.kaynak_durum,
                'kaynak_birim_maliyet': k.kaynak_birim_maliyet or 0,
                'kaynak_toplam_maliyet': k.kaynak_toplam_maliyet or 0,
                'kaynak_doviz': k.kaynak_doviz or 'USD',
                'uretim_blok_no': getattr(k, 'uretim_blok_no', None) or '',
                'fire_orani': k.fire_orani or 0,
                'fire_miktar': k.fire_miktar or 0,
                'aciklama': k.aciklama, 'kullanici': k.kullanici,
                'detaylar': detaylar_list  # KESIM OBJESININ ICINDE!
            }
        })

    @app.route('/api/kesim', methods=['POST'])
    def api_kesim_ekle():
        """Kesim islemi olustur.

        BLOK -> PLAKA/EBATLI (veya cokludan):
          - Kaynak: bir BLOK
          - Hedef: bir veya cok satir PLAKA/EBATLI
          - Her satir: ardisik aralik (basla-bitis adet) ile tek seferde N adet stok yaratir
          - Maliyet: kaynak_toplam_maliyet / hedef_toplam_m2 = birim_maliyet (fire YOK)

        PLAKA(LAR) -> EBATLI:
          - Kaynak: birden fazla PLAKA (ayni cins/blok tercihen)
          - Hedef: bir veya cok satir EBATLI
          - Maliyet: kaynak_toplam_maliyet / hedef_toplam_m2 (fire OTOMATIK: giren_m2 - cikan_m2)

        Input JSON:
        {
          "kesim_yon": "BLOK_PLAKA" | "BLOK_EBATLI" | "PLAKA_EBATLI",
          "kaynak_ids": ["BLK-001"] veya birden fazla PLK-XXX,
          "kesim_tarihi": "2026-05-30",
          "aciklama": "...",
          "hedefler": [
            {
              "hedef_tip": "PLAKA" | "EBATLI",
              "cins": "Calacatta",
              "boy": 280, "yukseklik": 180, "kalinlik": 2,
              "adet": 15,                   // bu satirdan kac plaka/kasa
              "ozellik": "1. Kalite",
              "miktar_m2_birim": 5.04       // istege bagli; bos ise boy*yukseklik/10000
              // EBATLI icin: kasa_ici_adet (1 kasada kac parca)
            }
          ]
        }
        """
        try:
            if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
            data = request.json or {}

            kesim_yon = (data.get('kesim_yon') or '').upper()
            kaynak_ids = data.get('kaynak_ids') or []
            hedefler = data.get('hedefler') or []

            if kesim_yon not in ('BLOK_PLAKA', 'BLOK_EBATLI', 'PLAKA_EBATLI'):
                return jsonify({'ok': False, 'mesaj': 'Gecersiz kesim yonu (BLOK_PLAKA, BLOK_EBATLI, PLAKA_EBATLI)'}), 400
            if not kaynak_ids:
                return jsonify({'ok': False, 'mesaj': 'En az bir kaynak stok secilmeli'}), 400
            if not hedefler:
                return jsonify({'ok': False, 'mesaj': 'En az bir hedef urun satiri eklenmelidir'}), 400

            # BLOK kaynaklar: tek BLOK olmali
            if kesim_yon in ('BLOK_PLAKA', 'BLOK_EBATLI'):
                if len(kaynak_ids) != 1:
                    return jsonify({'ok': False, 'mesaj': 'BLOK kaynaklarda sadece tek BLOK secilebilir'}), 400
                kaynak_tip_db = 'BLOK'
                kaynaklar = [BlokStok.query.get(kaynak_ids[0])]
            else:  # PLAKA_EBATLI
                kaynak_tip_db = 'PLAKA'
                kaynaklar = [PlakaStok.query.get(kid) for kid in kaynak_ids]

            if any(k is None for k in kaynaklar):
                return jsonify({'ok': False, 'mesaj': 'Bazi kaynak stoklar bulunamadi'}), 404
            # Durum kontrolü: Serbest her zaman kesilebilir.
            # Rezerve/Satildi ("müşteri için kes" senaryosu) sadece açık onayla kesilir —
            # kullanıcı bilerek onaylamalı, çünkü bu stok bir müşteriye ayrılmış.
            rezerve_onay = bool(data.get('rezerve_kesim_onay'))
            KESILEBILIR = ('Serbest', 'Rezerve', 'Satildi')
            for k in kaynaklar:
                if k.durum == 'Serbest':
                    continue
                if k.durum in ('Rezerve', 'Satildi'):
                    if not rezerve_onay:
                        return jsonify({'ok': False, 'error': 'rezerve_onay_gerekli',
                            'mesaj': f'Kaynak stok ({_baglanti_okunabilir("stok", k.id)}) "{k.durum}" durumunda — bir müşteriye ayrılmış. '
                                     f'Müşteri için kesim yapıyorsanız onaylayın.',
                            'stok_durum': k.durum}), 409
                    # Onaylandı: rezerve/satılmış stok müşteri için kesiliyor, devam
                else:
                    return jsonify({'ok': False,
                        'mesaj': f'Kaynak stok ({_baglanti_okunabilir("stok", k.id)}) kesilemez. Durum: {k.durum}'}), 400

            # Kaynak toplam miktar ve maliyet hesapla
            kaynak_toplam_maliyet = 0
            kaynak_toplam_miktar = 0  # BLOK icin m3, PLAKA icin m2
            kaynak_doviz = kaynaklar[0].doviz or 'USD'

            # Kaynaklara bağlı EK MALİYETLER (nakliye/gümrük vb.) — KDV kalemleri hariç.
            kaynak_id_listesi = [k.id for k in kaynaklar]
            ek_maliyet_map = {}
            if kaynak_id_listesi:
                for _bid, _t in db.session.query(
                    Maliyet.baglanti_id, db.func.sum(Maliyet.usd_karsilik)
                ).filter(
                    Maliyet.baglanti_id.in_(kaynak_id_listesi),
                    ~Maliyet.maliyet_tip.in_(['Devreden KDV', 'Iade KDV'])
                ).group_by(Maliyet.baglanti_id).all():
                    ek_maliyet_map[_bid] = _t or 0

            for k in kaynaklar:
                if kaynak_tip_db == 'BLOK':
                    miktar = k.hacim_m3 or 0
                else:
                    miktar = k.metraj_m2 or 0
                # GERÇEK alış bedeli: stok kaydındaki matrah (fiyat birimi ton/m3/m2 ne
                # olursa olsun doğru hesaplanmış halde). matrah yoksa eski yönteme düş.
                if k.matrah and k.matrah > 0:
                    maliyet = k.matrah
                elif kaynak_tip_db == 'BLOK' and (k.alis_fiyat_birim == 'ton') and k.tonaj:
                    maliyet = (k.tonaj or 0) * (k.alis_fiyati or 0)
                else:
                    maliyet = miktar * (k.alis_fiyati or 0)
                # Bloğa/plakaya eklenmiş ek maliyetleri (nakliye vb.) de dahil et (USD).
                maliyet += ek_maliyet_map.get(k.id, 0)
                kaynak_toplam_maliyet += maliyet
                kaynak_toplam_miktar += miktar

            # Hedef toplam m2 (her satir adet * m2_birim)
            hedef_toplam_m2 = 0
            for h in hedefler:
                try:
                    boy = float(h.get('boy') or 0)
                    yuk = float(h.get('yukseklik') or 0)
                    adet = int(h.get('adet') or 1)
                    m2_birim_str = h.get('miktar_m2_birim')
                    if m2_birim_str not in (None, ''):
                        m2_birim = float(m2_birim_str)
                    else:
                        # Otomatik: boy*yukseklik (cm) -> m2
                        m2_birim = (boy * yuk) / 10000 if boy > 0 and yuk > 0 else 0
                    if m2_birim <= 0 or adet <= 0:
                        return jsonify({'ok': False, 'mesaj': 'Hedef satirda olcu/adet sifirdan buyuk olmali'}), 400
                    hedef_toplam_m2 += m2_birim * adet
                    # form'a hesapladigi m2_birim'i geri yaz
                    h['_m2_birim'] = m2_birim
                except Exception as e:
                    return jsonify({'ok': False, 'mesaj': f'Hedef satir gecersiz: {e}'}), 400

            if hedef_toplam_m2 <= 0:
                return jsonify({'ok': False, 'mesaj': 'Toplam hedef m2 sifirdan buyuk olmali'}), 400

            # Birim maliyet: toplam kaynak maliyeti / hedef toplam m2
            # BLOK->PLAKA/EBATLI: Fire yok, basit bolme
            # PLAKA->EBATLI: Otomatik fire (giren_m2 - cikan_m2), maliyet net m2'ye bolunur (zaten cikan m2 net)
            birim_m2_maliyet = kaynak_toplam_maliyet / hedef_toplam_m2

            # Fire (otomatik)
            if kaynak_tip_db == 'PLAKA':
                # Plakada m2 cinsinden ortak birim var
                fire_miktar_m2 = kaynak_toplam_miktar - hedef_toplam_m2
                fire_orani_yuzde = (fire_miktar_m2 / kaynak_toplam_miktar * 100) if kaynak_toplam_miktar else 0
            else:
                # BLOK'ta m3->m2 cevirimi olmadigi icin fire hesabi YOK
                fire_miktar_m2 = 0
                fire_orani_yuzde = 0

            # Kesim kaydi (tek bir Kesim kaydi, birden fazla kaynak ise ilk kaynak referans)
            ana_kaynak = kaynaklar[0]
            kesim_id = _yeni_id('KSM')
            # ÜRETİM BLOK NO: kesilen bloktan üretilen plakaların yeni blok no'su.
            # Kullanıcı modal'dan girer; verilmediyse otomatik üret veya orijinali kullan.
            uretim_blok_no = (data.get('uretim_blok_no') or '').strip()
            if not uretim_blok_no:
                # Otomatik üretim no = orijinal_blok_no + '-U' (üretim) + kesim_id son 3 hane
                if kaynak_tip_db == 'BLOK' and ana_kaynak.blok_no:
                    uretim_blok_no = f"{ana_kaynak.blok_no}-U{kesim_id[-4:]}"
                else:
                    uretim_blok_no = ana_kaynak.blok_no or ana_kaynak.id

            # SLAB NUMARALANDIRMA: başlangıç no'dan başla, atlanan (eksik) no'ları atla.
            # Örn. başlangıç=1, atlanan=[3,5] → 1,2,4,6,7...
            try:
                _slab_sira = int(data.get('baslangic_slab') or 1)
                if _slab_sira < 1:
                    _slab_sira = 1
            except (ValueError, TypeError):
                _slab_sira = 1
            _atlanan_set = set()
            _atlanan_ham = data.get('atlanan_slablar') or ''
            if _atlanan_ham:
                for _parca in str(_atlanan_ham).replace(';', ',').split(','):
                    _parca = _parca.strip()
                    if _parca.isdigit():
                        _atlanan_set.add(int(_parca))
            def _sonraki_slab():
                nonlocal _slab_sira
                while _slab_sira in _atlanan_set:
                    _slab_sira += 1
                _deger = _slab_sira
                _slab_sira += 1
                return _deger

            kesim = Kesim(
                id=kesim_id,
                kesim_tarihi=_parse_date(data.get('kesim_tarihi')) or date.today(),
                kaynak_tip=kaynak_tip_db,
                kaynak_id=ana_kaynak.id,
                kaynak_ids_json=json.dumps([k.id for k in kaynaklar]),
                kaynak_no=ana_kaynak.blok_no,
                kaynak_cins=ana_kaynak.cins,
                kaynak_miktar_once=kaynak_toplam_miktar,
                kaynak_miktar_sonra=0,  # Tum kaynak tuketilir (yeni mantik)
                kaynak_durum='Tamamen',
                kaynak_birim_maliyet=ana_kaynak.alis_fiyati,
                kaynak_toplam_maliyet=kaynak_toplam_maliyet,
                kaynak_doviz=kaynak_doviz,
                uretim_blok_no=uretim_blok_no,
                fire_orani=fire_orani_yuzde,
                fire_miktar=fire_miktar_m2,
                aciklama=data.get('aciklama'),
                kullanici=session.get('kullanici')
            )
            db.session.add(kesim)

            # Otomatik kasa_no sayaci (EBATLI icin)
            # Mevcut son KSA-XXX numarasini bul
            son_ebt = db.session.query(EbatliStok.kasa_no).filter(
                EbatliStok.kasa_no.like('KSA-%')
            ).order_by(EbatliStok.kasa_no.desc()).first()
            kasa_sayac = 1
            if son_ebt and son_ebt[0]:
                try:
                    kasa_sayac = int(son_ebt[0].replace('KSA-', '')) + 1
                except Exception:
                    pass

            olusan_stoklar = []

            # Hedef stoklari yarat (her satir icin 'adet' kadar)
            for h in hedefler:
                hedef_tip = (h.get('hedef_tip') or '').upper()
                if hedef_tip not in ('PLAKA', 'EBATLI'):
                    continue

                cins = h.get('cins') or ana_kaynak.cins
                boy = float(h.get('boy') or 0)
                yukseklik = float(h.get('yukseklik') or 0)
                kalinlik = float(h.get('kalinlik') or 0)
                adet = int(h.get('adet') or 1)
                ozellik = h.get('ozellik') or getattr(ana_kaynak, 'ozellik', '') or ''
                m2_birim = h.get('_m2_birim') or 0
                kasa_ici_adet = int(h.get('kasa_ici_adet') or 1)

                if hedef_tip == 'PLAKA':
                    # PLAKA: her parca ayri bir plaka stok kaydi olur
                    for i in range(adet):
                        slab_int = _sonraki_slab()
                        etiket = f'{slab_int}/{adet}'
                        plk_id = _yeni_id('PLK')
                        plk = PlakaStok(
                            id=plk_id, cins=cins, uretici=ana_kaynak.uretici,
                            boy=boy, yukseklik=yukseklik, kalinlik=kalinlik,
                            metraj_m2=m2_birim, metraj_sqft=m2_birim * 10.7639 if m2_birim else 0,
                            blok_no=uretim_blok_no, ozellik=ozellik,
                            slab_no=slab_int,
                            alis_fiyati=birim_m2_maliyet,
                            doviz=kaynak_doviz, konum=ana_kaynak.konum,
                            durum='Serbest',
                            aciklama=f'Kesim {kesim_id} ({uretim_blok_no}-{slab_int}) | Orijinal Blok: {ana_kaynak.blok_no or ana_kaynak.id}',
                            kullanici=session.get('kullanici')
                        )
                        db.session.add(plk)
                        olusan_stoklar.append({'tip': 'PLAKA', 'id': plk_id, 'm2': m2_birim, 'etiket': f'{uretim_blok_no}-{slab_int}'})
                        kd = KesimDetay(
                            kesim_id=kesim_id, hedef_tip='PLAKA', hedef_stok_id=plk_id,
                            cins=cins, boy=boy, yukseklik=yukseklik, kalinlik=kalinlik,
                            miktar_m2=m2_birim, adet=1,
                            slab_no=str(slab_int), ozellik=ozellik,
                            birim_maliyet=birim_m2_maliyet,
                            toplam_maliyet=m2_birim * birim_m2_maliyet
                        )
                        db.session.add(kd)
                else:
                    # EBATLI: adet = toplam parca, kasa_ici_adet = bir kasada kac parca
                    # Kasa sayisi = ceil(adet / kasa_ici_adet)
                    # ornek: 6 parca, kasa_ici 10 -> 1 kasa (6 parca icinde)
                    #         18 parca, kasa_ici 10 -> 2 kasa (10 + 8)
                    if kasa_ici_adet < 1:
                        kasa_ici_adet = 1
                    import math as _math
                    kasa_sayisi = _math.ceil(adet / kasa_ici_adet)
                    kalan = adet  # dagitilacak toplam parca

                    # Tum kasalar icin tek seferde referans kodlari uret (sirali olsun)
                    try:
                        ref_kodlar = _referans_kodu_uret(
                            ana_kaynak.uretici or 'XXX',
                            cins or ana_kaynak.cins or 'XXX',
                            boy, yukseklik, kalinlik,
                            adet=kasa_sayisi
                        )
                    except Exception as ex:
                        app.logger.warning(f'Referans kodu uretilemedi: {type(ex).__name__}: {ex}')
                        ref_kodlar = []

                    for k_idx in range(kasa_sayisi):
                        # Bu kasada kac parca?
                        bu_kasa_parca = min(kasa_ici_adet, kalan)
                        kalan -= bu_kasa_parca

                        # Kasa no - referans kodu listesinden al, yoksa fallback
                        if k_idx < len(ref_kodlar) and ref_kodlar[k_idx]:
                            kasa_no = ref_kodlar[k_idx]
                        else:
                            try:
                                urt_kis = _uretici_kisaltma(ana_kaynak.uretici or 'XXX')
                                cks_kis = _cins_kisaltma(cins or 'XXX')
                                by_int = int(boy or 0)
                                yk_int = int(yukseklik or 0)
                                kk_int = int(round((kalinlik or 0) * 10))
                                kasa_no = f'{urt_kis}-{cks_kis}{by_int}{yk_int}{kk_int}A-{kasa_sayac}'
                            except Exception:
                                kasa_no = f'EBT-{kasa_sayac:04d}'
                            kasa_sayac += 1

                        kasa_m2 = m2_birim * bu_kasa_parca
                        ebt = EbatliStok(
                            id=_yeni_id('EBT'), cins=cins, uretici=ana_kaynak.uretici,
                            boy=boy, yukseklik=yukseklik, kalinlik=kalinlik,
                            metraj_m2=kasa_m2, kasa_no=kasa_no,
                            kasa_ici_adet=bu_kasa_parca,
                            ozellik=ozellik,
                            alis_fiyati=birim_m2_maliyet,
                            doviz=kaynak_doviz, konum=ana_kaynak.konum,
                            durum='Serbest',
                            aciklama=f'Kesim {kesim_id} - Üretim Blok: {uretim_blok_no} | Orijinal: {getattr(ana_kaynak, "blok_no", "") or ana_kaynak.id}',
                            kullanici=session.get('kullanici')
                        )
                        db.session.add(ebt)
                        olusan_stoklar.append({
                            'tip': 'EBATLI', 'id': ebt.id, 'kasa_no': kasa_no,
                            'kasa_ici_adet': bu_kasa_parca,
                            'm2_toplam': kasa_m2
                        })
                        kd = KesimDetay(
                            kesim_id=kesim_id, hedef_tip='EBATLI', hedef_stok_id=ebt.id,
                            cins=cins, boy=boy, yukseklik=yukseklik, kalinlik=kalinlik,
                            miktar_m2=m2_birim, adet=bu_kasa_parca,
                            kasa_no=kasa_no, ozellik=ozellik,
                            birim_maliyet=birim_m2_maliyet,
                            toplam_maliyet=kasa_m2 * birim_m2_maliyet
                        )
                        db.session.add(kd)

            # Kaynak stoklarini TUKETILDI durumuna al.
            # ÖNCE her kaynağın mevcut durumunu sakla (geri alma için) — müşteri için
            # kesilen rezerve/satılmış bloklar geri alınınca doğru duruma dönsün.
            onceki_durumlar = {}
            for k in kaynaklar:
                onceki_durumlar[k.id] = k.durum or 'Serbest'
            kesim.kaynak_onceki_durum = json.dumps(onceki_durumlar)

            for k in kaynaklar:
                k.durum = 'Tukendi'
                if kaynak_tip_db == 'BLOK':
                    k.hacim_m3 = 0
                else:
                    k.metraj_m2 = 0

            # MÜŞTERİYE REZERVE: kullanıcı üretilen ürünlerin de müşteriye rezerve
            # kalmasını seçtiyse, kaynağın rezervasyonundan müşteri/sipariş bilgisini
            # alıp üretilen her hedef stok için yeni rezervasyon açar.
            hedef_rez_sayisi = 0
            if data.get('uretilen_rezerve'):
                # Kaynağın aktif rezervasyonunu bul (müşteri/sipariş bilgisi için)
                kaynak_rez = None
                for k in kaynaklar:
                    kaynak_rez = Rezervasyon.query.filter_by(stok_id=k.id).filter(
                        (Rezervasyon.iptal_nedeni.is_(None)) | (Rezervasyon.iptal_nedeni == '')
                    ).first()
                    if kaynak_rez:
                        break
                # Rezervasyon yoksa ama kaynak Satildi/Rezerve idiyse, müşteri elle verilebilir
                musteri_ad = (kaynak_rez.musteri if kaynak_rez else None) or data.get('rezerve_musteri')
                if musteri_ad:
                    for os in olusan_stoklar:
                        h_tip = os.get('tip')
                        h_id = os.get('id')
                        if not h_id:
                            continue
                        h_stok = (BlokStok if h_tip == 'BLOK' else
                                  PlakaStok if h_tip == 'PLAKA' else EbatliStok).query.get(h_id)
                        if not h_stok:
                            continue
                        rez = Rezervasyon(
                            id=_yeni_id('REZ'),
                            musteri=musteri_ad,
                            siparis_id=(kaynak_rez.siparis_id if kaynak_rez else None),
                            siparis_kalem_id=(kaynak_rez.siparis_kalem_id if kaynak_rez else None),
                            proforma_id=(kaynak_rez.proforma_id if kaynak_rez else None),
                            stok_tip=h_tip,
                            cins=h_stok.cins,
                            ozellik=getattr(h_stok, 'ozellik', None),
                            stok_id=h_id,
                            miktar=getattr(h_stok, 'metraj_m2', None) or getattr(h_stok, 'hacim_m3', None),
                            rez_tip='Kesimden',
                            aciklama=f'Kesim {kesim_id} ile üretildi (kaynak müşteriye rezerve)',
                            kullanici=session.get('kullanici', 'sistem')
                        )
                        db.session.add(rez)
                        # Üretilen stoğun durumunu Rezerve yap (sipariş bağlıysa Satildi)
                        h_stok.durum = 'Satildi' if (kaynak_rez and kaynak_rez.siparis_id) else 'Rezerve'
                        hedef_rez_sayisi += 1

            _log_audit('KESIM', f'kesim/{kaynak_tip_db.lower()}', ana_kaynak.id,
                       yeni={'kesim_id': kesim_id, 'yon': kesim_yon,
                             'kaynak_adet': len(kaynaklar),
                             'olusan_stok_adet': len(olusan_stoklar),
                             'fire_m2': fire_miktar_m2},
                       aciklama=f'{kesim_yon}: {len(kaynaklar)} kaynak -> {len(olusan_stoklar)} hedef')

            ok, hata = _safe_commit('Kesim olusturma')
            if not ok:
                return jsonify({'ok': False, 'mesaj': f'Hata: {hata}'}), 500

            return jsonify({
                'ok': True, 'kesim_id': kesim_id,
                'olusan_stoklar': olusan_stoklar,
                'birim_maliyet': q3(birim_m2_maliyet),
                'kaynak_toplam_maliyet': q3(kaynak_toplam_maliyet),
                'hedef_toplam_m2': q3(hedef_toplam_m2),
                'fire_m2': q3(fire_miktar_m2),
                'fire_orani': q_oran(fire_orani_yuzde),
                'hedef_rez_sayisi': hedef_rez_sayisi,
                'mesaj': f'✅ Kesim tamam: {len(olusan_stoklar)} hedef stok olustu. Birim maliyet: {birim_m2_maliyet:.2f} {kaynak_doviz}/m²'
                         + (f' · {hedef_rez_sayisi} ürün müşteriye rezerve edildi.' if hedef_rez_sayisi else '')
            })
        except Exception as e:
            db.session.rollback()
            app.logger.exception('Kesim olusturma hatasi')
            return jsonify({'ok': False, 'mesaj': f'Hata: {type(e).__name__}: {str(e)}'}), 500


    @app.route('/api/kesim/<kesim_id>', methods=['DELETE'])
    def api_kesim_sil(kesim_id):
        """Kesimi geri al: hedef stoklar silinir, kaynak eski haline döner.
        Sadece hedef stoklar 'Serbest' durumda ise mümkün."""
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        k = Kesim.query.get(kesim_id)
        if not k:
            return jsonify({'ok': False, 'mesaj': 'Kesim bulunamadi'}), 404

        # Hedef stoklar serbest mi kontrol et.
        # İSTİSNA: kesimin kendi açtığı rezervasyonlar ('Kesimden') engel değil —
        # geri alırken onları otomatik iptal edip stoğu sileriz. Ama BAŞKA bir
        # rezervasyon/satış (kullanıcı sonradan eklemiş) varsa geri almayı engelleriz.
        for d in k.detaylar:
            if d.hedef_stok_id:
                if d.hedef_tip == 'PLAKA':
                    s = PlakaStok.query.get(d.hedef_stok_id)
                else:
                    s = EbatliStok.query.get(d.hedef_stok_id)
                if s and s.durum != 'Serbest':
                    # Bu stoğun rezervasyonu SADECE kesimden mi geliyor?
                    aktif_rezler = Rezervasyon.query.filter_by(stok_id=d.hedef_stok_id).filter(
                        (Rezervasyon.iptal_nedeni.is_(None)) | (Rezervasyon.iptal_nedeni == '')
                    ).all()
                    sadece_kesimden = aktif_rezler and all(r.rez_tip == 'Kesimden' for r in aktif_rezler)
                    if not sadece_kesimden:
                        return jsonify({'ok': False,
                            'mesaj': f'Hedef stok {_baglanti_okunabilir("stok", d.hedef_stok_id)} durumu "{s.durum}" ve kesim dışı bir '
                                     f'rezervasyon/satış var. Önce onu iptal edin.'}), 400

        # Hedef stokları sil (kesimden gelen rezervasyonları da temizle)
        silinen_sayi = 0
        iptal_rez = 0
        for d in k.detaylar:
            if d.hedef_stok_id:
                if d.hedef_tip == 'PLAKA':
                    s = PlakaStok.query.get(d.hedef_stok_id)
                else:
                    s = EbatliStok.query.get(d.hedef_stok_id)
                # Bu stoğa bağlı kesimden gelen rezervasyonları sil
                for r in Rezervasyon.query.filter_by(stok_id=d.hedef_stok_id, rez_tip='Kesimden').all():
                    db.session.delete(r)
                    iptal_rez += 1
                if s:
                    db.session.delete(s)
                    silinen_sayi += 1

        # Kaynakları eski haline döndür (çoklu kaynak destekli)
        kaynak_ids_listesi = []
        try:
            if k.kaynak_ids_json:
                kaynak_ids_listesi = json.loads(k.kaynak_ids_json)
        except Exception:
            pass
        if not kaynak_ids_listesi:
            kaynak_ids_listesi = [k.kaynak_id]  # fallback eski kayıtlar için

        # Kaynağın kesimden önceki durumunu oku (müşteri için kesilen rezerve/satılmış
        # bloklar körlemesine Serbest yapılmaz — eski durumuna döner).
        onceki_durumlar = {}
        try:
            if k.kaynak_onceki_durum:
                onceki_durumlar = json.loads(k.kaynak_onceki_durum)
        except Exception:
            pass
        def _geri_durum(kid):
            # Kayıtlı önceki durum varsa onu kullan, yoksa (eski kesimler) Serbest
            return onceki_durumlar.get(kid) or onceki_durumlar.get(str(kid)) or 'Serbest'

        if k.kaynak_tip == 'BLOK':
            for kid in kaynak_ids_listesi:
                kaynak = BlokStok.query.get(kid)
                if kaynak:
                    kaynak.hacim_m3 = k.kaynak_miktar_once  # tek BLOK için doğru
                    kaynak.durum = _geri_durum(kid)
        else:  # PLAKA - çoklu olabilir
            for kid in kaynak_ids_listesi:
                kaynak = PlakaStok.query.get(kid)
                if kaynak:
                    # Plakanın orijinal m²'sini boy*yuk'ten hesapla
                    if kaynak.boy and kaynak.yukseklik:
                        kaynak.metraj_m2 = (kaynak.boy * kaynak.yukseklik) / 10000
                    kaynak.durum = _geri_durum(kid)

        # Kesim ve detayları sil (cascade)
        db.session.delete(k)

        _log_audit('SIL', 'kesim', kesim_id,
                   eski={'silinen_hedef_sayi': silinen_sayi},
                   aciklama=f'Kesim geri alindi, kaynak {k.kaynak_tip} eski haline donduruldu')

        ok, hata = _safe_commit('Kesim silme')
        if not ok:
            return jsonify({'ok': False, 'mesaj': f'Hata: {hata}'}), 500
        return jsonify({'ok': True, 'mesaj': f'Kesim silindi, {silinen_sayi} hedef stok kaldirildi, kaynak geri yuklendi'})

    # ─── /KESİM SİSTEMİ ────────────────────────────────────────────────

    @app.route('/api/rapor/yaslandirilmis_cari', methods=['GET'])
    def api_rapor_yaslandirilmis_cari():
        """Yaşlandırılmış cari raporu - vadesi geçen alacaklar/borclar.
        Gruplar: Vadesi Gelmemiş, 0-30g, 31-60g, 61-90g, 90+g
        Query params:
          ?cari_tip=Musteri|Tedarikci|Her ikisi (default: hepsi)
          ?tarih=YYYY-MM-DD (default: bugün - baz alınan tarih)
        """
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401

        from datetime import date as _date_class
        baz_tarih_str = request.args.get('tarih')
        if baz_tarih_str:
            try:
                baz_tarih = _date_class.fromisoformat(baz_tarih_str)
            except Exception:
                baz_tarih = _date_class.today()
        else:
            baz_tarih = _date_class.today()

        cari_tip_filtre = request.args.get('cari_tip')

        # Tüm carileri al
        cari_q = Cari.query
        if cari_tip_filtre:
            cari_q = cari_q.filter(Cari.cari_tip == cari_tip_filtre)
        cariler = cari_q.all()

        # USD/EUR kurları (TRY bazında bakiyeler için)
        usd_kur = (DovizKur.query.filter_by(doviz='USD').order_by(DovizKur.tarih.desc()).first())
        eur_kur = (DovizKur.query.filter_by(doviz='EUR').order_by(DovizKur.tarih.desc()).first())
        kur_usd = (usd_kur.efektif if usd_kur and usd_kur.efektif else 45.07) or 45.07
        kur_eur = (eur_kur.efektif if eur_kur and eur_kur.efektif else 48.50) or 48.50

        def _try_karsilik(tutar, doviz, kur):
            """Hareket dovizini TRY karsiligina cevir"""
            if not tutar:
                return 0
            d = (doviz or 'TRY').upper()
            if d == 'TRY':
                return tutar
            k = kur or (kur_usd if d == 'USD' else kur_eur if d == 'EUR' else 1)
            return tutar * (k or 1)

        sonuc_cariler = []
        toplam_gruplar = {
            'vadesiz': {'borc': 0, 'alacak': 0, 'net': 0},
            'g_0_30': {'borc': 0, 'alacak': 0, 'net': 0},
            'g_31_60': {'borc': 0, 'alacak': 0, 'net': 0},
            'g_61_90': {'borc': 0, 'alacak': 0, 'net': 0},
            'g_90_plus': {'borc': 0, 'alacak': 0, 'net': 0}
        }

        for cari in cariler:
            hareketler = CariHareket.query.filter_by(cari_id=cari.id).all()
            if not hareketler:
                continue

            cari_gruplari = {
                'vadesiz': {'borc': 0, 'alacak': 0},
                'g_0_30': {'borc': 0, 'alacak': 0},
                'g_31_60': {'borc': 0, 'alacak': 0},
                'g_61_90': {'borc': 0, 'alacak': 0},
                'g_90_plus': {'borc': 0, 'alacak': 0}
            }

            for h in hareketler:
                borc_try = _try_karsilik(h.borc or 0, h.doviz, h.kur_uygulanan)
                alacak_try = _try_karsilik(h.alacak or 0, h.doviz, h.kur_uygulanan)

                # Vade tarihi yoksa veya henuz gelmedi -> vadesiz
                if not h.vade_tarihi or h.vade_tarihi > baz_tarih:
                    grup = 'vadesiz'
                else:
                    gun_farki = (baz_tarih - h.vade_tarihi).days
                    if gun_farki <= 30:
                        grup = 'g_0_30'
                    elif gun_farki <= 60:
                        grup = 'g_31_60'
                    elif gun_farki <= 90:
                        grup = 'g_61_90'
                    else:
                        grup = 'g_90_plus'

                cari_gruplari[grup]['borc'] += borc_try
                cari_gruplari[grup]['alacak'] += alacak_try
                toplam_gruplar[grup]['borc'] += borc_try
                toplam_gruplar[grup]['alacak'] += alacak_try

            # Cari toplam net
            toplam_borc = sum(g['borc'] for g in cari_gruplari.values())
            toplam_alacak = sum(g['alacak'] for g in cari_gruplari.values())
            net_bakiye = toplam_borc - toplam_alacak

            # Kapali cariler (net = 0) atla
            if abs(net_bakiye) < 0.01:
                continue

            # Cari net bakiyeleri grup bazında
            grup_netler = {}
            for grup_ad, gd in cari_gruplari.items():
                grup_netler[grup_ad] = gd['borc'] - gd['alacak']

            sonuc_cariler.append({
                'cari_id': cari.id,
                'unvan': cari.unvan,
                'cari_tip': cari.cari_tip,
                'ulke': cari.ulke,
                'toplam_borc_try': q3(toplam_borc),
                'toplam_alacak_try': q3(toplam_alacak),
                'net_bakiye_try': q3(net_bakiye),
                'vadesiz_try': q3(grup_netler['vadesiz']),
                'g_0_30_try': q3(grup_netler['g_0_30']),
                'g_31_60_try': q3(grup_netler['g_31_60']),
                'g_61_90_try': q3(grup_netler['g_61_90']),
                'g_90_plus_try': q3(grup_netler['g_90_plus'])
            })

        # Toplam net'ler
        for grup_ad in toplam_gruplar:
            toplam_gruplar[grup_ad]['net'] = q3(toplam_gruplar[grup_ad]['borc'] - toplam_gruplar[grup_ad]['alacak'])
            toplam_gruplar[grup_ad]['borc'] = q3(toplam_gruplar[grup_ad]['borc'])
            toplam_gruplar[grup_ad]['alacak'] = q3(toplam_gruplar[grup_ad]['alacak'])

        # Cari'leri net bakiyeye göre sırala (en yüksek borçlu üstte)
        sonuc_cariler.sort(key=lambda x: -x['net_bakiye_try'])

        return jsonify({
            'ok': True,
            'baz_tarih': baz_tarih.isoformat(),
            'kur_usd': q_kur(kur_usd),
            'kur_eur': q_kur(kur_eur),
            'toplam_gruplar': toplam_gruplar,
            'cariler': sonuc_cariler,
            'cari_sayisi': len(sonuc_cariler)
        })

    @app.route('/api/diag/durum_tutarsizliklari')
    def api_diag_durum_tutarsizlik():
        """Sistemin geneline bak ve durum tutarsizliklarini listele.
        Hangi siparislerde, proformalarda, faturalarda bagimli durumlar uyumsuz?"""
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401

        problemler = []

        # 1) Faturasi kesilmis ama siparis hala "Onaylandi" durumda olanlar
        kesilmis_faturalar = Fatura.query.filter(Fatura.durum == 'Kesildi').all()
        for f in kesilmis_faturalar:
            if not f.siparis_id:
                continue
            sip = Siparis.query.get(f.siparis_id)
            if sip and sip.durum in ('Onaylandi', 'Uretimde', 'Hazir', 'Teklif Asam.'):
                problemler.append({
                    'tip': 'FATURA_KESILDI_SIP_GERIDE',
                    'detay': f'Fatura {f.fatura_no} kesildi ama siparis {sip.id} hala "{sip.durum}"',
                    'fatura_id': f.id,
                    'siparis_id': sip.id,
                    'siparis_durum': sip.durum
                })

        # 2) Siparis "Teslim Edildi" ama bagli stoklar hala "Rezerve" veya "Serbest"
        teslim_siparisler = Siparis.query.filter(Siparis.durum == 'Teslim Edildi').all()
        for sip in teslim_siparisler:
            rezervasyonlar = Rezervasyon.query.filter_by(siparis_id=sip.id).filter(
                (Rezervasyon.iptal_nedeni.is_(None)) | (Rezervasyon.iptal_nedeni == '')
            ).all()
            for r in rezervasyonlar:
                stok = None
                for cls in (BlokStok, PlakaStok, EbatliStok):
                    s = cls.query.get(r.stok_id)
                    if s:
                        stok = s
                        break
                if stok and stok.durum not in ('Teslim Edildi', 'Satildi'):
                    problemler.append({
                        'tip': 'SIP_TESLIM_STOK_GERIDE',
                        'detay': f'Siparis {sip.id} teslim edildi ama stok {r.stok_id} hala "{stok.durum}"',
                        'siparis_id': sip.id,
                        'stok_id': r.stok_id,
                        'stok_durum': stok.durum
                    })

        # 3) Proforma "Onaylandi" ama sipariş baglantisi yok (sadece proforma var)
        onaylanmis_proformalar = Proforma.query.filter(Proforma.durum == 'Onaylandi').all()
        for p in onaylanmis_proformalar:
            if p.siparis_id:
                sip = Siparis.query.get(p.siparis_id)
                if sip and sip.durum == 'Teklif Asam.':
                    problemler.append({
                        'tip': 'PRF_ONAY_SIP_TEKLIF',
                        'detay': f'Proforma {p.proforma_no} onaylandi ama siparis {sip.id} hala "Teklif Asam."',
                        'proforma_id': p.id,
                        'siparis_id': sip.id
                    })

        # 4) Iptal edilmis fatura ama siparis hala "Hazir"
        iptal_faturalar = Fatura.query.filter(Fatura.durum == 'Iptal').all()
        for f in iptal_faturalar:
            if f.siparis_id:
                sip = Siparis.query.get(f.siparis_id)
                if sip and sip.durum in ('Hazir', 'Teslim Edildi'):
                    problemler.append({
                        'tip': 'FATURA_IPTAL_SIP_DEVAM',
                        'detay': f'Fatura {f.fatura_no} iptal ama siparis {sip.id} hala "{sip.durum}"',
                        'fatura_id': f.id,
                        'siparis_id': sip.id
                    })

        return jsonify({
            'ok': True,
            'tutarsizlik_sayisi': len(problemler),
            'problemler': problemler
        })

    @app.route('/api/diag/maliyet/tumu')
    def api_diag_maliyet_tumu():
        """Diagnostic: tum maliyet kayitlarini goster."""
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        kayitlar = Maliyet.query.order_by(Maliyet.maliyet_tarihi.desc()).all()
        return jsonify({
            'toplam': len(kayitlar),
            'aktif_sayi': sum(1 for m in kayitlar if m.aktif),
            'pasif_sayi': sum(1 for m in kayitlar if not m.aktif),
            'kayitlar': [{
                'id': m.id,
                'tarih': m.maliyet_tarihi.isoformat() if m.maliyet_tarihi else None,
                'maliyet_tip': m.maliyet_tip,
                'baglanti_tip': m.baglanti_tip,
                'baglanti_id': m.baglanti_id,
                'tutar': m.tutar, 'doviz': m.doviz,
                'try_karsilik': m.try_karsilik, 'usd_karsilik': m.usd_karsilik,
                'birim_maliyet': m.birim_maliyet,
                'fatura_no': m.fatura_no, 'aciklama': m.aciklama,
                'aktif': m.aktif,
                'donusum_id': m.donusum_id,
                'donusum_tarihi': m.donusum_tarihi.isoformat() if m.donusum_tarihi else None
            } for m in kayitlar]
        })

    @app.route('/api/diag/stoklar/tumu')
    def api_diag_stoklar_tumu():
        """Diagnostic: tum stoklari goster (BLOK + PLAKA + EBATLI)."""
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        result = {'blok': [], 'plaka': [], 'ebatli': []}
        for s in BlokStok.query.all():
            result['blok'].append({'id': s.id, 'cins': s.cins, 'durum': s.durum, 'tonaj': s.tonaj})
        for s in PlakaStok.query.all():
            result['plaka'].append({'id': s.id, 'cins': s.cins, 'durum': s.durum})
        for s in EbatliStok.query.all():
            result['ebatli'].append({'id': s.id, 'cins': s.cins, 'durum': s.durum,
                                     'boy': s.boy, 'yukseklik': s.yukseklik, 'kalinlik': s.kalinlik,
                                     'kasa_ici_adet': s.kasa_ici_adet})
        result['ozet'] = {
            'blok_sayi': len(result['blok']),
            'plaka_sayi': len(result['plaka']),
            'ebatli_sayi': len(result['ebatli']),
            'toplam': len(result['blok']) + len(result['plaka']) + len(result['ebatli'])
        }
        return jsonify(result)

    @app.route('/api/diag/siparisler/tumu')
    def api_diag_sip_tumu():
        """Diagnostic: tum siparisleri goster."""
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        sipler = Siparis.query.order_by(Siparis.siparis_tarihi.desc()).all()
        return jsonify({
            'toplam': len(sipler),
            'sipler': [{
                'id': s.id, 'musteri': s.musteri, 'cins': s.cins,
                'urun_tip': s.urun_tip, 'miktar': s.miktar,
                'satis_fiyati': s.satis_fiyati, 'doviz': s.doviz,
                'durum': s.durum,
                'siparis_tarihi': s.siparis_tarihi.isoformat() if s.siparis_tarihi else None,
                'olusturma': getattr(s, 'olusturma', None).isoformat() if getattr(s, 'olusturma', None) else None
            } for s in sipler]
        })

    @app.route('/api/diag/proforma/<proforma_id>/detay')
    def api_diag_proforma_detay(proforma_id):
        """Diagnostic: proforma kayitlarini detayli goster."""
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        p = Proforma.query.get(proforma_id)
        if not p:
            return jsonify({'ok': False, 'mesaj': 'Proforma bulunamadi'}), 404
        return jsonify({
            'id': p.id,
            'musteri': p.musteri,
            'siparis_id': p.siparis_id,
            'durum': p.durum,
            'olusturma': p.olusturma.isoformat() if p.olusturma else None
        })

    @app.route('/api/diag/proforma/<proforma_id>/siparis_bagla', methods=['POST', 'GET'])
    def api_diag_proforma_siparis_bagla(proforma_id):
        """Diagnostic: proforma'nin siparis_id'sini elle duzelt."""
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        p = Proforma.query.get(proforma_id)
        if not p:
            return jsonify({'ok': False, 'mesaj': 'Proforma bulunamadi'}), 404
        yeni_sip = request.args.get('siparis_id') or (request.json or {}).get('siparis_id')
        if not yeni_sip:
            return jsonify({'ok': False, 'mesaj': 'siparis_id parametresi gerekli'}), 400
        sip = Siparis.query.get(yeni_sip)
        if not sip:
            return jsonify({'ok': False, 'mesaj': f'Siparis bulunamadi: {yeni_sip}'}), 404
        eski = p.siparis_id
        p.siparis_id = yeni_sip
        db.session.commit()
        return jsonify({
            'ok': True, 'proforma_id': proforma_id,
            'eski_siparis_id': eski, 'yeni_siparis_id': yeni_sip,
            'mesaj': f'Proforma {proforma_id}: siparis {eski} -> {yeni_sip}'
        })

    @app.route('/api/diag/fatura/<fatura_id>/tip_duzelt', methods=['POST', 'GET'])
    def api_diag_fatura_tip_duzelt(fatura_id):
        """Fatura tipini manuel duzelt.
        ?tip=stoklu | transit | teklif"""
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        f = Fatura.query.get(fatura_id)
        if not f:
            return jsonify({'ok': False, 'mesaj': 'Fatura bulunamadi'}), 404

        yeni_tip = (request.args.get('tip') or (request.json or {}).get('tip') or '').strip()
        if yeni_tip not in ('stoklu', 'transit', 'teklif'):
            return jsonify({'ok': False, 'mesaj': 'Gecerli tip: stoklu, transit, teklif'}), 400

        eski_tip = f.fatura_tipi
        f.fatura_tipi = yeni_tip
        ok, hata = _safe_commit(f'Fatura tip duzelt: {eski_tip}->{yeni_tip}')
        if not ok:
            return jsonify({'ok': False, 'mesaj': f'Hata: {hata}'}), 500
        return jsonify({
            'ok': True,
            'fatura_id': fatura_id,
            'eski_tip': eski_tip,
            'yeni_tip': yeni_tip
        })

    @app.route('/api/diag/fatura/<fatura_id>/doviz_duzelt', methods=['POST', 'GET'])
    def api_diag_fatura_doviz_duzelt(fatura_id):
        """Diagnostic: Fatura'nin dovizini ve kur_farki_modu'nu degistir.
        ?doviz=TRY (varsayilan) - bagli cari hareketleri de gunceller."""
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        f = Fatura.query.get(fatura_id)
        if not f:
            return jsonify({'ok': False, 'mesaj': 'Fatura bulunamadi'}), 404
        yeni_doviz = (request.args.get('doviz') or (request.json or {}).get('doviz') or 'TRY').upper()
        eski_doviz = f.doviz
        f.doviz = yeni_doviz
        # Kur farki modu da otomatik
        f.kur_farki_modu = 'cari' if yeni_doviz == 'TRY' else 'gider'
        # Bagli cari hareketleri de guncelle
        guncellenen = 0
        hareketler = CariHareket.query.filter_by(
            baglanti_tip='fatura', baglanti_id=fatura_id).all()
        for h in hareketler:
            h.doviz = yeni_doviz
            if yeni_doviz == 'TRY':
                h.kur_uygulanan = 1.0
                # borc_try ve alacak_try'i sifirla, sistem yeniden hesaplayacak
                h.borc_try = h.borc or 0
                h.alacak_try = h.alacak or 0
            guncellenen += 1
        db.session.commit()
        return jsonify({
            'ok': True,
            'fatura_id': fatura_id,
            'eski_doviz': eski_doviz,
            'yeni_doviz': yeni_doviz,
            'kur_farki_modu': f.kur_farki_modu,
            'guncellenen_cari_hareket': guncellenen,
            'mesaj': f'Fatura dovizi {eski_doviz} -> {yeni_doviz}, kur_farki_modu={f.kur_farki_modu}, {guncellenen} cari hareket guncellendi'
        })

    @app.route('/api/diag/fatura/<fatura_id>/siparis_bagla', methods=['POST', 'GET'])
    def api_diag_fatura_siparis_bagla(fatura_id):
        """Diagnostic: fatura'nin siparis_id'sini elle duzelt."""
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        f = Fatura.query.get(fatura_id)
        if not f:
            return jsonify({'ok': False, 'mesaj': 'Fatura bulunamadi'}), 404
        yeni_sip = request.args.get('siparis_id') or (request.json or {}).get('siparis_id')
        if not yeni_sip:
            return jsonify({'ok': False, 'mesaj': 'siparis_id parametresi gerekli (?siparis_id=SIP-XXX)'}), 400
        sip = Siparis.query.get(yeni_sip)
        if not sip:
            return jsonify({'ok': False, 'mesaj': f'Siparis bulunamadi: {yeni_sip}'}), 404
        eski = f.siparis_id
        f.siparis_id = yeni_sip
        # Rezervasyon varsa otomatik stoklu yap
        rez = Rezervasyon.query.filter_by(siparis_id=yeni_sip, iptal_nedeni=None).first()
        if rez:
            f.fatura_tipi = 'stoklu'
        db.session.commit()
        return jsonify({
            'ok': True, 'fatura_id': fatura_id,
            'eski_siparis_id': eski, 'yeni_siparis_id': yeni_sip,
            'fatura_tipi': f.fatura_tipi,
            'mesaj': f'Fatura {fatura_id}: siparis {eski} -> {yeni_sip} (tip: {f.fatura_tipi})'
        })

    @app.route('/api/diag/rezervasyonlar/tumu')
    def api_diag_rez_tumu():
        """Diagnostic: DB'deki TUM rezervasyonlari goster."""
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        rezler = Rezervasyon.query.all()
        return jsonify({
            'toplam_sayi': len(rezler),
            'rezler': [{
                'id': r.id,
                'siparis_id': r.siparis_id,
                'stok_id': r.stok_id,
                'musteri': r.musteri,
                'stok_tip': r.stok_tip,
                'iptal_nedeni': r.iptal_nedeni,
                'olusturma': r.olusturma.isoformat() if r.olusturma else None
            } for r in rezler]
        })

    @app.route('/api/diag/fatura/<fatura_id>/tip_kontrol')
    def api_diag_fatura_tip(fatura_id):
        """Diagnostic: fatura tipi nasil tespit edilmis?"""
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        f = Fatura.query.get(fatura_id)
        if not f:
            return jsonify({'ok': False, 'mesaj': 'Fatura bulunamadi'}), 404
        rezler = []
        if f.siparis_id:
            rezler = Rezervasyon.query.filter_by(siparis_id=f.siparis_id).all()
        return jsonify({
            'fatura_id': f.id,
            'fatura_tipi_kaydedilen': f.fatura_tipi,
            'siparis_id': f.siparis_id,
            'proforma_id': f.proforma_id,
            'durum': f.durum,
            'rezervasyon_sayisi': len(rezler),
            'rezervasyonlar': [{
                'id': r.id, 'stok_id': r.stok_id, 'iptal_nedeni': r.iptal_nedeni
            } for r in rezler]
        })

    @app.route('/api/diag/cari/<cari_id>/hareketler_ham')
    def api_diag_cari_hareketler(cari_id):
        """Diagnostic: cari hareketlerinin DB'deki ham halini goster."""
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        hareketler = CariHareket.query.filter_by(cari_id=cari_id).order_by(
            CariHareket.hareket_tarihi.asc(), CariHareket.guncelleme.asc(), CariHareket.id.asc()).all()

        # ISLEM_TIP siralama oncelik atamasi: aynı timestamp icin mantiksal sira
        # Fatura/Acilis 1, Tahsilat/Odeme 2, Kur Farki 3 (hep sonra)
        def _islem_sirasi(h):
            t = (h.islem_tip or '').lower()
            if 'kur farki' in t or 'kur farkı' in t:
                return 3  # her zaman sonra
            if 'tahsilat' in t or 'odeme' in t or 'ödeme' in t:
                return 2  # ortada
            return 1  # fatura/acilis vs hep once
        # Stable sort - mevcut DB sirasini koruyup, sadece esit timestamp icin onceligi uygula
        from datetime import datetime as _dt
        hareketler.sort(key=lambda h: (
            h.hareket_tarihi or _dt.min.date(),
            _islem_sirasi(h),
            h.guncelleme or _dt.min,
            h.id or ''
        ))
        return jsonify({
            'cari_id': cari_id,
            'hareket_sayisi': len(hareketler),
            'hareketler': [{
                'id': h.id,
                'islem_tip': h.islem_tip,
                'borc': h.borc,
                'alacak': h.alacak,
                'doviz': h.doviz,
                'kur_uygulanan': h.kur_uygulanan,
                'kur_kaynak': h.kur_kaynak,
                'borc_try': h.borc_try,
                'alacak_try': h.alacak_try,
                'kapatildi': h.kapatildi,
                'kaynak': h.kaynak,
                'baglanti_tip': h.baglanti_tip,
                'baglanti_id': h.baglanti_id
            } for h in hareketler]
        })

    @app.route('/api/cari/hareket/try_yeniden_hesapla', methods=['POST'])
    def api_hareket_try_yeniden_hesapla():
        """
        TRY karsiligi 0 veya eksik olan tum hareketlerin kurunu ve TRY karsiligini
        DovizKur tablosundan yeniden hesaplar. Kur kayitlari girildikten SONRA
        bir kez calistirilir.
        """
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401

        hareketler = CariHareket.query.filter(
            (CariHareket.doviz != 'TRY') &
            ((CariHareket.borc_try == None) | (CariHareket.borc_try == 0)) &
            ((CariHareket.alacak_try == None) | (CariHareket.alacak_try == 0))
        ).all()

        guncellenen = 0
        bulunmayan = 0
        for h in hareketler:
            if not h.doviz or h.doviz == 'TRY':
                continue
            if (h.borc or 0) == 0 and (h.alacak or 0) == 0:
                continue
            kur = _kur_getir(h.doviz, h.hareket_tarihi)
            if not kur or kur <= 0:
                bulunmayan += 1
                continue
            h.kur_uygulanan = q_kur(kur)
            h.kur_kaynak = 'TCMB'
            if (h.borc or 0) > 0:
                h.borc_try = q3(h.borc * kur)
            if (h.alacak or 0) > 0:
                h.alacak_try = q3(h.alacak * kur)
            guncellenen += 1

        db.session.commit()
        msg = f'{guncellenen} hareket guncellendi'
        if bulunmayan > 0:
            msg += f', {bulunmayan} hareket icin kur bulunamadi (Doviz Kur tablosuna kur girin)'
        return jsonify({'ok': True, 'guncellenen': guncellenen,
                       'bulunmayan': bulunmayan, 'mesaj': msg})

    def _siparis_avans_bakiyesi(cari_id, siparis_id):
        """
        Belirli bir cari + sipariş için NET avans bakiyesini hesaplar.
        Avans tahsilatı (müşteriden aldığımız) → alacak; avans ödemesi (tedarikçiye verdiğimiz) → borç.
        Devir çıkışları da hesaba katılır (zaten devredilmişse kalan düşer).
        Döner: (net_avans, doviz)  — pozitif = kullanılabilir avans var
        """
        hrk = CariHareket.query.filter_by(cari_id=cari_id, siparis_id=siparis_id).filter(
            CariHareket.islem_tip.in_(['Avans Tahsilati', 'Avans Odemesi', 'Avans Devri (Giriş)', 'Avans Devri (Çıkış)'])
        ).all()
        net = 0.0
        doviz = 'USD'
        for h in hrk:
            doviz = h.doviz or doviz
            # Müşteri avansı: alacak artırır (bizim borcumuz/onun kredisi)
            # Devir çıkışı: borç (krediyi azaltır)
            net += (h.alacak or 0) - (h.borc or 0)
        return q3(net), doviz

    @app.route('/api/siparis/<siparis_id>/avans_bakiyesi', methods=['GET'])
    def api_siparis_avans_bakiyesi(siparis_id):
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        sip = Siparis.query.get(siparis_id)
        if not sip:
            return jsonify({'ok': False, 'mesaj': 'Sipariş bulunamadı'}), 404
        cari = Cari.query.filter_by(unvan=sip.musteri).first()
        if not cari:
            return jsonify({'ok': True, 'avans': 0, 'doviz': 'USD', 'cari_id': None})
        net, doviz = _siparis_avans_bakiyesi(cari.id, siparis_id)
        return jsonify({'ok': True, 'avans': abs(net), 'yon': 'alinan' if net >= 0 else 'verilen',
                        'doviz': doviz, 'cari_id': cari.id})

    @app.route('/api/avans/devret', methods=['POST'])
    def api_avans_devret():
        """
        Avansı bir siparişten diğerine devreder (Yöntem B: iki bağlı hareket).
        - Kaynak siparişe: 'Avans Devri (Çıkış)' (avansı sıfırlar yönünde)
        - Hedef siparişe:  'Avans Devri (Giriş)' (yeni siparişe avans olarak ekler)
        Net cari bakiye değişmez; sadece avansın bağlı olduğu sipariş değişir.
        Hem alınan (müşteri) hem verilen (tedarikçi) avanslar için çalışır.
        """
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        data = request.json or {}
        kaynak_sip_id = (data.get('kaynak_siparis_id') or '').strip()
        hedef_sip_id = (data.get('hedef_siparis_id') or '').strip()
        tutar = q3(float(data.get('tutar') or 0))
        if not kaynak_sip_id or not hedef_sip_id:
            return jsonify({'ok': False, 'mesaj': 'Kaynak ve hedef sipariş gerekli'}), 400
        if kaynak_sip_id == hedef_sip_id:
            return jsonify({'ok': False, 'mesaj': 'Kaynak ve hedef sipariş aynı olamaz'}), 400
        kaynak_sip = Siparis.query.get(kaynak_sip_id)
        hedef_sip = Siparis.query.get(hedef_sip_id)
        if not kaynak_sip or not hedef_sip:
            return jsonify({'ok': False, 'mesaj': 'Sipariş bulunamadı'}), 404
        # İki sipariş de aynı cariye mi ait? (müşteri avansı için müşteri aynı olmalı)
        cari = Cari.query.filter_by(unvan=kaynak_sip.musteri).first()
        if not cari:
            return jsonify({'ok': False, 'mesaj': 'Kaynak sipariş carisi bulunamadı'}), 404
        hedef_cari = Cari.query.filter_by(unvan=hedef_sip.musteri).first()
        if not hedef_cari or hedef_cari.id != cari.id:
            return jsonify({'ok': False, 'mesaj': 'İki sipariş de aynı cariye ait olmalı (avans aynı müşteri/tedarikçi içinde devredilir).'}), 400

        net, doviz = _siparis_avans_bakiyesi(cari.id, kaynak_sip_id)
        avans_yon = 'alinan' if net >= 0 else 'verilen'  # alinan: müşteri avansı, verilen: tedarikçi avansı
        mevcut = abs(net)
        if mevcut <= 0:
            return jsonify({'ok': False, 'mesaj': 'Kaynak siparişte devredilebilir avans yok'}), 400
        if tutar <= 0:
            tutar = mevcut  # tutar verilmezse tamamını devret
        if tutar > mevcut + 0.01:
            return jsonify({'ok': False, 'mesaj': f'Devir tutarı ({tutar}) mevcut avanstan ({mevcut}) fazla olamaz'}), 400

        kur_t = _kur_getir(doviz, date.today()) if doviz != 'TRY' else 1.0
        if not kur_t or kur_t <= 0:
            kur_t = 1.0
        try_kar, _ = _try_karsilik(tutar, doviz, kur_t)

        # ALINAN avans (müşteri): kaynak alacak→çıkışta borç, hedefe alacak (giriş)
        # VERİLEN avans (tedarikçi): kaynak borç→çıkışta alacak, hedefe borç (giriş)
        if avans_yon == 'alinan':
            cikis = CariHareket(id=_yeni_id('HR'), hareket_tarihi=date.today(), cari_id=cari.id,
                cari_unvan=cari.unvan, islem_tip='Avans Devri (Çıkış)',
                borc=tutar, alacak=0, doviz=doviz, kur_uygulanan=q_kur(kur_t),
                borc_try=q3(try_kar), alacak_try=0, vade_tarihi=date.today(),
                aciklama=f'Avans {hedef_sip_id} siparişine devredildi (iptal: {kaynak_sip_id})',
                kaynak='avans_devir', baglanti_tip='siparis', baglanti_id=kaynak_sip_id,
                siparis_id=kaynak_sip_id, kullanici=session['kullanici'])
            giris = CariHareket(id=_yeni_id('HR'), hareket_tarihi=date.today(), cari_id=cari.id,
                cari_unvan=cari.unvan, islem_tip='Avans Devri (Giriş)',
                borc=0, alacak=tutar, doviz=doviz, kur_uygulanan=q_kur(kur_t),
                borc_try=0, alacak_try=q3(try_kar), vade_tarihi=date.today(),
                aciklama=f'Avans {kaynak_sip_id} siparişinden devredildi',
                kaynak='avans_devir', baglanti_tip='siparis', baglanti_id=hedef_sip_id,
                siparis_id=hedef_sip_id, kullanici=session['kullanici'])
        else:
            cikis = CariHareket(id=_yeni_id('HR'), hareket_tarihi=date.today(), cari_id=cari.id,
                cari_unvan=cari.unvan, islem_tip='Avans Devri (Çıkış)',
                borc=0, alacak=tutar, doviz=doviz, kur_uygulanan=q_kur(kur_t),
                borc_try=0, alacak_try=q3(try_kar), vade_tarihi=date.today(),
                aciklama=f'Verilen avans {hedef_sip_id} siparişine devredildi (iptal: {kaynak_sip_id})',
                kaynak='avans_devir', baglanti_tip='siparis', baglanti_id=kaynak_sip_id,
                siparis_id=kaynak_sip_id, kullanici=session['kullanici'])
            giris = CariHareket(id=_yeni_id('HR'), hareket_tarihi=date.today(), cari_id=cari.id,
                cari_unvan=cari.unvan, islem_tip='Avans Devri (Giriş)',
                borc=tutar, alacak=0, doviz=doviz, kur_uygulanan=q_kur(kur_t),
                borc_try=q3(try_kar), alacak_try=0, vade_tarihi=date.today(),
                aciklama=f'Verilen avans {kaynak_sip_id} siparişinden devredildi',
                kaynak='avans_devir', baglanti_tip='siparis', baglanti_id=hedef_sip_id,
                siparis_id=hedef_sip_id, kullanici=session['kullanici'])
        db.session.add(cikis)
        db.session.add(giris)
        _log_audit('EKLE', 'avans_devir', kaynak_sip_id,
                   yeni={'hedef': hedef_sip_id, 'tutar': tutar, 'doviz': doviz})
        ok, hata = _safe_commit(f'Avans devri: {kaynak_sip_id}→{hedef_sip_id}')
        if not ok:
            return jsonify({'ok': False, 'mesaj': f'Hata: {hata}'}), 500
        return jsonify({'ok': True,
            'mesaj': f'{tutar:,.2f} {doviz} avans {kaynak_sip_id} → {hedef_sip_id} siparişine devredildi.'})

    @app.route('/api/cari/hareket', methods=['POST'])
    def api_hareket_ekle():
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        data = request.json
        if not data.get('cari_id') or not data.get('islem_tip') or not data.get('vade_tarihi'):
            return jsonify({'ok': False, 'mesaj': 'Cari, işlem tipi ve vade tarihi zorunlu'}), 400

        islem_tip = data['islem_tip']
        borc = q3(float(data.get('borc') or 0))
        alacak = q3(float(data.get('alacak') or 0))
        doviz = data.get('doviz', 'USD')
        siparis_id = data.get('siparis_id') or None
        stok_id = (data.get('stok_id') or '').strip() or None  # YENİ: stok bağlantısı
        vade = _parse_date(data['vade_tarihi'])
        evrak_no = data.get('evrak_no') or ''
        aciklama = data.get('aciklama') or ''
        # Mahsup / Bakiye Transfer için ek alanlar
        hedef_cari_id = (data.get('hedef_cari_id') or '').strip() or None
        hedef_kasa_id = data.get('hedef_kasa_id') or None
        transfer_yonu = (data.get('transfer_yonu') or 'cari').strip()  # 'cari' | 'kasa'
        mahsup_modu = (data.get('mahsup_modu') or 'cari').strip()      # 'cari' | 'fatura'
        karsilik_fatura_id = (data.get('karsilik_fatura_id') or '').strip() or None
        islem_tutar = q3(float(data.get('tutar') or data.get('borc') or data.get('alacak') or 0))

        # ═══ KDV AYRIMI (fatura nitelikli hareketlerde) ═══
        # Kullanıcı tutarı KDV dahil VEYA hariç girebilir:
        #   • kdv_dahil_mi=True  → girilen tutar genel toplamdır, matrah geriye hesaplanır
        #   • kdv_dahil_mi=False → girilen tutar matrahtır, KDV üstüne eklenir
        # borc/alacak HER ZAMAN genel toplam (KDV dahil) olarak yazılır —
        # cari bakiyesi müşterinin gerçek borcunu göstermeli.
        _fatura_tipleri = ('Fatura (Satis)', 'Fatura (Alis)',
                           'Alış Faturası', 'Satış Faturası',
                           'Alis Faturasi', 'Satis Faturasi')
        kdv_oran = q3(float(data.get('kdv_oran') or 0))
        kdv_dahil_mi = bool(data.get('kdv_dahil_mi'))
        kdv_tutar = 0.0
        matrah = 0.0
        if islem_tip in _fatura_tipleri and kdv_oran > 0:
            if kdv_dahil_mi:
                # Girilen tutar KDV dahil → matrahı ayır
                matrah = q3(islem_tutar / (1 + kdv_oran / 100.0))
                kdv_tutar = q3(islem_tutar - matrah)
                # borc/alacak zaten genel toplam, değişmez
            else:
                # Girilen tutar matrah → KDV'yi üste ekle, toplamı büyüt
                matrah = islem_tutar
                kdv_tutar = q3(matrah * kdv_oran / 100.0)
                islem_tutar = q3(matrah + kdv_tutar)
                if borc > 0:
                    borc = islem_tutar
                if alacak > 0:
                    alacak = islem_tutar
        elif islem_tip in _fatura_tipleri:
            # KDV'siz fatura (ihracat vb.) — matrah = toplam
            matrah = islem_tutar

        # ═══ FATURASIZ STOK FATURALANDIRMA (cari modülünden) ═══
        # İşlem tipi alış faturası + seçili stok FATURASIZ ise: cari modülünden
        # çıkmadan o stoğu faturalandır. Borç, faturalandırma akışında (doğru kur +
        # fatura tarihiyle) otomatik oluşur — burada ayrıca cari hareket eklenmez (mükerrer olmaz).
        _alis_fatura_tipleri = ('Fatura (Alis)', 'Alış Faturası', 'Alis Faturasi')
        if stok_id and islem_tip in _alis_fatura_tipleri:
            # Stoğun tipini ve fatura durumunu bul
            _hedef_stok = None
            _hedef_tip = None
            for _tip, _model in (('BLOK', BlokStok), ('PLAKA', PlakaStok), ('EBATLI', EbatliStok)):
                # grup_id formatları: BLOK→id, PLAKA→"BLOK:no", EBATLI→"EBATLI_GRP:prefix"
                if _tip == 'BLOK':
                    _s = _model.query.get(stok_id)
                    if _s:
                        _hedef_stok, _hedef_tip = _s, 'BLOK'
                        break
            # PLAKA/EBATLI grup ise: gruptaki ilk faturasız stoğu bul
            if not _hedef_stok and stok_id.startswith('BLOK:'):
                _bno = stok_id.split(':', 1)[1]
                _hedef_stok = PlakaStok.query.filter_by(blok_no=_bno, fatura_durumu='faturasiz').first()
                _hedef_tip = 'PLAKA'
            if not _hedef_stok and stok_id.startswith('EBATLI_GRP:'):
                _pref = stok_id.split(':', 1)[1]
                _adaylar = EbatliStok.query.filter_by(fatura_durumu='faturasiz').all()
                for _e in _adaylar:
                    _ekod = (_e.kasa_no or '')
                    if _ekod.startswith(_pref):
                        _hedef_stok, _hedef_tip = _e, 'EBATLI'
                        break

            if _hedef_stok and getattr(_hedef_stok, 'fatura_durumu', 'faturali') == 'faturasiz':
                # Bu bir faturalandırma: tüm gruptaki faturasız stokları faturalandır
                _fatura_tarihi = _parse_date(data.get('fatura_tarihi')) or vade or date.today()
                if stok_id.startswith('BLOK:'):
                    _grup = PlakaStok.query.filter_by(blok_no=stok_id.split(':',1)[1], fatura_durumu='faturasiz').all()
                elif stok_id.startswith('EBATLI_GRP:'):
                    _pref = stok_id.split(':',1)[1]
                    _grup = [e for e in EbatliStok.query.filter_by(fatura_durumu='faturasiz').all()
                             if (e.kasa_no or '').startswith(_pref)]
                else:
                    _grup = [_hedef_stok]
                _toplam_borc = 0
                for _st in _grup:
                    _st.fatura_no = evrak_no or _st.fatura_no
                    _st.alis_tarihi = _fatura_tarihi
                    _st.fatura_durumu = 'faturali'
                    if (_st.kdv_tutar or 0) > 0:
                        _devreden_kdv_kalemi_olustur(_st.id, _st.kdv_tutar, _st.doviz,
                            aciklama=f'{_hedef_tip} faturalandırma - {_st.cins}',
                            fatura_no=_st.fatura_no)
                    _ch = _stok_cari_hareket_olustur(_st.id, _st.uretici, q3((_st.matrah or 0)+(_st.kdv_tutar or 0)),
                        _st.doviz, fatura_no=evrak_no,
                        aciklama=f'{_hedef_tip} faturalandırma — {_st.cins} — {evrak_no}',
                        fatura_durumu='faturali', alis_tarihi=_fatura_tarihi)
                    if _ch:
                        _toplam_borc += (_ch.alacak or 0)
                _log_audit('GUNCELLE', 'stok_faturalandir_cari', stok_id,
                           yeni={'fatura_no': evrak_no, 'adet': len(_grup)})
                ok, hata = _safe_commit(f'Cari modülünden faturalandırma: {stok_id}')
                if not ok:
                    return jsonify({'ok': False, 'mesaj': f'Hata: {hata}'}), 500
                return jsonify({'ok': True,
                    'faturalandirma': True,
                    'mesaj': f'{len(_grup)} stok faturalandırıldı. Tedarikçi cariye {_toplam_borc:,.2f} {doviz} borç işlendi (fatura tarihi: {_fatura_tarihi.strftime("%d.%m.%Y")}).'})

        # ═══ BAKİYE TRANSFER (VİRMAN) ═══
        # Cari→Cari veya Cari→Kasa bakiye taşıma. Çift taraflı kayıt.
        if islem_tip == 'Bakiye Transfer':
            if islem_tutar <= 0:
                return jsonify({'ok': False, 'mesaj': 'Transfer tutarı gerekli (Tutar alanı).'}), 400
            kaynak = Cari.query.get(data['cari_id'])
            if not kaynak:
                return jsonify({'ok': False, 'mesaj': 'Kaynak cari bulunamadı'}), 404
            kur_t = _kur_getir(doviz, vade) if doviz != 'TRY' else 1.0
            if not kur_t or kur_t <= 0:
                kur_t = 1.0
            try_kar, _ = _try_karsilik(islem_tutar, doviz, kur_t)

            if transfer_yonu == 'kasa':
                # Cari → Kasa: cariden çıkış (borç azalır = alacak hareketi), kasaya giriş
                if not hedef_kasa_id:
                    return jsonify({'ok': False, 'mesaj': 'Hedef kasa seçilmeli'}), 400
                kasa = Kasa.query.get(int(hedef_kasa_id))
                if not kasa:
                    return jsonify({'ok': False, 'mesaj': 'Hedef kasa bulunamadı'}), 404
                # Cari tarafı: ödeme benzeri (cariye borcumuzu kapatıyoruz gibi → borç)
                h1 = CariHareket(
                    id=_yeni_id('HR'), hareket_tarihi=date.today(), cari_id=kaynak.id,
                    cari_unvan=kaynak.unvan, islem_tip='Bakiye Transfer (Kasaya)',
                    borc=islem_tutar, alacak=0, doviz=doviz, kur_uygulanan=q_kur(kur_t),
                    borc_try=q3(try_kar), alacak_try=0, vade_tarihi=vade,
                    evrak_no=evrak_no, aciklama=aciklama or f'Kasaya transfer: {kasa.ad}',
                    kaynak='virman', baglanti_tip='kasa', baglanti_id=str(kasa.id),
                    kullanici=session['kullanici'])
                db.session.add(h1)
                # Kasa tarafı: giriş hareketi.
                # DÖVİZ ÇEVRİMİ: işlem dövizi ile kasa dövizi farklıysa tutar,
                # TRY köprüsü üzerinden HEDEF KASANIN DÖVİZİNE çevrilir.
                # (Eski kod TRY karşılığını doğrudan yabancı para kasasına ekliyordu —
                #  10.000 EUR kasaya 5.000 USD transferi ~235.000 "EUR" yapıyordu.)
                if kasa.doviz == doviz:
                    hedef_tutar = islem_tutar
                elif kasa.doviz == 'TRY':
                    hedef_tutar = q3(try_kar)
                else:
                    kasa_kur = _kur_getir(kasa.doviz, vade)
                    if not kasa_kur or kasa_kur <= 0:
                        return jsonify({'ok': False,
                            'mesaj': f'{kasa.doviz} kuru alınamadı — farklı dövizli kasaya '
                                     f'transfer için kur gerekli. TCMB kurlarını güncelleyip tekrar dene.'}), 400
                    hedef_tutar = q3(try_kar / kasa_kur)
                kasa.bakiye = q3((kasa.bakiye or 0) + hedef_tutar)
                cevrim_notu = '' if kasa.doviz == doviz else f' ({islem_tutar:,.2f} {doviz} → {hedef_tutar:,.2f} {kasa.doviz})'
                kh = KasaHareket(kasa_id=kasa.id, tip='giris', tutar=hedef_tutar,
                    tarih=date.today(),
                    aciklama=(f'{kaynak.unvan} cari bakiye transferi' + (f' ({evrak_no})' if evrak_no else '') + cevrim_notu),
                    baglanti_tip='virman', baglanti_id=kaynak.id, cari_id=kaynak.id,
                    kullanici=session.get('kullanici'))
                db.session.add(kh)
                _log_audit('EKLE', 'virman_cari_kasa', kaynak.id,
                           yeni={'kasa': kasa.id, 'tutar': islem_tutar, 'kasa_tutar': hedef_tutar, 'kasa_doviz': kasa.doviz})
                ok, hata = _safe_commit(f'Virman cari→kasa: {kaynak.id}')
                if not ok:
                    return jsonify({'ok': False, 'mesaj': f'Hata: {hata}'}), 500
                return jsonify({'ok': True,
                    'mesaj': f'{islem_tutar:,.2f} {doviz} {kasa.ad} kasasına transfer edildi.{cevrim_notu}'})
            else:
                # Cari → Cari: kaynaktan çık, hedefe gir.
                # transfer_tip = 'alacak' (varsayılan): carinin bizden ALACAĞI taşınır
                #                → kaynağa borç (alacağı kapanır), hedefe alacak.
                # transfer_tip = 'borc': carinin bize BORCU taşınır
                #                → kaynağa alacak (borcu kapanır), hedefe borç.
                transfer_tip = (data.get('transfer_tip') or 'alacak').strip()
                if not hedef_cari_id or hedef_cari_id == kaynak.id:
                    return jsonify({'ok': False, 'mesaj': 'Geçerli bir hedef cari seçilmeli'}), 400
                hedef = Cari.query.get(hedef_cari_id)
                if not hedef:
                    return jsonify({'ok': False, 'mesaj': 'Hedef cari bulunamadı'}), 404
                borc_transferi = (transfer_tip == 'borc')
                tur_ek = 'borç' if borc_transferi else 'alacak'
                etiket_sip = (data.get('siparis_id') or '').strip() or None  # kaynak sipariş etiketi (ops.)
                # Kaynak cari: alacak transferinde borç yazılır (alacağı kapanır);
                # borç transferinde alacak yazılır (borcu kapanır).
                h1 = CariHareket(
                    id=_yeni_id('HR'), hareket_tarihi=date.today(), cari_id=kaynak.id,
                    cari_unvan=kaynak.unvan, islem_tip='Bakiye Transfer (Çıkış)',
                    borc=0 if borc_transferi else islem_tutar,
                    alacak=islem_tutar if borc_transferi else 0,
                    doviz=doviz, kur_uygulanan=q_kur(kur_t),
                    borc_try=0 if borc_transferi else q3(try_kar),
                    alacak_try=q3(try_kar) if borc_transferi else 0,
                    vade_tarihi=vade,
                    evrak_no=evrak_no,
                    aciklama=aciklama or f'{hedef.unvan} cariye {tur_ek} transferi',
                    kaynak='virman', baglanti_tip='cari', baglanti_id=hedef.id,
                    siparis_id=etiket_sip,
                    kullanici=session['kullanici'])
                db.session.add(h1)
                # Hedef cari: taşınan tür aynen hedefte doğar.
                h2 = CariHareket(
                    id=_yeni_id('HR'), hareket_tarihi=date.today(), cari_id=hedef.id,
                    cari_unvan=hedef.unvan, islem_tip='Bakiye Transfer (Giriş)',
                    borc=islem_tutar if borc_transferi else 0,
                    alacak=0 if borc_transferi else islem_tutar,
                    doviz=doviz, kur_uygulanan=q_kur(kur_t),
                    borc_try=q3(try_kar) if borc_transferi else 0,
                    alacak_try=0 if borc_transferi else q3(try_kar),
                    vade_tarihi=vade,
                    evrak_no=evrak_no,
                    aciklama=aciklama or f'{kaynak.unvan} cariden {tur_ek} transferi',
                    kaynak='virman', baglanti_tip='cari', baglanti_id=kaynak.id,
                    kullanici=session['kullanici'])
                db.session.add(h2)
                _log_audit('EKLE', 'virman_cari_cari', kaynak.id,
                           yeni={'hedef': hedef.id, 'tutar': islem_tutar, 'tip': transfer_tip})
                ok, hata = _safe_commit(f'Virman cari→cari: {kaynak.id}→{hedef.id}')
                if not ok:
                    return jsonify({'ok': False, 'mesaj': f'Hata: {hata}'}), 500
                return jsonify({'ok': True,
                    'mesaj': f'{islem_tutar:,.2f} {doviz} {tur_ek} bakiyesi {kaynak.unvan} → {hedef.unvan} transfer edildi.'})

        # ═══ MAHSUP ═══
        # Mod 'cari': bir carinin alacağını başka carinin borcuna say (iki cari arası netleştirme)
        # Mod 'fatura': belirli faturaya karşılık mahsup (faturayı kapatır)
        if islem_tip == 'Mahsup':
            if islem_tutar <= 0:
                return jsonify({'ok': False, 'mesaj': 'Mahsup tutarı gerekli (Tutar alanı).'}), 400
            kaynak = Cari.query.get(data['cari_id'])
            if not kaynak:
                return jsonify({'ok': False, 'mesaj': 'Cari bulunamadı'}), 404
            kur_t = _kur_getir(doviz, vade) if doviz != 'TRY' else 1.0
            if not kur_t or kur_t <= 0:
                kur_t = 1.0
            try_kar, _ = _try_karsilik(islem_tutar, doviz, kur_t)

            if mahsup_modu == 'fatura':
                # Belirli faturaya karşılık mahsup → faturayı kapatır gibi cariye karşı kayıt
                if not karsilik_fatura_id:
                    return jsonify({'ok': False, 'mesaj': 'Mahsup edilecek fatura seçilmeli'}), 400
                fat = Fatura.query.get(karsilik_fatura_id)
                if not fat:
                    return jsonify({'ok': False, 'mesaj': 'Fatura bulunamadı'}), 404
                # Faturanın yönüne göre ters kayıt: satış faturası ise alacak (tahsilat gibi), alış ise borç
                if (fat.yon or 'satis') == 'satis':
                    h = CariHareket(id=_yeni_id('HR'), hareket_tarihi=date.today(), cari_id=kaynak.id,
                        cari_unvan=kaynak.unvan, islem_tip='Mahsup (Fatura)',
                        borc=0, alacak=islem_tutar, doviz=doviz, kur_uygulanan=q_kur(kur_t),
                        borc_try=0, alacak_try=q3(try_kar), vade_tarihi=vade,
                        evrak_no=evrak_no or fat.fatura_no,
                        aciklama=aciklama or f'Fatura {fat.fatura_no or fat.id} mahsubu',
                        kaynak='mahsup', baglanti_tip='fatura', baglanti_id=fat.id,
                        kullanici=session['kullanici'])
                else:
                    h = CariHareket(id=_yeni_id('HR'), hareket_tarihi=date.today(), cari_id=kaynak.id,
                        cari_unvan=kaynak.unvan, islem_tip='Mahsup (Fatura)',
                        borc=islem_tutar, alacak=0, doviz=doviz, kur_uygulanan=q_kur(kur_t),
                        borc_try=q3(try_kar), alacak_try=0, vade_tarihi=vade,
                        evrak_no=evrak_no or fat.fatura_no,
                        aciklama=aciklama or f'Fatura {fat.fatura_no or fat.id} mahsubu',
                        kaynak='mahsup', baglanti_tip='fatura', baglanti_id=fat.id,
                        kullanici=session['kullanici'])
                db.session.add(h)
                db.session.flush()
                _fatura_tahsilat_durumu(fat.id)
                _log_audit('EKLE', 'mahsup_fatura', kaynak.id, yeni={'fatura': fat.id, 'tutar': islem_tutar})
                ok, hata = _safe_commit(f'Mahsup fatura: {fat.id}')
                if not ok:
                    return jsonify({'ok': False, 'mesaj': f'Hata: {hata}'}), 500
                return jsonify({'ok': True, 'mesaj': f'{islem_tutar:,.2f} {doviz} fatura {fat.fatura_no or fat.id} mahsup edildi.'})
            else:
                # İki cari arası: kaynak carinin alacağını hedef carinin borcuna say
                if not hedef_cari_id or hedef_cari_id == kaynak.id:
                    return jsonify({'ok': False, 'mesaj': 'Karşılık cari seçilmeli'}), 400
                hedef = Cari.query.get(hedef_cari_id)
                if not hedef:
                    return jsonify({'ok': False, 'mesaj': 'Karşılık cari bulunamadı'}), 404
                # Kaynak: alacak (alacağı azalır), Hedef: borç (borcu azalır)
                h1 = CariHareket(id=_yeni_id('HR'), hareket_tarihi=date.today(), cari_id=kaynak.id,
                    cari_unvan=kaynak.unvan, islem_tip='Mahsup (Karşılıklı)',
                    borc=0, alacak=islem_tutar, doviz=doviz, kur_uygulanan=q_kur(kur_t),
                    borc_try=0, alacak_try=q3(try_kar), vade_tarihi=vade,
                    evrak_no=evrak_no, aciklama=aciklama or f'{hedef.unvan} ile mahsup',
                    kaynak='mahsup', baglanti_tip='cari', baglanti_id=hedef.id,
                    kullanici=session['kullanici'])
                db.session.add(h1)
                h2 = CariHareket(id=_yeni_id('HR'), hareket_tarihi=date.today(), cari_id=hedef.id,
                    cari_unvan=hedef.unvan, islem_tip='Mahsup (Karşılıklı)',
                    borc=islem_tutar, alacak=0, doviz=doviz, kur_uygulanan=q_kur(kur_t),
                    borc_try=q3(try_kar), alacak_try=0, vade_tarihi=vade,
                    evrak_no=evrak_no, aciklama=aciklama or f'{kaynak.unvan} ile mahsup',
                    kaynak='mahsup', baglanti_tip='cari', baglanti_id=kaynak.id,
                    kullanici=session['kullanici'])
                db.session.add(h2)
                _log_audit('EKLE', 'mahsup_karsilikli', kaynak.id, yeni={'hedef': hedef.id, 'tutar': islem_tutar})
                ok, hata = _safe_commit(f'Mahsup karşılıklı: {kaynak.id}↔{hedef.id}')
                if not ok:
                    return jsonify({'ok': False, 'mesaj': f'Hata: {hata}'}), 500
                return jsonify({'ok': True, 'mesaj': f'{islem_tutar:,.2f} {doviz} {kaynak.unvan} ↔ {hedef.unvan} mahsup edildi.'})

        # FATURA AKIŞI: 'Fatura (Satis)' veya 'Fatura (Alis)' seçildiyse
        # otomatik Fatura kaydı oluştur + Faturalar sayfasında "Kesildi" durumunda görünsün.
        fatura_id = None
        fatura_yon = None
        if islem_tip in ('Fatura (Satis)', 'Fatura (Alis)'):
            fatura_yon = 'satis' if 'Satis' in islem_tip else 'alis'
            # Cari kartin bizim defterimiz oldugu muhasebe perspektifi:
            # Fatura (Satis) -> musteri bize borclanir -> BORC alani dolar
            # Fatura (Alis) -> biz tedarikciye borcluyuz -> ALACAK alani dolar
            tutar = borc if fatura_yon == 'satis' else alacak
            if tutar <= 0:
                return jsonify({'ok': False,
                    'mesaj': f"{islem_tip} icin tutar gerekli (Satis -> Borc, Alis -> Alacak)"}), 400

            # Cariden musteri unvani
            cari = Cari.query.get(data['cari_id'])
            if not cari:
                return jsonify({'ok': False, 'mesaj': 'Cari bulunamadi'}), 404

            # Siparis baglantisi varsa fatura tipi (stoklu/transit) tespiti
            fatura_tipi = 'transit'
            siparis_obj = None
            if siparis_id:
                siparis_obj = Siparis.query.get(siparis_id)
                rez_var = Rezervasyon.query.filter_by(siparis_id=siparis_id).first()
                if rez_var:
                    fatura_tipi = 'stoklu'

            fatura = Fatura(
                id=_yeni_id('FTR'),
                fatura_no=evrak_no,
                fatura_tarihi=date.today(),
                vade_tarihi=vade,
                siparis_id=siparis_id,
                musteri=cari.unvan,
                musteri_adres=getattr(cari, 'adres', None),
                musteri_ulke=getattr(cari, 'ulke', None) if hasattr(cari, 'ulke') else None,
                toplam=q3(tutar),
                ara_toplam=q3(tutar),
                kdv_tutar=0,
                kdv_oran=0,
                doviz=doviz,
                durum='Kesildi',  # Otomatik Kesildi
                yon=fatura_yon,
                fatura_tipi=fatura_tipi,
                aciklama=aciklama or f'Cari hareketten olusturuldu ({islem_tip})',
                kullanici=session['kullanici']
            )
            db.session.add(fatura)
            db.session.flush()  # fatura.id elde edilsin
            fatura_id = fatura.id

        # ÇOK DÖVİZLİ MUHASEBE: kur ve TRY karşılığı
        manuel_kur = data.get('kur_uygulanan')
        kur_kaynak = 'MANUEL' if manuel_kur and float(manuel_kur) > 0 else 'TCMB'
        if doviz == 'TRY':
            kullanilan_kur = 1.0
        elif manuel_kur and float(manuel_kur) > 0:
            kullanilan_kur = float(manuel_kur)
        else:
            kullanilan_kur = _kur_getir(doviz)
            # TCMB kuru bulunamadıysa hata - kullanıcı manuel girmeli ya da Doviz Kur'a kayıt eklemeli
            if not kullanilan_kur or kullanilan_kur <= 0:
                return jsonify({'ok': False,
                    'mesaj': f"{doviz} icin kur bulunamadi. Ya Islem Kuru alanina kuru elle girin, "
                            f"ya da Dashboard'dan 'Kuru Guncelle' ile TCMB'den cekin."}), 400
        borc_try, _ = _try_karsilik(borc, doviz, kullanilan_kur)
        alacak_try, _ = _try_karsilik(alacak, doviz, kullanilan_kur)

        # Tahsilat/odeme islem tipleri (acıklama ve kur farkı için)
        tahsilat_odeme_tipleri = ['Tahsilat', 'Avans Tahsilati', 'Odeme', 'Avans Odemesi']

        # AÇIKLAMA: Kullanıcı yazmadıysa, otomatik üret (sipariş/fatura bağlantısı varsa)
        if not aciklama or not aciklama.strip():
            parcalar = []
            if islem_tip in tahsilat_odeme_tipleri:
                parcalar.append(islem_tip)
            else:
                parcalar.append(islem_tip)
            if siparis_id:
                parcalar.append(f'Siparis {siparis_id}')
            if fatura_id:
                parcalar.append(f'Fatura {fatura_id}')
            elif stok_id:
                parcalar.append(f'Stok {stok_id}')
            if evrak_no:
                parcalar.append(f'(Evrak: {evrak_no})')
            aciklama = ' - '.join(parcalar)

        # CariHareket kaydi olustur
        hareket = CariHareket(
            id=_yeni_id('HR'), hareket_tarihi=date.today(),
            cari_id=data['cari_id'], islem_tip=islem_tip,
            evrak_no=evrak_no, aciklama=aciklama,
            borc=borc, alacak=alacak, doviz=doviz,
            kur_uygulanan=q_kur(kullanilan_kur),
            kur_kaynak=kur_kaynak,
            borc_try=q3(borc_try), alacak_try=q3(alacak_try),
            vade_tarihi=vade, siparis_id=siparis_id,
            kdv_dahil_mi=kdv_dahil_mi, kdv_oran=kdv_oran,
            kdv_tutar=q3(kdv_tutar), matrah=q3(matrah),
            kullanici=session['kullanici']
        )
        # Fatura olusturulduysa cari hareketi ona bagla (cift kayit korumasi icin)
        if fatura_id:
            hareket.baglanti_tip = 'fatura'
            hareket.baglanti_id = fatura_id
            hareket.kaynak = 'cari_fatura'
        elif stok_id:
            # Stok'a bagli cari hareketi (orn. tedarikciye alis faturasi karsiligi)
            hareket.baglanti_tip = 'stok'
            hareket.baglanti_id = stok_id

        db.session.add(hareket)
        db.session.flush()

        # Fatura tarafinda cari_hareket_id'yi bagla (geri donus icin)
        if fatura_id:
            fatura.cari_hareket_id = hareket.id

        # ═══ KASA ENTEGRASYONU (çift taraflı kayıt) ═══
        # Tahsilat türlerinde para kasaya GİRER, ödeme türlerinde kasadan ÇIKAR.
        # Döviz uyuşmazlığında tutar TRY köprüsüyle kasanın dövizine çevrilir.
        kasa_notu = ''
        _kasa_id = data.get('kasa_id')
        if _kasa_id and islem_tip in tahsilat_odeme_tipleri:
            e_kasa = Kasa.query.get(int(_kasa_id))
            if not e_kasa:
                db.session.rollback()
                return jsonify({'ok': False, 'mesaj': 'Seçilen kasa bulunamadı'}), 404
            islem_tut = q3(borc + alacak)          # biri her zaman 0
            tutar_try = q3(borc_try + alacak_try)  # TRY karşılığı (yukarıda hesaplandı)
            if e_kasa.doviz == doviz:
                k_tutar = islem_tut
            elif e_kasa.doviz == 'TRY':
                k_tutar = tutar_try
            else:
                k_kur = _kur_getir(e_kasa.doviz, vade)
                if not k_kur or k_kur <= 0:
                    db.session.rollback()
                    return jsonify({'ok': False,
                        'mesaj': f'{e_kasa.doviz} kuru alınamadı — farklı dövizli kasa için kur gerekli.'}), 400
                k_tutar = q3(tutar_try / k_kur)
            giris_mi = islem_tip in ('Tahsilat', 'Avans Tahsilati')
            if not giris_mi and q3((e_kasa.bakiye or 0)) < k_tutar:
                # Uyarı ver ama engelleme — eksi bakiye bilinçli olabilir
                kasa_notu_eksi = ' (DİKKAT: kasa bakiyesi eksiye düştü)'
            else:
                kasa_notu_eksi = ''
            e_kasa.bakiye = q3((e_kasa.bakiye or 0) + (k_tutar if giris_mi else -k_tutar))
            cevrim = '' if e_kasa.doviz == doviz else f' ({islem_tut:,.2f} {doviz} → {k_tutar:,.2f} {e_kasa.doviz})'
            _cari_unvan_k = getattr(hareket, 'cari_unvan', None) or (Cari.query.get(data['cari_id']).unvan if Cari.query.get(data['cari_id']) else data['cari_id'])
            kh = KasaHareket(
                kasa_id=e_kasa.id, tarih=date.today(),
                tip='giris' if giris_mi else 'cikis', tutar=k_tutar,
                aciklama=f'{_cari_unvan_k} — {islem_tip}' + (f' ({evrak_no})' if evrak_no else '') + cevrim,
                baglanti_tip='cari_hareket', baglanti_id=hareket.id,
                cari_id=data['cari_id'], kullanici=session.get('kullanici'))
            if hasattr(kh, 'siparis_id') and siparis_id:
                kh.siparis_id = siparis_id
            if hasattr(kh, 'evrak_no') and evrak_no:
                kh.evrak_no = evrak_no
            db.session.add(kh)
            kasa_notu = f'. {e_kasa.ad} kasasına {"giriş" if giris_mi else "çıkış"}: {k_tutar:,.2f} {e_kasa.doviz}{cevrim}{kasa_notu_eksi}'

        # ÇOK DÖVİZLİ: tahsilat/odeme ise otomatik kur farki hesabi
        kur_farki = None
        if islem_tip in tahsilat_odeme_tipleri:
            kur_farki = _kur_farki_hesapla_ve_olustur(hareket)

        db.session.commit()

        msg = 'Cari hareket eklendi' + kasa_notu
        if fatura_id:
            msg += f', Fatura {fatura_id} ({fatura_yon}) otomatik olusturuldu (durumu Kesildi)'
        if kur_farki:
            msg += f', otomatik kur farki kaydedildi ({kur_farki.islem_tip}: {kur_farki.borc + kur_farki.alacak:,.2f} TRY)'
        return jsonify({'ok': True, 'id': hareket.id, 'fatura_id': fatura_id,
                       'kur_farki_id': kur_farki.id if kur_farki else None,
                       'mesaj': msg})

    @app.route('/api/cari/hareket/<hareket_id>', methods=['PUT'])
    def api_hareket_guncelle(hareket_id):
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        hareket = CariHareket.query.get(hareket_id)
        if not hareket: return jsonify({'ok': False, 'mesaj': 'Hareket bulunamadı'}), 404
        data = request.json
        for key, val in data.items():
            if key == 'vade_tarihi': val = _parse_date(val)
            if hasattr(hareket, key): setattr(hareket, key, val)
        db.session.commit()
        return jsonify({'ok': True})

    @app.route('/api/cari/hareket/<hareket_id>', methods=['DELETE'])
    def api_hareket_sil(hareket_id):
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        hareket = CariHareket.query.get(hareket_id)
        if not hareket: return jsonify({'ok': False, 'mesaj': 'Bulunamadı'}), 404

        # KORUMA: Faturanın otomatik oluşturduğu BORÇ hareketi buradan silinemez.
        # Silinirse fatura "Kesildi" görünür ama cari hesapta borcu kalmaz (sessiz
        # tutarsızlık). Doğru yol: faturayı iptal etmek — o zaman borç da geri alınır.
        if (hareket.baglanti_tip == 'fatura' and hareket.kaynak == 'fatura'
                and (hareket.borc or 0) > 0):
            _f = Fatura.query.get(hareket.baglanti_id)
            if _f and _f.durum != 'Iptal':
                return jsonify({'ok': False, 'error': 'fatura_borcu',
                    'mesaj': f'Bu hareket {_f.fatura_no or _f.id} faturasinin otomatik borc kaydi. '
                             f'Dogrudan silinemez — faturayi iptal ederseniz borc da geri alinir.'}), 400

        silinen_fatura  = None
        silinen_maliyet = None

        # Bağlı Fatura varsa onu da sil
        if hareket.baglanti_tip == 'fatura' and hareket.baglanti_id and hareket.kaynak == 'cari_fatura':
            f = Fatura.query.get(hareket.baglanti_id)
            if f and f.cari_hareket_id == hareket_id:
                silinen_fatura = f.id
                db.session.delete(f)

        # Bağlı Maliyet varsa onu da sil (kaynak='maliyet')
        if hareket.baglanti_tip == 'maliyet' and hareket.baglanti_id and hareket.kaynak == 'maliyet':
            m = Maliyet.query.get(hareket.baglanti_id)
            if m:
                silinen_maliyet = m.id
                etkilenen_stok  = m.baglanti_id
                db.session.delete(m)
                db.session.flush()
                _satis_kaydi_maliyet_guncelle([etkilenen_stok] if etkilenen_stok else [])

        # Bağlı KASA hareketi varsa geri al (bakiyeyi düzelt + kaydı sil)
        silinen_kasa = None
        try:
            bagli_khs = KasaHareket.query.filter_by(
                baglanti_tip='cari_hareket', baglanti_id=hareket_id).all()
            for bkh in bagli_khs:
                bk = Kasa.query.get(bkh.kasa_id)
                if bk:
                    fark = q3(bkh.tutar or 0)
                    bk.bakiye = q3((bk.bakiye or 0) + (-fark if bkh.tip == 'giris' else fark))
                    silinen_kasa = bk.ad
                db.session.delete(bkh)
        except Exception:
            pass

        # KUR FARKI SİMETRİSİ: kapattığı hareketleri yeniden aç,
        # eşleşmeden doğan otomatik kur farkı kayıtlarını sil
        kf_acilan, kf_silinen = _kur_farki_geri_al(hareket_id)

        # Faturaya bağlı bir ödemeyse fatura durumunu yeniden hesapla
        _etkilenen_fatura = hareket.baglanti_id if hareket.baglanti_tip == 'fatura' and not silinen_fatura else None

        db.session.delete(hareket)
        db.session.flush()
        if _etkilenen_fatura:
            _fatura_tahsilat_durumu(_etkilenen_fatura)
        db.session.commit()
        msg = 'Cari hareket silindi'
        if silinen_kasa:    msg += f', bağlı kasa hareketi geri alındı ({silinen_kasa})'
        if silinen_fatura:  msg += f', bağlı Fatura ({silinen_fatura}) da silindi'
        if silinen_maliyet: msg += f', bağlı Maliyet ({silinen_maliyet}) da silindi'
        if kf_silinen:      msg += f', {kf_silinen} otomatik kur farkı kaydı geri alındı'
        if kf_acilan:       msg += f', {kf_acilan} kapatılmış hareket yeniden açıldı'
        return jsonify({'ok': True, 'silinen_fatura': silinen_fatura,
                        'silinen_maliyet': silinen_maliyet, 'mesaj': msg})

    @app.route('/api/cari/kisaltma_oner', methods=['GET'])
    def api_kisaltma_oner():
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        unvan = (request.args.get('unvan', '') or '').strip()
        if not unvan:
            return jsonify({'kisaltma': ''})

        # TAM 3 KARAKTERLIK kisaltma uret
        kelimeler = [k for k in unvan.split() if k]
        if len(kelimeler) >= 3:
            # 3+ kelime: ilk 3 kelimenin bas harfleri
            kisaltma = ''.join(k[0] for k in kelimeler[:3])
        elif len(kelimeler) == 2:
            # 2 kelime: 1. kelimenin ilk 2 harfi + 2. kelimenin ilk harfi
            kisaltma = (kelimeler[0][:2] + kelimeler[1][:1])
        else:
            # Tek kelime: ilk 3 harf
            kisaltma = kelimeler[0][:3]

        # 3 karaktere tamamla (kisa unvanlar icin)
        kisaltma = kisaltma.ljust(3, 'X')[:3]
        return jsonify({'kisaltma': kisaltma.upper()})

    # ---------- API: MÜŞTERİ / LOOKUP ----------
    @app.route('/api/musteri')
    def api_musteri():
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        # cari_tip virgullu string olabilir: "Musteri", "Musteri,Acente",
        # "Musteri,Tedarikci,Uretici" vb. Hem "Musteri" (u) hem "Müşteri" (ü),
        # hem büyük/küçük harf varyasyonlarını yakala. NULL/boş cari_tip'ler de
        # müşteri sayılır (eski kayıtlar için geri uyumluluk).
        from sqlalchemy import func as _sqlfunc
        q = Cari.query.filter(
            db.or_(
                Cari.cari_tip.like('%Musteri%'),
                Cari.cari_tip.like('%musteri%'),
                Cari.cari_tip.like('%Müşteri%'),
                Cari.cari_tip.like('%müşteri%'),
                Cari.cari_tip.like('%MUSTERI%'),
                Cari.cari_tip.like('%MÜŞTERİ%'),
                Cari.cari_tip == 'Her ikisi',  # eski deger, geri uyum
                Cari.cari_tip.is_(None),
                Cari.cari_tip == ''
            )
        ).with_entities(Cari.unvan).distinct().order_by(Cari.unvan).all()
        return jsonify([c.unvan for c in q])

    @app.route('/api/musteri/detay')
    def api_musteri_detay():
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        return jsonify([{'unvan': c.unvan, 'adres': c.adres, 'ulke': c.ulke} for c in Cari.query.all()])

    @app.route('/api/uretici')
    def api_uretici():
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        # Ureticiler artik Cari tablosundan gelir: urun_tedarikcisi=True olan cariler
        ureticiler = Cari.query.filter_by(urun_tedarikcisi=True).order_by(Cari.unvan).all()
        return jsonify([
            {'id': c.id, 'deger': c.unvan, 'kisaltma': c.uretici_kisaltma or c.unvan[:3].upper()}
            for c in ureticiler
        ])

    @app.route('/api/lookup/<kategori>')
    def api_lookup(kategori):
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        return jsonify([{'id': v.id, 'deger': v.deger, 'kisaltma': v.kisaltma} for v in Veriler.query.filter_by(kategori=kategori).all()])

    # ---------- API: SİPARİŞ ---------- # FAZ 16.3: SİPARİŞ KALEM-AWARE
    # FAZ 16: Çoklu kalem yapısı. Siparis = parent, SiparisKalem = child (1-N).

    def _kalem_hesapla(k_data):
        """Bir kalemin m2/m3/sqft/kg toplamlarını hesaplar. Inplace günceller."""
        import math
        boy = float(k_data.get('boy') or 0)
        yuk = float(k_data.get('yukseklik') or 0)
        en  = float(k_data.get('en') or 0)
        kal = float(k_data.get('kalinlik') or 0)
        adet = int(k_data.get('adet') or 1)
        kasa_ici = int(k_data.get('kasa_ici_adet') or 1)
        urun_tip = (k_data.get('urun_tip') or '').upper()

        m2 = 0; m3 = 0; sqft = 0
        if urun_tip == 'BLOK':
            # m³ = boy * yukseklik * en / 1.000.000 (cm³ → m³)
            if boy and yuk and en:
                m3 = (boy * yuk * en / 1_000_000) * adet
        else:
            # PLAKA/EBATLI: m² = boy * yukseklik / 10000 * adet * kasa_ici
            if boy and yuk:
                m2 = (boy * yuk / 10000) * adet * kasa_ici
                sqft = m2 * 10.7639

        # Ölçü string'i
        olcu_parts = []
        if boy: olcu_parts.append(str(int(boy) if boy == int(boy) else boy))
        if yuk: olcu_parts.append(str(int(yuk) if yuk == int(yuk) else yuk))
        if en and urun_tip == 'BLOK':
            olcu_parts.append(str(int(en) if en == int(en) else en))
        if kal and urun_tip != 'BLOK':
            olcu_parts.append(str(int(kal) if kal == int(kal) else kal))
        olcu_str = 'x'.join(olcu_parts) + ('cm' if olcu_parts else '')

        # Toplam fiyat
        birim_fiyat = float(k_data.get('birim_fiyat') or 0)
        miktar = float(k_data.get('miktar') or 0)
        toplam_fiyat = birim_fiyat * miktar

        return {
            'm2_toplam': round(m2, 4),
            'm3_toplam': round(m3, 6),
            'sqft_toplam': round(sqft, 2),
            'olcu': olcu_str,
            'toplam_fiyat': round(toplam_fiyat, 2)
        }


    def _kalem_to_dict(k):
        """SiparisKalem objesini JSON-friendly dict'e çevirir."""
        import json
        stok_ids = []
        if k.stok_ids_json:
            try: stok_ids = json.loads(k.stok_ids_json)
            except Exception: stok_ids = []
        return {
            'id': k.id, 'siparis_id': k.siparis_id, 'sira': k.sira,
            'urun_tip': k.urun_tip, 'cins': k.cins, 'ozellik': k.ozellik,
            'aciklama': k.aciklama,
            'boy': k.boy, 'yukseklik': k.yukseklik, 'en': k.en,
            'kalinlik': k.kalinlik, 'olcu': k.olcu,
            'adet': k.adet, 'kasa_ici_adet': k.kasa_ici_adet,
            'miktar': k.miktar, 'birim': k.birim,
            'm2_toplam': k.m2_toplam, 'm3_toplam': k.m3_toplam,
            'sqft_toplam': k.sqft_toplam, 'kg_toplam': k.kg_toplam,
            'birim_fiyat': k.birim_fiyat, 'toplam_fiyat': k.toplam_fiyat,
            'doviz': k.doviz,
            'stoktan_geldi': k.stoktan_geldi, 'stok_ids': stok_ids,
            'notlar': k.notlar
        }


    def _kalem_rezervasyonlari_olustur(siparis, kalem, stok_ids):
        """Bir kalem için verilen stok ID'lerine rezervasyon kaydı oluşturur.
        Stok durumunu sipariş durumuna göre Rezerve/Satildi yapar."""
        if not stok_ids:
            return 0
        olusturulan = 0
        sip_durum = siparis.durum or 'Teklif Asam.'
        # Sipariş Teklif aşamasındaysa stoklar Rezerve, onaylanmış vb. ise Satildi
        hedef_stok_durum = 'Rezerve' if sip_durum == 'Teklif Asam.' else 'Satildi'

        for sid in stok_ids:
            # Stok tipini kalemin urun_tip'inden tahmin et
            tip = (kalem.urun_tip or '').upper()
            if tip == 'BLOK':
                stok = BlokStok.query.get(sid)
            elif tip == 'PLAKA':
                stok = PlakaStok.query.get(sid)
            elif tip == 'EBATLI':
                stok = EbatliStok.query.get(sid)
            else:
                stok = None

            if not stok:
                continue
            if stok.durum not in ('Serbest', 'Rezerve'):
                continue  # Satilmış/Teslim Edilmiş stok atlanır

            rez = Rezervasyon(
                id=_yeni_id('REZ'),
                musteri=siparis.musteri,
                siparis_id=siparis.id,
                siparis_kalem_id=kalem.id,
                stok_tip=tip,
                cins=stok.cins,
                ozellik=getattr(stok, 'ozellik', None),
                stok_id=sid,
                miktar=getattr(stok, 'metraj_m2', None) or getattr(stok, 'hacim_m3', None),
                rez_tip='Siparis',
                kullanici=session.get('kullanici', 'sistem')
            )
            db.session.add(rez)
            stok.durum = hedef_stok_durum
            olusturulan += 1
        return olusturulan


    def _siparis_toplam_guncelle(siparis):
        """Bir siparişin toplam_tutar'ını kalemlerden hesaplar (sipariş döviziyle)."""
        toplam = 0
        for k in siparis.kalemler:
            # Şimdilik kalem dövizi sipariş dövizine eşit varsayılıyor.
            # Farklı dövizler için kur çevrimi eklenebilir.
            toplam += (k.toplam_fiyat or 0)
        siparis.toplam_tutar = round(toplam, 2)


    # ─── GET LİSTE ───
    @app.route('/api/siparis', methods=['GET'])
    def api_siparis_liste():
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        musteri = request.args.get('musteri')
        durum = request.args.get('durum')
        per_page = request.args.get('per_page', type=int, default=50)
        page = request.args.get('page', type=int, default=1)
        query = Siparis.query
        if musteri: query = query.filter_by(musteri=musteri)
        if durum:
            if durum == 'aktif':
                query = query.filter(Siparis.durum.notin_(['Iptal Edildi', 'Teslim Edildi']))
            else:
                query = query.filter_by(durum=durum)
        paginated = query.order_by(Siparis.siparis_tarihi.desc()).paginate(
            page=page, per_page=per_page, error_out=False)

        data = []
        for s in paginated.items:
            # Kalem özet bilgisi
            kalem_sayisi = len(s.kalemler) if s.kalemler else 0
            ilk_cins = s.kalemler[0].cins if s.kalemler else None
            ilk_tip = s.kalemler[0].urun_tip if s.kalemler else None
            # Birden fazla cins varsa "X + N daha" formatı
            cins_text = ilk_cins or '-'
            if kalem_sayisi > 1:
                cins_text = f'{ilk_cins} +{kalem_sayisi - 1} daha'
            data.append({
                'id': s.id,
                'tarih': s.siparis_tarihi.isoformat() if s.siparis_tarihi else None,
                'musteri': s.musteri,
                'urun_tip': ilk_tip,           # geri uyum (eski frontend)
                'cins': cins_text,              # özet
                'kalem_sayisi': kalem_sayisi,
                'toplam_tutar': s.toplam_tutar or 0,
                'doviz': s.doviz,
                'durum': s.durum,
                'termin': s.termin.isoformat() if s.termin else None
            })
        return jsonify({'data': data, 'meta': {
            'page': page, 'per_page': per_page, 'total': paginated.total}})


    # ─── GET DETAY ───
    @app.route('/api/siparis/<siparis_id>', methods=['GET'])
    def api_siparis_detay(siparis_id):
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        sip = Siparis.query.get(siparis_id)
        if not sip:
            return jsonify({'ok': False, 'mesaj': 'Sipariş bulunamadı'}), 404
        return jsonify({
            'ok': True,
            'siparis': {
                'id': sip.id,
                'siparis_tarihi': sip.siparis_tarihi.isoformat() if sip.siparis_tarihi else None,
                'musteri': sip.musteri,
                'doviz': sip.doviz,
                'odeme_sekli': sip.odeme_sekli,
                'teslim_sekli': sip.teslim_sekli,
                'termin': sip.termin.isoformat() if sip.termin else None,
                'durum': sip.durum,
                'aciklama': sip.aciklama,
                'toplam_tutar': sip.toplam_tutar or 0,
                'satis_tipi': sip.satis_tipi,
                'kdv_oran': sip.kdv_oran,
                'kdv_tutar': sip.kdv_tutar,
                'tevkifat_oran': sip.tevkifat_oran,
                'tevkifat_tutar': sip.tevkifat_tutar,
                'acente_cari_id': sip.acente_cari_id,
                'komisyon_yontem': sip.komisyon_yontem,
                'komisyon_deger': sip.komisyon_deger,
                'komisyon_tutar': sip.komisyon_tutar,
                'komisyon_doviz': sip.komisyon_doviz,
                'komisyon_aciklama': sip.komisyon_aciklama,
                'kullanici': sip.kullanici
            },
            'kalemler': [_kalem_to_dict(k) for k in sip.kalemler]
        })


    # ─── POST (YENİ SİPARİŞ + KALEMLER) ───
    @app.route('/api/siparis', methods=['POST'])
    def api_siparis_ekle():
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        import json as _json
        data = request.json or {}

        # Parent sipariş
        sip = Siparis(
            id=_yeni_id('SIP'),
            musteri=data.get('musteri'),
            doviz=data.get('doviz', 'USD'),
            odeme_sekli=data.get('odeme_sekli'),
            teslim_sekli=data.get('teslim_sekli'),
            termin=_parse_date(data.get('termin')),
            durum=data.get('durum') or 'Teklif Asam.',
            aciklama=data.get('aciklama'),
            satis_tipi=data.get('satis_tipi', 'ihracat'),
            kdv_oran=data.get('kdv_oran', 0),
            kdv_tutar=data.get('kdv_tutar', 0),
            tevkifat_oran=data.get('tevkifat_oran', ''),
            tevkifat_tutar=data.get('tevkifat_tutar', 0),
            acente_cari_id=data.get('acente_cari_id'),
            komisyon_yontem=data.get('komisyon_yontem'),
            komisyon_deger=data.get('komisyon_deger', 0),
            komisyon_tutar=data.get('komisyon_tutar', 0),
            komisyon_doviz=data.get('komisyon_doviz'),
            komisyon_aciklama=data.get('komisyon_aciklama'),
            kullanici=session.get('kullanici', 'sistem')
        )
        db.session.add(sip)
        db.session.flush()  # sip.id'yi al

        # Kalemler
        kalemler_data = data.get('kalemler', [])
        if not kalemler_data:
            db.session.rollback()
            return jsonify({'ok': False, 'mesaj': 'En az bir kalem girilmelidir'}), 400

        for idx, k_data in enumerate(kalemler_data):
            # Hesaplama
            hesap = _kalem_hesapla(k_data)
            stok_ids = k_data.get('stok_ids', [])
            stoktan_geldi = bool(stok_ids)

            kalem = SiparisKalem(
                siparis_id=sip.id,
                sira=k_data.get('sira', idx + 1),
                urun_tip=k_data.get('urun_tip'),
                cins=k_data.get('cins'),
                ozellik=k_data.get('ozellik'),
                aciklama=k_data.get('aciklama'),
                boy=k_data.get('boy'),
                yukseklik=k_data.get('yukseklik'),
                en=k_data.get('en'),
                kalinlik=k_data.get('kalinlik'),
                olcu=hesap['olcu'],
                adet=k_data.get('adet', 1),
                kasa_ici_adet=k_data.get('kasa_ici_adet', 1),
                miktar=k_data.get('miktar'),
                birim=k_data.get('birim'),
                m2_toplam=hesap['m2_toplam'],
                m3_toplam=hesap['m3_toplam'],
                sqft_toplam=hesap['sqft_toplam'],
                kg_toplam=k_data.get('kg_toplam', 0),
                birim_fiyat=k_data.get('birim_fiyat'),
                toplam_fiyat=hesap['toplam_fiyat'],
                doviz=k_data.get('doviz', sip.doviz),
                stoktan_geldi=stoktan_geldi,
                stok_ids_json=_json.dumps(stok_ids) if stok_ids else None,
                notlar=k_data.get('notlar')
            )
            db.session.add(kalem)
            db.session.flush()  # kalem.id'yi al

            # Stoktan rezervasyon oluştur
            if stok_ids:
                _kalem_rezervasyonlari_olustur(sip, kalem, stok_ids)

        # Toplam'ı güncelle
        _siparis_toplam_guncelle(sip)

        _log_audit('EKLE', 'siparis', sip.id, yeni={
            'musteri': sip.musteri, 'kalem_sayisi': len(kalemler_data),
            'toplam_tutar': sip.toplam_tutar
        })
        db.session.commit()
        return jsonify({'ok': True, 'id': sip.id, 'mesaj': f'Sipariş oluşturuldu: {sip.id}'})


    # ─── SICAK SATIŞ (depo anlık satışı) ───
    # Proforma akışını beklemeden, tek istekte:
    # sipariş (Onaylandı) + kalemler + stok satışı + fatura (Kesildi) + cari borç.
    @app.route('/api/sicak_satis', methods=['POST'])
    def api_sicak_satis():
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        import json as _json
        data = request.json or {}

        musteri = (data.get('musteri') or '').strip()
        if not musteri:
            return jsonify({'ok': False, 'mesaj': 'Müşteri (cari) seçilmelidir'}), 400
        kalemler_data = data.get('kalemler', [])
        if not kalemler_data:
            return jsonify({'ok': False, 'mesaj': 'En az bir kalem girilmelidir'}), 400

        doviz = data.get('doviz', 'USD')
        bugun = date.today()

        try:
            # ── 1) SİPARİŞ (doğrudan Onaylandı) ──
            sip = Siparis(
                id=_yeni_id('SIP'),
                musteri=musteri, doviz=doviz,
                odeme_sekli=data.get('odeme_sekli') or 'Pesin',
                teslim_sekli=data.get('teslim_sekli') or 'EXW',
                termin=bugun,
                durum='Onaylandi',
                aciklama=(data.get('aciklama') or 'Sıcak satış (depo anlık)'),
                satis_tipi=data.get('satis_tipi', 'yurtici'),
                kdv_oran=data.get('kdv_oran', 0),
                kdv_tutar=data.get('kdv_tutar', 0),
                kullanici=session.get('kullanici', 'sistem')
            )
            db.session.add(sip)
            db.session.flush()

            kalem_list_json = []  # fatura kalemler_json için
            for idx, k_data in enumerate(kalemler_data):
                hesap = _kalem_hesapla(k_data)
                stok_ids = k_data.get('stok_ids', [])
                stoktan_geldi = bool(stok_ids)
                kalem = SiparisKalem(
                    siparis_id=sip.id, sira=k_data.get('sira', idx + 1),
                    urun_tip=k_data.get('urun_tip'), cins=k_data.get('cins'),
                    ozellik=k_data.get('ozellik'), aciklama=k_data.get('aciklama'),
                    boy=k_data.get('boy'), yukseklik=k_data.get('yukseklik'),
                    en=k_data.get('en'), kalinlik=k_data.get('kalinlik'),
                    olcu=hesap['olcu'], adet=k_data.get('adet', 1),
                    kasa_ici_adet=k_data.get('kasa_ici_adet', 1),
                    miktar=k_data.get('miktar'), birim=k_data.get('birim'),
                    m2_toplam=hesap['m2_toplam'], m3_toplam=hesap['m3_toplam'],
                    sqft_toplam=hesap['sqft_toplam'], kg_toplam=k_data.get('kg_toplam', 0),
                    birim_fiyat=k_data.get('birim_fiyat'), toplam_fiyat=hesap['toplam_fiyat'],
                    doviz=k_data.get('doviz', doviz),
                    stoktan_geldi=stoktan_geldi,
                    stok_ids_json=_json.dumps(stok_ids) if stok_ids else None,
                    notlar=k_data.get('notlar')
                )
                db.session.add(kalem)
                db.session.flush()
                if stok_ids:
                    _kalem_rezervasyonlari_olustur(sip, kalem, stok_ids)
                kalem_list_json.append({
                    'cins': kalem.cins, 'ozellik': kalem.ozellik, 'urun_tip': kalem.urun_tip,
                    'olcu': kalem.olcu, 'adet': kalem.adet, 'miktar': kalem.miktar,
                    'birim': kalem.birim, 'm2_toplam': kalem.m2_toplam,
                    'birim_fiyat': kalem.birim_fiyat, 'toplam_fiyat': kalem.toplam_fiyat,
                    'doviz': kalem.doviz
                })

            _siparis_toplam_guncelle(sip)
            # NOT: Stoklar _kalem_rezervasyonlari_olustur içinde otomatik 'Satildi'
            # yapılır (sipariş 'Onaylandi' olduğu için). Ayrıca işleme gerek yok.

            # ── 3) FATURA (doğrudan Kesildi) ──
            ftr_no = data.get('fatura_no') or f'SF-{bugun.strftime("%Y%m%d")}-{sip.id[-4:]}'
            fatura = Fatura(
                id=_yeni_id('FTR'), fatura_no=ftr_no,
                fatura_tarihi=bugun,
                vade_tarihi=_parse_date(data.get('vade_tarihi')) or bugun,
                siparis_id=sip.id, musteri=musteri,
                musteri_adres=data.get('musteri_adres'),
                toplam=sip.toplam_tutar or 0,
                ara_toplam=sip.toplam_tutar or 0,
                kdv_oran=data.get('kdv_oran', 0), kdv_tutar=data.get('kdv_tutar', 0),
                doviz=doviz,
                odeme_sekli=sip.odeme_sekli, teslim_sekli=sip.teslim_sekli,
                durum='Taslak', yon='satis', fatura_tipi='stoklu',
                satis_tipi=sip.satis_tipi,
                kur_farki_modu=('cari' if doviz == 'TRY' else 'gider'),
                kalemler_json=_json.dumps(kalem_list_json, ensure_ascii=False, default=str),
                kullanici=session.get('kullanici', 'sistem')
            )
            db.session.add(fatura)
            db.session.flush()

            # ── 4) FATURAYI KESİLDİ yap → cariye BORÇ ──
            cari = Cari.query.filter_by(unvan=musteri).first()
            if not cari:
                db.session.rollback()
                return jsonify({'ok': False,
                    'mesaj': f'"{musteri}" carisi bulunamadı. Önce cari kaydı oluşturun.'}), 400

            ch = _cari_hareket_ekle(
                cari_unvan=musteri, islem_tip='Satış Faturası',
                borc=fatura.toplam or 0, alacak=0, doviz=doviz,
                aciklama=f'Sıcak satış faturası {fatura.fatura_no}',
                kaynak='fatura', baglanti_tip='fatura', baglanti_id=fatura.id,
                evrak_no=fatura.fatura_no,
                vade_tarihi=_parse_date(data.get('vade_tarihi'))
            )
            if ch:
                fatura.cari_hareket_id = ch.id
            fatura.durum = 'Kesildi'
            sip.durum = 'Hazir'
            db.session.flush()
            # SatisKaydi oluştur — karlılık ve satış raporları bu kayıtlardan beslenir.
            # (Fatura durumu doğrudan atandığı için api_fatura_durum'daki otomatik
            #  oluşturma tetiklenmez; burada açıkça çağırıyoruz.)
            try:
                sk_adet, sk_hata = _fatura_satis_kaydi_olustur(fatura.id)
                if sk_hata:
                    app.logger.warning(f'Sicak satis satis kaydi: {sk_hata}')
            except Exception as _e:
                app.logger.warning(f'Sicak satis satis kaydi hatasi: {_e}')

            # ── 5) ÇEK (ödeme şekli çek ise) ──
            # Müşteriden alınan çek olarak kaydet, faturaya bağla. Çek alındığı an
            # _cek_hareket + cari ALACAK otomatik oluşur (borç-alacak birbirini götürür).
            cek_bilgi = data.get('cek')
            cek_olusturuldu = None
            if cek_bilgi:
                cek_tutar = cek_bilgi.get('tutar') or fatura.toplam or 0
                try:
                    cek_tutar = float(cek_tutar)
                except (ValueError, TypeError):
                    cek_tutar = fatura.toplam or 0
                cek = Cek(
                    id=_yeni_id('CEK'), yon='alinan', tip='cek',
                    cek_no=(cek_bilgi.get('cek_no') or '').strip() or None,
                    banka_adi=(cek_bilgi.get('banka_adi') or '').strip() or None,
                    sube=(cek_bilgi.get('sube') or '').strip() or None,
                    hesap_sahibi=(cek_bilgi.get('hesap_sahibi') or '').strip() or None,
                    tutar=q3(cek_tutar), doviz=doviz,
                    keside_tarihi=_parse_date(cek_bilgi.get('keside_tarihi')),
                    vade_tarihi=_parse_date(data.get('vade_tarihi')) or bugun,
                    cari_id=cari.id, cari_unvan=musteri,
                    durum='Portfoyde', fatura_id=fatura.id,
                    aciklama=f'Sıcak satış {fatura.fatura_no} karşılığı alınan çek',
                    kullanici=session.get('kullanici', 'sistem'))
                db.session.add(cek)
                db.session.flush()
                _cek_hareket_ekle(cek, 'Alındı', None, 'Portfoyde', cek.aciklama or '')
                # Çek = ödeme: müşterinin borcu düşer → cari ALACAK
                _cari_hareket_ekle(
                    cari_unvan=musteri, islem_tip='Çek Tahsilatı',
                    borc=0, alacak=cek.tutar or 0, doviz=doviz,
                    aciklama=f'Çek alındı (No: {cek.cek_no or cek.id}) — {fatura.fatura_no}',
                    kaynak='cek', baglanti_tip='cek', baglanti_id=cek.id,
                    vade_tarihi=cek.vade_tarihi, evrak_no=cek.cek_no or cek.id)
                cek_olusturuldu = cek.id
                # Fatura tahsilat durumunu güncelle (çek ödeme sayılır)
                _fatura_tahsilat_durumu(fatura.id)

            # ── 6) PEŞİN TAHSİLAT → KASA (ödeme şekli peşin + kasa seçildiyse) ──
            # Nakit satışta para fiilen kasaya girer: cariye ALACAK (borç kapanır)
            # + seçilen kasaya GİRİŞ hareketi. Döviz farkı TRY köprüsüyle çevrilir.
            kasa_tahsilat_notu = ''
            _hs_kasa_id = data.get('kasa_id')
            if _hs_kasa_id and not cek_bilgi:
                hs_kasa = Kasa.query.get(int(_hs_kasa_id))
                if hs_kasa:
                    satis_tutar = q3(fatura.toplam or 0)
                    k_tutar = None
                    if hs_kasa.doviz == doviz:
                        k_tutar = satis_tutar
                    else:
                        _kur_d = _kur_getir(doviz, bugun) if doviz != 'TRY' else 1.0
                        _try_tut = q3(satis_tutar * (_kur_d or 0)) if doviz != 'TRY' else satis_tutar
                        if hs_kasa.doviz == 'TRY':
                            k_tutar = _try_tut if _try_tut else None
                        else:
                            _kur_k = _kur_getir(hs_kasa.doviz, bugun)
                            if _kur_k and _try_tut:
                                k_tutar = q3(_try_tut / _kur_k)
                    if k_tutar:
                        _cari_hareket_ekle(
                            cari_unvan=musteri, islem_tip='Tahsilat',
                            borc=0, alacak=satis_tutar, doviz=doviz,
                            aciklama=f'Sıcak satış peşin tahsilat — {fatura.fatura_no}',
                            kaynak='sicak_satis', baglanti_tip='fatura', baglanti_id=fatura.id,
                            vade_tarihi=bugun, evrak_no=fatura.fatura_no)
                        hs_kasa.bakiye = q3((hs_kasa.bakiye or 0) + k_tutar)
                        _cvr = '' if hs_kasa.doviz == doviz else f' ({satis_tutar:,.2f} {doviz} → {k_tutar:,.2f} {hs_kasa.doviz})'
                        hs_kh = KasaHareket(
                            kasa_id=hs_kasa.id, tarih=bugun, tip='giris', tutar=k_tutar,
                            aciklama=f'{musteri} — sıcak satış peşin tahsilat ({fatura.fatura_no}){_cvr}',
                            baglanti_tip='fatura', baglanti_id=fatura.id,
                            cari_id=cari.id, kullanici=session.get('kullanici'))
                        if hasattr(hs_kh, 'evrak_no'):
                            hs_kh.evrak_no = fatura.fatura_no
                        db.session.add(hs_kh)
                        _fatura_tahsilat_durumu(fatura.id)
                        kasa_tahsilat_notu = f' Peşin tahsilat {hs_kasa.ad} kasasına işlendi ({k_tutar:,.2f} {hs_kasa.doviz}).'
                    else:
                        kasa_tahsilat_notu = f' UYARI: {hs_kasa.doviz} kuru alınamadığı için kasa kaydı yapılamadı — tahsilatı elle gir.'

            _log_audit('EKLE', 'sicak_satis', sip.id, yeni={
                'musteri': musteri, 'fatura': fatura.id, 'toplam': fatura.toplam,
                'cek': cek_olusturuldu})

            ok, hata = _safe_commit(f'Sıcak satış: {sip.id} / {fatura.id}')
            if not ok:
                return jsonify({'ok': False, 'mesaj': f'Kayıt hatası: {hata}'}), 500

            mesaj = (f'Sıcak satış tamamlandı. Fatura {fatura.fatura_no} kesildi, '
                     f'{musteri} carisine {fatura.toplam:,.2f} {doviz} borç işlendi.')
            if cek_olusturuldu:
                mesaj += f' Çek ({cek_olusturuldu}) portföye eklendi, borç çekle kapatıldı.'
            mesaj += kasa_tahsilat_notu
            return jsonify({'ok': True, 'siparis_id': sip.id, 'fatura_id': fatura.id,
                'fatura_no': fatura.fatura_no, 'toplam': fatura.toplam, 'doviz': doviz,
                'cek_id': cek_olusturuldu, 'mesaj': mesaj})
        except Exception as e:
            db.session.rollback()
            app.logger.error(f'Sıcak satış hatası: {e}')
            return jsonify({'ok': False, 'mesaj': f'Sıcak satış hatası: {e}'}), 500




    # ─── PUT (GÜNCELLEME) ───
    @app.route('/api/siparis/<siparis_id>', methods=['PUT'])
    def api_siparis_guncelle(siparis_id):
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        import json as _json
        sip = Siparis.query.get(siparis_id)
        if not sip: return jsonify({'ok': False, 'mesaj': 'Bulunamadi'}), 404
        data = request.json or {}

        # Sipariş seviyesi alanlar
        for alan in ['musteri', 'doviz', 'odeme_sekli', 'teslim_sekli', 'aciklama',
                     'satis_tipi', 'kdv_oran', 'kdv_tutar', 'tevkifat_oran', 'tevkifat_tutar',
                     'acente_cari_id', 'komisyon_yontem', 'komisyon_deger', 'komisyon_tutar',
                     'komisyon_doviz', 'komisyon_aciklama']:
            if alan in data:
                setattr(sip, alan, data[alan])
        if 'termin' in data:
            sip.termin = _parse_date(data['termin'])

        # KALEM GÜNCELLEMESİ (eğer "kalemler" key'i gönderildiyse full replace)
        if 'kalemler' in data:
            # Önce mevcut kalemleri ve rezervasyonları temizle
            # Stokları serbest bırak (sipariş henüz Teslim Edildi değilse)
            for k in list(sip.kalemler):
                # Bu kaleme bağlı rezervasyonları iptal
                kalem_rezler = Rezervasyon.query.filter_by(
                    siparis_kalem_id=k.id, iptal_nedeni=None
                ).all()
                for r in kalem_rezler:
                    # Stoğu serbest bırak (Rezerve durumdaysa)
                    if r.stok_tip == 'BLOK':
                        stok = BlokStok.query.get(r.stok_id)
                    elif r.stok_tip == 'PLAKA':
                        stok = PlakaStok.query.get(r.stok_id)
                    else:
                        stok = EbatliStok.query.get(r.stok_id)
                    if stok and stok.durum == 'Rezerve':
                        stok.durum = 'Serbest'
                    r.iptal_nedeni = 'Kalem güncellendi'
                    r.iptal_tarihi = datetime.now()
                    r.iptal_eden = session.get('kullanici', 'sistem')
                db.session.delete(k)
            db.session.flush()

            # Yeni kalemleri ekle
            for idx, k_data in enumerate(data['kalemler']):
                hesap = _kalem_hesapla(k_data)
                stok_ids = k_data.get('stok_ids', [])
                stoktan_geldi = bool(stok_ids)
                kalem = SiparisKalem(
                    siparis_id=sip.id,
                    sira=k_data.get('sira', idx + 1),
                    urun_tip=k_data.get('urun_tip'),
                    cins=k_data.get('cins'),
                    ozellik=k_data.get('ozellik'),
                    aciklama=k_data.get('aciklama'),
                    boy=k_data.get('boy'),
                    yukseklik=k_data.get('yukseklik'),
                    en=k_data.get('en'),
                    kalinlik=k_data.get('kalinlik'),
                    olcu=hesap['olcu'],
                    adet=k_data.get('adet', 1),
                    kasa_ici_adet=k_data.get('kasa_ici_adet', 1),
                    miktar=k_data.get('miktar'),
                    birim=k_data.get('birim'),
                    m2_toplam=hesap['m2_toplam'],
                    m3_toplam=hesap['m3_toplam'],
                    sqft_toplam=hesap['sqft_toplam'],
                    kg_toplam=k_data.get('kg_toplam', 0),
                    birim_fiyat=k_data.get('birim_fiyat'),
                    toplam_fiyat=hesap['toplam_fiyat'],
                    doviz=k_data.get('doviz', sip.doviz),
                    stoktan_geldi=stoktan_geldi,
                    stok_ids_json=_json.dumps(stok_ids) if stok_ids else None,
                    notlar=k_data.get('notlar')
                )
                db.session.add(kalem)
                db.session.flush()
                if stok_ids:
                    _kalem_rezervasyonlari_olustur(sip, kalem, stok_ids)

            _siparis_toplam_guncelle(sip)

        # Durum geçişi kontrolü
        if 'durum' in data:
            yeni_durum = data['durum']
            gecerli = ['Teklif Asam.', 'Onaylandi', 'Uretimde', 'Hazir', 'Teslim Edildi', 'Iptal Edildi']
            if yeni_durum not in gecerli:
                return jsonify({'ok': False, 'mesaj': f'Gecersiz durum. Gecerli: {", ".join(gecerli)}'}), 400

            eski_durum = sip.durum or 'Teklif Asam.'
            gecisler = {
                'Teklif Asam.': ['Onaylandi', 'Iptal Edildi'],
                'Onaylandi':    ['Uretimde', 'Hazir', 'Iptal Edildi'],
                'Uretimde':     ['Hazir', 'Iptal Edildi'],
                'Hazir':        ['Teslim Edildi', 'Iptal Edildi'],
                'Teslim Edildi':['Hazir'],
                'Iptal Edildi': ['Teklif Asam.']
            }
            if eski_durum != yeni_durum and yeni_durum not in gecisler.get(eski_durum, []):
                return jsonify({
                    'ok': False,
                    'mesaj': f'"{eski_durum}" durumundan "{yeni_durum}" gecisi yapilamaz. Izinli: {", ".join(gecisler.get(eski_durum, [])) or "(yok)"}'
                }), 400

            # ═══ ÖDEME BEKLİYOR KAPISI ═══
            # Onaylı sipariş ÜRETİME geçmeden önce minimum avans tahsil edilmiş olmalı.
            # Eşik Ayarlar'dan gelir (0 = kapı kapalı). Yönetici zorla geçebilir (?zorla=1).
            if eski_durum == 'Onaylandi' and yeni_durum == 'Uretimde':
                esik = _uretim_avans_esigi()
                zorla = str((request.json or {}).get('zorla', '')).lower() in ('1', 'true', 'evet')
                if esik > 0 and not zorla:
                    td = _siparis_tahsilat_durumu(siparis_id)
                    if td and td['yuzde'] < esik:
                        return jsonify({
                            'ok': False, 'error': 'odeme_bekliyor',
                            'mesaj': f'Uretime gecmek icin en az %{esik:g} avans gerekli. '
                                     f'Tahsil edilen: %{td["yuzde"]:g} '
                                     f'({td["tahsil"]:,.0f}/{td["toplam"]:,.0f} {td["doviz"]}). '
                                     f'Once tahsilat girin ya da yonetici olarak zorla gecin.',
                            'tahsilat': td, 'esik': esik
                        }), 400
                elif esik > 0 and zorla and session.get('rol') not in ('admin', 'ADMIN'):
                    # Zorla geçiş yalnızca yöneticiye açık
                    return jsonify({'ok': False, 'error': 'zorla_yetkisi_yok',
                        'mesaj': 'Odeme kapisini zorla gecmek icin yonetici yetkisi gerekli.'}), 403
                elif esik > 0 and zorla:
                    ekstra_zorla = ' (yonetici odeme kapisini zorla gecti)'

            sip.durum = yeni_durum
            ekstra = locals().get('ekstra_zorla', '')

            # İPTAL: bağlı rezervasyonları iptal + stokları Serbest bırak + proformaları iptal
            # (_siparis_stoklarini_serbest_birak hem rezervasyonu iptal eder hem stoğu serbest yapar)
            if yeni_durum == 'Iptal Edildi' and eski_durum != 'Iptal Edildi':
                serbest = _siparis_stoklarini_serbest_birak(siparis_id, f'Siparis {siparis_id} iptal')
                if serbest > 0:
                    ekstra += f' {serbest} stok serbest birakildi.'
                aktif_proformalar = Proforma.query.filter(
                    Proforma.siparis_id == siparis_id,
                    ~Proforma.durum.in_(['Iptal', 'Faturalandi', 'Revize'])
                ).all()
                iptal_prf = 0
                for pf in aktif_proformalar:
                    pf.durum = 'Iptal'
                    iptal_prf += 1
                if iptal_prf > 0:
                    ekstra += f' {iptal_prf} bagli proforma iptal edildi.'
                logging.info(f"Siparis Iptal: {siparis_id}, {serbest} stok, {iptal_prf} proforma")

            if eski_durum == 'Teklif Asam.' and yeni_durum in ('Onaylandi', 'Uretimde', 'Hazir'):
                guncel = _siparis_stoklarini_durum_guncelle(siparis_id, 'Satildi')
                if guncel > 0:
                    ekstra += f' {guncel} stok Satildi durumuna gecti.'
            elif eski_durum in ('Onaylandi', 'Uretimde', 'Hazir') and yeni_durum == 'Teklif Asam.':
                guncel = _siparis_stoklarini_durum_guncelle(siparis_id, 'Rezerve')
                if guncel > 0:
                    ekstra += f' {guncel} stok Rezerve durumuna geri dondu.'

            if yeni_durum == 'Teslim Edildi' and eski_durum != 'Teslim Edildi':
                sat_sayisi = _siparis_teslim_edildi(siparis_id)
                if sat_sayisi > 0:
                    ekstra += f' {sat_sayisi} satis kaydi olusturuldu, stoklar Teslim Edildi durumuna gecti.'
            elif eski_durum == 'Teslim Edildi' and yeni_durum != 'Teslim Edildi':
                geri = _siparis_teslimi_iptal_et(siparis_id)
                if geri > 0:
                    ekstra += f' {geri} satis kaydi geri alindi.'

            _log_audit('DURUM', 'siparis', siparis_id, eski={'durum': eski_durum}, yeni={'durum': yeni_durum})
            db.session.commit()
            return jsonify({'ok': True, 'mesaj': f'Siparis durumu: {eski_durum} -> {yeni_durum}.{ekstra}'})

        db.session.commit()
        return jsonify({'ok': True, 'mesaj': 'Siparis guncellendi'})


    # ─── DELETE ───
    @app.route('/api/siparis/<siparis_id>/tahsilat_durumu', methods=['GET'])
    def api_siparis_tahsilat_durumu(siparis_id):
        """Siparişin ödeme/avans durumunu döner (UI ödeme kapısı göstergesi için)."""
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        td = _siparis_tahsilat_durumu(siparis_id)
        if td is None:
            return jsonify({'ok': False, 'mesaj': 'Siparis bulunamadi'}), 404
        esik = _uretim_avans_esigi()
        td['esik'] = esik
        td['kapi_acik'] = (esik <= 0) or (td['yuzde'] >= esik)
        td['odeme_bekliyor'] = (esik > 0) and (td['yuzde'] < esik)
        return jsonify({'ok': True, **td})

    @app.route('/api/ayarlar/smtp', methods=['GET'])
    def api_smtp_get():
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        if session.get('rol') not in ('admin', 'ADMIN'):
            return jsonify({'ok': False, 'mesaj': 'Yetkisiz'}), 403
        kayitlar = {v.deger: v.ek_bilgi for v in
                    Veriler.query.filter_by(kategori='smtp_ayar').all()}
        # Şifreyi asla düz döndürme — sadece "kayıtlı mı" bilgisi
        return jsonify({'ok': True,
            'sunucu': kayitlar.get('sunucu', ''),
            'port': kayitlar.get('port', '587'),
            'kullanici': kayitlar.get('kullanici', ''),
            'gonderen_ad': kayitlar.get('gonderen_ad', ''),
            'gonderen_email': kayitlar.get('gonderen_email', ''),
            'guvenlik': kayitlar.get('guvenlik', 'tls'),
            'sifre_kayitli': bool(kayitlar.get('sifre')),
        })

    @app.route('/api/ayarlar/smtp', methods=['POST'])
    def api_smtp_post():
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        if session.get('rol') not in ('admin', 'ADMIN'):
            return jsonify({'ok': False, 'mesaj': 'Yetkisiz'}), 403
        d = request.json or {}
        def _yaz(deger_ad, deger):
            k = Veriler.query.filter_by(kategori='smtp_ayar', deger=deger_ad).first()
            if not k:
                k = Veriler(kategori='smtp_ayar', deger=deger_ad)
                db.session.add(k)
            k.ek_bilgi = deger
        for alan in ['sunucu', 'port', 'kullanici', 'gonderen_ad', 'gonderen_email', 'guvenlik']:
            if alan in d:
                _yaz(alan, str(d.get(alan) or ''))
        # Şifre: sadece dolu gönderildiyse güncelle (boşsa mevcut korunur)
        if d.get('sifre'):
            _yaz('sifre', str(d.get('sifre')))
        db.session.commit()
        return jsonify({'ok': True, 'mesaj': 'SMTP ayarlari kaydedildi.'})

    @app.route('/api/ayarlar/smtp/test', methods=['POST'])
    def api_smtp_test():
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        if session.get('rol') not in ('admin', 'ADMIN'):
            return jsonify({'ok': False, 'mesaj': 'Yetkisiz'}), 403
        alici = (request.json or {}).get('alici', '').strip()
        if not alici or '@' not in alici:
            return jsonify({'ok': False, 'mesaj': 'Test icin gecerli bir alici e-posta girin.'}), 400

        # TEŞHİS MODU: SMTP sunucusuyla yapılan tüm konuşmayı yakala.
        # Sunucunun MAIL FROM / RCPT TO yanıtlarını görmek, "kabul etti ama
        # teslim etmedi" sorunlarını çözmenin tek yolu.
        import smtplib, ssl as _ssl, io as _io
        from email.mime.text import MIMEText
        from email.utils import formataddr

        ayar = _smtp_ayarlari_oku()
        if not ayar:
            return jsonify({'ok': False, 'mesaj': 'SMTP ayarlari eksik.'}), 400

        log_yakala = _io.StringIO()
        diyalog = []
        try:
            msg = MIMEText('Bu bir test e-postasidir. SMTP ayarlariniz calisiyor.', 'plain', 'utf-8')
            msg['From'] = formataddr((ayar['gonderen_ad'], ayar['gonderen_email']))
            msg['To'] = alici
            msg['Subject'] = 'Milestone ERP - SMTP Test'

            if ayar['guvenlik'] == 'ssl':
                sunucu = smtplib.SMTP_SSL(ayar['sunucu'], ayar['port'],
                                          context=_ssl.create_default_context(), timeout=30)
            else:
                sunucu = smtplib.SMTP(ayar['sunucu'], ayar['port'], timeout=30)
            sunucu.set_debuglevel(1)   # SMTP diyaloğunu aç
            # Debug çıktısını yakala (smtplib stderr'e yazar)
            import sys as _sys
            _eski_stderr = _sys.stderr
            _sys.stderr = log_yakala
            try:
                if ayar['guvenlik'] == 'tls':
                    sunucu.starttls(context=_ssl.create_default_context())
                    sunucu.ehlo()
                sunucu.login(ayar['kullanici'], ayar['sifre'])
                red = sunucu.send_message(msg)
                sunucu.quit()
            finally:
                _sys.stderr = _eski_stderr

            diyalog = log_yakala.getvalue().split('\n')
            if red:
                return jsonify({'ok': False,
                    'mesaj': f'Sunucu alicilari reddetti: {red}',
                    'diyalog': diyalog}), 200
            return jsonify({'ok': True,
                'mesaj': f'Test e-postasi kabul edildi. Sunucu diyalogu asagida. '
                         f'Alici kutusuna dusmezse, sunucunun teslim (relay/SPF) ayarlarini kontrol edin.',
                'diyalog': diyalog})
        except Exception as e:
            try: _sys.stderr = _eski_stderr
            except Exception: pass
            return jsonify({'ok': False,
                'mesaj': f'{type(e).__name__}: {e}',
                'diyalog': log_yakala.getvalue().split('\n')}), 200

    @app.route('/api/ayarlar/uretim_avans', methods=['GET'])
    def api_uretim_avans_get():
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        return jsonify({'ok': True, 'yuzde': _uretim_avans_esigi()})

    @app.route('/api/ayarlar/uretim_avans', methods=['POST'])
    def api_uretim_avans_post():
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        if session.get('rol') not in ('admin', 'ADMIN'):
            return jsonify({'ok': False, 'mesaj': 'Yetkisiz'}), 403
        try:
            yuzde = float((request.json or {}).get('yuzde', 0))
        except (ValueError, TypeError):
            yuzde = 0
        yuzde = max(0, min(100, yuzde))
        kayit = Veriler.query.filter_by(kategori='siparis_ayar', deger='uretim_avans_yuzdesi').first()
        if not kayit:
            kayit = Veriler(kategori='siparis_ayar', deger='uretim_avans_yuzdesi')
            db.session.add(kayit)
        kayit.kisaltma = str(yuzde)
        db.session.commit()
        return jsonify({'ok': True, 'yuzde': yuzde,
                        'mesaj': f'Uretim avans esigi %{yuzde:g} olarak ayarlandi'
                                 + (' (kapi kapali)' if yuzde == 0 else '')})

    @app.route('/api/siparis/<siparis_id>', methods=['DELETE'])
    def api_siparis_sil(siparis_id):
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        sip = Siparis.query.get(siparis_id)
        if not sip: return jsonify({'ok': False, 'mesaj': 'Bulunamadi'}), 404

        if sip.durum != 'Iptal Edildi':
            return jsonify({
                'ok': False,
                'mesaj': f'Bu siparis "{sip.durum}" durumunda. Silmek icin once IPTAL EDILDI durumuna almalisiniz.'
            }), 400

        serbest = _siparis_stoklarini_serbest_birak(siparis_id, 'Siparis silindi')
        rezler = Rezervasyon.query.filter_by(siparis_id=siparis_id).all()
        Maliyet.query.filter_by(baglanti_id=siparis_id, baglanti_tip='siparis').delete()

        _log_audit('SIL', 'siparis', siparis_id, eski={
            'musteri': sip.musteri, 'durum': sip.durum, 'toplam_tutar': sip.toplam_tutar
        })
        # cascade='all, delete-orphan' SiparisKalem'i otomatik silecek
        db.session.delete(sip)
        db.session.commit()
        logging.info(f"Siparis silindi: {siparis_id} ({len(rezler)} rezervasyon) - by {session.get('kullanici')}")
        return jsonify({'ok': True, 'mesaj': f'Siparis silindi, {len(rezler)} rezervasyon iptal edildi'})


    # ─── MÜŞTERİ BAZLI LİSTE (geri uyum için korundu) ───
    @app.route('/api/siparis/musteri/<musteri>', methods=['GET'])
    def api_siparis_musteri(musteri):
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        siparisler = Siparis.query.filter_by(musteri=musteri).order_by(Siparis.siparis_tarihi.desc()).all()
        result = []
        for s in siparisler:
            rez_sayisi = Rezervasyon.query.filter_by(siparis_id=s.id, iptal_nedeni=None).count()
            ilk_kalem = s.kalemler[0] if s.kalemler else None
            result.append({
                'id': s.id, 'tarih': s.siparis_tarihi.isoformat() if s.siparis_tarihi else None,
                'urun_tip': ilk_kalem.urun_tip if ilk_kalem else None,
                'cins': ilk_kalem.cins if ilk_kalem else None,
                'miktar': sum((k.miktar or 0) for k in s.kalemler),
                'birim': ilk_kalem.birim if ilk_kalem else None,
                'doviz': s.doviz, 'durum': s.durum,
                'toplam_tutar': s.toplam_tutar or 0,
                'rezervasyon_sayisi': rez_sayisi,
                'kalem_sayisi': len(s.kalemler)
            })
        return jsonify(result)

    @app.route('/api/siparis/<siparis_id>/stoklar', methods=['GET'])
    def api_siparis_stoklar(siparis_id):
        """Bir siparise bagli aktif rezervasyonlarin stok detaylarini dondurur."""
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        rezler = Rezervasyon.query.filter_by(siparis_id=siparis_id, iptal_nedeni=None).all()
        stoklar = []
        for r in rezler:
            if r.stok_tip == 'BLOK':
                s = BlokStok.query.get(r.stok_id)
                if not s: continue
                stoklar.append({
                    'stok_id': s.id, 'stok_tip': 'BLOK',
                    'cins': s.cins, 'ozellik': '',
                    'blok_no': s.blok_no,
                    'boy': s.boy, 'yukseklik': s.yukseklik, 'en': s.en,
                    'kalinlik': None, 'adet': 1,
                    'm2': None, 'sqft': None,
                    'hacim_m3': s.hacim_m3, 'tonaj': s.tonaj,
                    'agirlik_kg': (s.tonaj * 1000) if s.tonaj else None,
                    'birim_fiyat': s.alis_fiyati, 'doviz': s.doviz
                })
            elif r.stok_tip == 'PLAKA':
                s = PlakaStok.query.get(r.stok_id)
                if not s: continue
                stoklar.append({
                    'stok_id': s.id, 'stok_tip': 'PLAKA',
                    'cins': s.cins, 'ozellik': s.ozellik or '',
                    'blok_no': s.blok_no,
                    'boy': s.boy, 'yukseklik': s.yukseklik, 'kalinlik': s.kalinlik,
                    'adet': 1,
                    'm2': s.metraj_m2, 'sqft': s.metraj_sqft,
                    'agirlik_kg': s.m2_kg,
                    'birim_fiyat': s.alis_fiyati, 'doviz': s.doviz
                })
            else:
                s = EbatliStok.query.get(r.stok_id)
                if not s: continue
                stoklar.append({
                    'stok_id': s.id, 'stok_tip': 'EBATLI',
                    'cins': s.cins, 'ozellik': s.ozellik or '',
                    'blok_no': s.kasa_no,
                    'boy': s.boy, 'yukseklik': s.yukseklik, 'kalinlik': s.kalinlik,
                    'adet': s.kasa_ici_adet or 1,
                    'm2': s.metraj_m2, 'sqft': s.metraj_sqft,
                    'agirlik_kg': s.m2_kg,
                    'birim_fiyat': s.alis_fiyati, 'doviz': s.doviz
                })

        sip = Siparis.query.get(siparis_id)
        # FAZ 16: satis_fiyati Siparis'ten kalktı; ilk kalemin birim fiyati temsili.
        _ilk_kalem = (SiparisKalem.query.filter_by(siparis_id=siparis_id)
                      .order_by(SiparisKalem.sira).first()) if sip else None
        return jsonify({
            'ok': True,
            'stoklar': stoklar,
            'satis_fiyati': (_ilk_kalem.birim_fiyat if _ilk_kalem else 0) or 0,
            'doviz': sip.doviz if sip else 'USD',
            'odeme_sekli': sip.odeme_sekli if sip else '',
            'teslim_sekli': sip.teslim_sekli if sip else '',
            'termin': sip.termin.isoformat() if sip and sip.termin else None
        })

    # ---------- API: REZERVASYON ----------
    @app.route('/api/rezervasyon', methods=['GET'])
    def api_rezervasyon_liste():
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        sadece_aktif = request.args.get('sadece_aktif', '0') == '1'
        q = Rezervasyon.query
        if sadece_aktif:
            q = q.filter_by(iptal_nedeni=None)
        rezler = q.order_by(Rezervasyon.id.desc()).all()
        sonuc = []
        for r in rezler:
            cins = None
            ozellik = None
            blok_no = None
            boy = yukseklik = en = kalinlik = None
            kasa_no = None
            try:
                if r.stok_tip == 'BLOK':
                    s = BlokStok.query.get(r.stok_id)
                elif r.stok_tip == 'PLAKA':
                    s = PlakaStok.query.get(r.stok_id)
                else:
                    s = EbatliStok.query.get(r.stok_id)
                if s:
                    cins = s.cins
                    ozellik = getattr(s, 'ozellik', None)
                    blok_no = getattr(s, 'blok_no', None)
                    boy = getattr(s, 'boy', None)
                    yukseklik = getattr(s, 'yukseklik', None)
                    en = getattr(s, 'en', None)
                    kalinlik = getattr(s, 'kalinlik', None)
                    kasa_no = getattr(s, 'kasa_no', None)
            except Exception:
                pass
            sonuc.append({
                'id': r.id, 'musteri': r.musteri,
                'stok_tip': r.stok_tip, 'stok_id': r.stok_id,
                'siparis_id': getattr(r, 'siparis_id', None),
                'proforma_id': getattr(r, 'proforma_id', None),
                'iptal_nedeni': r.iptal_nedeni,
                'iptal_tarihi': r.iptal_tarihi.isoformat() if getattr(r, 'iptal_tarihi', None) else None,
                'olusturma_tarihi': r.olusturma_tarihi.isoformat() if getattr(r, 'olusturma_tarihi', None) else None,
                'kullanici': getattr(r, 'kullanici', None),
                'cins': cins, 'ozellik': ozellik, 'blok_no': blok_no,
                'boy': boy, 'yukseklik': yukseklik, 'en': en, 'kalinlik': kalinlik,
                'kasa_no': kasa_no,
                # Kullanıcıya gösterilecek okunabilir numara (ID değil)
                'stok_no': (kasa_no or blok_no or r.stok_id)
            })
        return jsonify(sonuc)

    @app.route('/api/rezervasyon', methods=['POST'])
    def api_rezervasyon_ekle():
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        data = request.json
        stok_ids = data.get('stok_idler', [])
        if not stok_ids:
            return jsonify({'ok': False, 'mesaj': 'Stok seçilmedi'}), 400

        siparis_id = data.get('siparis_id') or None
        musteri = data.get('musteri')
        stok_tip = data.get('stok_tip')
        if not stok_tip:
            return jsonify({'ok': False, 'mesaj': 'Stok tipi belirtilmedi'}), 400

        # Siparis varsa musteriyi ondan al + cins/ozellik bilgisi
        sip = None
        if siparis_id:
            sip = Siparis.query.get(siparis_id)
            if not sip:
                return jsonify({'ok': False, 'mesaj': f'Siparis bulunamadi: {siparis_id}'}), 404
            if not musteri:
                musteri = sip.musteri
        if not musteri:
            return jsonify({'ok': False, 'mesaj': 'Musteri belirtilmedi'}), 400

        olusturulan = []
        for sid in stok_ids:
            # Stogu bul + Serbest mi kontrol et
            if stok_tip == 'BLOK':
                stok = BlokStok.query.get(sid)
            elif stok_tip == 'PLAKA':
                stok = PlakaStok.query.get(sid)
            else:
                stok = EbatliStok.query.get(sid)
            if not stok:
                continue
            if stok.durum != 'Serbest':
                continue  # Zaten rezerve veya satıldı - atla

            # Bu stok bu siparise zaten aktif rezerve mi? (cift kayit korumasi)
            mevcut = Rezervasyon.query.filter_by(
                siparis_id=siparis_id, stok_id=sid, iptal_nedeni=None
            ).first()
            if mevcut:
                continue  # Zaten rezerve, atla

            # SİPARİŞ KALEMİ BAĞLANTISI (FAZ 16 zorunluluğu):
            # Teslimde SatisKaydi olusturulabilmesi icin rezervasyon bir sipariş
            # KALEMİNE bagli olmali. Cagiran acikca gonderdiyse onu kullan; yoksa
            # siparisin kalemleri icinde urun tipi + cins ile eslestir.
            kalem_id = data.get('siparis_kalem_id')
            if siparis_id and not kalem_id:
                _kalemler = SiparisKalem.query.filter_by(siparis_id=siparis_id).all()
                _stok_cins = (getattr(stok, 'cins', '') or '').strip().upper()
                # 1) Ürün tipi + cins birebir
                for _k in _kalemler:
                    if (_k.urun_tip or '').upper() == stok_tip.upper() and \
                       (_k.cins or '').strip().upper() == _stok_cins:
                        kalem_id = _k.id
                        break
                # 2) Sadece ürün tipi
                if not kalem_id:
                    for _k in _kalemler:
                        if (_k.urun_tip or '').upper() == stok_tip.upper():
                            kalem_id = _k.id
                            break
                # 3) Tek kalemli siparişte o kalem
                if not kalem_id and len(_kalemler) == 1:
                    kalem_id = _kalemler[0].id
                if not kalem_id:
                    app.logger.warning(
                        f'Rezervasyon: {sid} stogu icin siparis {siparis_id} kaleminde '
                        f'eslesme bulunamadi (tip={stok_tip}, cins={_stok_cins}). '
                        f'Teslimde satis kaydi olusmayabilir.')

            rez = Rezervasyon(
                id=_yeni_id('REZ'),
                musteri=musteri,
                siparis_id=siparis_id,
                siparis_kalem_id=kalem_id,
                stok_tip=stok_tip,
                stok_id=sid,
                cins=getattr(stok, 'cins', None),
                ozellik=getattr(stok, 'ozellik', None),
                miktar=(getattr(stok, 'metraj_m2', None) or getattr(stok, 'hacim_m3', None)),
                kullanici=session['kullanici']
            )
            db.session.add(rez)
            eski_durum = stok.durum
            stok.durum = 'Rezerve'
            # AUDIT: stok durum degisimi - hangi siparis icin
            _log_audit('DURUM', 'stok', sid,
                       eski={'durum': eski_durum},
                       yeni={'durum': 'Rezerve'},
                       aciklama=f'Siparis {siparis_id} icin rezerve edildi' if siparis_id else 'Rezerve edildi')
            olusturulan.append(sid)

        ok, hata = _safe_commit('Rezervasyon ekleme')
        if not ok:
            return jsonify({'ok': False, 'mesaj': f'Hata: {hata}'}), 500
        return jsonify({
            'ok': True,
            'olusturulan': olusturulan,
            'mesaj': f'{len(olusturulan)} stok rezerve edildi' + (f' (siparis: {siparis_id})' if siparis_id else '')
        })

    # ---------- API: MALİYET ----------
    @app.route('/api/stok/gruplar', methods=['GET'])
    def api_stok_gruplar():
        """Cari hareket icin gruplandirilmis stok listesi.
        BLOK: her stok ayri (id ile)
        PLAKA: blok_no'ya gore grup (ayni blok = ayni alim)
        EBATLI: kasa_no veya referans'a gore grup
        """
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        # Sadece aktif stoklari (Teslim Edildi haric) goster
        aktif_durumlar = ['Serbest', 'Rezerve', 'Satildi']

        gruplar = {'blok': [], 'plaka': [], 'ebatli': []}

        # BLOK - her stok ayri grup
        for s in BlokStok.query.filter(BlokStok.durum.in_(aktif_durumlar)).all():
            _fd = getattr(s, 'fatura_durumu', 'faturali') or 'faturali'
            _ek = ' · ⏳ FATURASIZ' if _fd == 'faturasiz' else ''
            _ref = s.blok_no or s.id
            _matrah = s.matrah or 0
            _kdv = s.kdv_tutar or 0
            gruplar['blok'].append({
                'grup_id': s.id,  # baglanti_id olarak kullanilacak
                'etiket': f"{_ref} · {s.cins or '-'} · {s.tonaj or 0:.1f} ton{_ek}",
                'tip': 'BLOK',
                'cins': s.cins,
                'blok_no': s.blok_no,
                'stok_idler': [s.id],
                'toplam_miktar': s.tonaj or 0,
                'birim': 'ton',
                'adet': 1,
                'fatura_durumu': _fd,
                'matrah': q3(_matrah),
                'kdv': q3(_kdv),
                'toplam': q3(_matrah + _kdv),
                'doviz': s.doviz or 'USD'
            })

        # PLAKA - blok_no'ya gore grupla
        from collections import defaultdict
        plaka_gruplar = defaultdict(list)
        for s in PlakaStok.query.filter(PlakaStok.durum.in_(aktif_durumlar)).all():
            anahtar = s.blok_no or s.id  # blok_no yoksa kendi id'si
            plaka_gruplar[anahtar].append(s)

        for blok_no, stoklar in plaka_gruplar.items():
            ilk = stoklar[0]
            toplam_m2 = sum(s.metraj_m2 or 0 for s in stoklar)
            _faturasiz_var = any(getattr(s, 'fatura_durumu', 'faturali') == 'faturasiz' for s in stoklar)
            _ek = ' · ⏳ FATURASIZ' if _faturasiz_var else ''
            # Sadece faturasız olanların tutarı (faturalandırma için)
            _fsiz = [s for s in stoklar if getattr(s, 'fatura_durumu', 'faturali') == 'faturasiz']
            _hedef = _fsiz if _fsiz else stoklar
            _matrah = sum(s.matrah or 0 for s in _hedef)
            _kdv = sum(s.kdv_tutar or 0 for s in _hedef)
            gruplar['plaka'].append({
                'grup_id': f"BLOK:{blok_no}",  # baglanti_id format
                'etiket': f"{blok_no} · {ilk.cins or '-'} · {len(stoklar)} plaka · {toplam_m2:.2f} m²{_ek}",
                'tip': 'PLAKA',
                'cins': ilk.cins,
                'blok_no': blok_no,
                'stok_idler': [s.id for s in stoklar],
                'toplam_miktar': toplam_m2,
                'birim': 'm²',
                'adet': len(stoklar),
                'fatura_durumu': 'faturasiz' if _faturasiz_var else 'faturali',
                'matrah': q3(_matrah),
                'kdv': q3(_kdv),
                'toplam': q3(_matrah + _kdv),
                'doviz': ilk.doviz or 'USD'
            })

        # EBATLI - kasa_no varsa onun ortak prefix'ine, yoksa cins+olcu kombinasyonuna gore
        # Aslinda kasa_no genelde "PREFIX-1", "PREFIX-2" formatinda
        import re as _re
        ebatli_gruplar = defaultdict(list)
        for s in EbatliStok.query.filter(EbatliStok.durum.in_(aktif_durumlar)).all():
            # Kasa_no'dan PREFIX cikar (sondaki -N)
            if s.kasa_no:
                m = _re.match(r'^(.*?)-\d+$', s.kasa_no)
                anahtar = m.group(1) if m else s.kasa_no
            else:
                # Kasa_no yok: cins + olcu kombinasyonu
                anahtar = f"{s.cins or '?'}_{s.boy or 0}x{s.yukseklik or 0}x{s.kalinlik or 0}"
            ebatli_gruplar[anahtar].append(s)

        for anahtar, stoklar in ebatli_gruplar.items():
            ilk = stoklar[0]
            toplam_m2 = sum(((s.boy or 0) * (s.yukseklik or 0) / 10000) * (s.kasa_ici_adet or 1)
                           for s in stoklar)
            _faturasiz_var = any(getattr(s, 'fatura_durumu', 'faturali') == 'faturasiz' for s in stoklar)
            _ek = ' · ⏳ FATURASIZ' if _faturasiz_var else ''
            _fsiz = [s for s in stoklar if getattr(s, 'fatura_durumu', 'faturali') == 'faturasiz']
            _hedef = _fsiz if _fsiz else stoklar
            _matrah = sum(s.matrah or 0 for s in _hedef)
            _kdv = sum(s.kdv_tutar or 0 for s in _hedef)
            gruplar['ebatli'].append({
                'grup_id': f"EBATLI_GRP:{anahtar}",
                'etiket': f"{anahtar} · {ilk.cins or '-'} · {len(stoklar)} kasa · {toplam_m2:.2f} m²{_ek}",
                'tip': 'EBATLI',
                'cins': ilk.cins,
                'kasa_prefix': anahtar,
                'olculer': f"{ilk.boy}x{ilk.yukseklik}x{ilk.kalinlik}" if ilk.boy else '',
                'stok_idler': [s.id for s in stoklar],
                'toplam_miktar': toplam_m2,
                'birim': 'm²',
                'adet': len(stoklar),
                'fatura_durumu': 'faturasiz' if _faturasiz_var else 'faturali',
                'matrah': q3(_matrah),
                'kdv': q3(_kdv),
                'toplam': q3(_matrah + _kdv),
                'doviz': ilk.doviz or 'USD'
            })

        return jsonify(gruplar)

    @app.route('/api/maliyet', methods=['GET'])
    def api_maliyet_liste():
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        # ?pasifler=1 ile pasif kayitlari da getir (default: sadece aktif)
        pasifler_dahil = request.args.get('pasifler', '0') in ('1', 'true', 'True')
        q = Maliyet.query
        if not pasifler_dahil:
            q = q.filter(Maliyet.aktif == True)
        # Kararlı sıralama: önce tarih, aynı tarihte kayıt zamanı (yeni üstte).
        # olusturma NULL olan eski kayıtlar en sona düşmesin diye id ile tamamlanır.
        maliyetler = q.order_by(Maliyet.maliyet_tarihi.desc(),
                                Maliyet.olusturma.desc().nullslast(),
                                Maliyet.id.desc()).all()

        # Stok ID -> miktar + birim ham haritası (birim başına maliyet hesabı için)
        # BLOK -> ton, PLAKA/EBATLI -> m²
        stok_idler = list({m.baglanti_id for m in maliyetler
                          if m.baglanti_tip and m.baglanti_tip.lower() == 'stok' and m.baglanti_id})
        # Stok ID -> (miktar, birim) + okunabilir no / alış bedeli bilgisi
        stok_birim_map = {}   # {stok_id: (miktar, birim_adi)}
        stok_bilgi_map = {}   # {stok_id: {...}} — okunabilir no + alış bedeli
        if stok_idler:
            # BLOK
            for s in BlokStok.query.filter(BlokStok.id.in_(stok_idler)).all():
                stok_birim_map[s.id] = (s.tonaj or 0, 'ton')
                stok_bilgi_map[s.id] = {
                    'no': s.blok_no or s.id, 'tip': 'BLOK', 'cins': s.cins,
                    'matrah': s.matrah or 0, 'doviz': s.doviz or 'USD',
                    'fatura_no': s.fatura_no,
                    'tarih': (s.alis_tarihi or s.giris_tarihi)}
            # PLAKA
            for s in PlakaStok.query.filter(PlakaStok.id.in_(stok_idler)).all():
                stok_birim_map[s.id] = (s.metraj_m2 or 0, 'm²')
                _pno = s.blok_no or s.id
                if getattr(s, 'slab_no', None):
                    _pno = f"{s.blok_no or ''}#{s.slab_no}".strip('#')
                stok_bilgi_map[s.id] = {
                    'no': _pno, 'tip': 'PLAKA', 'cins': s.cins,
                    'matrah': s.matrah or 0, 'doviz': s.doviz or 'USD',
                    'fatura_no': s.fatura_no,
                    'tarih': (s.alis_tarihi or s.giris_tarihi)}
            # EBATLI
            for s in EbatliStok.query.filter(EbatliStok.id.in_(stok_idler)).all():
                stok_birim_map[s.id] = (s.metraj_m2 or 0, 'm²')
                stok_bilgi_map[s.id] = {
                    'no': s.kasa_no or s.id, 'tip': 'EBATLI', 'cins': s.cins,
                    'matrah': s.matrah or 0, 'doviz': s.doviz or 'USD',
                    'fatura_no': s.fatura_no,
                    'tarih': (s.alis_tarihi or s.giris_tarihi)}

        # Stok dışı bağlantılar için okunabilir no (sipariş/fatura zaten anlamlı ID taşır)
        def _okunabilir_no(m):
            if m.baglanti_tip and m.baglanti_tip.lower() == 'stok':
                bilgi = stok_bilgi_map.get(m.baglanti_id)
                if bilgi:
                    return bilgi['no']
            return m.baglanti_id

        sonuc = []
        for m in maliyetler:
            birim_maliyet = None
            birim_ad = None
            if m.baglanti_tip and m.baglanti_tip.lower() == 'stok' and m.baglanti_id in stok_birim_map:
                miktar, birim_ad = stok_birim_map[m.baglanti_id]
                if miktar and miktar > 0:
                    birim_maliyet = q3((m.tutar or 0) / miktar)
            sonuc.append({
                'id': m.id, 'maliyet_tarihi': m.maliyet_tarihi.isoformat(),
                'maliyet_tip': m.maliyet_tip, 'baglanti_tip': m.baglanti_tip,
                'baglanti_id': m.baglanti_id,
                'baglanti_no': _okunabilir_no(m),
                'tutar': m.tutar, 'doviz': m.doviz,
                'usd_karsilik': m.usd_karsilik,
                'fatura_no': m.fatura_no,
                'aciklama': m.aciklama,
                'kur': m.kur,
                'birim_maliyet': birim_maliyet,
                'birim_ad': birim_ad,
                'aktif': m.aktif,
                'sanal': False,
                '_sira': (m.olusturma.isoformat() if m.olusturma else ''),
                'donusum_id': m.donusum_id,
                'donusum_tarihi': m.donusum_tarihi.isoformat() if m.donusum_tarihi else None
            })

        # ── SANAL "ALIŞ BEDELİ" SATIRLARI ──
        # Alış bedeli stok kartında tutulur (Maliyet tablosunda DEĞİL) — kâr hesabı
        # onu stoktan okur. Maliyet tablosuna eklemek ÇİFT SAYMA olurdu. Bu yüzden
        # listeye salt-görüntü (düzenlenemez/silinemez) sanal satır olarak eklenir;
        # kullanıcı bir stoğun toplam maliyet resmini tek yerde görür.
        if request.args.get('alis_dahil', '1') in ('1', 'true', 'True'):
            for stok_id, bilgi in stok_bilgi_map.items():
                if not bilgi['matrah'] or bilgi['matrah'] <= 0:
                    continue
                miktar, birim_ad2 = stok_birim_map.get(stok_id, (0, None))
                sonuc.append({
                    'id': f'ALIS:{stok_id}',
                    'maliyet_tarihi': (bilgi['tarih'] or date.today()).isoformat(),
                    'maliyet_tip': 'Alış Bedeli',
                    'baglanti_tip': 'stok', 'baglanti_id': stok_id,
                    'baglanti_no': bilgi['no'],
                    'tutar': q3(bilgi['matrah']), 'doviz': bilgi['doviz'],
                    'usd_karsilik': q3(bilgi['matrah']) if bilgi['doviz'] == 'USD' else None,
                    'fatura_no': bilgi['fatura_no'],
                    'aciklama': f"{bilgi['tip']} alış bedeli — {bilgi['cins'] or ''}".strip(),
                    'kur': None,
                    'birim_maliyet': q3(bilgi['matrah'] / miktar) if miktar else None,
                    'birim_ad': birim_ad2,
                    'aktif': True,
                    'sanal': True,       # UI: düzenle/sil butonu gösterme
                    '_sira': '',         # aynı tarihte alış bedeli en üstte kalsın
                    'donusum_id': None, 'donusum_tarihi': None
                })

        # KARARLI SIRALAMA — her durumda uygulanır (sanal satır olsun olmasın):
        # 1) tarih (yeni üstte)  2) kayıt zamanı  3) id (eşitlik bozucu)
        sonuc.sort(key=lambda x: (x['maliyet_tarihi'], x.get('_sira') or '', x['id']),
                   reverse=True)
        for x in sonuc:
            x.pop('_sira', None)

        return jsonify(sonuc)

    @app.route('/api/maliyet', methods=['POST'])
    def api_maliyet_ekle():
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        data = request.json
        doviz       = data.get('doviz', 'USD')
        maliyet_tip = data.get('maliyet_tip')
        # Bağlantı tipini küçük harfe normalleştir — bazı ekranlar 'Stok',
        # bazıları 'stok' gönderiyordu; sorgular kaçırıyordu.
        baglanti_tip = (data.get('baglanti_tip') or '').strip().lower() or None
        baglanti_id  = data.get('baglanti_id')
        if not maliyet_tip or not baglanti_tip or not baglanti_id:
            return jsonify({'ok': False, 'mesaj': 'maliyet_tip, baglanti_tip ve baglanti_id zorunludur'}), 400
        fatura_no   = (data.get('fatura_no') or '').strip()
        cari_id     = data.get('cari_id') or None
        kullanici   = session['kullanici']
        kdv_dahil_mi = bool(data.get('kdv_dahil_mi', False))
        kdv_oran     = float(data.get('kdv_oran', 0) or 0)
        tutar_ham    = float(data.get('tutar', 0) or 0)

        # Net matrah ve KDV hesapla
        if kdv_dahil_mi and kdv_oran > 0:
            net_tutar = round(tutar_ham / (1 + kdv_oran / 100), 4)
            kdv_tutar = round(tutar_ham - net_tutar, 4)
        elif not kdv_dahil_mi and kdv_oran > 0:
            net_tutar = tutar_ham
            kdv_tutar = round(tutar_ham * kdv_oran / 100, 4)
        else:
            net_tutar = tutar_ham
            kdv_tutar = 0.0

        def _usd(t, dv):
            if dv == 'USD': return t
            if dv == 'EUR': return t * 1.08
            k = DovizKur.query.filter_by(doviz='USD').order_by(DovizKur.tarih.desc()).first()
            return t / (k.efektif if k else 45.07)

        # Ana maliyet kaydı
        m = Maliyet(id=_yeni_id('MYT'), maliyet_tip=maliyet_tip,
                    baglanti_tip=baglanti_tip, baglanti_id=baglanti_id,
                    tutar=net_tutar, doviz=doviz, usd_karsilik=_usd(net_tutar, doviz),
                    fatura_no=fatura_no, kullanici=kullanici)
        db.session.add(m)

        # KDV maliyet kaydı
        kdv_m = None
        if kdv_tutar > 0:
            kdv_m = Maliyet(id=_yeni_id('MYT'), maliyet_tip='Devreden KDV',
                            baglanti_tip=baglanti_tip, baglanti_id=baglanti_id,
                            tutar=kdv_tutar, doviz=doviz, usd_karsilik=_usd(kdv_tutar, doviz),
                            fatura_no=fatura_no,
                            aciklama=f'KDV %{kdv_oran} — {maliyet_tip} ({fatura_no or "belgesiz"})',
                            kullanici=kullanici)
            db.session.add(kdv_m)

        # Cari entegrasyon: toplam tutar (net + KDV) alacak olarak cariye işle
        if cari_id:
            cari = Cari.query.get(cari_id)
            if cari:
                toplam_fatura = net_tutar + kdv_tutar
                ch = CariHareket(
                    id=_yeni_id('HR'),
                    hareket_tarihi=date.today(),
                    cari_id=cari_id, cari_unvan=cari.unvan,
                    islem_tip='Nakliye/Gider Faturası',
                    aciklama=f'{maliyet_tip} — {_baglanti_okunabilir(baglanti_tip, baglanti_id)} — {fatura_no or "belgesiz"}',
                    borc=0, alacak=toplam_fatura,
                    alacak_try=_usd(toplam_fatura, doviz) * (DovizKur.query.filter_by(doviz='USD').order_by(DovizKur.tarih.desc()).first().efektif if doviz != 'TRY' else 1),
                    doviz=doviz, kur_uygulanan=1.0,
                    evrak_no=fatura_no, baglanti_tip='maliyet', baglanti_id=m.id,
                    kaynak='maliyet', kullanici=kullanici)
                db.session.add(ch)

        db.session.flush()
        sk_guncel = _satis_kaydi_maliyet_guncelle([baglanti_id])
        # Maliyet bir BLOĞA eklendiyse ve blok kesilmişse, plakaları yeniden hesapla
        if baglanti_id and str(baglanti_id).startswith('BLK'):
            _kesilmis_blok_maliyet_yeniden_dagit(baglanti_id)
        db.session.commit()

        msg = f'Maliyet kaydedildi (Net: {net_tutar:,.2f} {doviz})'
        if kdv_tutar > 0: msg += f', KDV: {kdv_tutar:,.2f} {doviz}'
        if cari_id: msg += ' — cari alacak kaydı oluşturuldu'
        return jsonify({'ok': True, 'id': m.id,
                        'kdv_id': kdv_m.id if kdv_m else None,
                        'net_tutar': net_tutar, 'kdv_tutar': kdv_tutar,
                        'satis_kaydi_guncel': sk_guncel, 'mesaj': msg})

    @app.route('/api/maliyet/<maliyet_id>', methods=['PUT'])
    def api_maliyet_guncelle(maliyet_id):
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        # Sanal "Alış Bedeli" satırı düzenlenemez/silinemez — stok kartından gelir
        if (maliyet_id or '').startswith('ALIS:'):
            return jsonify({'ok': False, 'error': 'sanal_kalem',
                'mesaj': 'Alış bedeli stok kartında tutulur, maliyet kaydı degildir. '
                         'Degistirmek icin stok kartindaki alis fiyatini duzenleyin.'}), 400
        m = Maliyet.query.get(maliyet_id)
        if not m: return jsonify({'ok': False, 'mesaj': 'Bulunamadı'}), 404
        data = request.json
        if 'maliyet_tip' in data: m.maliyet_tip = data['maliyet_tip']
        if 'tutar' in data: m.tutar = data['tutar']
        if 'doviz' in data: m.doviz = data['doviz']
        usd_karsilik = m.tutar
        if m.doviz == 'EUR':
            usd_karsilik = m.tutar * 1.08
        elif m.doviz == 'TRY':
            usd_kur = DovizKur.query.filter_by(doviz='USD').order_by(DovizKur.tarih.desc()).first()
            usd_karsilik = m.tutar / (usd_kur.efektif if usd_kur else 45.07)
        m.usd_karsilik = usd_karsilik
        db.session.flush()
        sk_guncel = _satis_kaydi_maliyet_guncelle([m.baglanti_id])
        if m.baglanti_id and str(m.baglanti_id).startswith('BLK'):
            _kesilmis_blok_maliyet_yeniden_dagit(m.baglanti_id)
        db.session.commit()
        return jsonify({'ok': True, 'satis_kaydi_guncel': sk_guncel})

    @app.route('/api/maliyet/<maliyet_id>', methods=['DELETE'])
    def api_maliyet_sil(maliyet_id):
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        # Sanal "Alış Bedeli" satırı düzenlenemez/silinemez — stok kartından gelir
        if (maliyet_id or '').startswith('ALIS:'):
            return jsonify({'ok': False, 'error': 'sanal_kalem',
                'mesaj': 'Alış bedeli stok kartında tutulur, maliyet kaydı degildir. '
                         'Degistirmek icin stok kartindaki alis fiyatini duzenleyin.'}), 400
        m = Maliyet.query.get(maliyet_id)
        if not m: return jsonify({'ok': False, 'mesaj': 'Bulunamadı'}), 404
        etkilenen_stok = m.baglanti_id

        # KDV kardeş kaydını bul: aynı fatura_no + aynı stok + Devreden KDV tipi
        kardes_ids = []
        if m.fatura_no and m.baglanti_id:
            kardesler = Maliyet.query.filter(
                Maliyet.fatura_no == m.fatura_no,
                Maliyet.baglanti_id == m.baglanti_id,
                Maliyet.id != maliyet_id,
                Maliyet.maliyet_tip == 'Devreden KDV'
            ).all()
            for k in kardesler:
                kardes_ids.append(k.id)
                # Kardeşin cari hareketini de sil
                CariHareket.query.filter_by(
                    baglanti_tip='maliyet', baglanti_id=k.id
                ).delete()
                db.session.delete(k)

        # Ana maliyetin cari hareketlerini sil
        silinen_ch = CariHareket.query.filter_by(
            baglanti_tip='maliyet', baglanti_id=maliyet_id
        ).delete()

        db.session.delete(m)
        db.session.flush()
        sk_guncel = _satis_kaydi_maliyet_guncelle([etkilenen_stok] if etkilenen_stok else [])
        db.session.commit()
        msg = 'Maliyet silindi'
        if kardes_ids: msg += f', KDV kaydı da silindi ({", ".join(kardes_ids)})'
        if silinen_ch: msg += f', {silinen_ch} cari hareket de silindi'
        return jsonify({'ok': True, 'silinen_kdv': kardes_ids,
                        'silinen_cari_hareket': silinen_ch,
                        'satis_kaydi_guncel': sk_guncel, 'mesaj': msg})

    @app.route('/api/maliyet/blok_dagilim', methods=['POST'])
    def api_maliyet_blok_dagilim():
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        data = request.json
        urun_tipi   = data.get('urun_tipi')
        blok_no     = data.get('blok_no')
        maliyet_tip = data.get('maliyet_tip')
        doviz       = data.get('doviz', 'USD')
        hedef       = data.get('hedef', 'hepsi')
        fatura_no   = (data.get('fatura_no') or '').strip()
        cari_id     = data.get('cari_id') or None
        kdv_dahil_mi = bool(data.get('kdv_dahil_mi', False))
        kdv_oran     = float(data.get('kdv_oran', 0) or 0)
        tutar_ham    = float(data.get('tutar', 0) or 0)
        kullanici   = session['kullanici']

        if not urun_tipi or not blok_no or not tutar_ham:
            return jsonify({'ok': False, 'mesaj': 'Ürün Tipi, Blok No ve Tutar zorunludur'}), 400

        # Net matrah ve KDV
        if kdv_dahil_mi and kdv_oran > 0:
            toplam_net = round(tutar_ham / (1 + kdv_oran / 100), 4)
            toplam_kdv = round(tutar_ham - toplam_net, 4)
        elif not kdv_dahil_mi and kdv_oran > 0:
            toplam_net = tutar_ham
            toplam_kdv = round(tutar_ham * kdv_oran / 100, 4)
        else:
            toplam_net = tutar_ham
            toplam_kdv = 0.0

        def _usd(t, dv):
            if dv == 'USD': return t
            if dv == 'EUR': return t * 1.08
            k = DovizKur.query.filter_by(doviz='USD').order_by(DovizKur.tarih.desc()).first()
            return t / (k.efektif if k else 45.07)

        SATILMIS_DURUMLAR = ['Satildi', 'Sevkedildi', 'Teslim Edildi']
        def _durum_uygun(durum):
            if hedef == 'serbest': return durum in ('Serbest', 'Rezerve')
            if hedef == 'satilmis': return durum in SATILMIS_DURUMLAR
            return True

        if urun_tipi == 'BLOK':
            hepsi_stok = BlokStok.query.filter_by(blok_no=blok_no).all()
            miktar_f = lambda s: s.hacim_m3 or 1
        elif urun_tipi == 'PLAKA':
            hepsi_stok = PlakaStok.query.filter_by(blok_no=blok_no).all()
            miktar_f = lambda s: s.metraj_m2 or 1
        else:
            hepsi_stok = EbatliStok.query.filter_by(kasa_no=blok_no).all()
            miktar_f = lambda s: s.metraj_m2 or 1

        stoklar = [s for s in hepsi_stok if _durum_uygun(s.durum)]
        if not stoklar:
            return jsonify({'ok': False,
                'mesaj': f'Bu blok ({blok_no}) için "{hedef}" hedefine uygun {urun_tipi} stok bulunamadı'}), 404

        toplam_miktar = sum(miktar_f(s) for s in stoklar)
        if toplam_miktar <= 0:
            return jsonify({'ok': False, 'mesaj': 'Toplam miktar sıfır, dağıtım yapılamaz'}), 400

        eklenen = 0
        for stok in stoklar:
            miktar = miktar_f(stok)
            oran = miktar / toplam_miktar
            net_pay = q3(toplam_net * oran)
            kdv_pay = q3(toplam_kdv * oran) if toplam_kdv > 0 else 0
            if net_pay <= 0:
                continue

            m = Maliyet(id=_yeni_id('MYT'), maliyet_tip=maliyet_tip,
                        baglanti_tip='stok', baglanti_id=stok.id,
                        tutar=net_pay, doviz=doviz, usd_karsilik=_usd(net_pay, doviz),
                        fatura_no=fatura_no,
                        aciklama=f'Toplu dağıtım — Blok {blok_no} ({datetime.now().strftime("%Y-%m-%d")})',
                        kullanici=kullanici)
            db.session.add(m)

            if kdv_pay > 0:
                kdv_m = Maliyet(id=_yeni_id('MYT'), maliyet_tip='Devreden KDV',
                                baglanti_tip='stok', baglanti_id=stok.id,
                                tutar=kdv_pay, doviz=doviz, usd_karsilik=_usd(kdv_pay, doviz),
                                fatura_no=fatura_no,
                                aciklama=f'KDV %{kdv_oran} — {maliyet_tip} Blok {blok_no}',
                                kullanici=kullanici)
                db.session.add(kdv_m)
            eklenen += 1

        # Cari entegrasyon: toplam fatura tutarı (net + KDV) alacak kaydı
        if cari_id:
            cari = Cari.query.get(cari_id)
            if cari:
                toplam_fatura = toplam_net + toplam_kdv
                ch = CariHareket(
                    id=_yeni_id('HR'),
                    hareket_tarihi=date.today(),
                    cari_id=cari_id, cari_unvan=cari.unvan,
                    islem_tip='Nakliye/Gider Faturası',
                    aciklama=f'{maliyet_tip} — Blok {blok_no} ({eklenen} stok) — {fatura_no or "belgesiz"}',
                    borc=0, alacak=toplam_fatura,
                    doviz=doviz, kur_uygulanan=1.0,
                    evrak_no=fatura_no, baglanti_tip='maliyet',
                    kaynak='maliyet', kullanici=kullanici)
                db.session.add(ch)

        sk_guncel = 0
        satilmis_idler = [s.id for s in stoklar if s.durum in SATILMIS_DURUMLAR]
        if satilmis_idler:
            db.session.flush()
            sk_guncel = _satis_kaydi_maliyet_guncelle(satilmis_idler)

        db.session.commit()
        kdv_msg = f' + KDV {toplam_kdv:,.2f} {doviz}' if toplam_kdv > 0 else ''
        cari_msg = ' — cari alacak kaydı oluşturuldu' if cari_id else ''
        msg = f'{eklenen} stoka maliyet dağıtıldı (Net: {toplam_net:,.2f} {doviz}{kdv_msg}){cari_msg}'
        if sk_guncel > 0: msg += f', {sk_guncel} satış kaydının kârlılığı güncellendi'
        return jsonify({'ok': True, 'adet': eklenen,
                        'satis_kaydi_guncel': sk_guncel, 'mesaj': msg})

    @app.route('/api/sevkiyat', methods=['GET'])
    def api_sevkiyat_liste():
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        q = Sevkiyat.query
        durum = request.args.get('durum')
        if durum:
            q = q.filter_by(durum=durum)
        sevkiyatlar = q.order_by(Sevkiyat.sevk_tarihi.desc(), Sevkiyat.id.desc()).all()
        return jsonify([{
            'id': s.id,
            'sevk_tarihi': s.sevk_tarihi.isoformat() if s.sevk_tarihi else None,
            'sevk_tip': s.sevk_tip,
            'siparis_id': s.siparis_id,
            'musteri': s.musteri,
            'cikis_noktasi': s.cikis_noktasi,
            'varis_noktasi': s.varis_noktasi,
            'tah_yukleme': s.tah_yukleme.isoformat() if s.tah_yukleme else None,
            'tah_teslim': s.tah_teslim.isoformat() if s.tah_teslim else None,
            'gercek_teslim': s.gercek_teslim.isoformat() if s.gercek_teslim else None,
            'durum': s.durum,
            'nakliye_firma': s.nakliye_firma,
            'arac_plaka': s.arac_plaka,
            'sofor_adi': s.sofor_adi,
            'sofor_tc': s.sofor_tc,
            'konteyner_no': s.konteyner_no,
            'belge_no': s.belge_no,
            'belge_tip': s.belge_tip,
            'aciklama': s.aciklama,
            'hazirlama_tarihi': s.hazirlama_tarihi.isoformat() if s.hazirlama_tarihi else None,
            'cikis_tarihi': s.cikis_tarihi.isoformat() if s.cikis_tarihi else None,
            'gumruk_tarihi': s.gumruk_tarihi.isoformat() if s.gumruk_tarihi else None,
            'teslim_tarihi': s.teslim_tarihi.isoformat() if s.teslim_tarihi else None,
            'iptal_tarihi': s.iptal_tarihi.isoformat() if s.iptal_tarihi else None,
        } for s in sevkiyatlar])

    @app.route('/api/sevkiyat', methods=['POST'])
    def api_sevkiyat_ekle():
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        data = request.json
        # Siparise bagli sevkiyat olusturuluyorsa siparis Hazir/Uretimde olmali
        sip_id = data.get('siparis_id')
        if sip_id:
            sip = Siparis.query.get(sip_id)
            if sip and sip.durum in ('Teklif Asam.', 'Iptal Edildi'):
                return jsonify({
                    'ok': False,
                    'mesaj': f'Siparis "{sip.durum}" durumunda. Sevkiyat icin once Onaylandi/Uretimde/Hazir olmali.'
                }), 400
        s = Sevkiyat(id=_yeni_id('SEV'), sevk_tip=data['sevk_tip'], musteri=data['musteri'], siparis_id=data.get('siparis_id'),
                     cikis_noktasi=data.get('cikis_noktasi'), varis_noktasi=data.get('varis_noktasi'),
                     tah_yukleme=_parse_date(data.get('tah_yukleme')), tah_teslim=_parse_date(data.get('tah_teslim')),
                     nakliye_firma=data.get('nakliye_firma'), arac_plaka=data.get('arac_plaka'), konteyner_no=data.get('konteyner_no'),
                     belge_tip=data.get('belge_tip'), belge_no=data.get('belge_no'), durum=data.get('durum','Hazirlaniyor'),
                     aciklama=data.get('aciklama'), sofor_adi=data.get('sofor_adi'), sofor_tc=data.get('sofor_kimlik'), kullanici=session['kullanici'])
        db.session.add(s)
        db.session.flush()

        # Sevkiyat 'Hazirlaniyor' durumunda baslar - stoklar 'Satildi' kalir.
        # Stoklar ancak sevkiyat 'Yolda' durumuna gecince 'Sevkedildi' olur.
        db.session.commit()
        return jsonify({'ok': True, 'id': s.id,
                        'mesaj': 'Sevkiyat olusturuldu (Hazirlaniyor). "Yolda" yapinca stoklar Sevkedildi olacak.'})

    @app.route('/api/sevkiyat/<sevk_id>/durum', methods=['POST'])
    def api_sevkiyat_durum(sevk_id):
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        s = Sevkiyat.query.get(sevk_id)
        if not s: return jsonify({'ok': False, 'mesaj': 'Sevkiyat bulunamadi'}), 404

        yeni_durum = (request.json or {}).get('durum', '').strip()
        gecerli = ['Hazirlaniyor', 'Sevk Edildi', 'Yolda', 'Gumrukte', 'Teslim Edildi', 'Iptal']
        if yeni_durum not in gecerli:
            return jsonify({'ok': False, 'mesaj': f'Gecersiz durum. Gecerli: {", ".join(gecerli)}'}), 400

        eski_durum = s.durum or 'Hazirlaniyor'

        # ═══ INCOTERM'E GÖRE AKIŞ ═══
        # Grup A (EXW/FOB/CIF/CFR/CPT/CIP): yüklemede teslim → tek adım "Sevk Edildi"
        # Grup B (DAP/DDP/DPU): varışta teslim → iki aşamalı Yolda → Teslim Edildi
        teslim_sekli = None
        if s.siparis_id:
            _sip = Siparis.query.get(s.siparis_id)
            teslim_sekli = _sip.teslim_sekli if _sip else None
        varis_takibi = _teslimde_varis_takibi(teslim_sekli)  # True=Grup B, False=Grup A

        if varis_takibi:
            # GRUP B — iki aşamalı varış takibi
            gecisler = {
                'Hazirlaniyor':  ['Yolda', 'Iptal'],
                'Yolda':         ['Gumrukte', 'Teslim Edildi', 'Hazirlaniyor', 'Iptal'],
                'Gumrukte':      ['Teslim Edildi', 'Yolda', 'Iptal'],
                'Teslim Edildi': ['Gumrukte'],  # geri alma
                'Iptal':         ['Hazirlaniyor']
            }
        else:
            # GRUP A — tek adım: Hazirlaniyor → Sevk Edildi (her şey biter)
            gecisler = {
                'Hazirlaniyor':  ['Sevk Edildi', 'Iptal'],
                'Sevk Edildi':   ['Hazirlaniyor'],  # geri alma (yanlış işaretleme düzeltme)
                'Iptal':         ['Hazirlaniyor']
            }

        if eski_durum != yeni_durum and yeni_durum not in gecisler.get(eski_durum, []):
            _grup = 'varış takipli (DAP/DDP)' if varis_takibi else 'yüklemede teslim (FOB/CIF vb.)'
            return jsonify({
                'ok': False,
                'mesaj': f'"{eski_durum}" -> "{yeni_durum}" gecisi yapilamaz [{_grup}]. Izinli: {", ".join(gecisler.get(eski_durum, [])) or "(yok)"}'
            }), 400

        s.durum = yeni_durum
        ekstra = ''

        # ═══ GRUP A: "Sevk Edildi" = her şey biter (tek adımda teslim + satış) ═══
        if yeni_durum == 'Sevk Edildi' and eski_durum == 'Hazirlaniyor':
            if s.siparis_id:
                sip = Siparis.query.get(s.siparis_id)
                if sip and sip.durum in ('Hazir', 'Uretimde', 'Onaylandi'):
                    sip.durum = 'Teslim Edildi'
                    teslim_t = _parse_date((request.json or {}).get('gercek_teslim')) or date.today()
                    sat_sayisi = _siparis_teslim_edildi(s.siparis_id, teslim_t)
                    s.gercek_teslim = teslim_t
                    ekstra += f' Mal sevk edildi (yüklemede teslim). Siparis {sip.id} Teslim Edildi, {sat_sayisi} satis kaydi olustu.'
                elif sip and sip.durum == 'Teslim Edildi':
                    ekstra += ' (Siparis zaten Teslim Edildi.)'
                elif sip:
                    ekstra += f' (Siparis {sip.id} durumu: {sip.durum} - manuel kontrol edin.)'
        # GRUP A geri alma: Sevk Edildi → Hazirlaniyor (satış kaydını ve teslimi geri al)
        elif yeni_durum == 'Hazirlaniyor' and eski_durum == 'Sevk Edildi':
            if s.siparis_id:
                sip = Siparis.query.get(s.siparis_id)
                if sip and sip.durum == 'Teslim Edildi':
                    geri = _siparis_teslimi_iptal_et(s.siparis_id)
                    sip.durum = 'Hazir'
                    s.gercek_teslim = None
                    ekstra += f' Sevk geri alindi. Siparis {sip.id} Hazir durumuna dondu, {geri} satis kaydi silindi.'

        # ═══ GRUP B: iki aşamalı akış (Yolda → Sevkedildi, Teslim → satış) ═══
        # Hazirlaniyor -> Yolda: stoklar Satildi -> Sevkedildi (urun fiziksel yola cikti)
        elif yeni_durum == 'Yolda' and eski_durum == 'Hazirlaniyor':
            guncel = _sevkiyat_stoklarini_guncelle(sevk_id, 'Sevkedildi')
            if guncel > 0:
                ekstra = f' {guncel} stok Sevkedildi durumuna gecti (yolda).'
        elif yeni_durum == 'Hazirlaniyor' and eski_durum == 'Yolda':
            guncel = _sevkiyat_stoklarini_guncelle(sevk_id, 'Satildi')
            if guncel > 0:
                ekstra = f' {guncel} stok Satildi durumuna geri dondu.'
        # Gumrukte: urun zaten Sevkedildi, durum degismez (sadece operasyonel takip)

        # Grup B — Teslim Edildi'ye gecince: bagli siparisi de Teslim Edildi yap (SatisKaydi olusur)
        if yeni_durum == 'Teslim Edildi' and eski_durum != 'Teslim Edildi':
            if s.siparis_id:
                sip = Siparis.query.get(s.siparis_id)
                if sip and sip.durum not in ('Teslim Edildi', 'Iptal Edildi'):
                    if sip.durum in ('Hazir', 'Uretimde', 'Onaylandi'):
                        sip.durum = 'Teslim Edildi'
                        teslim_t = _parse_date((request.json or {}).get('gercek_teslim')) or date.today()
                        sat_sayisi = _siparis_teslim_edildi(s.siparis_id, teslim_t)
                        s.gercek_teslim = teslim_t
                        ekstra += f' Siparis {sip.id} Teslim Edildi, {sat_sayisi} satis kaydi olustu.'
                    else:
                        ekstra += f' (Siparis {sip.id} durumu: {sip.durum} - manuel kontrol edin.)'

        # Iptal'e gecince: stoklar Satildi'ya geri (siparis hala gecerli)
        if yeni_durum == 'Iptal' and eski_durum != 'Iptal':
            guncel = _sevkiyat_stoklarini_guncelle(sevk_id, 'Satildi')
            if guncel > 0:
                ekstra = f' {guncel} stok Satildi durumuna geri dondu.'

        _log_audit('DURUM', 'sevkiyat', sevk_id, eski={'durum': eski_durum}, yeni={'durum': yeni_durum})
        db.session.commit()
        logging.info(f"Sevkiyat durum: {sevk_id} {eski_durum} -> {yeni_durum}")
        return jsonify({'ok': True, 'mesaj': f'Sevkiyat durumu: {eski_durum} -> {yeni_durum}.{ekstra}'})

    @app.route('/api/sevkiyat/<sevk_id>', methods=['GET'])
    def api_sevkiyat_detay(sevk_id):
        """Sevkiyat + bagli siparis + kalemler detayi. # FAZ 16.5.3: SEVKIYAT DETAY KALEM-AWARE"""
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        s = Sevkiyat.query.get(sevk_id)
        if not s: return jsonify({'ok': False, 'mesaj': 'Sevkiyat bulunamadi'}), 404

        # FAZ 16: Kalem-aware siparis bilgisi
        siparis_info = None
        kalemler_data = []
        if s.siparis_id:
            sip = Siparis.query.get(s.siparis_id)
            if sip:
                siparis_info = {
                    'id': sip.id,
                    'musteri': sip.musteri,
                    'durum': sip.durum,
                    'doviz': sip.doviz,
                    'toplam_tutar': sip.toplam_tutar or 0,
                    'siparis_tarihi': sip.siparis_tarihi.isoformat() if sip.siparis_tarihi else None,
                    'termin': sip.termin.isoformat() if sip.termin else None,
                    'odeme_sekli': sip.odeme_sekli,
                    'teslim_sekli': sip.teslim_sekli
                }
                for k in sip.kalemler:
                    kalemler_data.append({
                        'sira': k.sira,
                        'urun_tip': k.urun_tip,
                        'cins': k.cins,
                        'ozellik': k.ozellik,
                        'olcu': k.olcu,
                        'adet': k.adet,
                        'kasa_ici_adet': k.kasa_ici_adet,
                        'miktar': k.miktar,
                        'birim': k.birim,
                        'm2_toplam': k.m2_toplam,
                        'm3_toplam': k.m3_toplam,
                        'sqft_toplam': k.sqft_toplam,
                        'birim_fiyat': k.birim_fiyat,
                        'toplam_fiyat': k.toplam_fiyat,
                        'doviz': k.doviz,
                        'stoktan_geldi': k.stoktan_geldi
                    })

        return jsonify({
            'ok': True,
            'sevkiyat': {
                'id': s.id,
                'sevk_tarihi': s.sevk_tarihi.isoformat() if s.sevk_tarihi else None,
                'sevk_tip': s.sevk_tip, 'siparis_id': s.siparis_id,
                'musteri': s.musteri,
                'cikis_noktasi': s.cikis_noktasi, 'varis_noktasi': s.varis_noktasi,
                'tah_yukleme': s.tah_yukleme.isoformat() if s.tah_yukleme else None,
                'tah_teslim': s.tah_teslim.isoformat() if s.tah_teslim else None,
                'gercek_teslim': s.gercek_teslim.isoformat() if s.gercek_teslim else None,
                'durum': s.durum,
                'nakliye_firma': s.nakliye_firma, 'arac_plaka': s.arac_plaka,
                'sofor_adi': s.sofor_adi, 'sofor_tc': s.sofor_tc,
                'konteyner_no': s.konteyner_no,
                'belge_no': s.belge_no, 'belge_tip': s.belge_tip,
                'aciklama': s.aciklama,
                'hazirlama_tarihi': s.hazirlama_tarihi.isoformat() if s.hazirlama_tarihi else None,
                'cikis_tarihi': s.cikis_tarihi.isoformat() if s.cikis_tarihi else None,
                'gumruk_tarihi': s.gumruk_tarihi.isoformat() if s.gumruk_tarihi else None,
                'teslim_tarihi': s.teslim_tarihi.isoformat() if s.teslim_tarihi else None,
                'iptal_tarihi': s.iptal_tarihi.isoformat() if s.iptal_tarihi else None,
                'kullanici': s.kullanici
            },
            'siparis': siparis_info,
            'kalemler': kalemler_data,
            # Incoterm grubu: True = varış takipli (DAP/DDP, iki aşamalı),
            # False = yüklemede teslim (FOB/CIF vb., tek adım "Sevk Edildi")
            'varis_takibi': _teslimde_varis_takibi(siparis_info['teslim_sekli'] if siparis_info else None),
            'teslim_sekli': siparis_info['teslim_sekli'] if siparis_info else None
        })

    @app.route('/api/sevkiyat/<sevk_id>', methods=['DELETE'])
    def api_sevkiyat_sil(sevk_id):
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        s = Sevkiyat.query.get(sevk_id)
        if not s:
            return jsonify({'ok': False, 'mesaj': 'Sevkiyat bulunamadi'}), 404

        # Teslim Edildi sevkiyat silinemez (satis kaydi olusmus olabilir)
        if s.durum == 'Teslim Edildi':
            return jsonify({
                'ok': False,
                'mesaj': 'Teslim edilmis sevkiyat silinemez. Bagli satis kayitlari korunmali.'
            }), 400

        # Silmeden once stoklari Satildi'ya geri dondur (siparis hala gecerli)
        geri = 0
        if s.durum != 'Iptal':
            geri = _sevkiyat_stoklarini_guncelle(sevk_id, 'Satildi')

        _log_audit('SIL', 'sevkiyat', sevk_id, eski={'musteri': s.musteri, 'durum': s.durum, 'siparis_id': s.siparis_id})
        db.session.delete(s)
        db.session.commit()
        logging.info(f"Sevkiyat silindi: {sevk_id} ({geri} stok Satildi'ya dondu) by {session.get('kullanici')}")
        msg = 'Sevkiyat silindi'
        if geri > 0:
            msg += f', {geri} stok Satildi durumuna geri dondu'
        return jsonify({'ok': True, 'mesaj': msg})

    # ---------- API: KARLILIK ----------
    @app.route('/api/karlilik', methods=['GET'])
    def api_karlilik():
        """
        Karlilik raporu - iki mod:
        mod='gerceklesen' (varsayilan): SatisKaydi'ndan teslim edilmis gercek satislar
        mod='acik': henuz teslim edilmemis aktif siparisler (tahmini)
        """
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        mod = request.args.get('mod', 'gerceklesen')
        musteri = request.args.get('musteri')
        cins = request.args.get('cins')
        baslangic = _parse_date(request.args.get('baslangic'))
        bitis = _parse_date(request.args.get('bitis'))

        result = []

        if mod == 'gerceklesen':
            # SatisKaydi - gercek teslim edilmis satislar
            q = SatisKaydi.query
            if musteri: q = q.filter(SatisKaydi.musteri == musteri)
            if cins: q = q.filter(SatisKaydi.cins == cins)
            if baslangic: q = q.filter(SatisKaydi.satis_tarihi >= baslangic)
            if bitis: q = q.filter(SatisKaydi.satis_tarihi <= bitis)

            for s in q.order_by(SatisKaydi.satis_tarihi.desc()).all():
                # Orijinal doviz ve tutarlar
                orj_doviz = s.doviz or 'USD'
                orj_satis = s.tutar or 0
                # Maliyet ve kar TRY karsiligi (orj_doviz'e cevirmek icin)
                if orj_doviz == 'TRY':
                    satis_orj = s.tutar_try or s.tutar or 0
                    maliyet_orj = s.maliyet_try or 0
                    if not maliyet_orj and s.maliyet_usd:
                        kur = s.kur_usd or 1
                        maliyet_orj = (s.maliyet_usd or 0) * kur
                    kar_orj = satis_orj - maliyet_orj
                elif orj_doviz == 'USD':
                    satis_orj = s.tutar_usd or s.tutar or 0
                    maliyet_orj = s.maliyet_usd or 0
                    kar_orj = s.kar_usd or 0
                else:
                    satis_orj = s.tutar or 0
                    maliyet_orj = 0
                    kar_orj = 0

                result.append({
                    'id': s.id,
                    'kaynak': 'satis',
                    'stok_id': s.stok_id,
                    'musteri': s.musteri,
                    'urun_tip': s.stok_tip,
                    'cins': s.cins,
                    'miktar': s.miktar,
                    'birim': s.birim,
                    'tarih': s.satis_tarihi.isoformat() if s.satis_tarihi else None,
                    'doviz': orj_doviz,
                    'satis': q3(satis_orj),
                    'maliyet': q3(maliyet_orj),
                    'kar': q3(kar_orj),
                    'satis_usd': q3(s.tutar_usd or 0),
                    'maliyet_usd': q3(s.maliyet_usd or 0),
                    'kar_usd': q3(s.kar_usd or 0),
                    'marj': q_oran(s.marj_yuzde or 0),
                    'siparis_id': s.siparis_id,
                    'proforma_id': s.proforma_id,
                    'fatura_no': s.fatura_no,
                    'durum': 'Teslim Edildi'
                })
        else:
            # Acik siparisler - henuz teslim edilmemis (tahmini karlilik)
            q = Siparis.query.filter(Siparis.durum.notin_(['Teslim Edildi', 'Iptal Edildi']))
            if musteri: q = q.filter_by(musteri=musteri)
            if cins: q = q.filter_by(cins=cins)
            if baslangic: q = q.filter(Siparis.siparis_tarihi >= baslangic)
            if bitis: q = q.filter(Siparis.siparis_tarihi <= bitis)

            # Kur tablosu (alım fiyatı çevrimi için)
            _usd_k = DovizKur.query.filter_by(doviz='USD').order_by(DovizKur.tarih.desc()).first()
            _eur_k = DovizKur.query.filter_by(doviz='EUR').order_by(DovizKur.tarih.desc()).first()
            _ku = (_usd_k.alis if _usd_k else 0) or 0
            _ke = (_eur_k.alis if _eur_k else 0) or 0
            def _ucevir(t, d):
                if not t: return 0
                if d == 'USD' or not d: return t
                if d == 'EUR': return (t * _ke / _ku) if _ku else 0
                if d == 'TRY': return (t / _ku) if _ku else 0
                return t

            for s in q.all():
                rezler_s = Rezervasyon.query.filter_by(siparis_id=s.id, iptal_nedeni=None).all()

                # SATIŞ TUTARI (birim sistemine göre)
                # Siparişin satış birimi normalize edilir
                satis_birim_s = (s.birim or '').lower()
                satis_birim_s = {'ton': 'ton', 'm3': 'm3', 'm2': 'm2', 'sqft': 'sqft'}.get(
                    satis_birim_s, 'ton' if s.urun_tip == 'BLOK' else 'm2')
                satis_fiyat_s = s.satis_fiyati or 0

                if rezler_s:
                    # Stoğa bağlı sipariş: satış tutarı = fiyat × stokların satış birimi ölçüsü
                    satis_olcu_top = 0
                    for rz in rezler_s:
                        stk = _stok_getir(rz.stok_id, rz.stok_tip)
                        if stk:
                            satis_olcu_top += _stok_olcu(stk, satis_birim_s)
                    satis_usd = _ucevir(satis_fiyat_s * satis_olcu_top, s.doviz)
                else:
                    # Stoksuz (manuel) sipariş: s.miktar zaten satış biriminde
                    satis_usd = _ucevir(satis_fiyat_s * (s.miktar or 0), s.doviz)

                # MALİYET = stok alım maliyeti (alış birimi × ölçü) + ek maliyetler
                alim_toplam_usd = 0
                for rz in rezler_s:
                    stk = _stok_getir(rz.stok_id, rz.stok_tip)
                    if stk:
                        # Alış fiyatı kendi biriminde
                        alis_birim_fiyat = getattr(stk, 'alis_fiyati', 0) or 0
                        alis_birim_kod = getattr(stk, 'alis_fiyat_birim', None) or (
                            'ton' if rz.stok_tip == 'BLOK' else 'm2')
                        alis_olcu_s = _stok_olcu(stk, alis_birim_kod)
                        alim_toplam_usd += _ucevir(alis_birim_fiyat * alis_olcu_s,
                                                   getattr(stk, 'doviz', 'USD') or 'USD')
                # Ek maliyetler (Maliyet tablosu)
                ek_maliyet_usd = db.session.query(db.func.sum(Maliyet.usd_karsilik)).filter(
                    Maliyet.baglanti_id == s.id
                ).scalar() or 0

                maliyet_usd = alim_toplam_usd + ek_maliyet_usd
                kar_usd = satis_usd - maliyet_usd
                marj = (kar_usd / satis_usd * 100) if satis_usd else 0
                result.append({
                    'id': s.id,
                    'kaynak': 'siparis',
                    'musteri': s.musteri,
                    'urun_tip': s.urun_tip,
                    'cins': s.cins,
                    'miktar': s.miktar,
                    'birim': s.birim,
                    'tarih': s.siparis_tarihi.isoformat() if s.siparis_tarihi else None,
                    'satis_usd': q3(satis_usd),
                    'maliyet_usd': q3(maliyet_usd),
                    'kar_usd': q3(kar_usd),
                    'marj': q_oran(marj),
                    'durum': s.durum
                })

        # Ozet
        toplam_satis = sum(r['satis_usd'] for r in result)
        toplam_maliyet = sum(r['maliyet_usd'] for r in result)
        toplam_kar = sum(r['kar_usd'] for r in result)
        ort_marj = (toplam_kar / toplam_satis * 100) if toplam_satis else 0

        return jsonify({
            'data': result,
            'ozet': {
                'adet': len(result),
                'toplam_satis_usd': q3(toplam_satis),
                'toplam_maliyet_usd': q3(toplam_maliyet),
                'toplam_kar_usd': q3(toplam_kar),
                'ort_marj': q_oran(ort_marj)
            },
            'mod': mod
        })

    @app.route('/api/karlilik/musteri_ozet', methods=['GET'])
    def api_karlilik_musteri_ozet():
        """Musteri bazli gerceklesnen karlilik ozeti (SatisKaydi'ndan)."""
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        sonuc = db.session.query(
            SatisKaydi.musteri,
            db.func.count(SatisKaydi.id).label('adet'),
            db.func.sum(SatisKaydi.tutar_usd).label('satis'),
            db.func.sum(SatisKaydi.maliyet_usd).label('maliyet'),
            db.func.sum(SatisKaydi.kar_usd).label('kar')
        ).group_by(SatisKaydi.musteri).all()

        liste = []
        for r in sonuc:
            satis = r.satis or 0
            kar = r.kar or 0
            liste.append({
                'musteri': r.musteri,
                'adet': r.adet,
                'satis_usd': q3(satis),
                'maliyet_usd': q3(r.maliyet or 0),
                'kar_usd': q3(kar),
                'marj': q_oran((kar / satis * 100) if satis else 0)
            })
        liste.sort(key=lambda x: x['kar_usd'], reverse=True)
        return jsonify(liste)

    @app.route('/api/karlilik/cins_ozet', methods=['GET'])
    def api_karlilik_cins_ozet():
        """Cins (urun) bazli gerceklesnen karlilik ozeti."""
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        sonuc = db.session.query(
            SatisKaydi.cins,
            db.func.count(SatisKaydi.id).label('adet'),
            db.func.sum(SatisKaydi.tutar_usd).label('satis'),
            db.func.sum(SatisKaydi.maliyet_usd).label('maliyet'),
            db.func.sum(SatisKaydi.kar_usd).label('kar')
        ).group_by(SatisKaydi.cins).all()

        liste = []
        for r in sonuc:
            satis = r.satis or 0
            kar = r.kar or 0
            liste.append({
                'cins': r.cins or '(belirsiz)',
                'adet': r.adet,
                'satis_usd': q3(satis),
                'maliyet_usd': q3(r.maliyet or 0),
                'kar_usd': q3(kar),
                'marj': q_oran((kar / satis * 100) if satis else 0)
            })
        liste.sort(key=lambda x: x['kar_usd'], reverse=True)
        return jsonify(liste)

    # ---------- API: BANKA HESAPLARI ----------
    @app.route('/api/banka', methods=['GET'])
    def api_banka_liste():
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        bankalar = Banka.query.filter_by(aktif=True).order_by(
            Banka.varsayilan.desc(), Banka.banka_adi).all()
        return jsonify([{
            'id': b.id, 'banka_adi': b.banka_adi, 'sube': b.sube,
            'hesap_no': b.hesap_no, 'iban': b.iban, 'swift': b.swift,
            'doviz': b.doviz, 'aciklama': b.aciklama,
            'varsayilan': b.varsayilan, 'aktif': b.aktif
        } for b in bankalar])

    @app.route('/api/banka', methods=['POST'])
    def api_banka_ekle():
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        d = request.json or {}
        if not d.get('banka_adi'):
            return jsonify({'ok': False, 'mesaj': 'Banka adi zorunlu'}), 400
        # Varsayilan secilmisse digerlerini sifirla
        if d.get('varsayilan'):
            Banka.query.update({'varsayilan': False})
        b = Banka(
            banka_adi=d['banka_adi'], sube=d.get('sube', ''),
            hesap_no=d.get('hesap_no', ''), iban=d.get('iban', ''),
            swift=d.get('swift', ''), doviz=d.get('doviz', 'USD'),
            aciklama=d.get('aciklama', ''),
            varsayilan=bool(d.get('varsayilan', False)),
            aktif=True
        )
        db.session.add(b)
        db.session.commit()
        return jsonify({'ok': True, 'id': b.id, 'mesaj': 'Banka eklendi'})

    @app.route('/api/banka/<int:banka_id>', methods=['PUT'])
    def api_banka_guncelle(banka_id):
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        b = Banka.query.get(banka_id)
        if not b: return jsonify({'ok': False, 'mesaj': 'Banka bulunamadi'}), 404
        d = request.json or {}
        if d.get('varsayilan') and not b.varsayilan:
            Banka.query.update({'varsayilan': False})
        for alan in ['banka_adi','sube','hesap_no','iban','swift','doviz','aciklama']:
            if alan in d:
                setattr(b, alan, d[alan])
        if 'varsayilan' in d:
            b.varsayilan = bool(d['varsayilan'])
        if 'aktif' in d:
            b.aktif = bool(d['aktif'])
        db.session.commit()
        return jsonify({'ok': True, 'mesaj': 'Banka guncellendi'})

    @app.route('/api/banka/<int:banka_id>', methods=['DELETE'])
    def api_banka_sil(banka_id):
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        b = Banka.query.get(banka_id)
        if not b: return jsonify({'ok': False, 'mesaj': 'Banka bulunamadi'}), 404
        # Soft delete - aktif=False
        b.aktif = False
        b.varsayilan = False
        db.session.commit()
        return jsonify({'ok': True, 'mesaj': 'Banka pasif yapildi'})

    @app.route('/api/cari/<cari_id>/acik_faturalar', methods=['GET'])
    def api_cari_acik_faturalar(cari_id):
        """Bir cariye ait açık (tahsil/ödeme bekleyen) faturaları kalan tutarıyla döner.
        Çek ve tahsilat eşleştirmesi için kullanılır."""
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        cari = Cari.query.get(cari_id)
        if not cari:
            return jsonify({'ok': False, 'mesaj': 'Cari bulunamadı', 'faturalar': []}), 404
        yon = request.args.get('yon', 'satis')  # satis | alis
        # Cari unvanına göre açık faturalar (Kesildi veya Kısmi Tahsil)
        faturalar = Fatura.query.filter(
            Fatura.musteri == cari.unvan,
            Fatura.yon == yon,
            Fatura.durum.in_(['Kesildi', 'Kismi Tahsil'])
        ).order_by(Fatura.fatura_tarihi).all()
        sonuc = []
        for f in faturalar:
            # Bu faturaya yapılmış tahsilat/ödeme toplamı
            if yon == 'satis':
                odenen = db.session.query(db.func.sum(CariHareket.alacak)).filter_by(
                    baglanti_tip='fatura', baglanti_id=f.id).scalar() or 0
            else:
                odenen = db.session.query(db.func.sum(CariHareket.borc)).filter_by(
                    baglanti_tip='fatura', baglanti_id=f.id).scalar() or 0
            kalan = q3((f.toplam or 0) - odenen)
            if kalan <= 0.01:
                continue  # tam kapanmış, atla
            sonuc.append({
                'id': f.id, 'fatura_no': f.fatura_no or f.id,
                'toplam': q3(f.toplam or 0), 'odenen': q3(odenen), 'kalan': kalan,
                'doviz': f.doviz or 'USD',
                'tarih': f.fatura_tarihi.strftime('%d.%m.%Y') if f.fatura_tarihi else '',
                'durum': f.durum,
            })
        return jsonify({'ok': True, 'faturalar': sonuc})

    @app.route('/api/cari/finansal_ozet', methods=['GET'])
    def api_cari_finansal_ozet():
        """Cari sayfası için finansal özet: toplam alacak/borç/net + yaklaşan ödeme/tahsilatlar."""
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        bugun = date.today()
        toplam_alacak_try = db.session.query(func.sum(CariHareket.borc_try)).scalar() or 0
        toplam_borc_try = db.session.query(func.sum(CariHareket.alacak_try)).scalar() or 0
        eski_borc = db.session.query(func.sum(CariHareket.borc)).filter(
            CariHareket.doviz == 'TRY',
            (CariHareket.borc_try.is_(None)) | (CariHareket.borc_try == 0),
            CariHareket.borc > 0
        ).scalar() or 0
        eski_alacak = db.session.query(func.sum(CariHareket.alacak)).filter(
            CariHareket.doviz == 'TRY',
            (CariHareket.alacak_try.is_(None)) | (CariHareket.alacak_try == 0),
            CariHareket.alacak > 0
        ).scalar() or 0
        toplam_alacak_try += eski_borc
        toplam_borc_try += eski_alacak
        net_bakiye = toplam_alacak_try - toplam_borc_try

        otuz = bugun + timedelta(days=30)
        odemeler = CariHareket.query.filter(
            CariHareket.vade_tarihi >= bugun, CariHareket.vade_tarihi <= otuz,
            CariHareket.alacak > 0
        ).order_by(CariHareket.vade_tarihi).all()
        yaklasan_odemeler = [{
            'tarih': h.vade_tarihi.strftime('%d.%m.%Y') if h.vade_tarihi else '',
            'cari': (Cari.query.get(h.cari_id).unvan if h.cari_id and Cari.query.get(h.cari_id) else 'Bilinmeyen'),
            'tutar': q3(h.alacak or 0), 'doviz': h.doviz or 'USD',
            'aciklama': (h.aciklama or '')[:30]
        } for h in odemeler]

        tahsilatlar = CariHareket.query.filter(
            CariHareket.vade_tarihi >= bugun, CariHareket.vade_tarihi <= otuz,
            CariHareket.borc > 0
        ).order_by(CariHareket.vade_tarihi).all()
        yaklasan_tahsilatlar = [{
            'tarih': h.vade_tarihi.strftime('%d.%m.%Y') if h.vade_tarihi else '',
            'cari': (Cari.query.get(h.cari_id).unvan if h.cari_id and Cari.query.get(h.cari_id) else 'Bilinmeyen'),
            'tutar': q3(h.borc or 0), 'doviz': h.doviz or 'USD',
            'aciklama': (h.aciklama or '')[:30]
        } for h in tahsilatlar]

        return jsonify({
            'ok': True,
            'toplam_alacak': q3(toplam_alacak_try), 'toplam_borc': q3(toplam_borc_try),
            'net_bakiye': q3(net_bakiye),
            'yaklasan_odemeler': yaklasan_odemeler, 'yaklasan_tahsilatlar': yaklasan_tahsilatlar,
        })

    # ═══════════════════════════════════════════════════════════════
    # ÇEK / SENET TAKİP
    # ═══════════════════════════════════════════════════════════════
    def _cek_kalan_gun(vade):
        if not vade:
            return None
        return (vade - date.today()).days

    def _cek_to_dict(c):
        kalan = _cek_kalan_gun(c.vade_tarihi)
        return {
            'id': c.id, 'yon': c.yon, 'tip': c.tip,
            'cek_no': c.cek_no, 'banka_adi': c.banka_adi, 'sube': c.sube,
            'hesap_sahibi': c.hesap_sahibi,
            'tutar': q3(c.tutar or 0), 'doviz': c.doviz or 'TRY',
            'keside_tarihi': c.keside_tarihi.strftime('%Y-%m-%d') if c.keside_tarihi else None,
            'vade_tarihi': c.vade_tarihi.strftime('%Y-%m-%d') if c.vade_tarihi else None,
            'vade_goster': c.vade_tarihi.strftime('%d.%m.%Y') if c.vade_tarihi else '',
            'kalan_gun': kalan,
            'cari_id': c.cari_id, 'cari_unvan': c.cari_unvan,
            'durum': c.durum,
            'tahsil_banka_id': c.tahsil_banka_id,
            'ciro_cari_id': c.ciro_cari_id, 'ciro_cari_unvan': c.ciro_cari_unvan,
            'fatura_id': c.fatura_id,
            'aciklama': c.aciklama,
        }

    @app.route('/api/cek', methods=['GET'])
    def api_cek_liste():
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        yon = request.args.get('yon', '')          # alinan | verilen
        durum = request.args.get('durum', '')
        cari_id = request.args.get('cari_id', '')
        q = Cek.query.filter_by(aktif=True)
        if yon: q = q.filter_by(yon=yon)
        if durum: q = q.filter_by(durum=durum)
        if cari_id: q = q.filter_by(cari_id=cari_id)
        cekler = q.order_by(Cek.vade_tarihi.asc()).all()
        return jsonify({'ok': True, 'data': [_cek_to_dict(c) for c in cekler]})

    @app.route('/api/cek/ozet', methods=['GET'])
    def api_cek_ozet():
        """Çek özeti + vade hatırlatmaları (dashboard/çek sayfası için)."""
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        bugun = date.today()
        otuz = bugun + timedelta(days=30)
        aktif_durumlar_alinan = ['Portfoyde', 'TahsildeBanka', 'Teminatta']
        aktif_durumlar_verilen = ['Verildi']

        # Portföydeki alınan çekler (bizim alacağımız, henüz tahsil edilmemiş)
        alinan = Cek.query.filter(Cek.aktif == True, Cek.yon == 'alinan',
                                  Cek.durum.in_(aktif_durumlar_alinan)).all()
        verilen = Cek.query.filter(Cek.aktif == True, Cek.yon == 'verilen',
                                   Cek.durum.in_(aktif_durumlar_verilen)).all()

        def _topla(cekler):
            t = {}
            for c in cekler:
                d = c.doviz or 'TRY'
                t[d] = q3(t.get(d, 0) + (c.tutar or 0))
            return t

        # Vade yaklaşan / geçmiş (hatırlatma)
        def _yaklasan(cekler):
            out = []
            for c in cekler:
                k = _cek_kalan_gun(c.vade_tarihi)
                if k is not None and k <= 30:
                    out.append(_cek_to_dict(c))
            return sorted(out, key=lambda x: x['kalan_gun'] if x['kalan_gun'] is not None else 9999)

        return jsonify({
            'ok': True,
            'alinan_adet': len(alinan), 'alinan_toplam': _topla(alinan),
            'verilen_adet': len(verilen), 'verilen_toplam': _topla(verilen),
            'yaklasan_alinan': _yaklasan(alinan),
            'yaklasan_verilen': _yaklasan(verilen),
        })

    @app.route('/api/cek/<cek_id>', methods=['GET'])
    def api_cek_detay(cek_id):
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        c = Cek.query.get(cek_id)
        if not c: return jsonify({'ok': False, 'mesaj': 'Çek bulunamadı'}), 404
        d = _cek_to_dict(c)
        d['hareketler'] = [{
            'tarih': h.tarih.strftime('%d.%m.%Y') if h.tarih else '',
            'islem': h.islem, 'onceki_durum': h.onceki_durum,
            'yeni_durum': h.yeni_durum, 'aciklama': h.aciklama,
            'kullanici': h.kullanici,
        } for h in sorted(c.hareketler, key=lambda x: x.olusturma or datetime.min)]
        return jsonify({'ok': True, 'data': d})

    def _cek_hareket_ekle(cek, islem, onceki, yeni, aciklama=''):
        h = CekHareket(cek_id=cek.id, tarih=date.today(), islem=islem,
                       onceki_durum=onceki, yeni_durum=yeni, aciklama=aciklama,
                       kullanici=session.get('kullanici'))
        db.session.add(h)

    @app.route('/api/cek', methods=['POST'])
    def api_cek_ekle():
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        d = request.json or {}
        yon = d.get('yon')
        if yon not in ('alinan', 'verilen'):
            return jsonify({'ok': False, 'mesaj': 'Yön "alinan" veya "verilen" olmalı'}), 400
        try:
            tutar = float(d.get('tutar') or 0)
        except (ValueError, TypeError):
            return jsonify({'ok': False, 'mesaj': 'Geçersiz tutar'}), 400
        if tutar <= 0:
            return jsonify({'ok': False, 'mesaj': 'Tutar 0\'dan büyük olmalı'}), 400
        vade = _parse_date(d.get('vade_tarihi'))
        if not vade:
            return jsonify({'ok': False, 'mesaj': 'Vade tarihi zorunlu'}), 400

        cari_id = (d.get('cari_id') or '').strip() or None
        cari_unvan = None
        if cari_id:
            cr = Cari.query.get(cari_id)
            cari_unvan = cr.unvan if cr else d.get('cari_unvan')
        else:
            cari_unvan = d.get('cari_unvan')

        # Başlangıç durumu yöne göre
        ilk_durum = 'Portfoyde' if yon == 'alinan' else 'Verildi'

        cek = Cek(
            id=_yeni_id('CEK'), yon=yon, tip=d.get('tip', 'cek'),
            cek_no=(d.get('cek_no') or '').strip() or None,
            banka_adi=(d.get('banka_adi') or '').strip() or None,
            sube=(d.get('sube') or '').strip() or None,
            hesap_sahibi=(d.get('hesap_sahibi') or '').strip() or None,
            tutar=q3(tutar), doviz=d.get('doviz', 'TRY'),
            keside_tarihi=_parse_date(d.get('keside_tarihi')),
            vade_tarihi=vade,
            cari_id=cari_id, cari_unvan=cari_unvan,
            durum=ilk_durum,
            fatura_id=(d.get('fatura_id') or '').strip() or None,
            aciklama=(d.get('aciklama') or '').strip() or None,
            kullanici=session.get('kullanici'))
        db.session.add(cek)
        db.session.flush()
        _cek_hareket_ekle(cek, 'Alındı' if yon == 'alinan' else 'Verildi',
                          None, ilk_durum, cek.aciklama or '')

        # ─── ÇEK = ÖDEME ARACI: alınınca cari/fatura bakiyesini düşür ───
        # (standart muhasebe: çek alındığı an müşteri borcu kapanır)
        cari_hareket_olustur = d.get('cari_hareket_olustur', True)  # varsayılan: oluştur
        if cari_id and cari_hareket_olustur:
            try:
                cdoviz = cek.doviz or 'TRY'
                # TRY karşılığı
                try:
                    kur = _kur_getir(cdoviz, date.today()) if cdoviz != 'TRY' else 1.0
                except Exception:
                    kur = 1.0
                try_karsilik = q3(tutar * (kur or 1.0))
                if yon == 'alinan':
                    # Müşteriden çek aldık → müşterinin borcu düşer → ALACAK (bizim defterimizde)
                    ch = CariHareket(
                        id=_yeni_id('HR'), hareket_tarihi=date.today(),
                        cari_id=cari_id, cari_unvan=cari_unvan,
                        islem_tip='Çek Tahsilatı',
                        evrak_no=cek.cek_no or cek.id,
                        aciklama=f'Çek ile tahsilat ({cek.cek_no or cek.id})' + (' - ' + cek.aciklama if cek.aciklama else ''),
                        borc=0, alacak=q3(tutar),
                        borc_try=0, alacak_try=try_karsilik,
                        doviz=cdoviz, vade_tarihi=vade,
                        kur_uygulanan=kur or 1.0,
                        kaynak='tahsilat',
                        baglanti_tip='fatura' if cek.fatura_id else 'cek',
                        baglanti_id=cek.fatura_id if cek.fatura_id else cek.id,
                        kullanici=session.get('kullanici'))
                else:
                    # Tedarikçiye çek verdik → borcumuz düşer → BORC (bizim defterimizde)
                    ch = CariHareket(
                        id=_yeni_id('HR'), hareket_tarihi=date.today(),
                        cari_id=cari_id, cari_unvan=cari_unvan,
                        islem_tip='Çek Ödemesi',
                        evrak_no=cek.cek_no or cek.id,
                        aciklama=f'Çek ile ödeme ({cek.cek_no or cek.id})' + (' - ' + cek.aciklama if cek.aciklama else ''),
                        borc=q3(tutar), alacak=0,
                        borc_try=try_karsilik, alacak_try=0,
                        doviz=cdoviz, vade_tarihi=vade,
                        kur_uygulanan=kur or 1.0,
                        kaynak='odeme',
                        baglanti_tip='fatura' if cek.fatura_id else 'cek',
                        baglanti_id=cek.fatura_id if cek.fatura_id else cek.id,
                        kullanici=session.get('kullanici'))
                db.session.add(ch)
                cek.cari_hareket_id = ch.id
                db.session.flush()
                # Faturaya bağlıysa fatura bakiyesini güncelle
                if cek.fatura_id:
                    try:
                        _fatura_tahsilat_durumu(cek.fatura_id)
                    except Exception:
                        pass
            except Exception as e:
                # Cari hareket oluşturulamazsa çek yine de kaydedilsin
                print(f'Çek cari hareket hatası: {e}')

        db.session.commit()
        return jsonify({'ok': True, 'id': cek.id, 'mesaj':
                        f'{"Alınan" if yon == "alinan" else "Verilen"} çek kaydedildi'})

    @app.route('/api/cek/<cek_id>', methods=['PUT'])
    def api_cek_guncelle(cek_id):
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        c = Cek.query.get(cek_id)
        if not c: return jsonify({'ok': False, 'mesaj': 'Çek bulunamadı'}), 404
        d = request.json or {}
        # Sadece temel bilgiler düzenlenebilir (durum ayrı endpoint'ten yönetilir)
        for alan in ['cek_no', 'banka_adi', 'sube', 'hesap_sahibi', 'aciklama']:
            if alan in d: setattr(c, alan, (d.get(alan) or '').strip() or None)
        if 'tutar' in d:
            try: c.tutar = q3(float(d['tutar']))
            except (ValueError, TypeError): pass
        if 'doviz' in d: c.doviz = d['doviz']
        if 'vade_tarihi' in d:
            v = _parse_date(d['vade_tarihi'])
            if v: c.vade_tarihi = v
        if 'keside_tarihi' in d: c.keside_tarihi = _parse_date(d['keside_tarihi'])
        if 'cari_id' in d:
            cid = (d.get('cari_id') or '').strip() or None
            c.cari_id = cid
            if cid:
                cr = Cari.query.get(cid)
                c.cari_unvan = cr.unvan if cr else d.get('cari_unvan')
            else:
                c.cari_unvan = d.get('cari_unvan')
        db.session.commit()
        return jsonify({'ok': True, 'mesaj': 'Çek güncellendi'})

    @app.route('/api/cek/<cek_id>/durum', methods=['POST'])
    def api_cek_durum_degistir(cek_id):
        """Çekin durumunu değiştirir (tahsile ver, tahsil et, ciro, teminat, karşılıksız, iade).
        İsteğe bağlı olarak kasa/banka kaydı ve cari hareket oluşturur."""
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        c = Cek.query.get(cek_id)
        if not c: return jsonify({'ok': False, 'mesaj': 'Çek bulunamadı'}), 404
        d = request.json or {}
        islem = d.get('islem')  # 'tahsile_ver','tahsil_et','ciro','teminat','karsiliksiz','iade','odendi'
        onceki = c.durum
        mesaj = ''

        # ═══ GEÇİŞ KONTROLÜ — çekin durumuna göre hangi işlemlere izin var ═══
        # (Nihai durumlar Tahsil Edildi/Odendi/Iade/Karsiliksiz'den sadece geri_al çıkabilir.)
        izinli_islemler = {
            # Alınan çek
            'Portfoyde':      ['tahsile_ver', 'teminat', 'ciro', 'tahsil_et', 'karsiliksiz', 'iade'],
            'TahsildeBanka':  ['tahsil_et', 'karsiliksiz', 'geri_al'],
            'Teminatta':      ['tahsil_et', 'karsiliksiz', 'geri_al', 'iade'],
            'Ciro Edildi':    ['geri_al', 'karsiliksiz'],
            'Tahsil Edildi':  ['geri_al'],
            'Karsiliksiz':    ['geri_al', 'iade'],
            'Iade Edildi':    ['geri_al'],
            # Verilen çek
            'Verildi':        ['odendi', 'karsiliksiz', 'iade'],
            'Odendi':         ['geri_al'],
            'Iade Alindi':    ['geri_al'],
        }
        gecerli = izinli_islemler.get(onceki, None)
        if gecerli is not None and islem not in gecerli and islem != 'geri_al':
            return jsonify({'ok': False, 'error': 'gecersiz_gecis',
                'mesaj': f'"{onceki}" durumundaki çeke "{islem}" işlemi yapılamaz. '
                         f'İzinli: {", ".join(gecerli) or "(yok)"}'}), 400

        if islem == 'tahsile_ver':
            # Alınan çeki bankaya tahsile ver
            banka_id = d.get('banka_id')
            c.durum = 'TahsildeBanka'
            c.tahsil_banka_id = banka_id
            _cek_hareket_ekle(c, 'Tahsile Verildi', onceki, c.durum,
                              d.get('aciklama', ''))
            mesaj = 'Çek bankaya tahsile verildi'

        elif islem == 'teminat':
            # Çeki bankaya teminat olarak ver
            banka_id = d.get('banka_id')
            c.durum = 'Teminatta'
            c.tahsil_banka_id = banka_id
            _cek_hareket_ekle(c, 'Teminata Verildi', onceki, c.durum,
                              d.get('aciklama', 'Kredi teminatı'))
            mesaj = 'Çek teminata verildi'

        elif islem == 'ciro':
            # Alınan çeki başka bir cariye ciro et
            ciro_cari_id = (d.get('ciro_cari_id') or '').strip() or None
            ciro_unvan = d.get('ciro_cari_unvan')
            if ciro_cari_id:
                cr = Cari.query.get(ciro_cari_id)
                if cr: ciro_unvan = cr.unvan
            c.durum = 'Ciro Edildi'
            c.ciro_cari_id = ciro_cari_id
            c.ciro_cari_unvan = ciro_unvan
            _cek_hareket_ekle(c, 'Ciro Edildi', onceki, c.durum,
                              f'Ciro: {ciro_unvan or ""}')
            # Ciro edilen cariye borç kapatma (opsiyonel cari hareket)
            if d.get('cari_hareket_olustur') and ciro_cari_id:
                try:
                    _ch = CariHareket(
                        id=_yeni_id('HR'), hareket_tarihi=date.today(),
                        cari_id=ciro_cari_id, cari_unvan=ciro_unvan,
                        islem_tip='Çek Cirosu',
                        aciklama=f'Çek cirosu ({c.cek_no or c.id})',
                        borc=q3(c.tutar), alacak=0, doviz=c.doviz or 'TRY',
                        kaynak='cek', baglanti_tip='cek', baglanti_id=c.id,
                        kullanici=session.get('kullanici'))
                    db.session.add(_ch)
                    c.cari_hareket_id = _ch.id
                except Exception:
                    pass
            mesaj = 'Çek ciro edildi'

        elif islem in ('tahsil_et', 'odendi'):
            # Alınan çek tahsil edildi VEYA verilen çek ödendi
            c.durum = 'Tahsil Edildi' if c.yon == 'alinan' else 'Odendi'
            _cek_hareket_ekle(c, 'Tahsil Edildi' if c.yon == 'alinan' else 'Ödendi',
                              onceki, c.durum, d.get('aciklama', ''))
            # İsteğe bağlı kasa/banka kaydı
            if d.get('kasa_id'):
                try:
                    tip = 'giris' if c.yon == 'alinan' else 'cikis'
                    kh = KasaHareket(
                        kasa_id=int(d['kasa_id']), tarih=date.today(), tip=tip,
                        tutar=q3(c.tutar),
                        aciklama=f'Çek {"tahsilatı" if c.yon=="alinan" else "ödemesi"} ({c.cek_no or c.id})',
                        baglanti_tip='cek', baglanti_id=c.id,
                        cari_id=c.cari_id, kullanici=session.get('kullanici'))
                    db.session.add(kh)
                    db.session.flush()
                    c.kasa_hareket_id = kh.id
                    # Kasa bakiyesini güncelle (alınan çek tahsili → giriş, verilen çek ödemesi → çıkış)
                    _kasa_obj = db.session.get(Kasa, int(d['kasa_id']))
                    if _kasa_obj:
                        if tip == 'giris':
                            _kasa_obj.bakiye = q3((_kasa_obj.bakiye or 0) + q3(c.tutar))
                        else:
                            _kasa_obj.bakiye = q3((_kasa_obj.bakiye or 0) - q3(c.tutar))
                except Exception:
                    pass
            mesaj = 'Çek tahsil edildi' if c.yon == 'alinan' else 'Çek ödendi'

        elif islem == 'karsiliksiz':
            c.durum = 'Karsiliksiz'
            _cek_hareket_ekle(c, 'Karşılıksız', onceki, c.durum, d.get('aciklama', ''))
            mesaj = 'Çek karşılıksız olarak işaretlendi'

        elif islem == 'iade':
            # Çeki iade et (müşteriye geri ver / tedarikçiden geri al)
            c.durum = 'Iade Edildi' if c.yon == 'alinan' else 'Iade Alindi'
            _cek_hareket_ekle(c, 'İade', onceki, c.durum, d.get('aciklama', ''))
            mesaj = 'Çek iade edildi'

        elif islem == 'geri_al':
            # Önceki duruma döndür (portföye geri al) + yan etkileri geri al
            geri_mesaj = ''
            # 1) Tahsil/ödeme sırasında oluşan KASA hareketini geri al (bakiye düzelt)
            if c.kasa_hareket_id:
                kh = db.session.get(KasaHareket, c.kasa_hareket_id)
                if kh:
                    _k = db.session.get(Kasa, kh.kasa_id)
                    if _k:
                        # Girişse bakiyeden düş, çıkışsa geri ekle (ters işlem)
                        if kh.tip == 'giris':
                            _k.bakiye = q3((_k.bakiye or 0) - (kh.tutar or 0))
                        else:
                            _k.bakiye = q3((_k.bakiye or 0) + (kh.tutar or 0))
                    db.session.delete(kh)
                    c.kasa_hareket_id = None
                    geri_mesaj += ' Kasa hareketi geri alındı.'
            # 2) Ciro sırasında oluşan CARİ hareketini geri al
            if c.cari_hareket_id:
                ch = CariHareket.query.get(c.cari_hareket_id)
                if ch and ch.baglanti_tip == 'cek':
                    db.session.delete(ch)
                    c.cari_hareket_id = None
                    geri_mesaj += ' Cari hareket geri alındı.'
            # Ciro bilgisini temizle
            if onceki == 'Ciro Edildi':
                c.ciro_cari_id = None
                c.ciro_cari_unvan = None
            c.durum = 'Portfoyde' if c.yon == 'alinan' else 'Verildi'
            c.tahsil_banka_id = None
            _cek_hareket_ekle(c, 'Geri Alındı', onceki, c.durum, 'Portföye geri alındı')
            mesaj = 'Çek portföye geri alındı.' + geri_mesaj
        else:
            return jsonify({'ok': False, 'mesaj': 'Geçersiz işlem'}), 400

        db.session.commit()
        return jsonify({'ok': True, 'durum': c.durum, 'mesaj': mesaj})

    @app.route('/api/cek/<cek_id>/faturaya_bagla', methods=['POST'])
    def api_cek_faturaya_bagla(cek_id):
        """Mevcut bir çeki bir faturaya bağlar ve fatura bakiyesini günceller.
        (Çek zaten cari hareketi oluşturmuşsa onu faturaya yeniden bağlar.)"""
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        c = Cek.query.get(cek_id)
        if not c: return jsonify({'ok': False, 'mesaj': 'Çek bulunamadı'}), 404
        d = request.json or {}
        fatura_id = (d.get('fatura_id') or '').strip() or None
        if not fatura_id:
            return jsonify({'ok': False, 'mesaj': 'Fatura seçilmedi'}), 400
        f = Fatura.query.get(fatura_id)
        if not f:
            return jsonify({'ok': False, 'mesaj': 'Fatura bulunamadı'}), 404

        eski_fatura = c.fatura_id
        c.fatura_id = fatura_id

        # Çekin cari hareketi varsa onu da faturaya bağla
        if c.cari_hareket_id:
            ch = CariHareket.query.get(c.cari_hareket_id)
            if ch:
                ch.baglanti_tip = 'fatura'
                ch.baglanti_id = fatura_id
                # tahsilat olarak işaretle ki fatura hesabına dahil olsun
                if c.yon == 'alinan':
                    ch.kaynak = 'tahsilat'
        db.session.flush()
        # Hem eski hem yeni faturanın durumunu güncelle
        try:
            if eski_fatura and eski_fatura != fatura_id:
                _fatura_tahsilat_durumu(eski_fatura)
            _fatura_tahsilat_durumu(fatura_id)
        except Exception:
            pass
        _cek_hareket_ekle(c, 'Faturaya Bağlandı', c.durum, c.durum,
                          f'Fatura: {f.fatura_no or fatura_id}')
        db.session.commit()
        return jsonify({'ok': True, 'mesaj': f'Çek {f.fatura_no or fatura_id} faturasına bağlandı'})

    @app.route('/api/cek/<cek_id>', methods=['DELETE'])
    def api_cek_sil(cek_id):
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        c = Cek.query.get(cek_id)
        if not c: return jsonify({'ok': False, 'mesaj': 'Çek bulunamadı'}), 404
        # Çekin oluşturduğu cari hareketi geri al (bakiye düzelsin)
        fatura_id_etkilenen = c.fatura_id
        if c.cari_hareket_id:
            ch = CariHareket.query.get(c.cari_hareket_id)
            if ch:
                db.session.delete(ch)
                c.cari_hareket_id = None
        # Bağlı kasa hareketi varsa geri al — BAKİYEYİ DE DÜZELT.
        # (Sadece hareketi silmek yetmez: kasa.bakiye alanı ayrıca tutuluyor,
        #  düzeltilmezse kasada karşılığı olmayan para görünür.)
        if c.kasa_hareket_id:
            kh = KasaHareket.query.get(c.kasa_hareket_id)
            if kh:
                _k = db.session.get(Kasa, kh.kasa_id)
                if _k:
                    if kh.tip == 'giris':
                        _k.bakiye = q3((_k.bakiye or 0) - (kh.tutar or 0))
                    else:
                        _k.bakiye = q3((_k.bakiye or 0) + (kh.tutar or 0))
                db.session.delete(kh)
                c.kasa_hareket_id = None
        c.aktif = False
        _cek_hareket_ekle(c, 'İptal', c.durum, 'Iptal', 'Çek kaydı iptal edildi')
        db.session.flush()
        # Faturaya bağlıysa fatura bakiyesini güncelle (tahsilat geri alındı)
        if fatura_id_etkilenen:
            try:
                _fatura_tahsilat_durumu(fatura_id_etkilenen)
            except Exception:
                pass
        db.session.commit()
        return jsonify({'ok': True, 'mesaj': 'Çek iptal edildi, bağlı kayıtlar geri alındı'})

    # ---------- API: AYARLAR ----------
    @app.route('/api/ayarlar/kdv_oran', methods=['POST'])
    def api_ayarlar_kdv_oran():
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        d = request.json or {}
        try:
            oran = float(d.get('oran', 20))
        except Exception:
            return jsonify({'ok': False, 'mesaj': 'Gecersiz oran'}), 400
        if oran < 0 or oran > 100:
            return jsonify({'ok': False, 'mesaj': 'Oran 0-100 arasinda olmali'}), 400

        # Mevcut kayit var mi
        kayit = Veriler.query.filter_by(
            kategori='kdv_ayar', deger='varsayilan_oran').first()
        if kayit:
            kayit.kisaltma = str(oran)
        else:
            kayit = Veriler(kategori='kdv_ayar', deger='varsayilan_oran',
                            kisaltma=str(oran))
            db.session.add(kayit)
        db.session.commit()
        return jsonify({'ok': True, 'oran': oran, 'mesaj': f'KDV varsayilan orani %{oran} olarak kaydedildi'})

    @app.route('/api/ayarlar/kdv_oran', methods=['GET'])
    def api_ayarlar_kdv_oran_get():
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        oran = _kdv_varsayilan_oran()
        return jsonify({'ok': True, 'oran': oran})

    @app.route('/api/ara', methods=['GET'])
    def api_global_ara():
        """
        GLOBAL ARAMA — tüm modüllerde tek sorgu.

        Bir stok numarası (blok no / kasa no) arandığında sadece stoğu değil,
        O STOĞUN TÜM YAŞAM DÖNGÜSÜNÜ döner: maliyetleri, rezervasyonu, kesimi,
        sevkiyatı, satışı, cari hareketleri. Kullanıcı "MLS-AG1301" yazınca
        o bloğa ait her işlemi tek ekranda görür.

        Yanıt: {'sorgu': ..., 'gruplar': [{'baslik','ikon','kayitlar':[...]}]}
        Her kayıt: {ad, ek, yol, ikon}
        """
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        from sqlalchemy import or_
        from urllib.parse import quote
        q = (request.args.get('q') or '').strip()
        if len(q) < 2:
            return jsonify({'sorgu': q, 'gruplar': []})
        like = f'%{q}%'
        gruplar = []

        def _gorebilir(modul):
            return session.get('rol') in ('admin', 'ADMIN') or \
                   _yetki_var_mi(modul, 'okuma')

        def _ekle(baslik, ikon, kayitlar):
            if kayitlar:
                gruplar.append({'baslik': baslik, 'ikon': ikon, 'kayitlar': kayitlar[:12]})

        # ── 1) STOK (blok no / kasa no / slab no) ──
        eslesen_stok_idler = []
        if _gorebilir('stok'):
            stok_kayitlari = []
            for s in BlokStok.query.filter(
                    or_(BlokStok.blok_no.ilike(like), BlokStok.id.ilike(like),
                        BlokStok.cins.ilike(like))).limit(12).all():
                eslesen_stok_idler.append(s.id)
                stok_kayitlari.append({
                    'ad': s.blok_no or s.id, 'ikon': '▦',
                    'ek': f"BLOK · {s.cins or ''} · {s.durum or ''}".strip(' ·'),
                    'yol': f'/stok?tip=BLOK&ara={quote(s.blok_no or s.id)}'})
            for s in PlakaStok.query.filter(
                    or_(PlakaStok.blok_no.ilike(like), PlakaStok.id.ilike(like),
                        PlakaStok.cins.ilike(like))).limit(12).all():
                eslesen_stok_idler.append(s.id)
                _no = f"{s.blok_no or ''}#{s.slab_no}".lstrip('#') if s.slab_no else (s.blok_no or s.id)
                stok_kayitlari.append({
                    'ad': _no, 'ikon': '▦',
                    'ek': f"PLAKA · {s.cins or ''} · {s.durum or ''}".strip(' ·'),
                    'yol': f'/stok?tip=PLAKA&ara={quote(s.blok_no or s.id)}'})
            for s in EbatliStok.query.filter(
                    or_(EbatliStok.kasa_no.ilike(like), EbatliStok.id.ilike(like),
                        EbatliStok.cins.ilike(like))).limit(12).all():
                eslesen_stok_idler.append(s.id)
                stok_kayitlari.append({
                    'ad': s.kasa_no or s.id, 'ikon': '▦',
                    'ek': f"EBATLI · {s.cins or ''} · {s.durum or ''}".strip(' ·'),
                    'yol': f'/stok?tip=EBATLI&ara={quote(s.kasa_no or s.id)}'})
            _ekle('Stok', '▦', stok_kayitlari)

        # ── 2) CARİ ──
        if _gorebilir('cari'):
            _ekle('Cari Hesaplar', '◉', [{
                'ad': c.unvan, 'ikon': '◉',
                'ek': f"{c.cari_tip or 'cari'}{' · ' + c.ulke if c.ulke else ''}",
                'yol': f'/cari?ara={quote(c.unvan or "")}'}
                for c in Cari.query.filter(
                    or_(Cari.unvan.ilike(like), Cari.id.ilike(like))).limit(12).all()])

        # ── 3) SİPARİŞ ──
        if _gorebilir('siparis'):
            _ekle('Siparişler', '⇗', [{
                'ad': s.id, 'ikon': '⇗',
                'ek': f"{s.musteri or ''} · {s.durum or ''}".strip(' ·'),
                'yol': f'/siparis?ara={quote(s.id)}'}
                for s in Siparis.query.filter(
                    or_(Siparis.id.ilike(like), Siparis.musteri.ilike(like))).limit(12).all()])

        # ── 4) PROFORMA ──
        if _gorebilir('proforma'):
            _ekle('Proformalar', '▤', [{
                'ad': p.id, 'ikon': '▤',
                'ek': f"{p.musteri or ''} · {p.durum or ''}".strip(' ·'),
                'yol': f'/proforma?ara={quote(p.id)}'}
                for p in Proforma.query.filter(
                    or_(Proforma.id.ilike(like), Proforma.musteri.ilike(like))).limit(12).all()])

        # ── 5) FATURA ──
        if _gorebilir('fatura'):
            _ekle('Faturalar', '▤', [{
                'ad': f.fatura_no or f.id, 'ikon': '▤',
                'ek': f"{f.musteri or ''} · {f.durum or ''}".strip(' ·'),
                'yol': f'/fatura?ara={quote(f.fatura_no or f.id)}'}
                for f in Fatura.query.filter(
                    or_(Fatura.fatura_no.ilike(like), Fatura.id.ilike(like),
                        Fatura.musteri.ilike(like))).limit(12).all()])

        # ── 6) ÇEK ──
        if _gorebilir('kasa'):
            _ekle('Çek / Senet', '▤', [{
                'ad': ck.cek_no or ck.id, 'ikon': '▤',
                'ek': f"{ck.cari_unvan or ''} · {ck.durum or ''}".strip(' ·'),
                'yol': f'/cek?ara={quote(ck.cek_no or ck.id)}'}
                for ck in Cek.query.filter(
                    or_(Cek.cek_no.ilike(like), Cek.cari_unvan.ilike(like))).limit(12).all()])

        # ── 7) BU STOĞUN YAŞAM DÖNGÜSÜ ──
        # Aranan bir stok numarasıysa: o stoğa bağlı TÜM işlemleri getir.
        if eslesen_stok_idler:
            sid_list = eslesen_stok_idler[:5]   # ilk 5 eşleşen stok için

            if _gorebilir('maliyet'):
                _ekle('Bu stoğun maliyetleri', '▤', [{
                    'ad': f"{m.maliyet_tip} · {q3(m.tutar or 0):,.0f} {m.doviz or ''}".strip(),
                    'ikon': '▤',
                    'ek': f"{_baglanti_okunabilir('stok', m.baglanti_id)}"
                          + (f" · {m.fatura_no}" if m.fatura_no else ''),
                    'yol': f'/maliyet?ara={quote(m.fatura_no or _baglanti_okunabilir("stok", m.baglanti_id) or "")}'}
                    for m in Maliyet.query.filter(
                        func.lower(Maliyet.baglanti_tip) == 'stok',
                        Maliyet.baglanti_id.in_(sid_list)).limit(12).all()])

            if _gorebilir('rezervasyon'):
                _ekle('Bu stoğun rezervasyonları', '▦', [{
                    'ad': r.musteri or '—', 'ikon': '▦',
                    'ek': (f"{r.siparis_id or 'siparişsiz'}"
                           + (' · İPTAL' if r.iptal_nedeni else ' · aktif')),
                    'yol': f'/rezervasyon?ara={quote(r.musteri or "")}'}
                    for r in Rezervasyon.query.filter(
                        Rezervasyon.stok_id.in_(sid_list)).limit(12).all()])

            if _gorebilir('kesim'):
                kesimler = []
                for k in Kesim.query.filter(Kesim.kaynak_id.in_(sid_list)).limit(6).all():
                    kesimler.append({'ad': k.uretim_blok_no or k.kaynak_no or k.id, 'ikon': '✂',
                                     'ek': f"kaynak · {k.kesim_tarihi.isoformat() if k.kesim_tarihi else ''}",
                                     'yol': f'/kesim?ara={quote(k.uretim_blok_no or k.id)}'})
                for kd in KesimDetay.query.filter(KesimDetay.hedef_stok_id.in_(sid_list)).limit(6).all():
                    _k = Kesim.query.get(kd.kesim_id)
                    kesimler.append({'ad': (_k.uretim_blok_no if _k else None) or kd.kesim_id, 'ikon': '✂',
                                     'ek': 'bu kesimden üretildi',
                                     'yol': f'/kesim?ara={quote((_k.uretim_blok_no if _k else "") or kd.kesim_id)}'})
                _ekle('Bu stoğun kesimleri', '✂', kesimler)

            if _gorebilir('satislar'):
                _ekle('Bu stoğun satışları', '▤', [{
                    'ad': sk.musteri or '—', 'ikon': '▤',
                    'ek': f"{sk.siparis_id or ''} · {q3(sk.satis_tutari or 0):,.0f} {sk.doviz or ''}".strip(' ·'),
                    'yol': f'/satislar?ara={quote(sk.blok_no or sk.stok_id or "")}'}
                    for sk in SatisKaydi.query.filter(
                        SatisKaydi.stok_id.in_(sid_list)).limit(12).all()])

            if _gorebilir('cari'):
                _ekle('Bu stoğun cari hareketleri', '◉', [{
                    'ad': h.cari_unvan or '—', 'ikon': '◉',
                    'ek': f"{h.islem_tip or ''} · {q3((h.borc or 0) or (h.alacak or 0)):,.0f} {h.doviz or ''}".strip(' ·'),
                    'yol': f'/cari?ara={quote(h.cari_unvan or "")}'}
                    for h in CariHareket.query.filter(
                        func.lower(CariHareket.baglanti_tip) == 'stok',
                        CariHareket.baglanti_id.in_(sid_list)).limit(12).all()])

        return jsonify({'sorgu': q, 'gruplar': gruplar,
                        'toplam': sum(len(g['kayitlar']) for g in gruplar)})

    @app.route('/api/ayarlar/lookup', methods=['GET'])
    def api_ayarlar_lookup_get():
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        kategori = request.args.get('kategori')
        if kategori:
            veriler = Veriler.query.filter_by(kategori=kategori).all()
        else:
            veriler = Veriler.query.all()
        return jsonify([{'id': v.id, 'kategori': v.kategori, 'deger': v.deger, 'kisaltma': v.kisaltma} for v in veriler])

    # ═══ FİRMA LOGOSU ═══
    # Logo veritabanında (Veriler tablosu, kategori='firma_logo') base64 olarak
    # saklanır. Dosya sistemi yerine veritabanı seçildi çünkü:
    #   • Yedekle birlikte taşınır (Pardus'a geçişte otomatik gelir)
    #   • Dosya izni / yol sorunu yaşanmaz
    #   • Tek yerden güncellenir, tüm belgelerde anında değişir
    def _firma_logo_al():
        """Kayıtlı logoyu döner: 'data:image/png;base64,...' veya None."""
        k = Veriler.query.filter_by(kategori='firma_logo', deger='logo').first()
        return (k.uzun_deger if k else None) or None

    @app.route('/api/ayarlar/logo', methods=['GET'])
    def api_logo_get():
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        logo = _firma_logo_al()
        k = Veriler.query.filter_by(kategori='firma_logo', deger='logo').first()
        return jsonify({
            'ok': True,
            'var_mi': bool(logo),
            'logo': logo,                      # data URI (önizleme için)
            'dosya_adi': (k.ek_bilgi if k else None) or '',
            'boyut_kb': round(len(logo) * 3 / 4 / 1024, 1) if logo else 0
        })

    @app.route('/api/ayarlar/logo', methods=['POST'])
    def api_logo_post():
        """
        Logo yükle. Body: {'logo': 'data:image/png;base64,...', 'dosya_adi': '...'}
        Sadece admin. PNG/JPG/SVG kabul edilir, en fazla 500 KB.
        """
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        if session.get('rol') not in ('admin', 'ADMIN'):
            return jsonify({'ok': False, 'mesaj': 'Bu islem icin yonetici yetkisi gerekli.'}), 403
        data = request.json or {}
        logo = (data.get('logo') or '').strip()
        if not logo:
            return jsonify({'ok': False, 'mesaj': 'Logo verisi bos.'}), 400
        # Biçim doğrulama — sadece görsel data URI'si
        izinli = ('data:image/png;base64,', 'data:image/jpeg;base64,',
                  'data:image/jpg;base64,', 'data:image/svg+xml;base64,',
                  'data:image/webp;base64,')
        if not logo.startswith(izinli):
            return jsonify({'ok': False, 'error': 'gecersiz_bicim',
                'mesaj': 'Yalnizca PNG, JPG, SVG veya WEBP gorseli yuklenebilir.'}), 400
        # Boyut sınırı (base64 ~%33 şişirir; 500 KB ham ≈ 683 KB base64)
        tahmini_kb = len(logo) * 3 / 4 / 1024
        if tahmini_kb > 500:
            return jsonify({'ok': False, 'error': 'cok_buyuk',
                'mesaj': f'Logo cok buyuk ({tahmini_kb:.0f} KB). En fazla 500 KB olmali. '
                         f'Gorseli kucultup tekrar deneyin.'}), 400

        k = Veriler.query.filter_by(kategori='firma_logo', deger='logo').first()
        if not k:
            k = Veriler(kategori='firma_logo', deger='logo')
            db.session.add(k)
        k.uzun_deger = logo
        k.ek_bilgi = (data.get('dosya_adi') or '')[:200]
        # Denetim izi commit'ten ÖNCE eklenmeli — _log_audit kendi commit'ini
        # yapmaz, çağıranın commit'ine yazılır. Sonra çağrılırsa kaydedilmez.
        _log_audit('GUNCELLE', 'ayarlar', 'firma_logo',
                   yeni={'dosya': k.ek_bilgi, 'boyut_kb': round(tahmini_kb, 1)})
        ok, hata = _safe_commit('Logo kaydet')
        if not ok:
            return jsonify({'ok': False, 'mesaj': f'Hata: {hata}'}), 500
        return jsonify({'ok': True, 'boyut_kb': round(tahmini_kb, 1),
                        'mesaj': 'Logo kaydedildi. Tum belgelerde kullanilacak.'})

    @app.route('/api/ayarlar/logo', methods=['DELETE'])
    def api_logo_sil():
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        if session.get('rol') not in ('admin', 'ADMIN'):
            return jsonify({'ok': False, 'mesaj': 'Bu islem icin yonetici yetkisi gerekli.'}), 403
        k = Veriler.query.filter_by(kategori='firma_logo', deger='logo').first()
        if k:
            _log_audit('SIL', 'ayarlar', 'firma_logo', eski={'dosya': k.ek_bilgi})
            db.session.delete(k)
            db.session.commit()
        return jsonify({'ok': True, 'mesaj': 'Logo kaldirildi. Belgelerde firma adi yazisi kullanilacak.'})

    @app.route('/api/ayarlar/firma', methods=['GET'])
    def api_ayarlar_firma_get():
        """Satıcı firma varsayılan bilgisi (proforma/fatura için). Veriler tablosunda kategori='firma'."""
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        kayit = Veriler.query.filter_by(kategori='firma').first()
        if kayit:
            return jsonify({'satici_firma': kayit.deger or '', 'kisaltma': kayit.kisaltma or ''})
        return jsonify({'satici_firma': '', 'kisaltma': ''})

    @app.route('/api/ayarlar/firma', methods=['POST'])
    def api_ayarlar_firma_post():
        """Satıcı firma varsayılan bilgisini kaydet/güncelle."""
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        data = request.json or {}
        satici = (data.get('satici_firma') or '').strip()
        kayit = Veriler.query.filter_by(kategori='firma').first()
        if not kayit:
            kayit = Veriler(kategori='firma', deger=satici, kisaltma=(data.get('kisaltma') or ''))
            db.session.add(kayit)
        else:
            kayit.deger = satici
            if data.get('kisaltma') is not None:
                kayit.kisaltma = data.get('kisaltma')
        ok, hata = _safe_commit('Firma ayarı kaydet')
        if not ok:
            return jsonify({'ok': False, 'mesaj': f'Hata: {hata}'}), 500
        return jsonify({'ok': True, 'satici_firma': satici})

    @app.route('/api/ayarlar/lookup', methods=['POST'])
    def api_ayarlar_lookup_post():
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        if session.get('rol') != 'ADMIN':
            return jsonify({'ok': False, 'mesaj': 'Yetkisiz'}), 403
        data = request.json
        v = Veriler(kategori=data['kategori'], deger=data['deger'], kisaltma=data.get('kisaltma'))
        db.session.add(v)
        db.session.commit()
        return jsonify({'ok': True})

    @app.route('/api/rezervasyon/<rez_id>', methods=['DELETE'])
    def api_rezervasyon_iptal(rez_id):
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        rez = Rezervasyon.query.get(rez_id)
        if not rez:
            return jsonify({'ok': False, 'mesaj': 'Rezervasyon bulunamadi'}), 404
        if rez.iptal_nedeni:
            return jsonify({'ok': False, 'mesaj': 'Bu rezervasyon zaten iptal edilmis'}), 400
        data = request.get_json(silent=True) or {}
        neden = (data.get('neden') or 'Manuel iptal').strip()[:200]
        eski_durum = None
        stok_id = rez.stok_id
        stok_tip = rez.stok_tip
        try:
            if stok_tip == 'BLOK':
                stok = BlokStok.query.get(stok_id)
            elif stok_tip == 'PLAKA':
                stok = PlakaStok.query.get(stok_id)
            else:
                stok = EbatliStok.query.get(stok_id)
            if stok:
                eski_durum = stok.durum
                # Rezerve VE Satildi durumundaki stoklar serbest bırakılır.
                # (Satildi: sipariş onaylanmış stok — rezervasyon iptal edilince
                #  hiçbir siparişe bağlı kalmaz, "öksüz" olmasın diye serbest döner.)
                # Sevkedildi/Teslim Edildi'ye DOKUNULMAZ — mal fiziksel olarak çıkmış.
                if stok.durum in ('Rezerve', 'Satildi'):
                    stok.durum = 'Serbest'
            rez.iptal_nedeni = neden
            if hasattr(rez, 'iptal_tarihi'):
                from datetime import datetime as _dt
                rez.iptal_tarihi = _dt.now()
            if hasattr(rez, 'iptal_eden'):
                rez.iptal_eden = session.get('kullanici', 'sistem')
            _log_audit(
                'SIL', 'rezervasyon', rez.id,
                eski={'musteri': rez.musteri, 'stok_tip': stok_tip, 'stok_id': stok_id,
                      'siparis_id': getattr(rez, 'siparis_id', None), 'stok_eski_durum': eski_durum},
                yeni={'iptal_nedeni': neden,
                      'stok_yeni_durum': 'Serbest' if eski_durum in ('Rezerve', 'Satildi') else eski_durum}
            )
            db.session.commit()
            logging.info(f"Rezervasyon iptal edildi: {rez_id} - by {session.get('kullanici')}")
            return jsonify({
                'ok': True,
                'mesaj': 'Rezervasyon iptal edildi' + (' ve stok serbest birakildi' if eski_durum in ('Rezerve', 'Satildi') else ''),
                'stok_serbest_birakildi': eski_durum in ('Rezerve', 'Satildi')
            })
        except Exception as e:
            db.session.rollback()
            logging.error(f"Rezervasyon iptal hatasi: {e}")
            return jsonify({'ok': False, 'mesaj': f'Iptal basarisiz: {str(e)}'}), 500


    @app.route('/api/ayarlar/lookup/<int:id>', methods=['DELETE'])
    def api_ayarlar_lookup_delete(id):
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        if session.get('rol') != 'ADMIN':
            return jsonify({'ok': False, 'mesaj': 'Yetkisiz'}), 403
        v = Veriler.query.get(id)
        if v:
            db.session.delete(v)
            db.session.commit()
        return jsonify({'ok': True})

    @app.route('/api/ayarlar/lookup/<int:id>', methods=['PUT'])
    def api_ayarlar_lookup_put(id):
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        if session.get('rol') != 'ADMIN':
            return jsonify({'ok': False, 'mesaj': 'Yetkisiz'}), 403
        v = Veriler.query.get(id)
        if not v:
            return jsonify({'ok': False, 'mesaj': 'Kayıt bulunamadı'}), 404
        data = request.json
        if data.get('deger'):
            v.deger = data['deger']
        if data.get('kisaltma') is not None:
            v.kisaltma = data['kisaltma']
        db.session.commit()
        return jsonify({'ok': True})

    # ---------- API: KULLANICI YÖNETİMİ ----------
    @app.route('/api/ayarlar/kullanici', methods=['GET'])
    def api_ayarlar_kullanici_liste():
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        if session.get('rol') not in ('ADMIN', 'admin'):
            return jsonify([])
        kullanicilar = Kullanici.query.all()
        result = []
        for k in kullanicilar:
            try:
                yetk = json.loads(k.yetkiler or '{}')
            except Exception:
                yetk = {}
            result.append({
                'id': k.id,
                'ad': k.ad,
                'rol': k.rol,
                'aktif': k.aktif,
                # GUVENLIK: tanimsiz modul 'gizli' sayilir — guard'larla (satir 230/301/353)
                # ayni varsayilan. Eskiden 'yazma' donuyordu; bu, kisitli bir kullaniciyi
                # duzenlemek icin acip kaydedince ona tum modullerde YAZMA yetkisi
                # veriyordu (yetki yukselmesi).
                'yetkiler': {m: yetk.get(m, 'gizli') for m in YETKI_MODULLERI},
                'proforma_onay': bool(yetk.get('proforma_onay', False)),
                'olusturma': k.olusturma.strftime('%Y-%m-%d %H:%M') if k.olusturma else None
            })
        return jsonify(result)

    @app.route('/api/ayarlar/kullanici', methods=['POST'])
    def api_ayarlar_kullanici_ekle():
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        if session.get('rol') not in ('ADMIN', 'admin'):
            return jsonify({'ok': False, 'mesaj': 'Yetkisiz'}), 403
        data = request.json
        if not data.get('ad'):
            return jsonify({'ok': False, 'mesaj': 'Kullanıcı adı zorunlu'}), 400
        if Kullanici.query.filter_by(ad=data['ad']).first():
            return jsonify({'ok': False, 'mesaj': 'Bu kullanıcı adı zaten mevcut'}), 400
        if not data.get('sifre'):
            return jsonify({'ok': False, 'mesaj': 'Şifre zorunlu'}), 400
        # Yetkiler
        gelen_yetki = data.get('yetkiler') or {}
        temiz_yetki = {m: gelen_yetki.get(m, 'gizli') for m in YETKI_MODULLERI
                       if gelen_yetki.get(m) in ('gizli', 'okuma', 'yazma')}
        # Özel yetki: proforma iç onay (çift kontrol için ayrı bayrak)
        if gelen_yetki.get('proforma_onay'):
            temiz_yetki['proforma_onay'] = True
        try:
            k = Kullanici(
                ad=data['ad'],
                sifre=generate_password_hash(data['sifre']),
                rol=data.get('rol', 'SATIS'),
                aktif=data.get('aktif', 'true') == 'true',
                yetkiler=json.dumps(temiz_yetki, ensure_ascii=False)
            )
            db.session.add(k)
            db.session.commit()
            return jsonify({'ok': True, 'id': k.id})
        except Exception as e:
            db.session.rollback()
            app.logger.exception('Kullanici ekleme hatasi')
            hata_str = str(e)
            if 'yetkiler' in hata_str or 'no column' in hata_str.lower() or 'has no column' in hata_str.lower():
                return jsonify({'ok': False, 'mesaj': 'VERİTABANI GÜNCEL DEĞİL: db_migrate_yetkiler.py calistirmaniz gerekiyor! Hata: ' + hata_str}), 500
            return jsonify({'ok': False, 'mesaj': 'Hata: ' + hata_str}), 500

    @app.route('/api/ayarlar/kullanici/<int:id>', methods=['PUT'])
    def api_ayarlar_kullanici_guncelle(id):
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        if session.get('rol') != 'ADMIN':
            return jsonify({'ok': False, 'mesaj': 'Yetkisiz'}), 403
        k = Kullanici.query.get(id)
        if not k:
            return jsonify({'ok': False, 'mesaj': 'Kullanıcı bulunamadı'}), 404
        data = request.json
        if data.get('sifre'):
            k.sifre = generate_password_hash(data['sifre'])
        if data.get('rol'):
            k.rol = data['rol']
        if data.get('aktif') is not None:
            k.aktif = data['aktif'] == 'true'
        # Yetkiler güncelle
        if 'yetkiler' in data:
            gelen = data.get('yetkiler') or {}
            temiz = {m: gelen.get(m, 'gizli') for m in YETKI_MODULLERI
                     if gelen.get(m) in ('gizli', 'okuma', 'yazma')}
            if gelen.get('proforma_onay'):
                temiz['proforma_onay'] = True
            k.yetkiler = json.dumps(temiz, ensure_ascii=False)
        db.session.commit()
        return jsonify({'ok': True})

    @app.route('/api/ayarlar/kullanici/<int:id>', methods=['DELETE'])
    def api_ayarlar_kullanici_sil(id):
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        if session.get('rol') not in ('ADMIN', 'admin'):
            return jsonify({'ok': False, 'mesaj': 'Yetkisiz'}), 403
        k = Kullanici.query.get(id)
        if not k:
            return jsonify({'ok': False, 'mesaj': 'Kullanıcı bulunamadı'}), 404
        if k.ad == 'admin':
            return jsonify({'ok': False, 'mesaj': 'Admin kullanıcısı silinemez'}), 400
        db.session.delete(k)
        db.session.commit()
        return jsonify({'ok': True})

    # ---------- API: PROFORMA ----------
    @app.route('/api/proforma', methods=['GET'])
    def api_proforma_liste():
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        # ?arsiv=1 → tüm sürümler; varsayılan → sadece aktif sürüm (arşivi gizle)
        arsiv_goster = request.args.get('arsiv') in ('1', 'true', 'evet')
        q = Proforma.query
        if not arsiv_goster:
            # aktif_surum None (eski kayıt) veya True olanları göster; sadece False'u gizle
            q = q.filter(db.or_(Proforma.aktif_surum == True, Proforma.aktif_surum.is_(None)))
        proformalar = q.order_by(Proforma.olusturma.desc()).all()
        # Zincir başına toplam sürüm sayısı (rozet için)
        from collections import Counter
        kok_sayac = Counter(p.ana_pi_id or p.id for p in Proforma.query.all())
        return jsonify([{'id': p.id, 'tarih': p.olusturma.strftime('%Y-%m-%d'), 'musteri': p.musteri,
                         'urun_tip': p.urun_tip, 'cins': p.cins, 'toplam': p.toplam, 'doviz': p.doviz,
                         'durum': p.durum, 'packing_list': p.packing_list,
                         'siparis_id': p.siparis_id,
                         'revizyon_no': p.revizyon_no or 0,
                         'aktif_surum': bool(p.aktif_surum) if p.aktif_surum is not None else True,
                         'surum_sayisi': kok_sayac.get(p.ana_pi_id or p.id, 1),
                         'onaya_gonderen': p.onaya_gonderen,
                         'onaylayan': p.onaylayan,
                         'onay_red_notu': p.onay_red_notu,
                         'proforma_tipi': getattr(p, 'proforma_tipi', None) or ('satis' if p.siparis_id else 'teklif')} for p in proformalar])

    @app.route('/api/proforma/<proforma_id>/detay_full', methods=['GET'])
    def api_proforma_detay(proforma_id):
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        p = Proforma.query.get(proforma_id)
        if not p: return jsonify({'error': 'Bulunamadı'}), 404
        kalemler = ProformaKalem.query.filter_by(proforma_id=proforma_id).order_by(ProformaKalem.sira).all()
        return jsonify({'id': p.id, 'musteri': p.musteri, 'musteri_adres': p.musteri_adres, 'musteri_ulke': p.musteri_ulke,
                        'tur': p.tur, 'proforma_tipi': getattr(p, 'proforma_tipi', None) or ('satis' if p.siparis_id else 'teklif'), 'kdv_oran': p.kdv_oran, 'packing_list': p.packing_list,
                        'genel_birim': 'm2',
                        'genel_bundle_sayisi': getattr(p, 'genel_bundle_sayisi', 10) or 10,
                        'karma_bundle': bool(getattr(p, 'karma_bundle', False)),
                        'iskonto': p.iskonto_sabit or 0,
                        'iskonto_tip': 'SAB', 'avans_deger': p.avans_tutari or 0, 'avans_tip': 'SAB',
                        'doviz': p.doviz, 'banka_adi': p.banka_adi, 'iban': p.iban, 'swift': p.swift,
                        'satici_firma': p.satici_firma, 'odeme_sekli': p.odeme_sekli, 'teslim_sekli': p.teslim_sekli,
                        'termin': p.termin.isoformat() if p.termin else '', 'ozel_sartlar': p.ozel_sartlar,
                        'notlar': p.notlar, 'fiyat_araliklari': [], 'kalemler': [{
                            'urun_tip': k.urun_tip, 'cins': k.cins, 'ozellik': k.yuzey_spec, 'blok_no': k.blok_no,
                            'boy': k.boy, 'yukseklik': k.yukseklik, 'kalinlik': k.kalinlik,
                            'adet': k.adet, 'kasa_ici_adet': getattr(k, 'kasa_ici_adet', 1) or 1,
                            'birim_fiyat': k.birim_fiyat, 'doviz': k.doviz, 'm2': k.m2_toplam, 'sqft': k.sqft_toplam,
                            'miktar': k.miktar, 'net_fiyat': k.net_fiyat, 'toplam_fiyat': k.toplam_fiyat,
                            'iskonto': k.iskonto, 'iskonto_tip': k.iskonto_tip, 'slab_no': k.slab_no,
                            'slab_baslangic': (k.slab_no or '').split('-')[0] if k.slab_no else '',
                            'bundle_no': k.bundle_no, 'agirlik_kg': k.agirlik, 'm2_kg': ''} for k in kalemler]})

    def _proforma_toplam_hesapla(p):
        """
        Proforma genel toplamını kalemlerden hesaplar (UI formülünün sunucu karşılığı):
          kalemToplam (net_fiyat) − iskonto → ara → +KDV (sadece yurt içi) = genel
        İstemci 'toplam' göndermezse sessizce 0 kalmasın diye savunma amaçlı kullanılır.
        """
        kalemler = ProformaKalem.query.filter_by(proforma_id=p.id).all()
        kalem_toplam = 0.0
        for k in kalemler:
            deger = k.net_fiyat
            if deger in (None, 0):
                deger = k.toplam_fiyat
            if deger in (None, 0):
                deger = (k.miktar or 0) * (k.birim_fiyat or 0)
            kalem_toplam += float(deger or 0)
        iskonto = float(p.iskonto_sabit or 0)
        ara = max(0.0, kalem_toplam - iskonto)
        kdv_oran = float(p.kdv_oran or 0) if (p.tur or '') == 'yurt_ici' else 0.0
        return q3(ara + (ara * kdv_oran / 100.0))

    @app.route('/api/proforma', methods=['POST'])
    def api_proforma_ekle():
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        data = request.json
        # siparis_id varsa - bu proforma mevcut bir siparise bagli
        siparis_id = data.get('siparis_id')
        p = Proforma(id=_yeni_id('PI'), musteri=data['musteri'], musteri_adres=data.get('musteri_adres'),
                     musteri_ulke=data.get('musteri_ulke'), tur=data.get('tur','ihracat'), kdv_oran=data.get('kdv_oran',0),
                     proforma_tipi=data.get('proforma_tipi', 'satis' if siparis_id else 'teklif'),
                     packing_list=data.get('packing_list',False), iskonto_sabit=data.get('iskonto',0),
                     genel_bundle_sayisi=data.get('genel_bundle_sayisi',10),
                     karma_bundle=bool(data.get('karma_bundle',False)),
                     avans_tutari=data.get('avans_deger',0), doviz=data.get('doviz','USD'), banka_adi=data.get('banka_adi'),
                     iban=data.get('iban'), swift=data.get('swift'), satici_firma=data.get('satici_firma'),
                     odeme_sekli=data.get('odeme_sekli'), teslim_sekli=data.get('teslim_sekli'),
                     termin=_parse_date(data.get('termin')), ozel_sartlar=data.get('ozel_sartlar'), notlar=data.get('notlar'),
                     toplam=data.get('toplam'), siparis_id=siparis_id,
                     kullanici=session['kullanici'], durum='Taslak',
                     revizyon_no=0, aktif_surum=True)
        db.session.add(p)
        db.session.flush()
        # Kok surum: ana_pi_id kendi id'sine esitlenir (revizyon zincirinin basi)
        p.ana_pi_id = p.id
        db.session.flush()
        for idx, k in enumerate(data.get('kalemler', [])):
            pk = ProformaKalem(proforma_id=p.id, urun_tip=k.get('urun_tip'), cins=k.get('cins'), yuzey_spec=k.get('ozellik'),
                               blok_no=k.get('blok_no'), en=k.get('en'),
                               boy=k.get('boy'), yukseklik=k.get('yukseklik'), kalinlik=k.get('kalinlik'),
                               adet=k.get('adet',1),
                               kasa_ici_adet=int(k.get('kasa_ici_adet') or 1),
                               miktar=k.get('miktar'), birim=k.get('birim','m2'), birim_fiyat=k.get('birim_fiyat'),
                               doviz=k.get('doviz','USD'), toplam_fiyat=k.get('toplam_fiyat'), net_fiyat=k.get('net_fiyat'),
                               iskonto=k.get('iskonto',0), iskonto_tip=k.get('iskonto_tip','%'), slab_no=k.get('slab_no'),
                               bundle_no=k.get('bundle_no'), m2_toplam=k.get('m2'), sqft_toplam=k.get('sqft'),
                               agirlik=k.get('agirlik_kg'), sira=idx)
            db.session.add(pk)
        db.session.flush()

        # SAVUNMA: istemci 'toplam' gondermediyse kalemlerden hesapla.
        # (Aksi halde proforma 0 tutarla kaydolur, faturaya donusunce cariye
        #  0 borc yazilir — sessiz mali hata.)
        if not p.toplam:
            p.toplam = _proforma_toplam_hesapla(p)
            app.logger.info(f'Proforma {p.id}: toplam istemciden gelmedi, '
                            f'kalemlerden hesaplandi: {p.toplam}')

        # SİPARİŞSİZ PROFORMA: kalemlerdeki Serbest stoklari Rezerve et
        rez_count = 0
        if not siparis_id:
            rez_count = _proforma_stoklarini_rezerve_et(p)

        db.session.commit()
        msg = 'Proforma kaydedildi'
        if rez_count > 0:
            msg += f' ({rez_count} stok rezerve edildi)'
        return jsonify({'ok': True, 'id': p.id, 'mesaj': msg, 'rezerve_sayisi': rez_count})

    @app.route('/api/proforma/<proforma_id>', methods=['PUT'])
    def api_proforma_guncelle(proforma_id):
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        p = Proforma.query.get(proforma_id)
        if not p:
            return jsonify({'ok': False, 'mesaj': 'Proforma bulunamadi'}), 404
        data = request.json or {}

        try:
            # Eski rezervasyonlari iptal et (siparise bagli degilse)
            if not p.siparis_id:
                eski_rezler = Rezervasyon.query.filter_by(
                    proforma_id=proforma_id, iptal_nedeni=None).all()
                for rz in eski_rezler:
                    stok = _stok_getir(rz.stok_id, rz.stok_tip)
                    if stok and stok.durum == 'Rezerve':
                        stok.durum = 'Serbest'
                    rz.iptal_nedeni = 'Proforma guncellendi'

            # Ana alanlari guncelle
            p.musteri = data.get('musteri', p.musteri)
            p.musteri_adres = data.get('musteri_adres')
            p.musteri_ulke = data.get('musteri_ulke')
            p.tur = data.get('tur', 'ihracat')
            if 'proforma_tipi' in data:
                p.proforma_tipi = data.get('proforma_tipi') or p.proforma_tipi
            p.kdv_oran = data.get('kdv_oran', 0)
            p.packing_list = data.get('packing_list', False)
            p.genel_bundle_sayisi = data.get('genel_bundle_sayisi', 10)
            p.karma_bundle = bool(data.get('karma_bundle', False))
            p.iskonto_sabit = data.get('iskonto', 0)
            p.avans_tutari = data.get('avans_deger', 0)
            p.doviz = data.get('doviz', 'USD')
            p.banka_adi = data.get('banka_adi')
            p.iban = data.get('iban')
            p.swift = data.get('swift')
            p.satici_firma = data.get('satici_firma')
            p.odeme_sekli = data.get('odeme_sekli')
            p.teslim_sekli = data.get('teslim_sekli')
            p.termin = _parse_date(data.get('termin'))
            p.ozel_sartlar = data.get('ozel_sartlar')
            p.notlar = data.get('notlar')
            p.toplam = data.get('toplam')
            # Siparis_id sadece dolu gelirse guncelle (bos string/None ile silmemek icin)
            yeni_sid = data.get('siparis_id')
            if yeni_sid:
                p.siparis_id = yeni_sid

            # Eski kalemleri sil, yenilerini ekle
            ProformaKalem.query.filter_by(proforma_id=proforma_id).delete()
            for idx, k in enumerate(data.get('kalemler', [])):
                pk = ProformaKalem(
                    proforma_id=p.id, urun_tip=k.get('urun_tip'), cins=k.get('cins'),
                    yuzey_spec=k.get('ozellik'), blok_no=k.get('blok_no'), en=k.get('en'),
                    boy=k.get('boy'), yukseklik=k.get('yukseklik'), kalinlik=k.get('kalinlik'),
                    adet=k.get('adet', 1), miktar=k.get('miktar'), birim=k.get('birim', 'm2'),
                    birim_fiyat=k.get('birim_fiyat'), doviz=k.get('doviz', 'USD'),
                    toplam_fiyat=k.get('toplam_fiyat'), net_fiyat=k.get('net_fiyat'),
                    iskonto=k.get('iskonto', 0), iskonto_tip=k.get('iskonto_tip', '%'),
                    slab_no=k.get('slab_no'), bundle_no=k.get('bundle_no'),
                    m2_toplam=k.get('m2'), sqft_toplam=k.get('sqft'),
                    agirlik=k.get('agirlik_kg'), sira=idx)
                db.session.add(pk)
            db.session.flush()

            # Siparissiz proforma: yeni kalemlerin stoklarini tekrar rezerve et
            rez_count = 0
            if not p.siparis_id:
                rez_count = _proforma_stoklarini_rezerve_et(p)

            _log_audit('GUNCELLE', 'proforma', proforma_id,
                       yeni={'musteri': p.musteri, 'toplam': p.toplam})
            db.session.commit()
            msg = 'Proforma guncellendi'
            if rez_count > 0:
                msg += f' ({rez_count} stok rezerve edildi)'
            return jsonify({'ok': True, 'id': p.id, 'mesaj': msg})
        except Exception as e:
            db.session.rollback()
            return jsonify({'ok': False, 'mesaj': str(e)}), 500

    @app.route('/api/proforma/<proforma_id>/revize', methods=['POST'])
    def api_proforma_revize(proforma_id):
        """
        Mevcut proformanın YENİ SÜRÜMÜNÜ oluşturur (müşteri değişiklik istediğinde).
        - Eski sürüm 'Revize' durumuna geçer ve arşivlenir (aktif_surum=False).
        - Yeni sürüm eski proformanın TAM KOPYASI olarak açılır (kalemler dahil),
          id = kök_id + '_Rev.N', durum='Taslak', aktif_surum=True.
        - Zincirdeki tüm sürümler ana_pi_id'yi paylaşır; numara kökten bağımsız artar.
        İstek gövdesi (opsiyonel): { "revizyon_notu": "fiyat guncellendi" }
        """
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        p = Proforma.query.get(proforma_id)
        if not p:
            return jsonify({'ok': False, 'mesaj': 'Proforma bulunamadi'}), 404

        # Faturalanmış proforma revize edilemez (fatura kesilmiş, belge dondu)
        if p.durum == 'Faturalandi':
            return jsonify({'ok': False,
                'mesaj': 'Faturalanmis proforma revize edilemez. Faturayi iptal edin ya da yeni proforma acin.'}), 400

        # Sadece AKTİF sürüm revize edilebilir (arşiv sürümü üzerinden dallanma olmaz)
        if p.aktif_surum is False:
            aktif = Proforma.query.filter_by(ana_pi_id=p.ana_pi_id or p.id, aktif_surum=True).first()
            return jsonify({'ok': False,
                'mesaj': f'Bu eski bir surum (arsiv). Guncel surum uzerinden revize edin: '
                         f'{aktif.id if aktif else "?"}'}), 400

        data = request.json or {}
        kok_id = p.ana_pi_id or p.id
        # Zincirdeki en yuksek revizyon no + 1
        mevcut_max = db.session.query(db.func.max(Proforma.revizyon_no)).filter(
            Proforma.ana_pi_id == kok_id).scalar()
        yeni_no = (mevcut_max or 0) + 1
        yeni_id = f'{kok_id}_Rev.{yeni_no}'

        # Çakışma güvenligi (teoride olmamali ama garanti)
        if Proforma.query.get(yeni_id):
            return jsonify({'ok': False, 'mesaj': f'{yeni_id} zaten var, tekrar deneyin'}), 409

        # ── Yeni sürüm: eski proformanın TÜM alanlarını kopyala ──
        yeni = Proforma(
            id=yeni_id, ana_pi_id=kok_id, revizyon_no=yeni_no, aktif_surum=True,
            durum='Taslak', proforma_tipi=p.proforma_tipi, siparis_id=p.siparis_id,
            musteri=p.musteri, musteri_adres=p.musteri_adres, musteri_ulke=p.musteri_ulke,
            tur=p.tur, kdv_oran=p.kdv_oran, packing_list=p.packing_list,
            genel_bundle_sayisi=p.genel_bundle_sayisi, karma_bundle=p.karma_bundle,
            iskonto=p.iskonto, iskonto_tip=p.iskonto_tip, iskonto_sabit=p.iskonto_sabit,
            avans_yuzdesi=p.avans_yuzdesi, avans_tutari=p.avans_tutari, avans_tip=p.avans_tip,
            avans_sabit=p.avans_sabit, doviz=p.doviz, toplam=p.toplam,
            banka_adi=p.banka_adi, iban=p.iban, swift=p.swift, ulke=p.ulke,
            satici_firma=p.satici_firma, satici_adres=p.satici_adres,
            satici_tel=p.satici_tel, satici_email=p.satici_email,
            odeme_sekli=p.odeme_sekli, teslim_sekli=p.teslim_sekli, termin=p.termin,
            yuklenme_limani=p.yuklenme_limani, varis_limani=p.varis_limani,
            hs_kodu=p.hs_kodu, konteyner_no=p.konteyner_no,
            ozel_sartlar=p.ozel_sartlar, notlar=p.notlar,
            revizyon_notu=(data.get('revizyon_notu') or '').strip() or None,
            kullanici=session.get('kullanici'))
        db.session.add(yeni)
        db.session.flush()

        # Kalemleri kopyala
        eski_kalemler = ProformaKalem.query.filter_by(proforma_id=p.id).order_by(
            ProformaKalem.sira, ProformaKalem.id).all()
        for k in eski_kalemler:
            yk = ProformaKalem(
                proforma_id=yeni.id, urun_tip=k.urun_tip, cins=k.cins, aciklama=k.aciklama,
                yuzey_spec=k.yuzey_spec, ozellik=k.ozellik, blok_no=k.blok_no,
                slab_no=k.slab_no, bundle_no=k.bundle_no, boy=k.boy, yukseklik=k.yukseklik,
                kalinlik=k.kalinlik, en=k.en, adet=k.adet, kasa_ici_adet=k.kasa_ici_adet,
                miktar=k.miktar, birim=k.birim, birim_fiyat=k.birim_fiyat, doviz=k.doviz,
                toplam_fiyat=k.toplam_fiyat, net_fiyat=k.net_fiyat, iskonto=k.iskonto,
                iskonto_tip=k.iskonto_tip, m2_toplam=k.m2_toplam, sqft_toplam=k.sqft_toplam,
                agirlik=k.agirlik, konteyner_no=k.konteyner_no, kap_no=k.kap_no,
                kap_tip=k.kap_tip, sira=k.sira)
            db.session.add(yk)

        # ── Eski sürümü arşivle ──
        # Rezervasyonları YENİ sürüme taşı (stok kilidi kopmasın, çift rezerve olmasın)
        tasinan = Rezervasyon.query.filter_by(proforma_id=p.id).update(
            {Rezervasyon.proforma_id: yeni.id}, synchronize_session=False)
        eski_durum = p.durum
        p.durum = 'Revize'
        p.aktif_surum = False

        _log_audit('REVIZE', 'proforma', p.id,
                   eski={'durum': eski_durum, 'aktif': True},
                   yeni={'yeni_surum': yeni.id, 'revizyon_no': yeni_no})
        db.session.commit()

        ek = f' {tasinan} rezervasyon yeni surume tasindi.' if tasinan else ''
        return jsonify({
            'ok': True, 'id': yeni.id, 'revizyon_no': yeni_no,
            'eski_id': p.id, 'ana_pi_id': kok_id,
            'mesaj': f'{p.id} arsivlendi (Revize). Yeni surum: {yeni.id}.{ek}'
        })

    @app.route('/api/proforma/<proforma_id>/risk')
    def api_proforma_risk(proforma_id):
        """
        Bir proformanın müşterisinin risk durumunu döner (onay öncesi uyarı için).
        Proforma tutarı müşterinin kullanılabilir risk limitini aşıyor mu?
        """
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        p = Proforma.query.get(proforma_id)
        if not p:
            return jsonify({'ok': False, 'mesaj': 'Proforma bulunamadi'}), 404
        cari = _cari_bul(p.musteri)
        if not cari:
            return jsonify({'ok': True, 'limit_var': False, 'cari_yok': True})
        # Proforma tutarını carinin risk dövizine çevirerek karşılaştır
        p_doviz = p.doviz or 'USD'
        risk = _cari_risk_durumu(cari, ek_tutar_doviz=(p.toplam or 0), ek_doviz=p_doviz)
        risk['ok'] = True
        risk['proforma_tutari'] = q3(p.toplam or 0)
        risk['proforma_doviz'] = p_doviz
        return jsonify(risk)

    @app.route('/api/proforma/<proforma_id>/surumler')
    def api_proforma_surumler(proforma_id):
        """Bir proformanın tüm revizyon zincirini (eski→yeni) döner."""
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        p = Proforma.query.get(proforma_id)
        if not p:
            return jsonify({'ok': False, 'mesaj': 'Proforma bulunamadi'}), 404
        kok = p.ana_pi_id or p.id
        zincir = Proforma.query.filter_by(ana_pi_id=kok).order_by(
            Proforma.revizyon_no.asc()).all()
        return jsonify({'ok': True, 'ana_pi_id': kok, 'surumler': [{
            'id': s.id, 'revizyon_no': s.revizyon_no, 'durum': s.durum,
            'aktif_surum': bool(s.aktif_surum), 'toplam': s.toplam, 'doviz': s.doviz,
            'olusturma': s.olusturma.isoformat() if s.olusturma else None,
            'revizyon_notu': s.revizyon_notu
        } for s in zincir]})

    @app.route('/api/proforma/<proforma_id>/durum', methods=['POST'])
    def api_proforma_durum(proforma_id):
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        p = Proforma.query.get(proforma_id)
        if not p:
            return jsonify({'ok': False, 'mesaj': 'Proforma bulunamadi'}), 404

        yeni_durum = (request.json or {}).get('durum', '').strip()
        gecerli_durumlar = ['Taslak', 'Ic Onay', 'Gonderildi', 'Onaylandi', 'Faturalandi', 'Iptal', 'Revize']
        if yeni_durum not in gecerli_durumlar:
            return jsonify({'ok': False, 'mesaj': f'Gecersiz durum. Gecerli: {", ".join(gecerli_durumlar)}'}), 400

        # Durum geçiş kontrolleri (is akisi)
        # 'Revize' bir manuel hedef DEGILDIR — yalnizca /revize endpoint'i uzerinden
        # otomatik atanir (eski surum arsivlenirken). Kullanici elle Revize yapamaz.
        #
        # İÇ ONAY AKIŞI (çift kontrol):
        #   Taslak → Ic Onay (hazırlayan onaya gönderir)
        #   Ic Onay → Onaylandi (YETKİLİ + FARKLI kişi onaylar — çift kontrol)
        #   Ic Onay → Taslak (yetkili reddeder, düzeltme için geri döner)
        #   Onaylandi → Gonderildi (HAZIRLAYAN müşteriye gönderir, "gönderdim" işareti)
        #   Gonderildi → Faturalandi (faturaya dönüşür)
        gecisler = {
            'Taslak':       ['Ic Onay', 'Iptal'],
            'Ic Onay':      ['Onaylandi', 'Taslak', 'Iptal'],
            'Onaylandi':    ['Gonderildi', 'Faturalandi', 'Ic Onay', 'Iptal'],
            'Gonderildi':   ['Faturalandi', 'Onaylandi', 'Iptal'],
            'Faturalandi':  ['Iptal'],
            'Iptal':        ['Taslak'],  # Tekrar acmak icin
            'Revize':       []           # Arsiv — durum degistirilemez (salt okunur)
        }
        mevcut = p.durum or 'Taslak'
        if yeni_durum == 'Revize':
            return jsonify({'ok': False,
                'mesaj': 'Revize durumu elle atanamaz. "Revize Et" ile yeni surum olusturun.'}), 400
        if mevcut == 'Revize':
            return jsonify({'ok': False,
                'mesaj': 'Bu bir arsiv surumu (Revize), durumu degistirilemez.'}), 400
        if mevcut == yeni_durum:
            return jsonify({'ok': False, 'mesaj': f'Proforma zaten {mevcut} durumunda'}), 400
        if yeni_durum not in gecisler.get(mevcut, []):
            return jsonify({
                'ok': False,
                'mesaj': f'"{mevcut}" durumundan "{yeni_durum}" durumuna gecis yapilamaz. Izinli gecisler: {", ".join(gecisler.get(mevcut, []))}'
            }), 400

        from datetime import datetime as _dt
        aktif_kullanici = session.get('kullanici')

        # ═══ İÇ ONAY ÇİFT KONTROL KURALLARI ═══
        # 1) Ic Onay → Onaylandi (ONAYLAMA): yetki + farklı kişi şart
        if mevcut == 'Ic Onay' and yeni_durum == 'Onaylandi':
            if not _proforma_onay_yetkisi_var_mi():
                return jsonify({'ok': False, 'error': 'onay_yetkisi_yok',
                    'mesaj': 'Proforma onaylama yetkiniz yok. Onay yetkisi olan bir kullanici onaylamali.'}), 403
            # Çift kontrol: onaya göndereni AYNI kişi onaylayamaz (admin hariç muaf değil — kural herkese)
            if p.onaya_gonderen and p.onaya_gonderen == aktif_kullanici:
                return jsonify({'ok': False, 'error': 'ayni_kisi',
                    'mesaj': f'Bu proformayi onaya siz gonderdiniz ({aktif_kullanici}). Cift kontrol geregi '
                             f'BASKA bir yetkili onaylamali.'}), 403
            p.onaylayan = aktif_kullanici
            p.onay_tarihi = _dt.now()

        # 2) Taslak → Ic Onay (ONAYA GÖNDERME): izi tut
        elif mevcut == 'Taslak' and yeni_durum == 'Ic Onay':
            p.onaya_gonderen = aktif_kullanici
            p.onaya_gonderme_tarihi = _dt.now()
            # Önceki red izini temizle (yeniden onaya gidiyor)
            p.onay_reddeden = None
            p.onay_red_notu = None

        # 3) Ic Onay → Taslak (RED): yetki + gerekçe
        elif mevcut == 'Ic Onay' and yeni_durum == 'Taslak':
            if not _proforma_onay_yetkisi_var_mi():
                return jsonify({'ok': False, 'error': 'onay_yetkisi_yok',
                    'mesaj': 'Onay reddetme yetkiniz yok.'}), 403
            red_notu = (request.json or {}).get('red_notu', '').strip()
            p.onay_reddeden = aktif_kullanici
            p.onay_red_notu = red_notu or 'Gerekce belirtilmedi'
            p.onaylayan = None  # önceki onay varsa temizle
            p.onay_tarihi = None

        eski_durum = p.durum
        p.durum = yeni_durum

        # Durum değişikliğinin stok/siparis tarafına etkisi
        ekstra_mesaj = ''

        # İç onay geçişleri için bilgilendirme mesajı
        if yeni_durum == 'Ic Onay':
            ekstra_mesaj = ' Ikinci bir yetkilinin onayi bekleniyor.'
        elif mevcut == 'Ic Onay' and yeni_durum == 'Onaylandi':
            ekstra_mesaj = f' {p.onaylayan} tarafindan onaylandi. Artik musteriye gonderilebilir.'
        elif mevcut == 'Ic Onay' and yeni_durum == 'Taslak':
            ekstra_mesaj = f' Onay reddedildi, taslaga geri dondu.'
        elif mevcut == 'Onaylandi' and yeni_durum == 'Gonderildi':
            ekstra_mesaj = ' Musteriye gonderildi olarak isaretlendi.'

        # ── ONAYLANDI (iç onay): teklif proforması da onaylanabilir, sipariş ZORUNLU DEĞİL.
        # (Yeni akışta "Onaylandi" = iç onaydan geçti demek, müşteri kabulü değil.)
        # Sipariş bağlama zorunluluğu faturaya dönüştürme anına taşındı.
        if yeni_durum == 'Onaylandi':
            if p.siparis_id:
                # Zaten bir siparişe bağlıysa, siparişin durumunu ve rezervasyonları güncelle
                sip = Siparis.query.get(p.siparis_id)
                if sip and sip.durum in ('Teklif Asam.', 'Onaylandi'):
                    sip.durum = 'Onaylandi'
                guncellenen = _proformayi_siparise_baglava(p, p.siparis_id)
                if guncellenen > 0:
                    ekstra_mesaj += f' {guncellenen} stok Satildi durumuna gecti.'

        # ── IPTAL: rezervasyonları iptal, stokları serbest bırak
        elif yeni_durum == 'Iptal':
            iptal_say = _proforma_rezervasyonlarini_iptal_et(p.id, f'Proforma {p.id} iptal')
            if iptal_say > 0:
                ekstra_mesaj = f' {iptal_say} stok rezervasyonu iptal edildi.'

            # Otomatik olusmus siparise bagliysa: baska aktif proforma yoksa siparisi de iptal et
            if p.siparis_id:
                sip = Siparis.query.get(p.siparis_id)
                if sip:
                    # Siparise bagli baska aktif (Iptal/Faturalandi olmayan) proforma var mi?
                    baska_aktif = Proforma.query.filter(
                        Proforma.siparis_id == sip.id,
                        Proforma.id != p.id,
                        ~Proforma.durum.in_(['Iptal', 'Faturalandi'])
                    ).count()
                    # Siparis otomatik olusmus mu? (aciklamada izi var)
                    otomatik_olusmus = sip.aciklama and 'Proforma' in sip.aciklama and 'onayindan' in sip.aciklama
                    # Siparis henuz uretim/sevkiyat asamasinda degilse iptal et
                    iptal_edilebilir_durumlar = ('Teklif Asam.', 'Onaylandi')
                    if baska_aktif == 0 and otomatik_olusmus and sip.durum in iptal_edilebilir_durumlar:
                        sip.durum = 'Iptal'
                        ekstra_mesaj += f' Bagli siparis {sip.id} de iptal edildi.'
                    elif baska_aktif > 0:
                        ekstra_mesaj += f' Siparis {sip.id} korunuyor ({baska_aktif} aktif proforma var).'
                    elif sip.durum not in iptal_edilebilir_durumlar:
                        ekstra_mesaj += f' Siparis {sip.id} ({sip.durum}) uretim/sevk asamasinda, manuel kontrol edin.'

        _log_audit('DURUM', 'proforma', proforma_id, eski={'durum': eski_durum}, yeni={'durum': yeni_durum})
        db.session.commit()
        logging.info(f"Proforma durum: {proforma_id} {eski_durum} -> {yeni_durum} (by {session.get('kullanici')})")
        return jsonify({'ok': True, 'mesaj': f'Durum guncellendi: {eski_durum} -> {yeni_durum}.{ekstra_mesaj}'})

    @app.route('/api/proforma/<proforma_id>', methods=['DELETE'])
    def api_proforma_sil(proforma_id):
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        p = Proforma.query.get(proforma_id)
        if not p:
            return jsonify({'ok': False, 'mesaj': 'Proforma bulunamadi'}), 404
        # Sadece TASLAK ve IPTAL durumundakiler silinebilir.
        # Onaylanmış/gönderilmiş/faturalanmış belgeler denetim izi için korunur —
        # bunlar önce İptal durumuna alınmalı (iz kalır), sonra silinebilir.
        if p.durum not in ('Taslak', 'Iptal'):
            return jsonify({
                'ok': False,
                'mesaj': f'Bu proforma "{p.durum}" durumunda ve silinemez. Onaylanmis/gonderilmis '
                         f'belgeler denetim izi icin korunur. Silmek icin once Iptal durumuna alin.'
            }), 400
        kalem_sayisi = ProformaKalem.query.filter_by(proforma_id=proforma_id).count()
        # Iptal durumunda zaten stoklar serbest olmus olmali, yine de garanti olsun
        iptal_say = _proforma_rezervasyonlarini_iptal_et(proforma_id, 'Proforma silindi')
        _log_audit('SIL', 'proforma', proforma_id, eski={'musteri': p.musteri, 'durum': p.durum, 'kalem': kalem_sayisi})
        db.session.delete(p)
        db.session.commit()
        logging.info(f"Proforma silindi: {proforma_id} ({kalem_sayisi} kalem, {iptal_say} rezervasyon) - by {session.get('kullanici')}")
        msg = f'Proforma silindi ({kalem_sayisi} kalem)'
        if iptal_say > 0:
            msg += f', {iptal_say} stok rezervasyonu iptal edildi'
        return jsonify({'ok': True, 'mesaj': msg})

    # ---------- PDF EKSTRE ----------
    @app.route('/api/cari/<cari_id>/ekstre_pdf')
    def api_ekstre_pdf(cari_id):
        cari = Cari.query.get(cari_id)
        if not cari: return "Cari bulunamadı", 404
        hareketler = CariHareket.query.filter_by(cari_id=cari_id).order_by(CariHareket.hareket_tarihi, CariHareket.guncelleme).all()
        # Cari'nin tercih ettigi para birimi (ya da varsayilan TRY)
        hedef_doviz = cari.para_birimi or 'TRY'

        # COKLU DOVIZ FARKINDALIGI: TRY uzerinden topla, sonra hedef dovize cevir
        # Her hareketin TRY karsiligi: borc_try / alacak_try kullanilir, yoksa hareket dovizi * kur ile cevrilir
        toplam_borc_try = 0
        toplam_alacak_try = 0
        kumulatif_try = 0
        for h in hareketler:
            # Borc TRY karsiligi
            if h.borc_try is not None:
                b_try = h.borc_try or 0
            else:
                kur = h.kur_uygulanan or (_kur_getir(h.doviz, h.hareket_tarihi) if h.doviz != 'TRY' else 1)
                b_try = (h.borc or 0) * (kur or 1)
            # Alacak TRY karsiligi
            if h.alacak_try is not None:
                a_try = h.alacak_try or 0
            else:
                kur = h.kur_uygulanan or (_kur_getir(h.doviz, h.hareket_tarihi) if h.doviz != 'TRY' else 1)
                a_try = (h.alacak or 0) * (kur or 1)

            toplam_borc_try += b_try
            toplam_alacak_try += a_try
            kumulatif_try += b_try - a_try
            # Goruntu icin: hedef doviz cinsinden bakiye
            if hedef_doviz == 'TRY':
                h.bakiye = kumulatif_try
            else:
                hedef_kur = _kur_getir(hedef_doviz, h.hareket_tarihi) or 1
                h.bakiye = kumulatif_try / hedef_kur if hedef_kur else 0

        # Toplamlari hedef doviz cinsine cevir
        if hedef_doviz == 'TRY':
            toplam_borc = toplam_borc_try
            toplam_alacak = toplam_alacak_try
            net_bakiye = toplam_borc_try - toplam_alacak_try
        else:
            son_kur = _kur_getir(hedef_doviz, date.today()) or 1
            toplam_borc = toplam_borc_try / son_kur if son_kur else 0
            toplam_alacak = toplam_alacak_try / son_kur if son_kur else 0
            net_bakiye = (toplam_borc_try - toplam_alacak_try) / son_kur if son_kur else 0

        return render_template('ekstre_print.html', cari=cari, hareketler=hareketler, baslik='Cari Ekstre', bugun=date.today(),
                               firma_adi='Milestone Mermer', toplam_borc=toplam_borc, toplam_alacak=toplam_alacak,
                               net_bakiye=net_bakiye, hedef_doviz=hedef_doviz)

    @app.route('/api/cari/<cari_id>/siparis_ekstre_pdf')
    def api_siparis_ekstre_pdf(cari_id):
        cari = Cari.query.get(cari_id)
        if not cari: return "Cari bulunamadı", 404
        siparis_id = request.args.get('siparis_id')
        if not siparis_id: return "Sipariş ID belirtilmedi", 400
        hareketler = CariHareket.query.filter_by(cari_id=cari_id, siparis_id=siparis_id).order_by(CariHareket.hareket_tarihi, CariHareket.guncelleme).all()
        hedef_doviz = cari.para_birimi or 'TRY'

        toplam_borc_try = 0
        toplam_alacak_try = 0
        kumulatif_try = 0
        for h in hareketler:
            if h.borc_try is not None:
                b_try = h.borc_try or 0
            else:
                kur = h.kur_uygulanan or (_kur_getir(h.doviz, h.hareket_tarihi) if h.doviz != 'TRY' else 1)
                b_try = (h.borc or 0) * (kur or 1)
            if h.alacak_try is not None:
                a_try = h.alacak_try or 0
            else:
                kur = h.kur_uygulanan or (_kur_getir(h.doviz, h.hareket_tarihi) if h.doviz != 'TRY' else 1)
                a_try = (h.alacak or 0) * (kur or 1)

            toplam_borc_try += b_try
            toplam_alacak_try += a_try
            kumulatif_try += b_try - a_try
            if hedef_doviz == 'TRY':
                h.bakiye = kumulatif_try
            else:
                hedef_kur = _kur_getir(hedef_doviz, h.hareket_tarihi) or 1
                h.bakiye = kumulatif_try / hedef_kur if hedef_kur else 0

        if hedef_doviz == 'TRY':
            toplam_borc = toplam_borc_try
            toplam_alacak = toplam_alacak_try
            net_bakiye = toplam_borc_try - toplam_alacak_try
        else:
            son_kur = _kur_getir(hedef_doviz, date.today()) or 1
            toplam_borc = toplam_borc_try / son_kur if son_kur else 0
            toplam_alacak = toplam_alacak_try / son_kur if son_kur else 0
            net_bakiye = (toplam_borc_try - toplam_alacak_try) / son_kur if son_kur else 0

        return render_template('ekstre_print.html', cari=cari, hareketler=hareketler, baslik=f'Sipariş Ekstresi - {siparis_id}',
                               bugun=date.today(), firma_adi='Milestone Mermer', toplam_borc=toplam_borc, toplam_alacak=toplam_alacak,
                               net_bakiye=net_bakiye, hedef_doviz=hedef_doviz)


    # ========== YEDEKLEME ENDPOINT'LERİ ==========
    # ─── KULLANICI YÖNETİMİ + YETKİLER ────────────────────────────────
    @app.route('/api/yetkilerim', methods=['GET'])
    def api_yetkilerim():
        """Aktif kullanıcının kendi yetkilerini döner (frontend menü/buton kontrolü için)."""
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        return jsonify({
            'ok': True,
            'kullanici': session.get('kullanici'),
            'rol': session.get('rol'),
            'admin': session.get('rol') in ('admin', 'ADMIN'),
            'yetkiler': _kullanici_yetkileri(),
            'proforma_onay': _proforma_onay_yetkisi_var_mi(),
            'moduller': YETKI_MODULLERI
        })



    @app.route('/api/yedek/liste', methods=['GET'])
    def api_yedek_liste():
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        if session.get('rol') != 'ADMIN':
            return jsonify({'error': 'Sadece admin yedek yönetebilir'}), 403
        return jsonify({'ok': True, 'yedekler': yedek_modul.yedek_listesi()})

    @app.route('/api/yedek/tani', methods=['GET'])
    def api_yedek_tani():
        """Yedekleme sisteminin durumunu kontrol eder (sorun giderme)."""
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        if session.get('rol') != 'ADMIN':
            return jsonify({'error': 'Sadece admin'}), 403
        try:
            return jsonify({'ok': True, 'tani': yedek_modul.tani()})
        except Exception as e:
            return jsonify({'ok': False, 'mesaj': str(e)}), 500

    @app.route('/api/yedek/al', methods=['POST'])
    def api_yedek_al():
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        if session.get('rol') != 'ADMIN':
            return jsonify({'error': 'Sadece admin yedek alabilir'}), 403
        yol = yedek_modul.yedek_al('manual')
        if yol:
            return jsonify({'ok': True, 'mesaj': 'Yedek alındı', 'dosya': os.path.basename(yol)})
        # Başarısızsa neden başarısız olduğunu tanı ile açıkla
        try:
            t = yedek_modul.tani()
            return jsonify({'ok': False, 'mesaj': 'Yedekleme başarısız: ' + t.get('mesaj', 'bilinmeyen hata'),
                            'tani': t}), 500
        except Exception:
            return jsonify({'ok': False, 'mesaj': 'Yedekleme başarısız'}), 500

    @app.route('/api/yedek/geri_yukle', methods=['POST'])
    def api_yedek_geri_yukle():
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        if session.get('rol') != 'ADMIN':
            return jsonify({'error': 'Sadece admin geri yükleyebilir'}), 403
        dosya = request.json.get('dosya')
        if not dosya: return jsonify({'ok': False, 'mesaj': 'Dosya adı gerekli'}), 400
        basarili, mesaj = yedek_modul.yedek_geri_yukle(dosya)
        return jsonify({'ok': basarili, 'mesaj': mesaj})

    @app.route('/api/yedek/sil', methods=['POST'])
    def api_yedek_sil():
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        if session.get('rol') != 'ADMIN':
            return jsonify({'error': 'Sadece admin silebilir'}), 403
        dosya = request.json.get('dosya')
        if not dosya: return jsonify({'ok': False, 'mesaj': 'Dosya adı gerekli'}), 400
        basarili, mesaj = yedek_modul.yedek_sil(dosya)
        return jsonify({'ok': basarili, 'mesaj': mesaj})

    @app.route('/api/yedek/indir/<dosya>', methods=['GET'])
    def api_yedek_indir(dosya):
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        if session.get('rol') != 'ADMIN':
            return jsonify({'error': 'Yetkisiz'}), 403
        # Güvenlik: sadece milestone_*.dump formatında olanlar
        if not dosya.startswith('milestone_') or not dosya.endswith('.dump'):
            return jsonify({'error': 'Geçersiz dosya'}), 400
        yedek_yolu = yedek_modul.BACKUP_DIR / dosya
        if not yedek_yolu.exists():
            return jsonify({'error': 'Bulunamadı'}), 404
        return send_file(str(yedek_yolu), as_attachment=True, download_name=dosya)



    # ════════════════════════════════════════════════════════
    # SATIŞ KAYDI ENDPOINT'LERİ
    # ════════════════════════════════════════════════════════
    @app.route('/satislar')
    def satislar_sayfa():
        if 'kullanici' not in session: return redirect(url_for('giris'))
        if not _yetki_var_mi('satislar', 'okuma'):
            return redirect(url_for('dashboard'))
        return render_template('satislar.html')

    @app.route('/api/satislar', methods=['GET'])
    def api_satis_liste():
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401

        # Filtreler
        musteri  = request.args.get('musteri')
        stok_tip = request.args.get('stok_tip')
        cins     = request.args.get('cins')
        bas_tarih = request.args.get('bas_tarih')
        son_tarih = request.args.get('son_tarih')
        fatura_durum = request.args.get('fatura_durum')  # 'var', 'yok'
        ara      = request.args.get('ara', '').strip()

        page     = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 25))

        q = SatisKaydi.query

        if musteri:
            q = q.filter(SatisKaydi.musteri == musteri)
        if stok_tip:
            q = q.filter(SatisKaydi.stok_tip == stok_tip)
        if cins:
            q = q.filter(SatisKaydi.cins == cins)
        if bas_tarih:
            try: q = q.filter(SatisKaydi.satis_tarihi >= _parse_date(bas_tarih))
            except: pass
        if son_tarih:
            try: q = q.filter(SatisKaydi.satis_tarihi <= _parse_date(son_tarih))
            except: pass
        if fatura_durum == 'var':
            q = q.filter(SatisKaydi.fatura_no.isnot(None), SatisKaydi.fatura_no != '')
        elif fatura_durum == 'yok':
            q = q.filter((SatisKaydi.fatura_no.is_(None)) | (SatisKaydi.fatura_no == ''))
        if ara:
            like = f'%{ara}%'
            q = q.filter(db.or_(
                SatisKaydi.stok_id.ilike(like),
                SatisKaydi.cins.ilike(like),
                SatisKaydi.blok_no.ilike(like),
                SatisKaydi.musteri.ilike(like),
                SatisKaydi.fatura_no.ilike(like),
                SatisKaydi.siparis_id.ilike(like),
                SatisKaydi.proforma_id.ilike(like)
            ))

        paginated = q.order_by(SatisKaydi.satis_tarihi.desc(), SatisKaydi.olusturma.desc()).paginate(
            page=page, per_page=per_page, error_out=False
        )

        rows = []
        for s in paginated.items:
            rows.append({
                'id': s.id,
                'stok_id': s.stok_id, 'stok_tip': s.stok_tip,
                'cins': s.cins, 'ozellik': s.ozellik, 'blok_no': s.blok_no,
                # Kullanıcıya gösterilecek okunabilir numara (ID değil)
                'stok_no': (s.blok_no or s.stok_id),
                'boy': s.boy, 'yukseklik': s.yukseklik, 'kalinlik': s.kalinlik,
                'metraj_m2': s.metraj_m2, 'metraj_sqft': s.metraj_sqft,
                'hacim_m3': s.hacim_m3, 'tonaj': s.tonaj,
                'siparis_id': s.siparis_id,
                'proforma_id': s.proforma_id,
                'sevkiyat_id': s.sevkiyat_id,
                'musteri': s.musteri, 'musteri_ulke': s.musteri_ulke,
                'satis_tarihi': s.satis_tarihi.isoformat() if s.satis_tarihi else None,
                'teslim_tarihi': s.teslim_tarihi.isoformat() if s.teslim_tarihi else None,
                'fatura_no': s.fatura_no,
                'fatura_tarihi': s.fatura_tarihi.isoformat() if s.fatura_tarihi else None,
                'birim_fiyat': s.birim_fiyat, 'miktar': s.miktar, 'birim': s.birim,
                'doviz': s.doviz, 'tutar': s.tutar,
                'kur_usd': s.kur_usd, 'kur_eur': s.kur_eur,
                'tutar_usd': s.tutar_usd, 'tutar_try': s.tutar_try,
                'maliyet_usd': s.maliyet_usd, 'maliyet_try': s.maliyet_try,
                'kar_usd': s.kar_usd, 'marj_yuzde': s.marj_yuzde,
                'notlar': s.notlar
            })

        # Özet (filtreli sorgu üzerinden)
        ozet_q = q.with_entities(
            db.func.count(SatisKaydi.id).label('adet'),
            db.func.sum(SatisKaydi.tutar_usd).label('toplam_usd'),
            db.func.sum(SatisKaydi.tutar_try).label('toplam_try'),
            db.func.sum(SatisKaydi.maliyet_usd).label('maliyet_usd'),
            db.func.sum(SatisKaydi.kar_usd).label('kar_usd')
        ).first()

        toplam_usd = ozet_q.toplam_usd or 0
        kar_usd = ozet_q.kar_usd or 0
        ort_marj = (kar_usd / toplam_usd * 100) if toplam_usd > 0 else 0

        return jsonify({
            'data': rows,
            'meta': {
                'page': paginated.page,
                'pages': paginated.pages,
                'total': paginated.total,
                'per_page': paginated.per_page
            },
            'ozet': {
                'adet': ozet_q.adet or 0,
                'toplam_usd': toplam_usd,
                'toplam_try': ozet_q.toplam_try or 0,
                'maliyet_usd': ozet_q.maliyet_usd or 0,
                'kar_usd': kar_usd,
                'ort_marj': ort_marj
            }
        })

    @app.route('/api/satislar/<satis_id>', methods=['GET'])
    def api_satis_detay(satis_id):
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        s = SatisKaydi.query.get(satis_id)
        if not s:
            return jsonify({'ok': False, 'mesaj': 'Bulunamadi'}), 404
        return jsonify({
            'ok': True,
            'satis': {
                'id': s.id,
                'stok_id': s.stok_id, 'stok_tip': s.stok_tip,
                'cins': s.cins, 'ozellik': s.ozellik, 'blok_no': s.blok_no,
                # Kullanıcıya gösterilecek okunabilir numara (ID değil)
                'stok_no': (s.blok_no or s.stok_id),
                'boy': s.boy, 'yukseklik': s.yukseklik, 'kalinlik': s.kalinlik, 'en': s.en,
                'metraj_m2': s.metraj_m2, 'metraj_sqft': s.metraj_sqft,
                'hacim_m3': s.hacim_m3, 'tonaj': s.tonaj, 'agirlik_kg': s.agirlik_kg,
                'siparis_id': s.siparis_id, 'proforma_id': s.proforma_id, 'sevkiyat_id': s.sevkiyat_id,
                'musteri': s.musteri, 'musteri_ulke': s.musteri_ulke,
                'satis_tarihi': s.satis_tarihi.isoformat() if s.satis_tarihi else None,
                'teslim_tarihi': s.teslim_tarihi.isoformat() if s.teslim_tarihi else None,
                'fatura_no': s.fatura_no,
                'fatura_tarihi': s.fatura_tarihi.isoformat() if s.fatura_tarihi else None,
                'birim_fiyat': s.birim_fiyat, 'miktar': s.miktar, 'birim': s.birim,
                'doviz': s.doviz, 'tutar': s.tutar,
                'kur_usd': s.kur_usd, 'kur_eur': s.kur_eur,
                'tutar_usd': s.tutar_usd, 'tutar_try': s.tutar_try,
                'maliyet_usd': s.maliyet_usd, 'maliyet_try': s.maliyet_try,
                'kar_usd': s.kar_usd, 'marj_yuzde': s.marj_yuzde,
                'notlar': s.notlar,
                'kullanici': s.kullanici,
                'olusturma': s.olusturma.isoformat() if s.olusturma else None
            }
        })

    @app.route('/api/satislar/<satis_id>/fatura', methods=['POST'])
    def api_satis_fatura_guncelle(satis_id):
        """
        Fatura numarası, fatura tarihi, teslim tarihi ve notları günceller.
        Proforma_id'ye DOKUNMAZ — kayıtlarda korunur.
        Diğer kritik alanlar (tutar, maliyet, kar) değiştirilemez.
        """
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        s = SatisKaydi.query.get(satis_id)
        if not s:
            return jsonify({'ok': False, 'mesaj': 'Bulunamadi'}), 404

        data = request.json or {}
        if 'fatura_no' in data:
            s.fatura_no = (data.get('fatura_no') or '').strip() or None
        if 'fatura_tarihi' in data:
            s.fatura_tarihi = _parse_date(data.get('fatura_tarihi')) if data.get('fatura_tarihi') else None
        if 'teslim_tarihi' in data:
            s.teslim_tarihi = _parse_date(data.get('teslim_tarihi')) if data.get('teslim_tarihi') else None
        if 'notlar' in data:
            s.notlar = data.get('notlar')

        db.session.commit()
        logging.info(f"Satis fatura guncelleme: {satis_id} fatura={s.fatura_no} by {session.get('kullanici')}")
        return jsonify({'ok': True, 'mesaj': 'Satis kaydi guncellendi'})

    @app.route('/api/satislar/musteri/<musteri>/ozet', methods=['GET'])
    def api_satis_musteri_ozet(musteri):
        """Müşteri bazlı satış özeti."""
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        sklar = SatisKaydi.query.filter_by(musteri=musteri).all()
        if not sklar:
            return jsonify({'ok': True, 'adet': 0})

        toplam_usd = sum(s.tutar_usd or 0 for s in sklar)
        toplam_try = sum(s.tutar_try or 0 for s in sklar)
        maliyet_usd = sum(s.maliyet_usd or 0 for s in sklar)
        kar_usd = toplam_usd - maliyet_usd

        return jsonify({
            'ok': True,
            'adet': len(sklar),
            'toplam_usd': toplam_usd,
            'toplam_try': toplam_try,
            'maliyet_usd': maliyet_usd,
            'kar_usd': kar_usd,
            'marj_yuzde': (kar_usd / toplam_usd * 100) if toplam_usd > 0 else 0,
            'ilk_satis': min(s.satis_tarihi for s in sklar if s.satis_tarihi).isoformat() if any(s.satis_tarihi for s in sklar) else None,
            'son_satis': max(s.satis_tarihi for s in sklar if s.satis_tarihi).isoformat() if any(s.satis_tarihi for s in sklar) else None
        })

    @app.route('/api/satislar/lookup', methods=['GET'])
    def api_satis_lookup():
        """Filtre dropdown'lari icin distinct musteri ve cins."""
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        musteriler = [r[0] for r in db.session.query(SatisKaydi.musteri).distinct().all() if r[0]]
        cinsler = [r[0] for r in db.session.query(SatisKaydi.cins).distinct().all() if r[0]]
        return jsonify({'musteriler': sorted(musteriler), 'cinsler': sorted(cinsler)})


    # ════════════════════════════════════════════════════════
    # AUDIT LOG (Denetim Kaydı)
    # ════════════════════════════════════════════════════════
    @app.route('/denetim')
    def denetim_sayfa():
        if 'kullanici' not in session: return redirect(url_for('giris'))
        if session.get('rol') not in ('ADMIN', 'admin') and not _yetki_var_mi('denetim', 'okuma'):
            return render_template('denetim.html', yetkisiz=True)
        return render_template('denetim.html')

    @app.route('/api/denetim', methods=['GET'])
    def api_denetim_liste():
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        if session.get('rol') != 'ADMIN':
            return jsonify({'error': 'Sadece admin'}), 403

        # Filtreler
        islem    = request.args.get('islem')
        tablo    = request.args.get('tablo')
        kullanici = request.args.get('kullanici')
        bas_tarih = request.args.get('bas_tarih')
        son_tarih = request.args.get('son_tarih')

        page     = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 50))

        q = AuditLog.query
        if islem:    q = q.filter(AuditLog.islem_tipi == islem)
        if tablo:    q = q.filter(AuditLog.tablo_adi == tablo)
        if kullanici: q = q.filter(AuditLog.kullanici == kullanici)
        if bas_tarih:
            try: q = q.filter(AuditLog.tarih >= datetime.strptime(bas_tarih, '%Y-%m-%d'))
            except: pass
        if son_tarih:
            try:
                son = datetime.strptime(son_tarih, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
                q = q.filter(AuditLog.tarih <= son)
            except: pass

        paginated = q.order_by(AuditLog.tarih.desc()).paginate(page=page, per_page=per_page, error_out=False)

        rows = []
        for a in paginated.items:
            eski = yeni = None
            try:
                if a.eski_veri: eski = json.loads(a.eski_veri)
            except: pass
            try:
                if a.yeni_veri: yeni = json.loads(a.yeni_veri)
            except: pass
            rows.append({
                'id': a.id,
                'tarih': a.tarih.strftime('%d.%m.%Y %H:%M:%S') if a.tarih else None,
                'kullanici': a.kullanici,
                'islem': a.islem_tipi,
                'tablo': a.tablo_adi,
                'kayit_id': a.kayit_id,
                'eski': eski,
                'yeni': yeni,
                'ip': a.ip_adresi
            })

        return jsonify({
            'data': rows,
            'meta': {'page': paginated.page, 'pages': paginated.pages, 'total': paginated.total}
        })

    @app.route('/api/denetim/lookup', methods=['GET'])
    def api_denetim_lookup():
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        islemler = [r[0] for r in db.session.query(AuditLog.islem_tipi).distinct().all() if r[0]]
        tablolar = [r[0] for r in db.session.query(AuditLog.tablo_adi).distinct().all() if r[0]]
        kullanicilar = [r[0] for r in db.session.query(AuditLog.kullanici).distinct().all() if r[0]]
        return jsonify({
            'islemler': sorted(islemler),
            'tablolar': sorted(tablolar),
            'kullanicilar': sorted(kullanicilar)
        })


    # ════════════════════════════════════════════════════════
    # DASHBOARD SATIŞ İSTATİSTİKLERİ
    # ════════════════════════════════════════════════════════
    @app.route('/api/dashboard/satis_istatistik', methods=['GET'])
    def api_dashboard_satis():
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401

        bugun = date.today()

        # Son 6 ayın aylık satış toplamı
        aylik = []
        for i in range(5, -1, -1):
            ay_basi = (bugun.replace(day=1) - timedelta(days=i*30)).replace(day=1)
            # Bir sonraki ayin ilk gunu
            if ay_basi.month == 12:
                ay_sonu = ay_basi.replace(year=ay_basi.year+1, month=1, day=1)
            else:
                ay_sonu = ay_basi.replace(month=ay_basi.month+1, day=1)

            satislar = SatisKaydi.query.filter(
                SatisKaydi.satis_tarihi >= ay_basi,
                SatisKaydi.satis_tarihi < ay_sonu
            ).all()
            toplam_usd = sum(s.tutar_usd or 0 for s in satislar)
            toplam_kar = sum(s.kar_usd or 0 for s in satislar)
            aylar_tr = ['Oca','Şub','Mar','Nis','May','Haz','Tem','Ağu','Eyl','Eki','Kas','Ara']
            aylik.append({
                'ay': f'{aylar_tr[ay_basi.month-1]} {ay_basi.year}',
                'satis_usd': q3(toplam_usd),
                'kar_usd': q3(toplam_kar),
                'adet': len(satislar)
            })

        # En karlı 5 müşteri (SatisKaydi'ndan)
        musteri_q = db.session.query(
            SatisKaydi.musteri,
            db.func.sum(SatisKaydi.tutar_usd).label('satis'),
            db.func.sum(SatisKaydi.kar_usd).label('kar'),
            db.func.count(SatisKaydi.id).label('adet')
        ).group_by(SatisKaydi.musteri).order_by(db.func.sum(SatisKaydi.kar_usd).desc()).limit(5).all()
        en_karli_musteri = [{
            'musteri': m.musteri,
            'satis_usd': q3(m.satis or 0),
            'kar_usd': q3(m.kar or 0),
            'adet': m.adet
        } for m in musteri_q]

        # En karlı 5 ürün cinsi
        cins_q = db.session.query(
            SatisKaydi.cins,
            db.func.sum(SatisKaydi.tutar_usd).label('satis'),
            db.func.sum(SatisKaydi.kar_usd).label('kar'),
            db.func.count(SatisKaydi.id).label('adet')
        ).group_by(SatisKaydi.cins).order_by(db.func.sum(SatisKaydi.kar_usd).desc()).limit(5).all()
        en_karli_cins = [{
            'cins': c.cins or '(belirsiz)',
            'satis_usd': q3(c.satis or 0),
            'kar_usd': q3(c.kar or 0),
            'adet': c.adet
        } for c in cins_q]

        # Genel toplam (tüm zamanlar)
        tum = SatisKaydi.query.all()
        genel = {
            'toplam_satis': q3(sum(s.tutar_usd or 0 for s in tum)),
            'toplam_kar': q3(sum(s.kar_usd or 0 for s in tum)),
            'toplam_adet': len(tum),
            'faturasiz': SatisKaydi.query.filter(
                (SatisKaydi.fatura_no.is_(None)) | (SatisKaydi.fatura_no == '')
            ).count()
        }
        genel['ort_marj'] = q_oran(
            (genel['toplam_kar'] / genel['toplam_satis'] * 100) if genel['toplam_satis'] else 0
        )

        # Bu ay
        ay_basi = bugun.replace(day=1)
        bu_ay = SatisKaydi.query.filter(SatisKaydi.satis_tarihi >= ay_basi).all()
        bu_ay_ozet = {
            'satis': q3(sum(s.tutar_usd or 0 for s in bu_ay)),
            'kar': q3(sum(s.kar_usd or 0 for s in bu_ay)),
            'adet': len(bu_ay)
        }

        return jsonify({
            'aylik': aylik,
            'en_karli_musteri': en_karli_musteri,
            'en_karli_cins': en_karli_cins,
            'genel': genel,
            'bu_ay': bu_ay_ozet
        })


    # ════════════════════════════════════════════════════════
    # PROFORMA / PACKING LIST / TİCARİ FATURA ÇIKTI
    # ════════════════════════════════════════════════════════
    def _sayi_yaziya_en(tutar, doviz='USD'):
        """Tutarı İngilizce yazıya çevirir (fatura için). Basit versiyon."""
        try:
            tam = int(tutar)
            kurus = int(round((tutar - tam) * 100))
        except (ValueError, TypeError):
            return ''
        birler = ['', 'ONE', 'TWO', 'THREE', 'FOUR', 'FIVE', 'SIX', 'SEVEN', 'EIGHT', 'NINE',
                  'TEN', 'ELEVEN', 'TWELVE', 'THIRTEEN', 'FOURTEEN', 'FIFTEEN', 'SIXTEEN',
                  'SEVENTEEN', 'EIGHTEEN', 'NINETEEN']
        onlar = ['', '', 'TWENTY', 'THIRTY', 'FORTY', 'FIFTY', 'SIXTY', 'SEVENTY', 'EIGHTY', 'NINETY']

        def uc_haneli(n):
            if n == 0: return ''
            sonuc = ''
            if n >= 100:
                sonuc += birler[n // 100] + ' HUNDRED '
                n %= 100
            if n >= 20:
                sonuc += onlar[n // 10] + ' '
                n %= 10
            if n > 0:
                sonuc += birler[n] + ' '
            return sonuc

        if tam == 0:
            yazi = 'ZERO '
        else:
            yazi = ''
            milyar = tam // 1_000_000_000
            milyon = (tam % 1_000_000_000) // 1_000_000
            bin = (tam % 1_000_000) // 1000
            kalan = tam % 1000
            if milyar: yazi += uc_haneli(milyar) + 'BILLION '
            if milyon: yazi += uc_haneli(milyon) + 'MILLION '
            if bin: yazi += uc_haneli(bin) + 'THOUSAND '
            if kalan: yazi += uc_haneli(kalan)

        doviz_ad = {'USD': 'US DOLLARS', 'EUR': 'EUROS', 'TRY': 'TURKISH LIRA'}.get(doviz, doviz)
        sonuc = yazi.strip() + ' ' + doviz_ad
        if kurus > 0:
            sonuc += ' AND ' + uc_haneli(kurus).strip() + ' CENTS'
        return sonuc.strip() + ' ONLY'

    def _packing_list_bundle_bol(proforma, kalemler):
        """
        Packing list icin kalemleri bundle'lara boler.
        Her PLAKA kalemi, bundle limitine gore parcalanir; her parca ayri satir.
        m2/agirlik plaka sayisina gore orantili dagitilir.
        slab_no = gercek plaka araligi, bundle_no = tek bundle numarasi.
        Doner: bolunmus kalem listesi (dict)
        """
        limit = getattr(proforma, 'genel_bundle_sayisi', 10) or 10
        karma = bool(getattr(proforma, 'karma_bundle', False))

        sonuc = []
        bundle_no = 1      # mevcut bundle numarasi
        bundle_dolu = 0    # mevcut bundle'da kac plaka var (karma icin)
        blok_sayac = {}    # blok_no -> sonraki bos plaka no

        for k in kalemler:
            tip = (k.urun_tip or 'PLAKA')
            adet = int(k.adet or 0)

            # PLAKA disi (BLOK/EBATLI) -> tek satir, bundle bolme yok
            if tip != 'PLAKA' or adet <= 0 or limit <= 0:
                sonuc.append({
                    'urun_tip': tip, 'cins': k.cins, 'yuzey_spec': k.yuzey_spec,
                    'ozellik': k.yuzey_spec, 'blok_no': k.blok_no,
                    'boy': k.boy, 'yukseklik': k.yukseklik, 'kalinlik': k.kalinlik,
                    'adet': adet, 'miktar': k.miktar or 0, 'agirlik': k.agirlik or 0,
                    'bundle_no': k.bundle_no or '', 'slab_no': k.slab_no or '',
                    'birim': k.birim
                })
                continue

            # Slab baslangici: slab_no'da '3' veya '3-17' olabilir, ilk sayiyi al
            bno = (k.blok_no or 'BILINMEYEN').strip()
            slab_raw = (k.slab_no or '').strip()
            slab_bas = None
            if slab_raw:
                ilk = slab_raw.split('-')[0].strip()
                if ilk.isdigit():
                    slab_bas = int(ilk)
            if slab_bas is None:
                slab_bas = blok_sayac.get(bno, 1)

            # Orantili dagitim icin plaka basi degerler
            top_m2 = k.miktar or 0
            top_agirlik = k.agirlik or 0
            m2_birim = top_m2 / adet if adet else 0
            agirlik_birim = top_agirlik / adet if adet else 0

            # Bundle'lara bol
            kalan = adet
            slab_imlec = slab_bas
            if not karma:
                # Karma degil: her kalem yeni bundle'dan baslar, bundle_dolu sifirla
                if bundle_dolu > 0:
                    bundle_no += 1
                    bundle_dolu = 0

            while kalan > 0:
                bosluk = limit - bundle_dolu
                if bosluk <= 0:
                    bundle_no += 1
                    bundle_dolu = 0
                    bosluk = limit
                bu_parca = min(kalan, bosluk)

                slab_son = slab_imlec + bu_parca - 1
                sonuc.append({
                    'urun_tip': tip, 'cins': k.cins, 'yuzey_spec': k.yuzey_spec,
                    'ozellik': k.yuzey_spec, 'blok_no': k.blok_no,
                    'boy': k.boy, 'yukseklik': k.yukseklik, 'kalinlik': k.kalinlik,
                    'adet': bu_parca,
                    'miktar': q3(m2_birim * bu_parca),
                    'agirlik': q3(agirlik_birim * bu_parca),
                    'bundle_no': str(bundle_no),
                    'slab_no': str(slab_imlec) if bu_parca == 1 else f'{slab_imlec}-{slab_son}',
                    'birim': k.birim
                })

                bundle_dolu += bu_parca
                kalan -= bu_parca
                slab_imlec = slab_son + 1

            blok_sayac[bno] = slab_imlec

        return sonuc

    def _proforma_cikti_data(proforma_id):
        """Proforma + kalemlerini şablona uygun şekilde hazırlar."""
        p = Proforma.query.get(proforma_id)
        if not p:
            return None, None, None
        kalemler = ProformaKalem.query.filter_by(proforma_id=proforma_id).order_by(
            ProformaKalem.sira, ProformaKalem.id).all()
        return p, kalemler, None

    def _proforma_belge_html(proforma_id, mod):
        """Bir proforma belgesinin (pi/pl/ci) HTML çıktısını render eder (PDF için)."""
        p, kalemler, _ = _proforma_cikti_data(proforma_id)
        if not p:
            return None
        toplam_adet = sum((k.adet or 0) for k in kalemler)
        toplam_agirlik = sum((k.agirlik or 0) for k in kalemler)
        toplam_yazili = _sayi_yaziya_en(p.toplam or 0, p.doviz or 'USD')
        if mod == 'pl':
            sablon = 'packing_list_print.html'
            cikti_kalemler = _packing_list_bundle_bol(p, kalemler)
        elif mod == 'ci':
            sablon = 'commercial_invoice_print.html'
            cikti_kalemler = kalemler
        else:
            sablon = 'proforma_print.html'
            cikti_kalemler = kalemler
        return render_template(sablon, p=p, kalemler=cikti_kalemler,
                               toplam_adet=toplam_adet, toplam_agirlik=toplam_agirlik,
                               toplam_yazili=toplam_yazili)

    def _html_to_pdf(html_str):
        """HTML string'i PDF bytes'a çevirir (weasyprint). Kurulu değilse None."""
        try:
            from weasyprint import HTML as _WeasyHTML
            return _WeasyHTML(string=html_str, base_url=request.url_root).write_pdf()
        except ImportError:
            app.logger.error("weasyprint kurulu degil — PDF uretilemiyor")
            return None
        except Exception as e:
            app.logger.error(f"PDF uretim hatasi: {e}")
            return None

    def _smtp_ayarlari_oku():
        """SMTP ayarlarını Veriler tablosundan okur. Döner dict veya None (eksikse)."""
        kayitlar = {v.deger: v.ek_bilgi for v in
                    Veriler.query.filter_by(kategori='smtp_ayar').all()}
        gerekli = ['sunucu', 'port', 'kullanici', 'sifre']
        if not all(kayitlar.get(k) for k in gerekli):
            return None
        return {
            'sunucu': kayitlar.get('sunucu'),
            'port': int(kayitlar.get('port') or 587),
            'kullanici': kayitlar.get('kullanici'),
            'sifre': kayitlar.get('sifre'),
            'gonderen_ad': kayitlar.get('gonderen_ad') or kayitlar.get('kullanici'),
            'gonderen_email': kayitlar.get('gonderen_email') or kayitlar.get('kullanici'),
            'guvenlik': (kayitlar.get('guvenlik') or 'tls').lower(),  # tls | ssl | yok
        }

    def _eposta_gonder(alici, konu, govde_text, ekler=None):
        """
        SMTP üzerinden e-posta gönderir. ekler: [(dosya_adi, bytes, mime_alt), ...]
        Döner: (basari: bool, mesaj: str)
        """
        import smtplib, ssl as _ssl
        from email.mime.multipart import MIMEMultipart
        from email.mime.text import MIMEText
        from email.mime.application import MIMEApplication
        from email.utils import formataddr

        ayar = _smtp_ayarlari_oku()
        if not ayar:
            return False, 'SMTP ayarlari eksik. Ayarlar > E-posta bolumunden yapilandirin.'

        msg = MIMEMultipart()
        msg['From'] = formataddr((ayar['gonderen_ad'], ayar['gonderen_email']))
        msg['To'] = alici
        msg['Subject'] = konu
        msg.attach(MIMEText(govde_text or '', 'plain', 'utf-8'))

        for (dosya_adi, icerik, mime_alt) in (ekler or []):
            if not icerik:
                continue
            ek = MIMEApplication(icerik, _subtype=mime_alt or 'pdf')
            ek.add_header('Content-Disposition', 'attachment', filename=dosya_adi)
            msg.attach(ek)

        try:
            reddedilenler = {}
            if ayar['guvenlik'] == 'ssl':
                ctx = _ssl.create_default_context()
                with smtplib.SMTP_SSL(ayar['sunucu'], ayar['port'], context=ctx, timeout=30) as sunucu:
                    sunucu.login(ayar['kullanici'], ayar['sifre'])
                    reddedilenler = sunucu.send_message(msg)
            else:
                with smtplib.SMTP(ayar['sunucu'], ayar['port'], timeout=30) as sunucu:
                    if ayar['guvenlik'] == 'tls':
                        sunucu.starttls(context=_ssl.create_default_context())
                    sunucu.login(ayar['kullanici'], ayar['sifre'])
                    reddedilenler = sunucu.send_message(msg)
            # send_message, teslim edilemeyen alıcıları dict olarak döner ({adres: (kod, mesaj)})
            if reddedilenler:
                detay = '; '.join(f"{a}: {r}" for a, r in reddedilenler.items())
                app.logger.error(f"E-posta reddedilen alicilar: {detay}")
                return False, f'Sunucu bazi alicilari reddetti: {detay}'
            app.logger.info(f"E-posta gonderildi: {alici} (konu: {konu})")
            return True, 'E-posta gonderildi'
        except smtplib.SMTPAuthenticationError as e:
            return False, f'SMTP kimlik dogrulama hatasi (kullanici/sifre): {e}'
        except smtplib.SMTPRecipientsRefused as e:
            return False, f'Alici adresi sunucu tarafindan reddedildi: {e.recipients}'
        except smtplib.SMTPSenderRefused as e:
            return False, f'Gonderen adresi reddedildi (From ile SMTP kullanici uyumsuz olabilir): {e}'
        except smtplib.SMTPResponseException as e:
            return False, f'SMTP sunucu hatasi [{e.smtp_code}]: {e.smtp_error}'
        except Exception as e:
            app.logger.error(f"E-posta gonderim hatasi: {type(e).__name__}: {e}")
            return False, f'E-posta gonderilemedi: {type(e).__name__}: {e}'

    @app.route('/api/proforma/<proforma_id>/eposta_gonder', methods=['POST'])
    def api_proforma_eposta(proforma_id):
        """
        Proforma belgelerini (PI/PL/CI seçmeli) müşterinin kayıtlı e-postasına PDF ekli gönderir.
        Body: { belgeler: ['pi','pl','ci'], on_metin: '...', alici: '(ops. override)' }
        """
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        p = Proforma.query.get(proforma_id)
        if not p:
            return jsonify({'ok': False, 'mesaj': 'Proforma bulunamadi'}), 404

        d = request.json or {}
        belgeler = d.get('belgeler') or []
        on_metin = (d.get('on_metin') or '').strip()
        if not belgeler:
            return jsonify({'ok': False, 'mesaj': 'En az bir belge secmelisiniz (PI/PL/CI).'}), 400

        # CI seçildiyse onay kontrolü (onaysız CI gönderilemez)
        if 'ci' in belgeler and p.durum not in ('Onaylandi', 'Gonderildi', 'Faturalandi'):
            return jsonify({'ok': False,
                'mesaj': 'Commercial Invoice yalnizca onaylanmis proformada gonderilebilir. '
                         'CI secimini kaldirin veya proformayi onaylatin.'}), 400

        # Alıcı: override varsa onu, yoksa carinin kayıtlı e-postası
        alici = (d.get('alici') or '').strip()
        cari = _cari_bul(p.musteri)
        if not alici:
            alici = (cari.email if cari else '') or ''
        if not alici or '@' not in alici:
            return jsonify({'ok': False, 'error': 'alici_yok',
                'mesaj': f'Musteri "{p.musteri}" icin kayitli gecerli e-posta yok. '
                         f'Cari kartina e-posta ekleyin veya elle girin.'}), 400

        # SMTP kontrolü (önden — PDF üretmeden boşuna uğraşma)
        if not _smtp_ayarlari_oku():
            return jsonify({'ok': False, 'error': 'smtp_yok',
                'mesaj': 'SMTP ayarlari yapilmamis. Ayarlar > E-posta bolumunden yapilandirin.'}), 400

        # Belgeleri PDF olarak hazırla
        belge_adlari = {'pi': 'Proforma_Invoice', 'pl': 'Packing_List', 'ci': 'Commercial_Invoice'}
        ekler = []
        uretilemedi = []
        for b in belgeler:
            if b not in belge_adlari:
                continue
            html_str = _proforma_belge_html(proforma_id, b)
            if not html_str:
                uretilemedi.append(b)
                continue
            pdf = _html_to_pdf(html_str)
            if not pdf:
                uretilemedi.append(b)
                continue
            dosya_adi = f"{belge_adlari[b]}_{p.id}.pdf"
            ekler.append((dosya_adi, pdf, 'pdf'))

        if not ekler:
            return jsonify({'ok': False,
                'mesaj': 'Belgeler PDF olarak uretilemedi. weasyprint kurulu olmayabilir.'}), 500

        # Konu ve gövde
        konu = f"{p.musteri} - Proforma {p.id}"
        govde = on_metin or f"Sayin yetkili,\n\nEkte {p.id} numarali proforma belgelerini bulabilirsiniz.\n\nSaygilarimizla."

        basari, mesaj = _eposta_gonder(alici, konu, govde, ekler)
        if not basari:
            return jsonify({'ok': False, 'mesaj': mesaj}), 500

        # Gönderim izini denetime yaz
        _log_audit('EPOSTA', 'proforma', proforma_id,
                   yeni={'alici': alici, 'belgeler': ','.join(belgeler)})
        # Proforma Onaylandı durumundaysa ve henüz gönderilmediyse Gönderildi'ye çek
        ekstra = ''
        if p.durum == 'Onaylandi':
            p.durum = 'Gonderildi'
            db.session.commit()
            ekstra = ' Proforma "Gonderildi" durumuna gecti.'

        return jsonify({'ok': True,
            'mesaj': f'{len(ekler)} belge {alici} adresine gonderildi.{ekstra}',
            'alici': alici, 'belge_sayisi': len(ekler)})

    @app.route('/api/proforma/<proforma_id>/eposta_bilgi')
    def api_proforma_eposta_bilgi(proforma_id):
        """Gönder penceresi için: müşterinin e-postası + SMTP hazır mı + CI gönderilebilir mi."""
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        p = Proforma.query.get(proforma_id)
        if not p:
            return jsonify({'ok': False, 'mesaj': 'Proforma bulunamadi'}), 404
        cari = _cari_bul(p.musteri)
        return jsonify({
            'ok': True,
            'musteri': p.musteri,
            'alici_email': (cari.email if cari else '') or '',
            'smtp_hazir': _smtp_ayarlari_oku() is not None,
            'ci_gonderilebilir': p.durum in ('Onaylandi', 'Gonderildi', 'Faturalandi'),
            'proforma_id': p.id,
        })

    @app.route('/api/proforma/<proforma_id>/html')
    def api_proforma_html(proforma_id):
        if _auth_required(): return "Unauthorized", 401
        mod = request.args.get('mod', 'pi')  # pi | pl | ci

        p, kalemler, _ = _proforma_cikti_data(proforma_id)
        if not p:
            return "Proforma bulunamadi", 404

        # ONAY KONTROLÜ: Commercial Invoice (ticari fatura) yalnızca ONAYLANMIŞ
        # proformadan üretilebilir. Onaysız proforma için CI oluşturmak çelişkidir
        # (henüz onaylanmamış bir belge ticari faturaya dönüşemez). PI/PL serbesttir.
        if mod == 'ci' and p.durum not in ('Onaylandi', 'Gonderildi', 'Faturalandi'):
            return (f"<div style='font-family:sans-serif;padding:40px;text-align:center;color:#b91c1c'>"
                    f"<h2>Commercial Invoice oluşturulamaz</h2>"
                    f"<p>Bu proforma henüz onaylanmadı (durum: <b>{p.durum}</b>).</p>"
                    f"<p>Ticari fatura yalnızca iç onaydan geçmiş proformalar için oluşturulabilir.</p>"
                    f"<p style='color:#666;font-size:.9rem'>Proforma Invoice (PI) ve Packing List (PL) her durumda alınabilir.</p>"
                    f"</div>"), 403

        # Toplamlar
        toplam_adet = sum((k.adet or 0) for k in kalemler)
        toplam_agirlik = sum((k.agirlik or 0) for k in kalemler)
        toplam_yazili = _sayi_yaziya_en(p.toplam or 0, p.doviz or 'USD')

        # Şablon seçimi
        if mod == 'pl':
            sablon = 'packing_list_print.html'
            # Packing list: kalemleri bundle'lara böl (limit + karma ayarina gore)
            cikti_kalemler = _packing_list_bundle_bol(p, kalemler)
        elif mod == 'ci':
            sablon = 'commercial_invoice_print.html'
            cikti_kalemler = kalemler
        else:
            sablon = 'proforma_print.html'
            cikti_kalemler = kalemler

        return render_template(sablon, p=p, kalemler=cikti_kalemler,
                               toplam_adet=toplam_adet, toplam_agirlik=toplam_agirlik,
                               toplam_yazili=toplam_yazili)


    @app.route('/api/siparis/<siparis_id>/finansal_ozet')
    def api_siparis_finansal_ozet(siparis_id):
        """
        Bir siparisin finansal ozeti: siparis tutari + bagli proforma(lar)
        + bagli fatura(lar) + cari hareket durumu.
        Cari modulunde hareket girilirken bilgi amacli kullanilir.
        """
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        sip = Siparis.query.get(siparis_id)
        if not sip:
            return jsonify({'ok': False, 'mesaj': 'Siparis bulunamadi'}), 404

        # FAZ 16: tutar artik kalemlerden gelir (toplam_tutar alani da tutulur)
        _kalemler = SiparisKalem.query.filter_by(siparis_id=siparis_id).order_by(SiparisKalem.sira).all()
        siparis_tutari = (sip.toplam_tutar
                          or sum((k.toplam_fiyat or 0) for k in _kalemler)
                          or 0)
        _ilk_kalem = _kalemler[0] if _kalemler else None

        # Bagli proformalar
        proformalar = Proforma.query.filter_by(siparis_id=siparis_id).all()
        proforma_list = [{
            'id': p.id, 'toplam': p.toplam or 0, 'doviz': p.doviz or 'USD',
            'durum': p.durum
        } for p in proformalar]

        # Bagli faturalar
        faturalar = Fatura.query.filter_by(siparis_id=siparis_id).all()
        fatura_list = [{
            'id': f.id, 'fatura_no': f.fatura_no, 'toplam': f.toplam or 0,
            'doviz': f.doviz or 'USD', 'durum': f.durum
        } for f in faturalar]

        # Bu siparise bagli cari hareketler (varsa)
        hareketler = CariHareket.query.filter_by(siparis_id=siparis_id).all()
        hareket_borc = sum(h.borc or 0 for h in hareketler)
        hareket_alacak = sum(h.alacak or 0 for h in hareketler)

        # SatisKaydi (karlilik) - siparise bagli
        satis_kayitlari = SatisKaydi.query.filter_by(siparis_id=siparis_id).all()
        toplam_satis_usd = sum(s.tutar_usd or s.tutar or 0 for s in satis_kayitlari)
        toplam_maliyet_usd = sum(s.maliyet_usd or 0 for s in satis_kayitlari)
        toplam_kar_usd = sum(s.kar_usd or 0 for s in satis_kayitlari)
        ort_marj = (toplam_kar_usd / toplam_satis_usd * 100) if toplam_satis_usd else 0

        # Fatura bazli tahsilat durumu
        fatura_ids = [f.id for f in faturalar]
        tahsil_toplam = 0
        if fatura_ids:
            tahsil_toplam = db.session.query(db.func.sum(CariHareket.alacak)).filter(
                CariHareket.baglanti_tip == 'fatura',
                CariHareket.baglanti_id.in_(fatura_ids),
                CariHareket.kaynak == 'tahsilat').scalar() or 0

        # TUTAR UYUMSUZLUGU tespiti
        uyarilar = []
        # Proforma tutari siparisten farkli mi?
        for p in proformalar:
            if p.toplam and abs((p.toplam or 0) - siparis_tutari) > 1:
                uyarilar.append(
                    f'Proforma {p.id} tutari ({p.toplam:,.2f}) siparis tutarindan '
                    f'({siparis_tutari:,.2f}) farkli')
        # Fatura tutari proformadan farkli mi?
        for f2 in faturalar:
            for p in proformalar:
                if f2.proforma_id == p.id and f2.toplam and p.toplam:
                    if abs((f2.toplam or 0) - (p.toplam or 0)) > 1:
                        uyarilar.append(
                            f'Fatura {f2.id} tutari ({f2.toplam:,.2f}) proforma '
                            f'{p.id} tutarindan ({p.toplam:,.2f}) farkli')

        return jsonify({
            'ok': True,
            'siparis': {
                'id': sip.id, 'musteri': sip.musteri,
                'tutar': q3(siparis_tutari),
                'doviz': sip.doviz or 'USD',
                'durum': sip.durum,
                'cins': (_ilk_kalem.cins if _ilk_kalem else None) or '-',
                'miktar': sum((k.miktar or 0) for k in _kalemler),
                'birim': (_ilk_kalem.birim if _ilk_kalem else None) or '',
                'kalem_sayisi': len(_kalemler),
                'tarih': sip.siparis_tarihi.isoformat() if sip.siparis_tarihi else None
            },
            'proformalar': proforma_list,
            'faturalar': fatura_list,
            'cari_hareket': {
                'borc': q3(hareket_borc),
                'alacak': q3(hareket_alacak),
                'sayi': len(hareketler)
            },
            'karlilik': {
                'satis_usd': q3(toplam_satis_usd),
                'maliyet_usd': q3(toplam_maliyet_usd),
                'kar_usd': q3(toplam_kar_usd),
                'marj': q_oran(ort_marj),
                'satis_kaydi_sayisi': len(satis_kayitlari)
            },
            'tahsilat': {
                'tahsil_edilen': q3(tahsil_toplam),
                'fatura_toplam': q3(sum(f.toplam or 0 for f in faturalar))
            },
            'rezervasyon': {
                'sayi': Rezervasyon.query.filter_by(siparis_id=siparis_id, iptal_nedeni=None).count(),
                'urun_tip': (_ilk_kalem.urun_tip if _ilk_kalem else None) or '-'
            },
            'uyarilar': uyarilar
        })

    @app.route('/api/diag/siparis/<siparis_id>/rezervasyonlari_temizle', methods=['POST', 'GET'])
    def api_diag_rez_temizle(siparis_id):
        """Diagnostic: bir siparisin TUM rezervasyonlarini sil + stoklari Serbest yap."""
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        rezler = Rezervasyon.query.filter_by(siparis_id=siparis_id).all()
        silinen = 0
        for r in rezler:
            # Stogu Serbest yap
            if r.stok_id:
                if r.stok_tip == 'BLOK':
                    s = BlokStok.query.get(r.stok_id)
                elif r.stok_tip == 'PLAKA':
                    s = PlakaStok.query.get(r.stok_id)
                else:
                    s = EbatliStok.query.get(r.stok_id)
                if s and s.durum == 'Rezerve':
                    s.durum = 'Serbest'
            db.session.delete(r)
            silinen += 1
        db.session.commit()
        return jsonify({'ok': True, 'silinen': silinen, 'siparis_id': siparis_id,
                       'mesaj': f'{silinen} rezervasyon silindi, stoklar Serbest yapildi'})

    @app.route('/api/diag/siparis/<siparis_id>/rezervasyonlar')
    def api_diag_rezervasyonlar(siparis_id):
        """Diagnostic: bir siparisin tum rezervasyonlarini listele."""
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        rezler = Rezervasyon.query.filter_by(siparis_id=siparis_id).all()
        return jsonify({
            'siparis_id': siparis_id,
            'toplam_rez': len(rezler),
            'aktif_rez': len([r for r in rezler if r.iptal_nedeni is None]),
            'iptal_rez': len([r for r in rezler if r.iptal_nedeni]),
            'rezler': [{
                'id': r.id, 'stok_id': r.stok_id, 'stok_tip': r.stok_tip,
                'iptal_nedeni': r.iptal_nedeni, 'olusturma': r.olusturma.isoformat() if r.olusturma else None
            } for r in rezler]
        })

    @app.route('/api/siparis/<siparis_id>/kalemler')
    def api_siparis_kalemler(siparis_id):
        """
        Bir siparişin bağlı stoklarını PROFORMA KALEMİ formatında döndürür.
        Proforma formunda 'siparişi seç → kalemlere yaz' için kullanılır.
        """
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        sip = Siparis.query.get(siparis_id)
        if not sip:
            return jsonify({'ok': False, 'mesaj': 'Siparis bulunamadi'}), 404

        # FAZ 16: birim/fiyat artik SiparisKalem'de. Ilk kalem varsayilan referanstir.
        _ik = (SiparisKalem.query.filter_by(siparis_id=siparis_id)
               .order_by(SiparisKalem.sira).first())
        _sip_birim_vars = (_ik.birim if _ik else None)
        _sip_fiyat_vars = (_ik.birim_fiyat if _ik else None) or 0

        # Siparişe bağlı iptal edilmemiş rezervasyonlar
        rez_q = Rezervasyon.query.filter_by(siparis_id=siparis_id, iptal_nedeni=None).all()
        # Aynı stok_id birden fazla rezerve edilmişse (eski bug) tekleştir
        gorulen_stoklar = set()
        rezler = []
        for r in rez_q:
            if r.stok_id and r.stok_id in gorulen_stoklar:
                continue  # Bu stok zaten kalem listesinde
            if r.stok_id:
                gorulen_stoklar.add(r.stok_id)
            rezler.append(r)

        # PLAKA stoklarini blok+cins+yuzey+olcu bazinda GRUPLA.
        # Boylece 50 plaka -> tek kalem (adet=50), bundle bolme dogru calisir.
        # BLOK/EBATLI ayri kalemler olarak kalir.
        gruplar = {}     # anahtar -> kalem dict
        grup_sira = []   # ekleme sirasi
        for r in rezler:
            stok = _stok_getir(r.stok_id, r.stok_tip)
            if not stok:
                continue

            if r.stok_tip == 'BLOK':
                m2 = 0
                boy = getattr(stok, 'boy', 0) or 0
                yuk = getattr(stok, 'yukseklik', 0) or 0
                kal = getattr(stok, 'kalinlik', 0) or getattr(stok, 'en', 0) or 0
                # Sipariş'in birim'i ne ise o ölçü: ton -> tonaj, m3 -> hacim
                _sip_birim = (_sip_birim_vars or 'ton').lower()
                if _sip_birim == 'ton':
                    miktar = getattr(stok, 'tonaj', 0) or 0
                elif _sip_birim == 'm3':
                    miktar = getattr(stok, 'hacim_m3', 0) or 0
                else:
                    # Bilinmeyen birim: tonaj fallback
                    miktar = getattr(stok, 'tonaj', 0) or 0
            else:
                m2 = getattr(stok, 'metraj_m2', 0) or 0
                boy = getattr(stok, 'boy', 0) or 0
                yuk = getattr(stok, 'yukseklik', 0) or 0
                kal = getattr(stok, 'kalinlik', 0) or 0
                miktar = m2

            cins = getattr(stok, 'cins', '') or ''
            ozellik = getattr(stok, 'ozellik', '') or ''
            # EBATLI'da blok_no yok, kasa_no (referans kodu) var
            blok_no = getattr(stok, 'blok_no', None) or getattr(stok, 'kasa_no', None) or ''
            slab_no = getattr(stok, 'slab_no', '') or ''

            if r.stok_tip == 'PLAKA':
                # Gruplama anahtari: ayni blok/cins/yuzey/olcu -> tek kalem
                anahtar = (blok_no, cins, ozellik, boy, yuk, kal)
                if anahtar not in gruplar:
                    gruplar[anahtar] = {
                        'stok_id': stok.id,        # ilk stok (referans)
                        'stok_idler': [],          # gruptaki tum stoklar
                        'urun_tip': 'PLAKA',
                        'cins': cins, 'ozellik': ozellik, 'blok_no': blok_no,
                        'slab_no': slab_no,        # ilk plakanin slab no'su -> baslangic
                        'boy': boy, 'yukseklik': yuk, 'kalinlik': kal,
                        'm2': 0, 'sqft': 0, 'miktar': 0, 'adet': 0,
                        'birim': _sip_birim_vars or 'm2',
                        'birim_fiyat': _sip_fiyat_vars,
                        'doviz': sip.doviz or 'USD'
                    }
                    grup_sira.append(anahtar)
                g = gruplar[anahtar]
                g['stok_idler'].append(stok.id)
                g['adet'] += 1
                g['m2'] = q3(g['m2'] + m2)
                g['miktar'] = q3(g['miktar'] + miktar)
                g['sqft'] = q3(g['sqft'] + (m2 * M2_TO_SQFT if m2 else 0))
                # Slab baslangici: gruptaki en kucuk slab no
                try:
                    if slab_no and str(slab_no).isdigit():
                        mevcut = g.get('slab_no')
                        if not mevcut or not str(mevcut).isdigit() or int(slab_no) < int(mevcut):
                            g['slab_no'] = slab_no
                except Exception:
                    pass
            else:
                # BLOK / EBATLI: tek tek kalem
                sqft = q3(m2 * M2_TO_SQFT) if m2 else 0
                # EBATLI'da kasa_ici_adet onemli (proformada hesaplama için)
                kasa_ici = getattr(stok, 'kasa_ici_adet', 1) or 1
                anahtar = ('_TEKIL_', stok.id)
                gruplar[anahtar] = {
                    'stok_id': stok.id,
                    'stok_idler': [stok.id],
                    'urun_tip': r.stok_tip,
                    'cins': cins, 'ozellik': ozellik, 'blok_no': blok_no,
                    'slab_no': slab_no,
                    'boy': boy, 'yukseklik': yuk, 'kalinlik': kal,
                    'm2': q3(m2), 'sqft': sqft,
                    'miktar': q3(miktar), 'adet': 1,
                    'kasa_ici_adet': kasa_ici,  # EBATLI icin kasa basina plaka adedi
                    'tonaj': q3(getattr(stok, 'tonaj', 0) or 0) if r.stok_tip == 'BLOK' else None,
                    'hacim_m3': q3(getattr(stok, 'hacim_m3', 0) or 0) if r.stok_tip == 'BLOK' else None,
                    'birim': _sip_birim_vars or ('ton' if r.stok_tip == 'BLOK' else 'm2'),
                    'birim_fiyat': _sip_fiyat_vars,
                    'doviz': sip.doviz or 'USD'
                }
                grup_sira.append(anahtar)

        kalemler = [gruplar[a] for a in grup_sira]

        # FALLBACK: Siparişe bağlı stok rezervasyonu yoksa (henüz stok ayrılmamış
        # taze sipariş) proforma kalemlerini SiparisKalem satırlarından üret.
        # "Sipariş oluşturdum ama proforma yapamıyorum" akışının çözümü budur.
        if not kalemler:
            sk_list = (SiparisKalem.query.filter_by(siparis_id=siparis_id)
                       .order_by(SiparisKalem.sira).all())
            for sk in sk_list:
                u_tip = (sk.urun_tip or 'PLAKA').upper()
                m2 = q3(sk.m2_toplam or 0)
                miktar = q3(sk.miktar or 0)
                kalemler.append({
                    'stok_id': None,
                    'stok_idler': [],
                    'urun_tip': u_tip,
                    'cins': sk.cins or '', 'ozellik': sk.ozellik or '',
                    'blok_no': '', 'slab_no': '',
                    'boy': sk.boy or 0, 'yukseklik': sk.yukseklik or 0,
                    'kalinlik': sk.kalinlik or (sk.en or 0),
                    'm2': m2, 'sqft': q3(m2 * M2_TO_SQFT) if m2 else 0,
                    'miktar': miktar,
                    'adet': sk.adet or 1,
                    'kasa_ici_adet': getattr(sk, 'kasa_ici_adet', 1) or 1,
                    'tonaj': miktar if (u_tip == 'BLOK' and (sk.birim or '').lower() == 'ton') else None,
                    'hacim_m3': miktar if (u_tip == 'BLOK' and (sk.birim or '').lower() == 'm3') else None,
                    'birim': sk.birim or ('ton' if u_tip == 'BLOK' else 'm2'),
                    'birim_fiyat': sk.birim_fiyat or 0,
                    'doviz': sk.doviz or sip.doviz or 'USD'
                })

        # MUSTERI BILGILERI: Cari'den adres, ulke vs al
        m_adres = ''
        m_ulke = ''
        try:
            mus = Cari.query.filter_by(unvan=sip.musteri).first()
            if mus:
                m_adres = getattr(mus, 'adres', '') or ''
                m_ulke = getattr(mus, 'ulke', '') or ''
        except Exception:
            pass

        # SATIS TIPI -> Proforma 'tur' alanina cevir
        # ihracat -> 'ihracat', yurtici_kdvli/yurtici_kdvsiz -> 'yurt_ici', ihrac_kayitli -> 'ihrac_kayitli'
        satis_tipi = getattr(sip, 'satis_tipi', 'ihracat') or 'ihracat'
        if satis_tipi in ('yurtici_kdvli', 'yurtici_kdvsiz'):
            proforma_tur = 'yurt_ici'
        elif satis_tipi == 'ihrac_kayitli':
            proforma_tur = 'ihrac_kayitli'
        else:
            proforma_tur = 'ihracat'

        return jsonify({
            'ok': True,
            'siparis_id': siparis_id,
            'musteri': sip.musteri,
            'musteri_adres': m_adres,
            'musteri_ulke': m_ulke,
            'doviz': sip.doviz or 'USD',
            'satis_tipi': satis_tipi,
            'proforma_tur': proforma_tur,
            'kdv_oran': float(sip.kdv_oran or 0),
            'odeme_sekli': sip.odeme_sekli or '',
            'teslim_sekli': sip.teslim_sekli or '',
            'termin': sip.termin.isoformat() if sip.termin else None,
            'kalemler': kalemler,
            'adet': len(kalemler)
        })


    # ════════════════════════════════════════════════════════
    # FATURA - Proforma'dan dönüştürme + yönetim
    # ════════════════════════════════════════════════════════
    @app.route('/api/proforma/<proforma_id>/faturaya_donustur', methods=['POST'])
    def api_proforma_faturaya(proforma_id):
        """Proformayı yeni bir Fatura kaydına kopyalar."""
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        p = Proforma.query.get(proforma_id)
        if not p:
            return jsonify({'ok': False, 'mesaj': 'Proforma bulunamadi'}), 404

        # Arşiv (eski revizyon) sürümünden fatura kesilemez — güncel sürüm kullanılmalı
        if p.durum == 'Revize' or p.aktif_surum is False:
            aktif = Proforma.query.filter_by(ana_pi_id=p.ana_pi_id or p.id, aktif_surum=True).first()
            return jsonify({'ok': False,
                'mesaj': f'Bu eski bir revizyon surumu. Guncel surumden fatura kesin: '
                         f'{aktif.id if aktif else "?"}'}), 400

        # ONAY KONTROLÜ: sadece onaylanmış (ve sonrası) proformadan fatura kesilebilir.
        # Onaylanmamış proformanın cariye borç yazması engellenir.
        if p.durum not in ('Onaylandi', 'Gonderildi', 'Faturalandi'):
            return jsonify({'ok': False, 'error': 'onaysiz',
                'mesaj': f'Bu proforma henuz onaylanmadi (durum: {p.durum}). '
                         f'Faturaya donusturmek icin once ic onaydan gecmeli.'}), 400

        # SİPARİŞ KONTROLÜ: fatura kesiliyorsa iş kesinleşmiş demektir — sipariş şart.
        # (Teklif proforması onaylanabilir ama faturaya dönüşürken bir siparişe bağlı olmalı.)
        if not p.siparis_id:
            return jsonify({'ok': False, 'error': 'siparis_yok',
                'mesaj': 'Bu proforma bir siparise bagli degil. Faturaya donusturmek icin '
                         'once proforma duzenleme formundan bir siparis secin.'}), 400
        mevcut = Fatura.query.filter_by(proforma_id=proforma_id).filter(
            Fatura.durum != 'Iptal').first()
        if mevcut:
            return jsonify({'ok': False,
                'mesaj': f'Bu proformadan zaten fatura olusturulmus: {mevcut.id}'}), 400

        data = request.json or {}

        # Kalemleri JSON snapshot olarak sakla
        kalemler = ProformaKalem.query.filter_by(proforma_id=proforma_id).order_by(
            ProformaKalem.sira, ProformaKalem.id).all()
        kalem_list = []
        for k in kalemler:
            kalem_list.append({
                'urun_tip': k.urun_tip, 'cins': k.cins, 'ozellik': k.yuzey_spec,
                'blok_no': k.blok_no, 'slab_no': k.slab_no, 'bundle_no': k.bundle_no,
                'boy': k.boy, 'yukseklik': k.yukseklik, 'kalinlik': k.kalinlik,
                'adet': k.adet, 'miktar': k.miktar, 'birim': k.birim,
                'birim_fiyat': k.birim_fiyat, 'toplam_fiyat': k.toplam_fiyat,
                'net_fiyat': k.net_fiyat, 'doviz': k.doviz, 'agirlik': k.agirlik
            })

        # KDV hesabı - SADECE yurt_ici turunde KDV uygula
        toplam = p.toplam or 0
        # Ihracat ve ihrac_kayitli'da KDV YOK (kdv_oran 0'a zorlanir)
        if (p.tur or 'ihracat') == 'yurt_ici':
            kdv_oran = p.kdv_oran or 0
        else:
            kdv_oran = 0  # Ihracat / ihrac_kayitli
        if kdv_oran > 0:
            ara_toplam = toplam / (1 + kdv_oran / 100)
            kdv_tutar = toplam - ara_toplam
        else:
            ara_toplam = toplam
            kdv_tutar = 0

        # FATURA TİPİ tespiti:
        # 1) Proforma_tipi == 'teklif' -> fatura da 'teklif' (proforma seviyesinde belirlenmis)
        # 2) Siparise bagli rezervasyon var -> 'stoklu'
        # 3) Siparise bagli ama rezervasyon yok -> 'transit'
        # 4) Siparise hic bagli degil -> 'teklif' (fallback)
        # Kullanici manuel olarak data.get('fatura_tipi') ile override edebilir
        manuel_tip = (data.get('fatura_tipi') or '').strip()
        proforma_tipi_alan = getattr(p, 'proforma_tipi', None) or ('satis' if p.siparis_id else 'teklif')
        if manuel_tip in ('stoklu', 'transit', 'teklif'):
            fatura_tipi = manuel_tip
        elif proforma_tipi_alan == 'teklif':
            fatura_tipi = 'teklif'
        elif not p.siparis_id:
            fatura_tipi = 'teklif'
        else:
            rez_var = Rezervasyon.query.filter_by(siparis_id=p.siparis_id).first()
            fatura_tipi = 'stoklu' if rez_var else 'transit'

        fatura = Fatura(
            id=_yeni_id('FTR'),
            fatura_no=data.get('fatura_no') or '',
            fatura_tarihi=_parse_date(data.get('fatura_tarihi')) or date.today(),
            vade_tarihi=_parse_date(data.get('vade_tarihi')),
            proforma_id=p.id,
            siparis_id=p.siparis_id,
            musteri=p.musteri,
            musteri_adres=p.musteri_adres,
            musteri_ulke=p.musteri_ulke,
            toplam=toplam,
            ara_toplam=q3(ara_toplam),
            kdv_oran=kdv_oran,
            kdv_tutar=q3(kdv_tutar),
            doviz=p.doviz or 'USD',
            odeme_sekli=p.odeme_sekli,
            teslim_sekli=p.teslim_sekli,
            durum='Taslak',
            fatura_tipi=fatura_tipi,
            kur_farki_modu=('cari' if (p.doviz or 'USD') == 'TRY' else 'gider'),
            satis_tipi=(p.tur or 'ihracat'),
            kalemler_json=json.dumps(kalem_list, ensure_ascii=False, default=str),
            kullanici=session.get('kullanici')
        )
        db.session.add(fatura)
        _log_audit('EKLE', 'fatura', fatura.id,
                   yeni={'proforma': proforma_id, 'musteri': p.musteri,
                         'tutar': toplam, 'tip': fatura_tipi})
        ok, hata = _safe_commit(f'Proforma->Fatura: {proforma_id}->{fatura.id}')
        if not ok:
            return jsonify({'ok': False, 'mesaj': f'Hata: {hata}. Tum degisiklikler geri alindi.'}), 500
        tip_not = ' (Transit/dis alim - alis maliyeti girilmeli)' if fatura_tipi == 'transit' else ''
        return jsonify({'ok': True, 'id': fatura.id,
                        'fatura_tipi': fatura_tipi,
                        'mesaj': f'Fatura olusturuldu: {fatura.id}.{tip_not}'})

    @app.route('/api/fatura', methods=['GET'])
    def api_fatura_liste():
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        musteri = request.args.get('musteri')
        durum = request.args.get('durum')
        yon = request.args.get('yon')  # satis | alis | (boş = hepsi)
        q = Fatura.query
        if musteri: q = q.filter_by(musteri=musteri)
        if durum: q = q.filter_by(durum=durum)
        if yon: q = q.filter_by(yon=yon)
        faturalar = q.order_by(Fatura.olusturma.desc()).all()
        return jsonify({'data': [{
            'id': f.id, 'fatura_no': f.fatura_no,
            'fatura_tarihi': f.fatura_tarihi.isoformat() if f.fatura_tarihi else None,
            'vade_tarihi': f.vade_tarihi.isoformat() if f.vade_tarihi else None,
            'proforma_id': f.proforma_id, 'siparis_id': f.siparis_id,
            'musteri': f.musteri, 'toplam': f.toplam, 'doviz': f.doviz,
            'durum': f.durum,
            'yon': getattr(f, 'yon', 'satis'),
            'fatura_tipi': getattr(f, 'fatura_tipi', 'stoklu'),
            'satis_tipi': getattr(f, 'satis_tipi', 'ihracat'),
            'kdv_oran': getattr(f, 'kdv_oran', 0) or 0,
            'kdv_tutar': getattr(f, 'kdv_tutar', 0) or 0,
            'tevkifat_oran': getattr(f, 'tevkifat_oran', '') or '',
            'tevkifat_tutar': getattr(f, 'tevkifat_tutar', 0) or 0,
            'odeme_sekli': getattr(f, 'odeme_sekli', '') or '',
            'kur_farki_modu': getattr(f, 'kur_farki_modu', 'gider') or 'gider',
            'teslim_sekli': getattr(f, 'teslim_sekli', '') or '',
            'musteri_adres': getattr(f, 'musteri_adres', '') or '',
            'musteri_ulke': getattr(f, 'musteri_ulke', '') or '',
            'aciklama': getattr(f, 'aciklama', '') or '',
            'alis_maliyeti': getattr(f, 'alis_maliyeti', 0),
            'cari_hareket_id': getattr(f, 'cari_hareket_id', None)
        } for f in faturalar]})

    @app.route('/api/fatura/<fatura_id>', methods=['GET'])
    def api_fatura_detay(fatura_id):
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        f = Fatura.query.get(fatura_id)
        if not f:
            return jsonify({'ok': False, 'mesaj': 'Fatura bulunamadi'}), 404
        kalemler = []
        try:
            if f.kalemler_json:
                kalemler = json.loads(f.kalemler_json)
        except Exception:
            pass
        # Maliyet kalemleri (transit fatura)
        maliyet_kalemleri = []
        try:
            if getattr(f, 'maliyet_kalemleri_json', None):
                maliyet_kalemleri = json.loads(f.maliyet_kalemleri_json)
        except Exception:
            pass
        return jsonify({'ok': True, 'fatura': {
            'id': f.id, 'fatura_no': f.fatura_no,
            'fatura_tarihi': f.fatura_tarihi.isoformat() if f.fatura_tarihi else None,
            'vade_tarihi': f.vade_tarihi.isoformat() if f.vade_tarihi else None,
            'proforma_id': f.proforma_id, 'siparis_id': f.siparis_id,
            'musteri': f.musteri, 'musteri_adres': f.musteri_adres, 'musteri_ulke': f.musteri_ulke,
            'toplam': f.toplam, 'ara_toplam': f.ara_toplam,
            'kdv_oran': f.kdv_oran, 'kdv_tutar': f.kdv_tutar, 'doviz': f.doviz,
            'odeme_sekli': f.odeme_sekli, 'teslim_sekli': f.teslim_sekli,
            'durum': f.durum, 'aciklama': f.aciklama, 'kalemler': kalemler,
            'fatura_tipi': getattr(f, 'fatura_tipi', 'stoklu'),
            'alis_maliyeti': getattr(f, 'alis_maliyeti', 0),
            'maliyet_doviz': getattr(f, 'maliyet_doviz', 'USD'), 'kur_farki_modu': getattr(f, 'kur_farki_modu', 'gider'),
            'maliyet_kalemleri': maliyet_kalemleri
        }})

    def _fatura_no_mukerrer_mi(fatura_no, haric_fatura_id=None):
        """Fatura no zaten kullaniliyor mu? haric: bu ID'yi sayma (kendisi)."""
        if not fatura_no or not str(fatura_no).strip():
            return False
        q = Fatura.query.filter(Fatura.fatura_no == str(fatura_no).strip())
        if haric_fatura_id:
            q = q.filter(Fatura.id != haric_fatura_id)
        return q.first() is not None

    @app.route('/api/fatura/<fatura_id>', methods=['PUT'])
    def api_fatura_guncelle(fatura_id):
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        f = Fatura.query.get(fatura_id)
        if not f:
            return jsonify({'ok': False, 'mesaj': 'Fatura bulunamadi'}), 404
        data = request.json or {}
        # Mukerrer fatura no kontrolu
        eski_fn = (f.fatura_no or '').strip()
        yeni_fn = (data.get('fatura_no') or '').strip()
        fatura_no_degisti = bool(yeni_fn and yeni_fn != eski_fn)
        if fatura_no_degisti:
            if _fatura_no_mukerrer_mi(yeni_fn, haric_fatura_id=fatura_id):
                return jsonify({'ok': False, 'mesaj': f'Fatura no "{yeni_fn}" zaten kullanimda. Lutfen farkli bir numara girin.'}), 400

        for alan in ['fatura_no', 'odeme_sekli', 'teslim_sekli', 'aciklama', 'maliyet_doviz', 'musteri_adres', 'musteri_ulke', 'kur_farki_modu', 'fatura_tipi']:
            if alan in data:
                setattr(f, alan, data[alan])
        for tarih_alan in ['fatura_tarihi', 'vade_tarihi']:
            if tarih_alan in data:
                setattr(f, tarih_alan, _parse_date(data[tarih_alan]))
        # Transit fatura maliyet kalemleri
        if 'maliyet_kalemleri' in data:
            mk = data.get('maliyet_kalemleri') or []
            f.maliyet_kalemleri_json = json.dumps(mk, ensure_ascii=False)
            # Toplam alis maliyetini hesapla
            toplam_mal = sum((float(k.get('miktar') or 0) * float(k.get('birim_fiyat') or 0))
                             for k in mk)
            f.alis_maliyeti = q3(toplam_mal)

        # FATURA NO DEĞİŞTİYSE: sistem genelinde ilişkili kayıtları güncelle
        if fatura_no_degisti:
            # Cari hareketlerde bu faturaya bağlı evrak_no'ları güncelle
            for ch in CariHareket.query.filter_by(baglanti_tip='fatura', baglanti_id=fatura_id).all():
                if (ch.evrak_no or '') == eski_fn or not ch.evrak_no:
                    ch.evrak_no = yeni_fn
                if ch.aciklama and eski_fn and eski_fn in ch.aciklama:
                    ch.aciklama = ch.aciklama.replace(eski_fn, yeni_fn)
            # Bu faturaya bağlı çeklerin açıklama/evrak referanslarını güncelle
            for cek in Cek.query.filter_by(fatura_id=fatura_id).all():
                if cek.aciklama and eski_fn and eski_fn in cek.aciklama:
                    cek.aciklama = cek.aciklama.replace(eski_fn, yeni_fn)
                # Çeke bağlı cari hareketin evrak/açıklamasını da güncelle
                for ch in CariHareket.query.filter_by(baglanti_tip='cek', baglanti_id=cek.id).all():
                    if ch.aciklama and eski_fn and eski_fn in ch.aciklama:
                        ch.aciklama = ch.aciklama.replace(eski_fn, yeni_fn)
            app.logger.info(f'Fatura no degisti: {eski_fn} -> {yeni_fn}, iliskili kayitlar guncellendi')

        try:
            db.session.commit()
            return jsonify({'ok': True, 'mesaj': 'Fatura guncellendi' + (f' (no: {yeni_fn}, ilişkili kayıtlar güncellendi)' if fatura_no_degisti else '')})
        except Exception as e:
            db.session.rollback()
            return jsonify({'ok': False, 'mesaj': str(e)}), 500

    @app.route('/api/fatura/<fatura_id>', methods=['DELETE'])
    def api_fatura_sil(fatura_id):
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        f = Fatura.query.get(fatura_id)
        if not f:
            return jsonify({'ok': False, 'mesaj': 'Fatura bulunamadi'}), 404
        if f.durum in ('Kismi Tahsil', 'Tahsil Edildi'):
            return jsonify({'ok': False,
                'mesaj': 'Tahsilati olan fatura silinemez. Once tahsilatlari iptal edin.'}), 400
        _log_audit('SIL', 'fatura', fatura_id,
                   eski={'musteri': f.musteri, 'tutar': f.toplam, 'durum': f.durum})

        # Bagli cari hareket varsa (cari'den olusturulmus fatura) onu da sil
        silinen_hareket = None
        if f.cari_hareket_id:
            h = CariHareket.query.get(f.cari_hareket_id)
            if h:
                silinen_hareket = h.id
                db.session.delete(h)

        # Bu faturaya bagli tum SatisKaydi'lari da sil (karlilik tutarliligi)
        silinen_sk = SatisKaydi.query.filter_by(fatura_id=fatura_id).all()
        sk_say = len(silinen_sk)
        for sk in silinen_sk:
            db.session.delete(sk)

        # Bu faturaya bagli, fatura kaynakli cari hareketleri de sil
        # (Kesildi durumunda olusturulan otomatik borc kaydi)
        bagli_hareketler = CariHareket.query.filter_by(
            baglanti_tip='fatura', baglanti_id=fatura_id).all()
        hareket_say = 0
        for h in bagli_hareketler:
            db.session.delete(h)
            hareket_say += 1

        db.session.delete(f)
        db.session.commit()

        msg = 'Fatura silindi'
        if silinen_hareket:
            msg += f', bagli cari hareket ({silinen_hareket}) da silindi'
        if hareket_say > 0:
            msg += f', {hareket_say} bagli cari hareket silindi'
        if sk_say > 0:
            msg += f', {sk_say} satis kaydi silindi'
        return jsonify({'ok': True, 'mesaj': msg})

    @app.route('/api/fatura/<fatura_id>/durum', methods=['POST'])
    def api_fatura_durum(fatura_id):
        """
        Fatura durumu değiştir. 'Kesildi' olunca cariye otomatik BORÇ işlenir.
        Çift kayıt koruması: aynı fatura için ikinci borç oluşmaz.
        """
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        f = Fatura.query.get(fatura_id)
        if not f:
            return jsonify({'ok': False, 'mesaj': 'Fatura bulunamadi'}), 404

        data = request.json or {}
        yeni_durum = data.get('durum')
        gecerli = ['Taslak', 'Kesildi', 'Kismi Tahsil', 'Tahsil Edildi', 'Iptal']
        if yeni_durum not in gecerli:
            return jsonify({'ok': False, 'mesaj': 'Gecersiz durum'}), 400

        eski_durum = f.durum
        ekstra = ''

        # ── KURAL: Tahsilat durumlarina gecis icin once Kesildi olmali ──
        tahsilat_durumlari = ['Kismi Tahsil', 'Tahsil Edildi']
        if yeni_durum in tahsilat_durumlari and eski_durum not in (['Kesildi'] + tahsilat_durumlari):
            return jsonify({'ok': False,
                'mesaj': 'Tahsilat durumuna gecmek icin fatura once "Kesildi" durumunda olmali.'}), 400

        # ── KESİLDİ: cariye otomatik borç + SatisKaydi (kar hesabi) ──
        if yeni_durum == 'Kesildi' and eski_durum != 'Kesildi':
            # Fatura no zorunlu
            if not f.fatura_no:
                return jsonify({'ok': False,
                    'mesaj': 'Fatura kesmek icin once Fatura No girin'}), 400

            # Transit fatura tespiti: faturanin KAYITLI tipini kullan.
            # (Rezervasyonlar teslim/sevk sonrasi kapanmis olabilir,
            #  o yuzden yeniden sorgulama guvenilir degil.)
            # FATURA TIPI YENIDEN KONTROL (yanlis kayit varsa duzelt)
            # Siparise bagli AKTIF rezervasyon varsa -> stoklu olmali (teklif/transit'ten zorla)
            if f.siparis_id:
                aktif_rez = Rezervasyon.query.filter_by(
                    siparis_id=f.siparis_id, iptal_nedeni=None).first()
                if aktif_rez and f.fatura_tipi in ('transit', 'teklif'):
                    # Yanlis kaydedilmis - duzelt
                    f.fatura_tipi = 'stoklu'
                    db.session.flush()
                    app.logger.info(f'Fatura tipi otomatik duzeltildi: {f.id} transit -> stoklu')

            transit_mi = (getattr(f, 'fatura_tipi', 'stoklu') == 'transit')
            teklif_mi = (getattr(f, 'fatura_tipi', 'stoklu') == 'teklif')

            # TRANSIT FATURA: alis maliyeti ZORUNLU
            # TEKLIF FATURA: alis maliyeti opsiyonel (uyari verir ama izin verir)
            if transit_mi and not (f.alis_maliyeti and f.alis_maliyeti > 0):
                return jsonify({'ok': False,
                    'mesaj': 'Bu transit/dis alim faturasi. Fatura kesmek icin once Alis Maliyeti girin.'}), 400

            # Bu fatura için zaten borç işlenmiş mi? (çift kayıt koruması)
            mevcut_borc = CariHareket.query.filter_by(
                baglanti_tip='fatura', baglanti_id=fatura_id).filter(
                CariHareket.borc > 0).first()
            if not mevcut_borc:
                try:
                    _cari_hareket_ekle(
                        cari_unvan=f.musteri,
                        islem_tip='Satis Faturasi',
                        borc=f.toplam or 0,
                        doviz=f.doviz or 'USD',
                        aciklama=f'Fatura {f.fatura_no} ({fatura_id})',
                        kaynak='fatura',
                        baglanti_tip='fatura',
                        baglanti_id=fatura_id,
                        vade_tarihi=f.vade_tarihi,
                        evrak_no=f.fatura_no
                    )
                    ekstra = f' Cariye {f.toplam:,.2f} {f.doviz} borc islendi.'
                except ValueError as e:
                    # Cari bulunamadi - fatura kesilemez (borc kritik)
                    db.session.rollback()
                    return jsonify({'ok': False, 'mesaj': str(e)}), 400

            # SatisKaydi olustur (kar hesabi - karlilik bu kayitlardan beslenir)
            sk_adet, sk_hata = _fatura_satis_kaydi_olustur(fatura_id)
            if sk_hata:
                ekstra += f' (Satis kaydi: {sk_hata})'
            elif sk_adet > 0:
                ekstra += f' {sk_adet} satis kaydi olusturuldu (karlilik guncellendi).'

            # ── DURUM SENKRONIZASYONU: Fatura Kesildi -> bagli siparis "Hazir" ──
            if f.siparis_id:
                sip = Siparis.query.get(f.siparis_id)
                # Iptal/Teslim/Hazir disinda her durumdan Hazir'a gec
                if sip and sip.durum not in ('Hazir', 'Teslim Edildi', 'Iptal Edildi'):
                    eski_sip_durum = sip.durum
                    sip.durum = 'Hazir'
                    ekstra += f' Bagli siparis {sip.id} durumu: {eski_sip_durum} -> Hazir.'
                    app.logger.info(f'[SENKRO] Fatura {fatura_id} Kesildi -> Siparis {sip.id}: {eski_sip_durum} -> Hazir')

            # ── DURUM SENKRONIZASYONU: Fatura Kesildi -> bagli proforma "Faturalandi" ──
            if f.proforma_id:
                pf = Proforma.query.get(f.proforma_id)
                # Iptal disinda her durumdan Faturalandi'ya gec (proforma kendi akisindan bagimsiz)
                if pf and pf.durum not in ('Faturalandi', 'Iptal'):
                    eski_pf_durum = pf.durum
                    pf.durum = 'Faturalandi'
                    ekstra += f' Bagli proforma {pf.id} durumu: {eski_pf_durum} -> Faturalandi.'
                    app.logger.info(f'[SENKRO] Fatura {fatura_id} Kesildi -> Proforma {pf.id}: {eski_pf_durum} -> Faturalandi')

        # ── İPTAL: borç hareketini + SatisKaydi'ni geri al ──
        # Kesildi VEYA tahsilat durumlarından (Kismi/Tam) iptal edilebilir.
        # Ama önce: tahsilatı olan fatura doğrudan iptal edilemez — kullanıcı
        # önce tahsilatları geri almalı (müşteri ödemesi cari hesapta asılı kalmasın).
        if yeni_durum == 'Iptal' and eski_durum in ('Kesildi', 'Kismi Tahsil', 'Tahsil Edildi'):
            # Faturaya bağlı aktif tahsilat (alacak hareketi) var mı?
            aktif_tahsilat = CariHareket.query.filter(
                CariHareket.baglanti_tip == 'fatura',
                CariHareket.baglanti_id == fatura_id,
                CariHareket.alacak > 0).first()
            # Çek yoluyla tahsilat var mı?
            cek_tahsilat = Cek.query.filter_by(fatura_id=fatura_id).filter(
                Cek.durum.notin_(['Iptal', 'Karsiliksiz'])).first()
            if aktif_tahsilat or cek_tahsilat:
                return jsonify({'ok': False, 'error': 'tahsilat_var',
                    'mesaj': 'Bu faturaya tahsilat yapilmis. Iptal etmeden once tahsilatlari '
                             '(ve varsa cekleri) geri almalisiniz. Boylece musteri odemesi '
                             'cari hesapta asili kalmaz.'}), 400

            borc_hareket = CariHareket.query.filter_by(
                baglanti_tip='fatura', baglanti_id=fatura_id, kaynak='fatura').first()
            if borc_hareket:
                db.session.delete(borc_hareket)
                ekstra = ' Borc hareketi geri alindi.'
            # Bu faturadan olusan SatisKaydi'lari sil (karlilikten dus)
            sat_kayitlar = SatisKaydi.query.filter_by(fatura_id=fatura_id).all()
            for sk in sat_kayitlar:
                # Stoklu satista stok durumunu geri al
                if sk.stok_tip in ('BLOK', 'PLAKA', 'EBATLI'):
                    stok = _stok_getir(sk.stok_id, sk.stok_tip)
                    if stok and stok.durum == 'Teslim Edildi':
                        stok.durum = 'Satildi'
                db.session.delete(sk)
            if sat_kayitlar:
                ekstra += f' {len(sat_kayitlar)} satis kaydi geri alindi.'

            # ── DURUM SENKRONIZASYONU: Fatura Iptal -> bagli siparis geri "Onaylandi" ──
            if f.siparis_id:
                sip = Siparis.query.get(f.siparis_id)
                if sip and sip.durum in ('Hazir', 'Teslim Edildi'):
                    eski_sip_durum = sip.durum
                    sip.durum = 'Onaylandi'
                    ekstra += f' Bagli siparis {sip.id} durumu: {eski_sip_durum} -> Onaylandi.'
                    app.logger.info(f'[SENKRO] Fatura {fatura_id} Iptal -> Siparis {sip.id}: {eski_sip_durum} -> Onaylandi')

            # ── DURUM SENKRONIZASYONU: Fatura Iptal -> bagli proforma geri "Onaylandi" ──
            if f.proforma_id:
                pf = Proforma.query.get(f.proforma_id)
                if pf and pf.durum == 'Faturalandi':
                    pf.durum = 'Onaylandi'
                    ekstra += f' Bagli proforma {pf.id} durumu: Faturalandi -> Onaylandi.'
                    app.logger.info(f'[SENKRO] Fatura {fatura_id} Iptal -> Proforma {pf.id}: Faturalandi -> Onaylandi')

        f.durum = yeni_durum
        _log_audit('DURUM', 'fatura', fatura_id,
                   eski={'durum': eski_durum}, yeni={'durum': yeni_durum})
        ok, hata = _safe_commit(f'Fatura durum: {eski_durum}->{yeni_durum}')
        if not ok:
            return jsonify({'ok': False, 'mesaj': f'Hata: {hata}. Tum degisiklikler geri alindi.'}), 500
        return jsonify({'ok': True,
                        'mesaj': f'Fatura durumu: {eski_durum} -> {yeni_durum}.{ekstra}'})

    # ════════════════════════════════════════════════════════
    # TAHSİLAT - Faturaya bağlı müşteri ödemesi
    # ════════════════════════════════════════════════════════
    def _hareket_fatura_esdegeri(h, f_doviz):
        """
        Bir ödeme hareketinin (alacak) FATURA DÖVİZİ cinsinden eşdeğerini döner.
        Aynı dövizde: ham tutar. Farklı dövizde (çapraz): TRY köprüsü —
        hareketin TRY karşılığı, hareket tarihindeki fatura-dövizi kuruna bölünür.
        """
        tutar = float(h.alacak or 0)
        if (h.doviz or 'USD') == f_doviz:
            return tutar
        k = 1.0 if f_doviz == 'TRY' else _kur_getir(f_doviz, h.hareket_tarihi)
        if not k or k <= 0:
            return 0.0
        tutar_try = float(h.alacak_try or 0)
        if tutar_try <= 0:
            tutar_try = tutar * float(h.kur_uygulanan or _kur_getir(h.doviz, h.hareket_tarihi) or 0)
        return tutar_try / float(k)

    def _fatura_odenen_esdeger(f):
        """
        Faturaya bağlı TÜM ödemeleri (tahsilat + çek, farklı dövizler dahil)
        fatura dövizi eşdeğeri olarak toplar. Çapraz döviz kapatma için temel.
        """
        f_doviz = f.doviz or 'USD'
        hs = CariHareket.query.filter(
            CariHareket.baglanti_tip == 'fatura', CariHareket.baglanti_id == f.id,
            CariHareket.alacak > 0).all()
        cek_hs = CariHareket.query.filter(
            CariHareket.baglanti_tip == 'cek', CariHareket.alacak > 0,
            CariHareket.baglanti_id.in_(
                db.session.query(Cek.id).filter_by(fatura_id=f.id))).all()
        gorulen, toplam = set(), 0.0
        for h in hs + cek_hs:
            if h.id in gorulen:
                continue
            gorulen.add(h.id)
            toplam += _hareket_fatura_esdegeri(h, f_doviz)
        return q3(toplam)

    def _fatura_tahsilat_durumu(fatura_id):
        """Faturanın tahsilat durumunu hesaplar ve durumunu günceller.
        Çapraz dövizli ödemeler fatura dövizi eşdeğeri üzerinden sayılır."""
        f = Fatura.query.get(fatura_id)
        if not f:
            return
        tahsil = _fatura_odenen_esdeger(f)
        toplam = f.toplam or 0
        # Durum: tam / kismi / kesildi
        if f.durum == 'Iptal':
            return
        if tahsil <= 0:
            if f.durum in ('Kismi Tahsil', 'Tahsil Edildi'):
                f.durum = 'Kesildi'
        elif tahsil + 0.01 < toplam:
            f.durum = 'Kismi Tahsil'
        else:
            f.durum = 'Tahsil Edildi'

    @app.route('/api/fatura/<fatura_id>/tahsilat', methods=['POST'])
    def api_fatura_tahsilat(fatura_id):
        """Faturaya tahsilat (müşteri ödemesi) ekler -> cariye alacak."""
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        f = Fatura.query.get(fatura_id)
        if not f:
            return jsonify({'ok': False, 'mesaj': 'Fatura bulunamadi'}), 404
        if f.durum == 'Iptal':
            return jsonify({'ok': False, 'mesaj': 'Iptal fatura tahsil edilemez'}), 400
        if f.durum == 'Taslak':
            return jsonify({'ok': False,
                'mesaj': 'Once faturayi "Kesildi" durumuna alin'}), 400

        data = request.json or {}
        try:
            tutar = float(data.get('tutar') or 0)
        except (ValueError, TypeError):
            tutar = 0
        if tutar <= 0:
            return jsonify({'ok': False, 'mesaj': 'Gecerli bir tutar girin'}), 400

        # ═══ ÇAPRAZ DÖVİZ TAHSİLAT ═══
        # Ödeme dövizi fatura dövizinden farklı olabilir (örn. USD faturaya TRY
        # tahsilat). Ödeme TRY köprüsüyle fatura dövizi eşdeğerine çevrilir;
        # aşım kontrolü ve kapatma bu eşdeğer üzerinden yapılır.
        f_doviz = f.doviz or 'USD'
        odeme_doviz = ((data.get('doviz') or f_doviz).strip() or f_doviz).upper()
        try:
            manuel_kur = float(data.get('kur') or 0)
        except (ValueError, TypeError):
            manuel_kur = 0

        if odeme_doviz == 'TRY':
            odeme_kur = 1.0
        elif manuel_kur > 0:
            odeme_kur = q_kur(manuel_kur)
        else:
            odeme_kur = _kur_getir(odeme_doviz)
            if not odeme_kur or odeme_kur <= 0:
                return jsonify({'ok': False,
                    'mesaj': f'{odeme_doviz} icin kur bulunamadi. Kur alanina elle girin '
                             f"ya da Dashboard'dan kurlari guncelleyin."}), 400

        f_kur = None
        if odeme_doviz == f_doviz:
            esdeger = q3(tutar)
        else:
            f_kur = 1.0 if f_doviz == 'TRY' else _kur_getir(f_doviz)
            if not f_kur or f_kur <= 0:
                return jsonify({'ok': False,
                    'mesaj': f'{f_doviz} icin guncel kur bulunamadi (çapraz döviz çevrimi için gerekli).'}), 400
            esdeger = q3((tutar * odeme_kur) / f_kur)

        # Mevcut tahsilat + yeni tutar faturayı aşıyor mu? (çek + çapraz döviz dahil)
        mevcut_tahsil = _fatura_odenen_esdeger(f)
        kalan = q3((f.toplam or 0) - mevcut_tahsil)
        tolerans = 0.01 if odeme_doviz == f_doviz else max(0.01, kalan * 0.005)
        if esdeger > kalan + tolerans:
            msj = f'Tutar kalan bakiyeyi asiyor. Kalan: {kalan:,.2f} {f_doviz}'
            if odeme_doviz != f_doviz:
                msj += f' (girilen {tutar:,.2f} {odeme_doviz} ≈ {esdeger:,.2f} {f_doviz})'
            return jsonify({'ok': False, 'mesaj': msj}), 400

        # ── ÇEK İLE TAHSİLAT ──
        # Çek alınınca: alınan çek olarak kaydet, faturaya bağla, cariye ALACAK.
        # Çek kasaya girmez (portföye girer, tahsil edilince kasaya girer).
        tahsilat_tipi = data.get('tahsilat_tipi', 'nakit')
        if tahsilat_tipi == 'cek':
            if odeme_doviz != f_doviz:
                return jsonify({'ok': False,
                    'mesaj': f'Çek tahsilatı fatura dövizinde ({f_doviz}) olmalı. '
                             f'Çapraz döviz için nakit/havale kullanın.'}), 400
            cek_bilgi = data.get('cek') or {}
            cek_vade = _parse_date(cek_bilgi.get('vade_tarihi'))
            if not cek_vade:
                return jsonify({'ok': False, 'mesaj': 'Çek için vade tarihi zorunlu'}), 400
            cari = Cari.query.filter_by(unvan=f.musteri).first()
            cek = Cek(
                id=_yeni_id('CEK'), yon='alinan', tip='cek',
                cek_no=(cek_bilgi.get('cek_no') or '').strip() or None,
                banka_adi=(cek_bilgi.get('banka_adi') or '').strip() or None,
                sube=(cek_bilgi.get('sube') or '').strip() or None,
                hesap_sahibi=(cek_bilgi.get('hesap_sahibi') or '').strip() or None,
                tutar=q3(tutar), doviz=f.doviz or 'USD',
                keside_tarihi=_parse_date(cek_bilgi.get('keside_tarihi')),
                vade_tarihi=cek_vade,
                cari_id=cari.id if cari else None, cari_unvan=f.musteri,
                durum='Portfoyde', fatura_id=fatura_id,
                aciklama=data.get('aciklama') or f'Fatura {f.fatura_no or fatura_id} karşılığı alınan çek',
                kullanici=session.get('kullanici', 'sistem'))
            db.session.add(cek)
            db.session.flush()
            _cek_hareket_ekle(cek, 'Alındı', None, 'Portfoyde', cek.aciklama or '')
            _cari_hareket_ekle(
                cari_unvan=f.musteri, islem_tip='Çek Tahsilatı',
                borc=0, alacak=tutar, doviz=f.doviz or 'USD',
                aciklama=f'Çek alındı (No: {cek.cek_no or cek.id}) — {f.fatura_no or fatura_id}',
                kaynak='cek', baglanti_tip='cek', baglanti_id=cek.id,
                vade_tarihi=cek_vade, evrak_no=cek.cek_no or cek.id)
            _fatura_tahsilat_durumu(fatura_id)
            _log_audit('EKLE', 'tahsilat_cek', fatura_id, yeni={'tutar': tutar, 'cek': cek.id})
            ok, hata = _safe_commit(f'Çek tahsilatı: {fatura_id} / {cek.id}')
            if not ok:
                return jsonify({'ok': False, 'mesaj': f'Kayıt hatası: {hata}'}), 500
            return jsonify({'ok': True, 'cek_id': cek.id,
                'mesaj': f'Çek ({cek.cek_no or cek.id}) portföye eklendi, {tutar:,.2f} {f.doviz} tahsilat işlendi.'})

        # Cariye ALACAK hareketi (ödeme dövizinde — gerçek para hareketi budur)
        _capraz = (odeme_doviz != f_doviz)
        try:
            hareket = _cari_hareket_ekle(
                cari_unvan=f.musteri,
                islem_tip='Tahsilat',
                alacak=tutar,
                doviz=odeme_doviz,
                kur=odeme_kur,
                aciklama=data.get('aciklama') or (
                    f'Tahsilat - Fatura {f.fatura_no or fatura_id}'
                    + (f' ({tutar:,.2f} {odeme_doviz} ≈ {esdeger:,.2f} {f_doviz})' if _capraz else '')),
                kaynak='tahsilat',
                baglanti_tip='fatura',
                baglanti_id=fatura_id,
                evrak_no=data.get('evrak_no')
            )
        except ValueError as e:
            db.session.rollback()
            return jsonify({'ok': False, 'mesaj': str(e)}), 400
        db.session.flush()
        # ÇOK DÖVİZLİ: otomatik kur farkı eşleştirmesi (çapraz döviz dahil —
        # hareket faturaya bağlı olduğundan faturanın açılış borcunu hedefler)
        kur_farki = _kur_farki_hesapla_ve_olustur(hareket)
        # Fatura durumunu güncelle (kismi/tam — eşdeğer toplam üzerinden)
        _fatura_tahsilat_durumu(fatura_id)
        _log_audit('EKLE', 'tahsilat', fatura_id,
                   yeni={'tutar': tutar, 'musteri': f.musteri})

        # KASA ENTEGRASYONU: Seçili kasaya giriş kaydı
        kasa_mesaj = ''
        kasa_id = data.get('kasa_id')
        if kasa_id:
            try:
                kasa = Kasa.query.get(int(kasa_id))
                if kasa and kasa.aktif:
                    # Kasa dövizi ödeme dövizinden farklıysa TRY köprüsüyle çevir
                    cevrim = ''
                    if kasa.doviz == odeme_doviz:
                        k_tutar = q3(tutar)
                    elif kasa.doviz == 'TRY':
                        k_tutar = q3(tutar * odeme_kur)
                        cevrim = f' ({tutar:,.2f} {odeme_doviz} → {k_tutar:,.2f} TRY)'
                    else:
                        kk = _kur_getir(kasa.doviz)
                        if not kk or kk <= 0:
                            raise ValueError(f'{kasa.doviz} kuru alınamadı')
                        k_tutar = q3((tutar * odeme_kur) / kk)
                        cevrim = f' ({tutar:,.2f} {odeme_doviz} → {k_tutar:,.2f} {kasa.doviz})'
                    kh = KasaHareket(
                        kasa_id=kasa.id, tip='giris', tutar=k_tutar,
                        aciklama=f'Tahsilat: {f.musteri} - Fatura {f.fatura_no or fatura_id}{cevrim}',
                        baglanti_tip='tahsilat', baglanti_id=fatura_id,
                        cari_id=hareket.cari_id,
                        kullanici=session.get('kullanici')
                    )
                    db.session.add(kh)
                    kasa.bakiye = q3((kasa.bakiye or 0) + k_tutar)
                    kasa_mesaj = f' Kasa: {kasa.ad} +{k_tutar:,.2f} {kasa.doviz} (bakiye {q3(kasa.bakiye):,.2f}){cevrim}'
            except Exception as e:
                app.logger.warning(f'Kasa entegrasyonu hatası: {e}')
                kasa_mesaj = f' ⚠️ Kasa kaydı yapılamadı: {e}'

        db.session.commit()

        yeni_kalan = q3(kalan - esdeger)
        msj = f'{tutar:,.2f} {odeme_doviz} tahsil edildi'
        if _capraz:
            _kur_goster = f_kur if odeme_doviz == 'TRY' else odeme_kur
            _kur_doviz = f_doviz if odeme_doviz == 'TRY' else odeme_doviz
            msj += f' (≈ {esdeger:,.2f} {f_doviz}, {_kur_doviz} kuru {_kur_goster:,.4f})'
        msj += f'. Kalan: {yeni_kalan:,.2f} {f_doviz}.'
        if kur_farki:
            msj += f' Otomatik kur farkı: {kur_farki.islem_tip} {q3((kur_farki.borc or 0) + (kur_farki.alacak or 0)):,.2f} TRY.'
        msj += kasa_mesaj
        return jsonify({'ok': True, 'mesaj': msj, 'durum': f.durum,
                        'kalan': yeni_kalan, 'esdeger': esdeger,
                        'kur_farki_id': kur_farki.id if kur_farki else None})

    @app.route('/api/fatura/<fatura_id>/tahsilatlar', methods=['GET'])
    def api_fatura_tahsilatlar(fatura_id):
        """Faturanın tahsilat geçmişi + özet."""
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        f = Fatura.query.get(fatura_id)
        if not f:
            return jsonify({'ok': False, 'mesaj': 'Fatura bulunamadi'}), 404

        # Faturaya bağlı TÜM alacak hareketleri = ödemeler (tahsilat, çek, vb.)
        # Çek de bir ödeme aracıdır; kaynak ne olursa olsun alacak (>0) say.
        hareketler = CariHareket.query.filter(
            CariHareket.baglanti_tip == 'fatura',
            CariHareket.baglanti_id == fatura_id,
            (CariHareket.alacak != None),
            CariHareket.alacak > 0
        ).order_by(CariHareket.hareket_tarihi, CariHareket.guncelleme).all()
        # Ayrıca: çek faturaya fatura_id ile bağlıysa, çek tahsilat hareketi
        # baglanti_tip='cek' olabilir. Onları da yakala.
        cek_hareketler = CariHareket.query.filter(
            CariHareket.baglanti_tip == 'cek',
            (CariHareket.alacak != None),
            CariHareket.alacak > 0,
            CariHareket.baglanti_id.in_(
                db.session.query(Cek.id).filter_by(fatura_id=fatura_id)
            )
        ).all()
        # İkisini birleştir (tekrarsız)
        gorulen = {h.id for h in hareketler}
        for ch in cek_hareketler:
            if ch.id not in gorulen:
                hareketler.append(ch)

        _f_doviz = f.doviz or 'USD'
        tahsilatlar = [{
            'id': h.id,
            'tarih': h.hareket_tarihi.isoformat() if h.hareket_tarihi else None,
            'tutar': h.alacak,
            'doviz': h.doviz,
            # Çapraz dövizli ödemenin fatura dövizindeki eşdeğeri (aynı dövizde tutara eşit)
            'esdeger': q3(_hareket_fatura_esdegeri(h, _f_doviz)),
            'evrak_no': h.evrak_no,
            'tip': h.islem_tip or ('Çek' if h.kaynak == 'cek' else 'Tahsilat'),
            'aciklama': h.aciklama
        } for h in hareketler]

        toplam_tahsil = sum(t['esdeger'] or 0 for t in tahsilatlar)
        return jsonify({
            'ok': True,
            'fatura_toplam': f.toplam or 0,
            'doviz': _f_doviz,
            'tahsil_edilen': q3(toplam_tahsil),
            'kalan': q3((f.toplam or 0) - toplam_tahsil),
            'tahsilatlar': tahsilatlar
        })

    @app.route('/api/tahsilat/<hareket_id>', methods=['DELETE'])
    def api_tahsilat_sil(hareket_id):
        """Bir tahsilat kaydını siler, fatura durumunu geri günceller."""
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        h = CariHareket.query.get(hareket_id)
        if not h or h.kaynak != 'tahsilat':
            return jsonify({'ok': False, 'mesaj': 'Tahsilat bulunamadi'}), 404
        fatura_id = h.baglanti_id
        # KUR FARKI SİMETRİSİ: kapattığı hareketleri aç, kur farkı kayıtlarını sil
        kf_acilan, kf_silinen = _kur_farki_geri_al(hareket_id)
        db.session.delete(h)
        db.session.flush()
        if fatura_id:
            _fatura_tahsilat_durumu(fatura_id)
        _log_audit('SIL', 'tahsilat', hareket_id, eski={'tutar': h.alacak})
        db.session.commit()
        msg = 'Tahsilat silindi'
        if kf_silinen: msg += f', {kf_silinen} otomatik kur farkı kaydı geri alındı'
        if kf_acilan:  msg += f', kapatılmış fatura borcu yeniden açıldı'
        return jsonify({'ok': True, 'mesaj': msg})

# ════════════════════════════════════════════════════════
    # KASA YÖNETİMİ (Banka hesapları + nakit kasalar + hareketleri)
    # Frontend: templates/kasa.html
    # Yetki: /api/kasa -> 'kasa' modülü (URL_MODUL_MAP'te zaten var)
    # ════════════════════════════════════════════════════════

    @app.route('/api/kasa', methods=['GET'])
    def api_kasa_liste():
        """Tüm kasaları + giriş/çıkış özetleriyle döner.

        Ana kasalar (ana_kasa=True): doğrudan hareket girilmez, bakiyesi
        aynı döviz koduna sahip alt kasaların (ana_kasa=False) bakiye
        toplamından canlı hesaplanır.
        """
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401

        if hasattr(Kasa, 'aktif'):
            kasalar = Kasa.query.filter_by(aktif=True).order_by(Kasa.id).all()
        else:
            kasalar = Kasa.query.order_by(Kasa.id).all()

        # Banka bagli kasalar icin banka bilgisi (tek sorgu)
        _banka_map = {b.id: {'banka_adi': b.banka_adi, 'iban': b.iban} for b in Banka.query.all()}

        alt_kasa_toplam = {}
        if hasattr(Kasa, 'ana_kasa'):
            for k in kasalar:
                if not k.ana_kasa:
                    alt_kasa_toplam[k.doviz] = alt_kasa_toplam.get(k.doviz, 0) + (k.bakiye or 0)

        sonuc = []
        for k in kasalar:
            is_ana = bool(getattr(k, 'ana_kasa', False))
            bakiye_goster = q3(alt_kasa_toplam.get(k.doviz, 0)) if is_ana else q3(k.bakiye or 0)

            giris = db.session.query(db.func.sum(KasaHareket.tutar)).filter_by(
                kasa_id=k.id, tip='giris').scalar() or 0
            cikis = db.session.query(db.func.sum(KasaHareket.tutar)).filter_by(
                kasa_id=k.id, tip='cikis').scalar() or 0

            sonuc.append({
                'id': k.id,
                'ad': k.ad,
                'doviz': k.doviz,
                'bakiye': bakiye_goster,
                'baslangic_bakiye': q3(getattr(k, 'baslangic_bakiye', 0) or 0),
                'aciklama': getattr(k, 'aciklama', '') or '',
                'giris_toplam': q3(giris),
                'cikis_toplam': q3(cikis),
                'ana_kasa': is_ana,
                # Banka baglantisi: bagliysa banka adi/IBAN da doner
                'banka_id': getattr(k, 'banka_id', None),
                'banka_adi': (_banka_map.get(getattr(k, 'banka_id', None)) or {}).get('banka_adi'),
                'banka_iban': (_banka_map.get(getattr(k, 'banka_id', None)) or {}).get('iban'),
            })
        return jsonify(sonuc)

    @app.route('/api/kasa', methods=['POST'])
    def api_kasa_ekle():
        """Yeni kasa oluşturur. Başlangıç bakiyesi varsa otomatik 'giris' hareketi açar."""
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        data = request.json or {}

        ad = (data.get('ad') or '').strip()
        if not ad:
            return jsonify({'ok': False, 'mesaj': 'Kasa adı zorunlu'}), 400

        is_ana_kasa = bool(data.get('ana_kasa', False))

        try:
            baslangic = float(data.get('baslangic_bakiye') or 0)
        except (ValueError, TypeError):
            baslangic = 0

        if is_ana_kasa and baslangic > 0:
            return jsonify({
                'ok': False,
                'mesaj': 'Ana kasaya başlangıç bakiyesi girilemez; bakiyesi alt kasalardan otomatik hesaplanır.'
            }), 400

        # Aynı isimde aktif kasa var mı?
        mevcut = Kasa.query.filter_by(ad=ad)
        if hasattr(Kasa, 'aktif'):
            mevcut = mevcut.filter_by(aktif=True)
        if mevcut.first():
            return jsonify({'ok': False, 'mesaj': f'"{ad}" adında bir kasa zaten var'}), 400

        try:
            k = Kasa(
                ad=ad,
                doviz=(data.get('doviz') or 'TRY').upper(),
                bakiye=q3(baslangic),
            )
            # Opsiyonel alanlar
            if hasattr(Kasa, 'baslangic_bakiye'):
                k.baslangic_bakiye = q3(baslangic)
            if hasattr(Kasa, 'aciklama'):
                k.aciklama = data.get('aciklama') or ''
            if hasattr(Kasa, 'aktif'):
                k.aktif = True
            if hasattr(Kasa, 'kullanici'):
                k.kullanici = session.get('kullanici', 'sistem')
            if hasattr(Kasa, 'ana_kasa'):
                k.ana_kasa = is_ana_kasa

            db.session.add(k)
            db.session.flush()

            # Başlangıç bakiyesi > 0 ise otomatik açılış hareketi
            if baslangic > 0:
                kh = KasaHareket(
                    kasa_id=k.id,
                    tip='giris',
                    tutar=q3(baslangic),
                    aciklama='Başlangıç bakiyesi (açılış)',
                )
                if hasattr(KasaHareket, 'tarih'):
                    kh.tarih = date.today()
                if hasattr(KasaHareket, 'kaynak'):
                    kh.kaynak = 'acilis'
                if hasattr(KasaHareket, 'kullanici'):
                    kh.kullanici = session.get('kullanici', 'sistem')
                db.session.add(kh)

            _log_audit('EKLE', 'kasa', k.id,
                       yeni={'ad': k.ad, 'doviz': k.doviz, 'baslangic': baslangic})

            ok, hata = _safe_commit('Kasa ekleme')
            if not ok:
                return jsonify({'ok': False, 'mesaj': f'Hata: {hata}'}), 500

            return jsonify({'ok': True, 'id': k.id, 'mesaj': 'Kasa oluşturuldu'})
        except Exception as e:
            db.session.rollback()
            app.logger.exception('Kasa ekleme hatası')
            return jsonify({'ok': False, 'mesaj': f'Hata: {str(e)}'}), 500

    @app.route('/api/kasa/<int:kasa_id>', methods=['PUT'])
    def api_kasa_guncelle(kasa_id):
        """Kasa bilgilerini günceller. Başlangıç bakiyesi değişirse fark bakiye'ye yansır."""
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        k = Kasa.query.get(kasa_id)
        if not k:
            return jsonify({'ok': False, 'mesaj': 'Kasa bulunamadı'}), 404

        data = request.json or {}
        eski = {'ad': k.ad, 'doviz': k.doviz,
                'baslangic_bakiye': getattr(k, 'baslangic_bakiye', 0)}

        if 'ad' in data and (data['ad'] or '').strip():
            k.ad = data['ad'].strip()
        if 'doviz' in data and data['doviz']:
            k.doviz = data['doviz'].upper()
        if 'aciklama' in data and hasattr(k, 'aciklama'):
            k.aciklama = data['aciklama'] or ''
        if 'ana_kasa' in data and hasattr(k, 'ana_kasa'):
            k.ana_kasa = bool(data['ana_kasa'])

        # Başlangıç bakiyesi değişirse: fark bakiyeye yansır
        if 'baslangic_bakiye' in data and hasattr(k, 'baslangic_bakiye'):
            try:
                yeni_bas = float(data['baslangic_bakiye'] or 0)
                eski_bas = getattr(k, 'baslangic_bakiye', 0) or 0
                fark = yeni_bas - eski_bas
                k.baslangic_bakiye = q3(yeni_bas)
                k.bakiye = q3((k.bakiye or 0) + fark)
            except (ValueError, TypeError):
                pass

        _log_audit('GUNCELLE', 'kasa', k.id, eski=eski,
                   yeni={'ad': k.ad, 'doviz': k.doviz})

        ok, hata = _safe_commit('Kasa güncelleme')
        if not ok:
            return jsonify({'ok': False, 'mesaj': f'Hata: {hata}'}), 500
        return jsonify({'ok': True, 'mesaj': 'Kasa güncellendi'})

    @app.route('/api/kasa/<int:kasa_id>', methods=['DELETE'])
    def api_kasa_sil(kasa_id):
        """Kasayı siler. Hareketi varsa engellenir (manuel silinmeli önce)."""
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        k = Kasa.query.get(kasa_id)
        if not k:
            return jsonify({'ok': False, 'mesaj': 'Kasa bulunamadı'}), 404

        hareket_sayisi = KasaHareket.query.filter_by(kasa_id=kasa_id).count()
        if hareket_sayisi > 0:
            return jsonify({
                'ok': False,
                'mesaj': f'Bu kasada {hareket_sayisi} hareket var, silinemez. Önce hareketleri silin.'
            }), 400

        _log_audit('SIL', 'kasa', kasa_id, eski={'ad': k.ad, 'doviz': k.doviz})
        db.session.delete(k)

        ok, hata = _safe_commit('Kasa silme')
        if not ok:
            return jsonify({'ok': False, 'mesaj': f'Hata: {hata}'}), 500
        return jsonify({'ok': True, 'mesaj': 'Kasa silindi'})

    @app.route('/api/kasa/<int:kasa_id>/hareket', methods=['GET'])
    def api_kasa_hareket_liste(kasa_id):
        """Bir kasanın tüm hareketlerini tarih DESC sırada döner.
        Ana kasa (konsolide) ise aynı dövizdeki alt kasaların hareketlerini birleştirir."""
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401

        k = Kasa.query.get(kasa_id)
        if not k:
            return jsonify([])

        is_ana = bool(getattr(k, 'ana_kasa', False))
        kasa_adi_map = {}
        if is_ana:
            # Aynı dövizdeki alt kasalar (ana_kasa olmayanlar)
            alt_q = Kasa.query.filter_by(doviz=k.doviz)
            if hasattr(Kasa, 'ana_kasa'):
                alt_q = alt_q.filter_by(ana_kasa=False)
            alt_kasalar = alt_q.all()
            kasa_ids = [a.id for a in alt_kasalar]
            kasa_adi_map = {a.id: a.ad for a in alt_kasalar}
            if kasa_ids:
                hareket_filtresi = KasaHareket.kasa_id.in_(kasa_ids)
            else:
                hareket_filtresi = (KasaHareket.kasa_id == -1)  # hiç eşleşmeyen
        else:
            hareket_filtresi = (KasaHareket.kasa_id == kasa_id)
            kasa_adi_map = {k.id: k.ad}

        # Tarih alanı varsa ona göre, yoksa id'ye göre sırala
        if hasattr(KasaHareket, 'tarih'):
            hareketler = KasaHareket.query.filter(hareket_filtresi).order_by(
                KasaHareket.tarih.desc(), KasaHareket.id.desc()).all()
        else:
            hareketler = KasaHareket.query.filter(hareket_filtresi).order_by(
                KasaHareket.id.desc()).all()

        sonuc = []
        for h in hareketler:
            tarih_iso = None
            if hasattr(h, 'tarih') and h.tarih:
                tarih_iso = h.tarih.isoformat()
            elif hasattr(h, 'olusturma') and h.olusturma:
                tarih_iso = h.olusturma.isoformat()[:10]

            sonuc.append({
                'id': h.id,
                'tarih': tarih_iso,
                'tip': h.tip,
                'tutar': q3(h.tutar or 0),
                'evrak_no': getattr(h, 'evrak_no', '') or '',
                'siparis_id': getattr(h, 'siparis_id', '') or '',
                'aciklama': h.aciklama or '',
                'baglanti_tip': getattr(h, 'baglanti_tip', None),
                'baglanti_id': getattr(h, 'baglanti_id', None),
                'kasa_id': h.kasa_id,
                'kasa_adi': kasa_adi_map.get(h.kasa_id, ''),
                'kullanici': getattr(h, 'kullanici', '') or ''
            })
        return jsonify(sonuc)

    @app.route('/api/kasa/hareket', methods=['POST'])
    def api_kasa_hareket_ekle():
        """Manuel kasa hareketi (giriş/çıkış). Bakiye otomatik güncellenir."""
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        data = request.json or {}

        kasa_id = data.get('kasa_id')
        tip = (data.get('tip') or '').strip().lower()

        if not kasa_id:
            return jsonify({'ok': False, 'mesaj': 'Kasa seçilmedi'}), 400
        if tip not in ('giris', 'cikis'):
            return jsonify({'ok': False, 'mesaj': 'Tip "giris" veya "cikis" olmalı'}), 400

        try:
            tutar = float(data.get('tutar') or 0)
        except (ValueError, TypeError):
            tutar = 0
        if tutar <= 0:
            return jsonify({'ok': False, 'mesaj': 'Geçerli bir tutar girin'}), 400

        k = Kasa.query.get(int(kasa_id))
        if not k:
            return jsonify({'ok': False, 'mesaj': 'Kasa bulunamadı'}), 404

        if getattr(k, 'ana_kasa', False):
            return jsonify({
                'ok': False,
                'mesaj': 'Ana kasaya doğrudan hareket girilemez. Hareketi ilgili alt kasaya işleyin.'
            }), 400

        # Çıkışta bakiye yetersizliği uyarısı (eksiye düşmesin)
        if tip == 'cikis' and (k.bakiye or 0) < tutar:
            return jsonify({
                'ok': False,
                'mesaj': f'Yetersiz bakiye. Mevcut: {q3(k.bakiye or 0):,.2f} {k.doviz}'
            }), 400

        try:
            kh = KasaHareket(
                kasa_id=k.id,
                tip=tip,
                tutar=q3(tutar),
                aciklama=data.get('aciklama') or '',
            )
            if hasattr(KasaHareket, 'tarih'):
                kh.tarih = _parse_date(data.get('tarih')) or date.today()
            if hasattr(KasaHareket, 'evrak_no'):
                kh.evrak_no = data.get('evrak_no') or ''
            if hasattr(KasaHareket, 'kaynak'):
                kh.kaynak = 'manuel'
            if hasattr(KasaHareket, 'kullanici'):
                kh.kullanici = session.get('kullanici', 'sistem')
            # Cari bağlantısı: tahsilat / tediye için cari_id verilebilir,
            # integration engine bunu CariHareket'e otomatik aktarır.
            if hasattr(KasaHareket, 'cari_id') and data.get('cari_id'):
                kh.cari_id = data.get('cari_id')
            # Sipariş bağlantısı: sipariş bazlı takip/ekstre için
            if hasattr(KasaHareket, 'siparis_id') and data.get('siparis_id'):
                kh.siparis_id = data.get('siparis_id')

            db.session.add(kh)

            # Bakiye güncelle
            if tip == 'giris':
                k.bakiye = q3((k.bakiye or 0) + tutar)
            else:
                k.bakiye = q3((k.bakiye or 0) - tutar)

            db.session.flush()
            _log_audit('EKLE', 'kasa_hareket', kh.id,
                       yeni={'kasa': k.ad, 'tip': tip, 'tutar': tutar})

            ok, hata = _safe_commit(f'Kasa hareketi: {tip}')
            if not ok:
                return jsonify({'ok': False, 'mesaj': f'Hata: {hata}'}), 500

            return jsonify({
                'ok': True,
                'id': kh.id,
                'yeni_bakiye': q3(k.bakiye),
                'mesaj': f'{tip.capitalize()} kaydedildi. Yeni bakiye: {k.bakiye:,.2f} {k.doviz}'
            })
        except Exception as e:
            db.session.rollback()
            app.logger.exception('Kasa hareket ekleme hatası')
            return jsonify({'ok': False, 'mesaj': f'Hata: {str(e)}'}), 500

    @app.route('/api/kasa/virman', methods=['POST'])
    def api_kasa_virman():
        """Kasa <-> Kasa (veya Kasa <-> Banka kasasi) para transferi.
        CIFT TARAFLI: kaynaktan CIKIS + hedefe GIRIS, tek islemde (atomik).
        Ornek: nakit kasadan bankaya para yatirma / bankadan nakit cekme."""
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        if _yazma_yetki_guard(): return _yazma_yetki_guard()
        data = request.json or {}
        kaynak_id = data.get('kaynak_kasa_id')
        hedef_id = data.get('hedef_kasa_id')
        try:
            tutar = float(data.get('tutar') or 0)
        except (TypeError, ValueError):
            tutar = 0

        if not kaynak_id or not hedef_id:
            return jsonify({'ok': False, 'mesaj': 'Kaynak ve hedef kasa gerekli'}), 400
        if int(kaynak_id) == int(hedef_id):
            return jsonify({'ok': False, 'mesaj': 'Kaynak ve hedef kasa ayni olamaz'}), 400
        if tutar <= 0:
            return jsonify({'ok': False, 'mesaj': 'Tutar sifirdan buyuk olmali'}), 400

        kaynak = db.session.get(Kasa, int(kaynak_id))
        hedef = db.session.get(Kasa, int(hedef_id))
        if not kaynak or not hedef:
            return jsonify({'ok': False, 'mesaj': 'Kasa bulunamadi'}), 404
        if kaynak.ana_kasa or hedef.ana_kasa:
            return jsonify({'ok': False, 'mesaj': 'Ana kasaya dogrudan hareket girilemez'}), 400
        if (kaynak.doviz or '') != (hedef.doviz or ''):
            return jsonify({'ok': False,
                            'mesaj': f'Doviz uyusmuyor: {kaynak.doviz} -> {hedef.doviz}. '
                                     f'Farkli dovizler arasi transfer icin once bozdurma yapin.'}), 400
        if (kaynak.bakiye or 0) < tutar:
            return jsonify({'ok': False,
                            'mesaj': f'Yetersiz bakiye. {kaynak.ad}: {kaynak.bakiye:,.2f} {kaynak.doviz}'}), 400

        _tarih = _parse_date(data.get('tarih')) or date.today()
        _aciklama = (data.get('aciklama') or '').strip()
        _evrak = (data.get('evrak_no') or '').strip()

        try:
            # Aciklamada evrak no da gorunsun (KasaHareket'te ayri evrak_no alani yok)
            _metin = _aciklama or f'Virman: {kaynak.ad} -> {hedef.ad}'
            if _evrak:
                _metin = f'{_metin} (Evrak: {_evrak})'

            # 1) Kaynaktan CIKIS
            cikis = KasaHareket(
                kasa_id=kaynak.id, tip='cikis', tutar=q3(tutar), tarih=_tarih,
                aciklama=_metin,
                baglanti_tip='virman', baglanti_id=str(hedef.id),
                kullanici=session.get('kullanici')
            )
            kaynak.bakiye = q3((kaynak.bakiye or 0) - tutar)
            db.session.add(cikis)

            # 2) Hedefe GIRIS
            giris = KasaHareket(
                kasa_id=hedef.id, tip='giris', tutar=q3(tutar), tarih=_tarih,
                aciklama=_metin,
                baglanti_tip='virman', baglanti_id=str(kaynak.id),
                kullanici=session.get('kullanici')
            )
            hedef.bakiye = q3((hedef.bakiye or 0) + tutar)
            db.session.add(giris)

            # Denetim izi commit'ten ÖNCE — _log_audit kendi commit'ini yapmaz
            _log_audit('EKLE', 'kasa_virman', f'{kaynak.id}->{hedef.id}',
                       yeni={'kaynak': kaynak.ad, 'hedef': hedef.ad,
                             'tutar': tutar, 'doviz': kaynak.doviz})
            db.session.commit()
            return jsonify({
                'ok': True,
                'mesaj': f'{tutar:,.2f} {kaynak.doviz} transfer edildi: '
                         f'{kaynak.ad} -> {hedef.ad}. '
                         f'Yeni bakiyeler: {kaynak.ad} {kaynak.bakiye:,.2f} / '
                         f'{hedef.ad} {hedef.bakiye:,.2f}',
                'kaynak_bakiye': kaynak.bakiye,
                'hedef_bakiye': hedef.bakiye
            })
        except Exception as hata:
            db.session.rollback()
            return jsonify({'ok': False, 'mesaj': f'Hata: {hata}'}), 500

    @app.route('/api/kasa/hareket/<int:hareket_id>', methods=['DELETE'])
    def api_kasa_hareket_sil(hareket_id):
        """Hareketi siler ve bakiyeden geri alır. Tahsilat bağlantısı varsa engellenir."""
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        h = KasaHareket.query.get(hareket_id)
        if not h:
            return jsonify({'ok': False, 'mesaj': 'Hareket bulunamadı'}), 404

        # Tahsilat kaynaklı (fatura) hareketi silmeyi engelle
        if hasattr(h, 'baglanti_tip') and h.baglanti_tip == 'tahsilat':
            return jsonify({
                'ok': False,
                'mesaj': 'Bu hareket fatura tahsilatından oluştu. Önce fatura tahsilatını silin.'
            }), 400

        k = Kasa.query.get(h.kasa_id)
        if k:
            # Bakiyeyi geri al
            if h.tip == 'giris':
                k.bakiye = q3((k.bakiye or 0) - (h.tutar or 0))
            else:
                k.bakiye = q3((k.bakiye or 0) + (h.tutar or 0))

        _log_audit('SIL', 'kasa_hareket', hareket_id,
                   eski={'kasa_id': h.kasa_id, 'tip': h.tip, 'tutar': h.tutar})
        db.session.delete(h)

        ok, hata = _safe_commit('Kasa hareketi silme')
        if not ok:
            return jsonify({'ok': False, 'mesaj': f'Hata: {hata}'}), 500
        return jsonify({'ok': True, 'mesaj': 'Hareket silindi'})

    @app.route('/api/kasa/<int:kasa_id>/hareket/excel', methods=['GET'])
    def api_kasa_hareket_excel(kasa_id):
        """Bir kasanın (veya ana kasaysa konsolide alt kasaların) hareket
        dökümünü Excel (.xlsx) olarak indirir."""
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        import io
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            return jsonify({'ok': False,
                'mesaj': 'Excel çıktısı için "openpyxl" paketi gerekli. '
                         'Kurulum: pip install openpyxl'}), 500

        k = Kasa.query.get(kasa_id)
        if not k:
            return jsonify({'ok': False, 'mesaj': 'Kasa bulunamadı'}), 404

        is_ana = bool(getattr(k, 'ana_kasa', False))

        if is_ana:
            alt_kasalar = Kasa.query.filter_by(doviz=k.doviz)
            if hasattr(Kasa, 'ana_kasa'):
                alt_kasalar = alt_kasalar.filter_by(ana_kasa=False)
            alt_kasa_ids = [a.id for a in alt_kasalar.all()]
            kasa_adi_map = {a.id: a.ad for a in alt_kasalar.all()}
            hareketler_q = KasaHareket.query.filter(KasaHareket.kasa_id.in_(alt_kasa_ids)) if alt_kasa_ids else KasaHareket.query.filter(False)
        else:
            hareketler_q = KasaHareket.query.filter_by(kasa_id=k.id)
            kasa_adi_map = {k.id: k.ad}

        if hasattr(KasaHareket, 'tarih'):
            hareketler = hareketler_q.order_by(KasaHareket.tarih.asc(), KasaHareket.id.asc()).all()
        else:
            hareketler = hareketler_q.order_by(KasaHareket.id.asc()).all()

        wb = Workbook()
        sheet = wb.active
        sheet.title = 'Kasa Hareketleri'

        baslik = f"{'Konsolide Kasa Raporu' if is_ana else 'Kasa Hareket Raporu'} — {k.ad} ({k.doviz})"
        sheet.merge_cells('A1:G1')
        sheet['A1'] = baslik
        sheet['A1'].font = Font(bold=True, size=13)
        sheet['A1'].alignment = Alignment(horizontal='left')

        headers = ['Tarih', 'Kasa', 'Tip', 'Tutar', 'Evrak No', 'Açıklama', 'Bakiye']
        if not is_ana:
            headers.remove('Kasa')
        header_row = 3
        for col_idx, h in enumerate(headers, start=1):
            c = sheet.cell(row=header_row, column=col_idx, value=h)
            c.font = Font(bold=True, color='FFFFFF')
            c.fill = PatternFill('solid', start_color='232824')
            c.alignment = Alignment(horizontal='center')

        # AÇILIŞ BAKİYESİ (kendini doğrulayan): güncel bakiye − hareketlerin neti.
        # Başlangıç bakiyesi hem kasa alanında hem açılış hareketi olarak tutulduğu
        # için baslangic_bakiye'den başlamak çift sayıma yol açıyordu; bu formülle
        # son satır her zaman güncel bakiyeye eşit çıkar.
        _net = sum(q3(h.tutar or 0) if h.tip == 'giris' else -q3(h.tutar or 0) for h in hareketler)
        if is_ana:
            _guncel = sum(q3(a.bakiye or 0) for a in alt_kasalar.all())
        else:
            _guncel = q3(k.bakiye or 0)
        yuruyen = q3(_guncel - _net)
        row_idx = header_row + 1
        for h in hareketler:
            tutar = q3(h.tutar or 0)
            yuruyen = q3(yuruyen + tutar) if h.tip == 'giris' else q3(yuruyen - tutar)
            tarih_val = h.tarih.strftime('%d.%m.%Y') if getattr(h, 'tarih', None) else ''

            col = 1
            sheet.cell(row=row_idx, column=col, value=tarih_val); col += 1
            if is_ana:
                sheet.cell(row=row_idx, column=col, value=kasa_adi_map.get(h.kasa_id, '-')); col += 1
            sheet.cell(row=row_idx, column=col, value='Giriş' if h.tip == 'giris' else 'Çıkış'); col += 1
            tutar_cell = sheet.cell(row=row_idx, column=col, value=tutar)
            tutar_cell.number_format = '#,##0.00'
            col += 1
            sheet.cell(row=row_idx, column=col, value=getattr(h, 'evrak_no', '') or '-'); col += 1
            sheet.cell(row=row_idx, column=col, value=h.aciklama or '-'); col += 1
            bakiye_cell = sheet.cell(row=row_idx, column=col, value=yuruyen)
            bakiye_cell.number_format = '#,##0.00'
            row_idx += 1

        if hareketler:
            son_row = sheet.cell(row=row_idx, column=1, value='Güncel Bakiye:')
            son_row.font = Font(bold=True)
            bakiye_son_col = len(headers)
            son_bakiye_cell = sheet.cell(row=row_idx, column=bakiye_son_col,
                                          value=q3(k.bakiye if not is_ana else yuruyen))
            son_bakiye_cell.font = Font(bold=True)
            son_bakiye_cell.number_format = '#,##0.00'

        for col_idx in range(1, len(headers) + 1):
            sheet.column_dimensions[get_column_letter(col_idx)].width = 20

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        dosya_adi = f"kasa_{k.ad}_{date.today().isoformat()}.xlsx".replace(' ', '_')
        return send_file(
            buf,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=dosya_adi,
        )

    @app.route('/api/kasa/konsolide/<doviz>/excel', methods=['GET'])
    def api_kasa_konsolide_excel(doviz):
        """Bir dövizdeki TÜM alt kasaların birleşik hareket dökümü (.xlsx).
        Ana kasa tanımlı olmasa da çalışır — doğrudan döviz üzerinden toplar."""
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        import io
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            return jsonify({'ok': False,
                'mesaj': 'Excel çıktısı için "openpyxl" paketi gerekli. '
                         'Kurulum: pip install openpyxl'}), 500

        doviz = (doviz or '').upper().strip()
        kq = Kasa.query.filter_by(doviz=doviz)
        if hasattr(Kasa, 'ana_kasa'):
            kq = kq.filter_by(ana_kasa=False)
        if hasattr(Kasa, 'aktif'):
            kq = kq.filter_by(aktif=True)
        kasalar = kq.all()
        if not kasalar:
            return jsonify({'ok': False, 'mesaj': f'{doviz} dövizinde kasa bulunamadı'}), 404

        kasa_ids = [a.id for a in kasalar]
        kasa_adi_map = {a.id: a.ad for a in kasalar}
        hq = KasaHareket.query.filter(KasaHareket.kasa_id.in_(kasa_ids))
        if hasattr(KasaHareket, 'tarih'):
            hareketler = hq.order_by(KasaHareket.tarih.asc(), KasaHareket.id.asc()).all()
        else:
            hareketler = hq.order_by(KasaHareket.id.asc()).all()

        wb = Workbook()
        sheet = wb.active
        sheet.title = 'Konsolide Hareketler'

        sheet.merge_cells('A1:G1')
        sheet['A1'] = f"Konsolide Kasa Raporu — {doviz} ({len(kasalar)} kasa: {', '.join(kasa_adi_map.values())})"
        sheet['A1'].font = Font(bold=True, size=13)
        sheet['A1'].alignment = Alignment(horizontal='left')

        headers = ['Tarih', 'Kasa', 'Tip', 'Tutar', 'Evrak No', 'Açıklama', 'Bakiye']
        header_row = 3
        for col_idx, h in enumerate(headers, start=1):
            c = sheet.cell(row=header_row, column=col_idx, value=h)
            c.font = Font(bold=True, color='FFFFFF')
            c.fill = PatternFill('solid', start_color='232824')
            c.alignment = Alignment(horizontal='center')

        # Açılış = güncel toplam bakiye − hareketlerin neti (kendini doğrulayan)
        _net = sum(q3(h.tutar or 0) if h.tip == 'giris' else -q3(h.tutar or 0) for h in hareketler)
        _guncel = sum(q3(a.bakiye or 0) for a in kasalar)
        yuruyen = q3(_guncel - _net)

        row_idx = header_row + 1
        for h in hareketler:
            tutar = q3(h.tutar or 0)
            yuruyen = q3(yuruyen + tutar) if h.tip == 'giris' else q3(yuruyen - tutar)
            tarih_val = h.tarih.strftime('%d.%m.%Y') if getattr(h, 'tarih', None) else ''
            sheet.cell(row=row_idx, column=1, value=tarih_val)
            sheet.cell(row=row_idx, column=2, value=kasa_adi_map.get(h.kasa_id, '-'))
            sheet.cell(row=row_idx, column=3, value='Giriş' if h.tip == 'giris' else 'Çıkış')
            c4 = sheet.cell(row=row_idx, column=4, value=tutar); c4.number_format = '#,##0.00'
            sheet.cell(row=row_idx, column=5, value=getattr(h, 'evrak_no', '') or '-')
            sheet.cell(row=row_idx, column=6, value=h.aciklama or '-')
            c7 = sheet.cell(row=row_idx, column=7, value=yuruyen); c7.number_format = '#,##0.00'
            row_idx += 1

        # Kasa bazlı alt özet
        row_idx += 1
        oz = sheet.cell(row=row_idx, column=1, value='Kasa Bazlı Özet'); oz.font = Font(bold=True)
        row_idx += 1
        for a in kasalar:
            sheet.cell(row=row_idx, column=2, value=a.ad)
            cb = sheet.cell(row=row_idx, column=7, value=q3(a.bakiye or 0)); cb.number_format = '#,##0.00'
            row_idx += 1
        sr = sheet.cell(row=row_idx, column=1, value='Konsolide Güncel Bakiye:'); sr.font = Font(bold=True)
        sc = sheet.cell(row=row_idx, column=7, value=q3(_guncel))
        sc.font = Font(bold=True); sc.number_format = '#,##0.00'

        for col_idx in range(1, len(headers) + 1):
            sheet.column_dimensions[get_column_letter(col_idx)].width = 20

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(
            buf,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'kasa_konsolide_{doviz}_{date.today().isoformat()}.xlsx',
        )

    # ═══════════════════════════════════════════════════════════
    #  JENERİK DIŞA AKTARMA (xlsx + pdf) — tüm liste modülleri
    #  /api/export/<modul>?format=xlsx|pdf
    # ═══════════════════════════════════════════════════════════
    @app.route('/api/export/<modul>', methods=['GET'])
    def api_export(modul):
        if _auth_required(): return jsonify({'error': 'Unauthorized'}), 401
        try:
            from export_utils import liste_xlsx, liste_pdf
        except ImportError as e:
            eksik = 'openpyxl' if 'openpyxl' in str(e) else ('reportlab' if 'reportlab' in str(e) else str(e))
            return jsonify({'ok': False,
                'mesaj': f'Dışa aktarma için "{eksik}" paketi gerekli. '
                         f'Kurulum: pip install openpyxl reportlab'}), 500
        fmt = (request.args.get('format') or 'xlsx').lower()

        def _f(deger, ondalik=False):
            if deger is None: return ''
            if ondalik:
                try: return f'{float(deger):,.2f}'
                except (ValueError, TypeError): return str(deger)
            return str(deger)

        def _tarih(d):
            # date/datetime ise formatla; string ise olduğu gibi döndür; None ise boş.
            if not d:
                return ''
            if hasattr(d, 'strftime'):
                try:
                    return d.strftime('%d.%m.%Y')
                except Exception:
                    return str(d)
            return str(d)

        try:
            return _export_uret(modul, fmt, _f, _tarih, liste_xlsx, liste_pdf)
        except Exception as e:
            import traceback
            app.logger.error(f'Export hatası ({modul}/{fmt}): {e}\n{traceback.format_exc()}')
            return jsonify({'ok': False,
                'mesaj': f'Dışa aktarma hatası ({modul}): {str(e)}'}), 500

    def _export_uret(modul, fmt, _f, _tarih, liste_xlsx, liste_pdf):
        baslik = 'Liste'
        headers = []
        rows = []
        sayisal = []
        dosya = modul

        if modul == 'siparis':
            baslik = 'Sipariş Listesi'
            headers = ['Sipariş No', 'Tarih', 'Müşteri', 'Döviz', 'Tutar', 'Durum', 'Ödeme', 'Termin']
            sayisal = [4]
            q = Siparis.query.order_by(Siparis.id.desc()).all()
            for s in q:
                rows.append([s.id, _tarih(getattr(s, 'siparis_tarihi', None) or getattr(s, 'olusturma', None)),
                             s.musteri, s.doviz, _f(s.toplam_tutar, True), s.durum,
                             s.odeme_sekli or '', _tarih(s.termin)])
            dosya = 'siparis_listesi'

        elif modul == 'fatura':
            baslik = 'Fatura Listesi'
            headers = ['Fatura No', 'Tarih', 'Vade', 'Müşteri', 'Döviz', 'Tutar', 'Durum', 'Yön']
            sayisal = [5]
            q = Fatura.query.order_by(Fatura.id.desc()).all()
            for f in q:
                rows.append([f.fatura_no or f.id, _tarih(f.fatura_tarihi), _tarih(f.vade_tarihi),
                             f.musteri, f.doviz, _f(f.toplam, True), f.durum, f.yon or 'satis'])
            dosya = 'fatura_listesi'

        elif modul == 'cek':
            baslik = 'Çek Listesi'
            headers = ['Çek No', 'Yön', 'Banka', 'Keşideci', 'Tutar', 'Döviz', 'Vade', 'Durum', 'Cari']
            sayisal = [4]
            q = Cek.query.filter_by(aktif=True).order_by(Cek.vade_tarihi).all()
            for c in q:
                rows.append([c.cek_no or c.id, 'Alınan' if c.yon == 'alinan' else 'Verilen',
                             c.banka_adi or '', c.hesap_sahibi or '', _f(c.tutar, True), c.doviz,
                             _tarih(c.vade_tarihi), c.durum, c.cari_unvan or ''])
            dosya = 'cek_listesi'

        elif modul == 'cari':
            baslik = 'Cari Listesi'
            headers = ['Ünvan', 'Tip', 'Ülke', 'Telefon', 'E-posta', 'Vergi No']
            q = Cari.query.order_by(Cari.unvan).all()
            for c in q:
                rows.append([c.unvan, c.cari_tip or '', getattr(c, 'ulke', '') or '',
                             getattr(c, 'telefon', '') or '', getattr(c, 'email', '') or '',
                             getattr(c, 'vergi_no', '') or ''])
            dosya = 'cari_listesi'

        elif modul == 'stok':
            tip = (request.args.get('tip') or 'BLOK').upper()
            baslik = f'{tip} Stok Listesi'
            if tip == 'BLOK':
                headers = ['Stok No', 'Blok No', 'Cins', 'Üretici', 'Hacim m³', 'Tonaj', 'Durum', 'Konum']
                sayisal = [4, 5]
                for s in BlokStok.query.order_by(BlokStok.id.desc()).all():
                    rows.append([s.id, s.blok_no or '', s.cins or '', s.uretici or '',
                                 _f(s.hacim_m3, True), _f(s.tonaj, True), s.durum or '', s.konum or ''])
            elif tip == 'PLAKA':
                headers = ['Stok No', 'Blok-Slab', 'Cins', 'Üretici', 'm²', 'Kalınlık', 'Durum', 'Konum']
                sayisal = [4]
                for s in PlakaStok.query.order_by(PlakaStok.id.desc()).all():
                    blok_slab = f"{s.blok_no or ''}-{s.slab_no or ''}" if s.blok_no else (s.slab_no or '')
                    rows.append([s.id, blok_slab, s.cins or '', s.uretici or '',
                                 _f(s.metraj_m2, True), _f(s.kalinlik, True), s.durum or '', s.konum or ''])
            else:  # EBATLI
                headers = ['Stok No', 'Kasa No', 'Cins', 'Üretici', 'm²', 'Durum', 'Konum']
                sayisal = [4]
                for s in EbatliStok.query.order_by(EbatliStok.id.desc()).all():
                    rows.append([s.id, s.kasa_no or '', s.cins or '', s.uretici or '',
                                 _f(s.metraj_m2, True), s.durum or '', s.konum or ''])
            dosya = f'stok_{tip.lower()}'

        elif modul == 'proforma':
            baslik = 'Proforma Listesi'
            headers = ['Proforma No', 'Tarih', 'Müşteri', 'Döviz', 'Tutar', 'Durum']
            sayisal = [4]
            for p in Proforma.query.order_by(Proforma.id.desc()).all():
                rows.append([p.id, _tarih(getattr(p, 'proforma_tarihi', None) or getattr(p, 'olusturma', None)),
                             p.musteri, p.doviz, _f(getattr(p, 'toplam_tutar', None) or getattr(p, 'toplam', None), True),
                             p.durum])
            dosya = 'proforma_listesi'

        elif modul == 'cari_hareket':
            cari_id = request.args.get('cari_id')
            baslik = 'Cari Hareketler'
            headers = ['Tarih', 'İşlem', 'Evrak No', 'Borç', 'Alacak', 'Döviz', 'Açıklama']
            sayisal = [3, 4]
            q = CariHareket.query
            if cari_id:
                q = q.filter_by(cari_id=cari_id)
                cr = Cari.query.get(cari_id)
                if cr: baslik = f'Cari Ekstre — {cr.unvan}'
            for h in q.order_by(CariHareket.hareket_tarihi).all():
                rows.append([_tarih(h.hareket_tarihi), h.islem_tip or '', h.evrak_no or '',
                             _f(h.borc, True), _f(h.alacak, True), h.doviz or '', h.aciklama or ''])
            dosya = 'cari_hareket'

        elif modul == 'sevkiyat':
            baslik = 'Sevkiyat Listesi'
            headers = ['Sevkiyat No', 'Müşteri', 'Çıkış', 'Varış', 'Nakliye', 'Konteyner', 'Durum', 'Çıkış Tarihi']
            q = Sevkiyat.query.order_by(Sevkiyat.id.desc()).all()
            for s in q:
                rows.append([s.id, s.musteri or '', s.cikis_noktasi or '', s.varis_noktasi or '',
                             s.nakliye_firma or '', s.konteyner_no or '', s.durum or '',
                             _tarih(getattr(s, 'cikis_tarihi', None))])
            dosya = 'sevkiyat_listesi'

        elif modul == 'satislar':
            baslik = 'Satış Listesi'
            headers = ['Satış No', 'Tarih', 'Müşteri', 'Cins', 'Miktar', 'Birim',
                       'Tutar (USD)', 'Maliyet (USD)', 'Kâr (USD)', 'Marj %', 'Fatura No']
            sayisal = [4, 6, 7, 8, 9]
            q = SatisKaydi.query.order_by(SatisKaydi.satis_tarihi.desc()).all()
            for s in q:
                rows.append([s.id, _tarih(s.satis_tarihi), s.musteri or '', s.cins or '',
                             _f(s.miktar, True), s.birim or '',
                             _f(s.tutar_usd, True), _f(s.maliyet_usd, True), _f(s.kar_usd, True),
                             _f(s.marj_yuzde, True), s.fatura_no or ''])
            dosya = 'satis_listesi'

        elif modul == 'karlilik':
            baslik = 'Karlılık Analizi'
            headers = ['Müşteri', 'Cins', 'Ciro (USD)', 'Maliyet (USD)', 'Kâr (USD)', 'Marj %', 'Satış Adedi']
            sayisal = [2, 3, 4, 5, 6]
            _grup = {}
            for s in SatisKaydi.query.all():
                anahtar = (s.musteri or '-', s.cins or '-')
                g = _grup.setdefault(anahtar, {'ciro': 0.0, 'mal': 0.0, 'kar': 0.0, 'adet': 0})
                g['ciro'] += (s.tutar_usd or 0)
                g['mal'] += (s.maliyet_usd or 0)
                g['kar'] += (s.kar_usd or 0)
                g['adet'] += 1
            for (mus, cins), g in sorted(_grup.items(), key=lambda x: -x[1]['kar']):
                marj = (g['kar'] / g['ciro'] * 100) if g['ciro'] else 0
                rows.append([mus, cins, _f(g['ciro'], True), _f(g['mal'], True),
                             _f(g['kar'], True), _f(marj, True), g['adet']])
            dosya = 'karlilik_analizi'

        elif modul == 'maliyet':
            baslik = 'Maliyet Listesi'
            headers = ['Maliyet No', 'Tarih', 'Tip', 'Bağlantı', 'Kayıt', 'Tutar', 'Döviz', 'USD Karşılık', 'Fatura No']
            sayisal = [5, 7]
            q = Maliyet.query.filter(Maliyet.aktif == True).order_by(Maliyet.maliyet_tarihi.desc()).all()
            for m in q:
                rows.append([m.id, _tarih(m.maliyet_tarihi), m.maliyet_tip or '',
                             m.baglanti_tip or '', m.baglanti_id or '',
                             _f(m.tutar, True), m.doviz or '', _f(m.usd_karsilik, True),
                             m.fatura_no or ''])
            dosya = 'maliyet_listesi'

        elif modul == 'kesim':
            baslik = 'Kesim Listesi'
            headers = ['Kesim No', 'Tarih', 'Kaynak Tip', 'Kaynak', 'Üretilen Adet', 'Fire', 'Fire %', 'Üretim Blok No']
            sayisal = [4, 5, 6]
            q = Kesim.query.order_by(Kesim.kesim_tarihi.desc()).all()
            for k in q:
                _det = KesimDetay.query.filter_by(kesim_id=k.id).count()
                rows.append([k.id, _tarih(k.kesim_tarihi), getattr(k, 'kaynak_tip', '') or '',
                             getattr(k, 'kaynak_id', '') or '', _det,
                             _f(getattr(k, 'fire_miktar', None), True),
                             _f(getattr(k, 'fire_orani', None), True),
                             getattr(k, 'uretim_blok_no', '') or ''])
            dosya = 'kesim_listesi'

        elif modul == 'rezervasyon':
            baslik = 'Rezervasyon Listesi'
            headers = ['Rezervasyon No', 'Tarih', 'Müşteri', 'Stok', 'Tip', 'Cins', 'Sipariş', 'Durum']
            q = Rezervasyon.query.order_by(Rezervasyon.olusturma.desc()).all()
            for r in q:
                rows.append([r.id, _tarih(getattr(r, 'olusturma', None)), r.musteri or '',
                             r.stok_id or '', r.stok_tip or '', getattr(r, 'cins', '') or '',
                             r.siparis_id or '',
                             'İptal' if getattr(r, 'iptal_tarihi', None) else 'Aktif'])
            dosya = 'rezervasyon_listesi'

        elif modul == 'denetim':
            baslik = 'Denetim Kaydı'
            headers = ['Zaman', 'Kullanıcı', 'İşlem', 'Tablo', 'Kayıt', 'IP']
            q = AuditLog.query.order_by(AuditLog.id.desc()).limit(5000).all()
            for a in q:
                _t = a.tarih
                rows.append([_t.strftime('%d.%m.%Y %H:%M') if hasattr(_t, 'strftime') else str(_t or ''),
                             a.kullanici or '', a.islem_tipi or '', a.tablo_adi or '',
                             str(a.kayit_id or ''), a.ip_adresi or ''])
            dosya = 'denetim_kaydi'

        else:
            return jsonify({'ok': False, 'mesaj': f'Bilinmeyen modül: {modul}'}), 400

        dosya = f'{dosya}_{date.today().isoformat()}'
        if fmt == 'pdf':
            return liste_pdf(baslik, headers, rows, dosya_adi=dosya, sayisal_sutunlar=sayisal)
        return liste_xlsx(baslik, headers, rows, dosya_adi=dosya, sayisal_sutunlar=sayisal)

    return app

app = create_app()
if __name__ == '__main__':
    # ── ÇALIŞTIRMA MODU ──
    # debug=True SADECE geliştirme içindir. Üretimde açık kalırsa:
    #   • Hata sayfasından sunucuda kod çalıştırılabilir (ciddi güvenlik açığı)
    #   • Performans düşer, bellek sızıntısı olur
    # Bu yüzden ortam değişkenine bağlandı. Üretimde zaten gunicorn kullanılır,
    # bu blok hiç çalışmaz.
    _debug = os.environ.get('FLASK_DEBUG', '0').lower() in ('1', 'true', 'yes')
    _port = int(os.environ.get('PORT', '5000'))
    print("=" * 54)
    print("  Milestone ERP")
    print(f"  Adres : http://0.0.0.0:{_port}")
    print(f"  Mod   : {'GELİŞTİRME (debug açık)' if _debug else 'ÜRETİM (debug kapalı)'}")
    if not _debug:
        print("  NOT   : Üretimde gunicorn ile çalıştırın:")
        print(f"          gunicorn -w 3 -b 127.0.0.1:{_port} flask_app:app")
    print("=" * 54)
    app.run(debug=_debug, host='0.0.0.0', port=_port)
