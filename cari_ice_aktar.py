#!/usr/bin/env python3
# ══════════════════════════════════════════════════════════════════════
#  Milestone ERP — CARİ İÇE AKTARMA (Excel)  ·  CI1
#
#  Baska bir programdan disa aktarilmis cari listesini sisteme alir.
#
#  ── BEKLENEN SÜTUNLAR (basliklar 1. satirda) ──
#    ÜNVAN · Üretici · Tedarikçi · Acente · Müşteri · Ülke
#    VERGİ NO · VERGİ DAİRESİ · DÖVİZ · TELEFON
#
#  ── DÖNÜŞÜMLER ──
#    · Isaretli tip sutunlari virgullu `cari_tip` olur:
#      "Müşteri,Tedarikçi" (sistemin kendi bicimi)
#    · TL   → TRY      (sistem ISO kodu kullaniyor)
#    · EURO → EUR
#    · Bos ulke → bos birakilir, UYDURULMAZ
#
#  ── AYNI ÜNVAN İKİ KEZ GELİRSE ──
#    Var olan cari GUNCELLENMEZ, ATLANIR ve raporlanir. Ustune
#    yazmak, sistemde elle duzeltilmis bir kaydi sessizce bozardi.
#
#  ── VARSAYILAN: RAPOR ──
#    --uygula demeden hicbir kayit olusturulmaz.
#
#  KULLANIM (proje klasöründe):
#      venv/bin/python cari_ice_aktar.py DOSYA.xlsx
#      venv/bin/python cari_ice_aktar.py DOSYA.xlsx --uygula
# ══════════════════════════════════════════════════════════════════════
import os
import sys
from pathlib import Path

if not Path('flask_app.py').exists():
    print("HATA: flask_app.py bu klasörde yok. Proje klasöründe çalıştırın.")
    sys.exit(1)

try:
    from dotenv import load_dotenv
    load_dotenv('.env')
except ImportError:
    pass
os.environ['MILESTONE_ACILIS_ATLA'] = '1'
sys.path.insert(0, str(Path('.').resolve()))

try:
    import openpyxl
except ImportError:
    print("HATA: openpyxl kurulu değil.")
    print("  venv/bin/pip install openpyxl")
    sys.exit(1)

import flask_app  # noqa: E402
from models import Cari, db  # noqa: E402

UYGULA = '--uygula' in sys.argv
dosyalar = [a for a in sys.argv[1:] if not a.startswith('--')]
if not dosyalar:
    print("KULLANIM: venv/bin/python cari_ice_aktar.py DOSYA.xlsx [--uygula]")
    sys.exit(1)
DOSYA = Path(dosyalar[0])
if not DOSYA.exists():
    print(f"HATA: {DOSYA} bulunamadı.")
    sys.exit(1)

DOVIZ_HARITA = {'TL': 'TRY', 'TRY': 'TRY', 'USD': 'USD',
                'EURO': 'EUR', 'EUR': 'EUR'}

print("═" * 70)
print(" CARİ İÇE AKTARMA")
print("═" * 70)
print(f" Dosya: {DOSYA.name}")
print()

kitap = openpyxl.load_workbook(DOSYA, data_only=True)
sayfa = kitap[kitap.sheetnames[0]]

def _norm(s):
    """Turkce duyarsiz normalize.

    DIKKAT: Python'da 'Üretici'.upper() → 'ÜRETICI' (NOKTASIZ I),
    ama 'ÜRETİCİ'.upper() → 'ÜRETİCİ' (noktali). Ikisi ESLESMEZ.
    Ilk surumde sutunlar bu yuzden hic bulunamadi ve 154 kayit
    "tipsiz" gorundu — oysa Excel'de isaret vardi.

    Cozum: once noktali/noktasiz i'leri tek bicime indir.
    """
    if s is None:
        return ''
    t = str(s).strip()
    t = (t.replace('İ', 'i').replace('I', 'i').replace('ı', 'i')
          .replace('Ş', 's').replace('ş', 's')
          .replace('Ç', 'c').replace('ç', 'c')
          .replace('Ğ', 'g').replace('ğ', 'g')
          .replace('Ü', 'u').replace('ü', 'u')
          .replace('Ö', 'o').replace('ö', 'o'))
    return t.lower()


# Basliklari OKU — sutun sirasina guvenmeyelim, ad ile eslesitirelim.
basliklar = {}
for i, h in enumerate(next(sayfa.iter_rows(min_row=1, max_row=1,
                                           values_only=True))):
    if h:
        basliklar[_norm(h)] = i


def sutun(*adlar):
    for a in adlar:
        if _norm(a) in basliklar:
            return basliklar[_norm(a)]
    return None


S_UNVAN = sutun('ÜNVAN', 'UNVAN')
if S_UNVAN is None:
    print(" ✗ 'ÜNVAN' sütunu bulunamadı. Başlıklar 1. satırda mı?")
    print(f"   Bulunanlar: {list(basliklar)[:10]}")
    sys.exit(1)
S_URETICI = sutun('ÜRETİCİ', 'URETICI')
S_TEDARIK = sutun('TEDARİKÇİ', 'TEDARIKCI')
S_ACENTE = sutun('ACENTE')
S_MUSTERI = sutun('MÜŞTERİ', 'MUSTERI')
S_ULKE = sutun('ÜLKE', 'ULKE')
S_VNO = sutun('VERGİ NO', 'VERGI NO')
S_VD = sutun('VERGİ DAİRESİ', 'VERGI DAIRESI')
S_DOVIZ = sutun('DÖVİZ', 'DOVIZ')
S_TEL = sutun('TELEFON')


def metin(satir, idx, uzunluk=None):
    if idx is None or idx >= len(satir):
        return ''
    d = satir[idx]
    if d is None:
        return ''
    s = str(d).strip()
    return s[:uzunluk] if uzunluk else s


def isaretli(satir, idx):
    """Excel'de ✔ / X / 1 gibi degerler 'isaretli' sayilir."""
    if idx is None or idx >= len(satir):
        return False
    d = satir[idx]
    return bool(d) and str(d).strip() not in ('', '0', 'False')


with flask_app.app.app_context():
    mevcut = {}
    for c in Cari.query.all():
        # Mukerrer kontrolu de Turkce duyarsiz — "ŞTİ" ile "STI"
        # ayni firmayi gosterebilir.
        mevcut[_norm(c.unvan)] = c.id

    yeni, atlanan, tipsiz = [], [], 0
    for satir in sayfa.iter_rows(min_row=2, values_only=True):
        unvan = metin(satir, S_UNVAN, 200)
        if not unvan:
            continue
        if _norm(unvan) in mevcut:
            atlanan.append((unvan, mevcut[_norm(unvan)]))
            continue

        tipler = []
        if isaretli(satir, S_MUSTERI):
            tipler.append('Müşteri')
        if isaretli(satir, S_TEDARIK):
            tipler.append('Tedarikçi')
        if isaretli(satir, S_ACENTE):
            tipler.append('Acente')
        if isaretli(satir, S_URETICI):
            tipler.append('Üretici')
        if not tipler:
            # Tip UYDURULMAZ; bos birakmak yerine raporlanir.
            tipsiz += 1
            tipler = ['Müşteri']

        _dv = metin(satir, S_DOVIZ).upper()
        yeni.append({
            'unvan': unvan,
            'cari_tip': ','.join(tipler),
            'ulke': metin(satir, S_ULKE, 80),
            'vergi_no': metin(satir, S_VNO, 20),
            'vergi_dairesi': metin(satir, S_VD, 100),
            'para_birimi': DOVIZ_HARITA.get(_dv, 'USD'),
            'telefon': metin(satir, S_TEL, 30),
        })

    print(f" Eklenecek : {len(yeni)}")
    print(f" Atlanan   : {len(atlanan)}  (ünvan zaten kayıtlı)")
    if tipsiz:
        print(f" ⚠ {tipsiz} kayıtta tip işaretli değil → 'Müşteri' varsayıldı")
    print()

    if yeni:
        print(" ── İLK 5 ÖRNEK ──")
        for k in yeni[:5]:
            print(f"   {k['unvan'][:34]:<34} {k['cari_tip'][:22]:<22} "
                  f"{k['para_birimi']} {k['ulke'][:12]}")
        print()
    if atlanan:
        print(" ── ATLANANLAR (ilk 5) ──")
        for u, i in atlanan[:5]:
            print(f"   {u[:44]:<44} → {i}")
        print()

    if not UYGULA:
        print("─" * 70)
        print(" RAPOR MODU — hiçbir kayıt oluşturulmadı.")
        print()
        print(" Uygulamak için:")
        print(f"   venv/bin/python cari_ice_aktar.py {DOSYA.name} --uygula")
        print()
        print(" ⚠ Önce yedek alın:")
        print("   sudo /usr/local/bin/milestone-yedek.sh")
        print("─" * 70)
        sys.exit(0)

    if not yeni:
        print(" ✓ Eklenecek yeni cari yok.")
        sys.exit(0)

    # ID uretimi sistemin kendi yardimcisiyla — elle sayac tutmak
    # cakisma riski dogururdu.
    _yeni_id = getattr(flask_app, '_yeni_id', None)
    eklendi = 0
    for k in yeni:
        try:
            if _yeni_id:
                cid = _yeni_id('CR')
            else:
                cid = f"CR-{eklendi + 1:05d}"
            db.session.add(Cari(id=cid, gorunurluk='ortak', **k))
            eklendi += 1
        except Exception as exc:
            print(f"   ✗ {k['unvan'][:40]}: {str(exc)[:60]}")
    db.session.commit()
    print("═" * 70)
    print(f" ✓ {eklendi} cari eklendi.")
    print()
    print(" SONRAKİ ADIM: Cari ekranından sorumlu ve görünürlük atayın;")
    print(" atanmayan cariler satış ekibine kapalı görünebilir.")
    print("═" * 70)
