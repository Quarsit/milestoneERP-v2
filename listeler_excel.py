#!/usr/bin/env python3
# ══════════════════════════════════════════════════════════════════════
#  Milestone ERP — AYARLAR → LİSTELER  ·  EXCEL DIŞA AKTARIM
#
#  NE YAPAR:
#    Ayarlar ekranındaki "Listeler" sekmesinde tuttuğunuz tüm kayıtlı
#    değerleri veritabanından okur ve tek bir Excel dosyasına yazar.
#    Her liste AYRI SAYFA olur; başta bir "Özet" sayfası bulunur.
#
#  NEDEN İŞE YARAR:
#    • Yedek/arşiv: listeler veritabanında; Excel'de elde taşınabilir
#    • Gözden geçirme: mükerrer veya yazım hatalı kayıtları görmek kolay
#    • Paylaşım: mali müşavir / ekip arkadaşına gönderilebilir
#
#  HANGİ LİSTELER:
#    Ayarlar ekranında tanımlı altı liste:
#      cins       Mermer Cinsleri
#      ozellik    Yüzey İşlemleri
#      ulke       Ülkeler
#      liman      Limanlar
#      konteyner  Konteyner Tipleri
#      gtip       GTİP / HS Kodları
#
#    Bunların dışında veritabanında BAŞKA bir liste kategorisi varsa
#    (ileride eklenirse) o da otomatik bulunur ve dışa aktarılır —
#    sessizce atlanmaz. Ayar kayıtları (firma bilgisi, logo, SMTP,
#    KDV oranı vb.) liste DEĞİLDİR; dışarıda bırakılır.
#
#  BU BETİK VERİTABANINI DEĞİŞTİRMEZ — yalnızca okur.
#
#  KULLANIM (proje dizininde):
#      venv/bin/python listeler_excel.py
#      venv/bin/python listeler_excel.py --cikti /home/mermer/listeler.xlsx
#
#  Varsayılan çıktı: ~/milestone_listeler_<tarih>.xlsx
# ══════════════════════════════════════════════════════════════════════
import os
import sys
from datetime import datetime
from pathlib import Path

from dotenv import load_dotenv
from sqlalchemy import create_engine, text

load_dotenv()

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("HATA: openpyxl kurulu değil.")
    print("  venv/bin/pip install openpyxl")
    sys.exit(1)

# ── Bağlantı ──
VT_URL = None
CIKTI = None
args = sys.argv[1:]
for i, a in enumerate(args):
    if a.startswith('--url='):
        VT_URL = a.split('=', 1)[1].strip().strip('"').strip("'")
    if a == '--cikti' and i + 1 < len(args):
        CIKTI = args[i + 1]
    if a.startswith('--cikti='):
        CIKTI = a.split('=', 1)[1]

VT_URL = VT_URL or os.environ.get('DATABASE_URL')
if not VT_URL:
    print("HATA: DATABASE_URL bulunamadı (.env okunamadı).")
    print("  Elle vermek için: --url=\"postgresql://kullanici:sifre@localhost:5432/vt\"")
    sys.exit(1)
if VT_URL.startswith('postgres://'):
    VT_URL = VT_URL.replace('postgres://', 'postgresql://', 1)

# ── Ayarlar ekranındaki liste kategorileri (templates/ayarlar.html:LOOKUPLAR) ──
LISTELER = [
    ('cins',      'Mermer Cinsleri',    'Stok ve proforma formlarında cins önerisi'),
    ('ozellik',   'Yüzey İşlemleri',    'Polisaj, honlu, eskitme vb.'),
    ('ulke',      'Ülkeler',            'Cari ve belge ülke alanları (ISO listesi)'),
    ('liman',     'Limanlar',           'Yükleme / varış limanı — belgelerde basılır'),
    ('konteyner', 'Konteyner Tipleri',  "20' DC, 40' HC vb."),
    ('gtip',      'GTİP / HS Kodları',  'Gümrük tarife istatistik pozisyonu'),
]

# Ayarlar ekranında GÖRÜNMEYEN ama sistemin kullandığı listeler.
# Proforma/fatura/sipariş formlarındaki açılır kutuları besler; ilk
# kurulumda tohumlanır. Ekranda düzenlenemedikleri için gözden kaçarlar,
# bu yüzden dışa aktarımda ADLARIYLA yer alırlar.
SISTEM_LISTELERI = {
    'odeme':         ('Ödeme Şekilleri',    'Proforma/fatura ödeme şekli açılır kutusu'),
    'teslim':        ('Teslim Şekilleri',   'Incoterm — EXW, FOB, CIF vb.'),
    'durum':         ('Stok Durumları',     'Serbest, Rezerve, Satıldı, Teslim Edildi…'),
    'siparis_durum': ('Sipariş Durumları',  'Teklif Aşaması → Onaylandı → … → Teslim'),
}

# Liste DEĞİL — bunlar ayar kayıtlarıdır, dışa aktarılmaz.
AYAR_KATEGORILERI = {
    'firma', 'firma_logo', 'smtp_ayar', 'smtp_kullanici',
    'kdv_ayar', 'siparis_ayar', 'muhasebe',
}

# ── Biçim ──
YAZI = 'Arial'
BASLIK_DOLGU = PatternFill('solid', fgColor='1F4E5F')
BASLIK_YAZI = Font(name=YAZI, size=11, bold=True, color='FFFFFF')
NORMAL = Font(name=YAZI, size=10)
KALIN = Font(name=YAZI, size=10, bold=True)
BUYUK = Font(name=YAZI, size=14, bold=True, color='1F4E5F')
SOLUK = Font(name=YAZI, size=9, color='808080')
INCE = Side(style='thin', color='D0D0D0')
CERCEVE = Border(left=INCE, right=INCE, top=INCE, bottom=INCE)


def sutun_genislet(sayfa, veriler, min_g=10, max_g=60):
    """Sütun genişliklerini içeriğe göre ayarlar."""
    if not veriler:
        return
    for idx in range(len(veriler[0])):
        en_uzun = max((len(str(s[idx])) if s[idx] is not None else 0)
                      for s in veriler)
        sayfa.column_dimensions[get_column_letter(idx + 1)].width = \
            max(min_g, min(max_g, en_uzun + 4))


print("═" * 68)
print(" MILESTONE ERP — AYARLAR → LİSTELER · EXCEL DIŞA AKTARIM")
print("═" * 68)

motor = create_engine(VT_URL)
with motor.connect() as b:
    satirlar = b.execute(text(
        'SELECT id, kategori, deger, kisaltma, ek_bilgi FROM veriler '
        'ORDER BY kategori, deger'
    )).fetchall()

# Kategoriye göre grupla
gruplar = {}
for r in satirlar:
    gruplar.setdefault(r[1] or '(kategorisiz)', []).append(r)

# Bilinen listeler + veritabanında olup listede olmayan yeni kategoriler
bilinen = {k for k, _, _ in LISTELER}
ekstra = sorted(k for k in gruplar
                if k not in bilinen and k not in AYAR_KATEGORILERI)
if ekstra:
    tanidik = [k for k in ekstra if k in SISTEM_LISTELERI]
    yabanci = [k for k in ekstra if k not in SISTEM_LISTELERI]
    if tanidik:
        print(f"\n ℹ Ayarlar ekranında görünmeyen {len(tanidik)} sistem listesi")
        print(f"   de aktarılacak: {', '.join(tanidik)}")
    if yabanci:
        print(f"\n ℹ Tanımsız {len(yabanci)} kategori bulundu, onlar da")
        print(f"   aktarılacak: {', '.join(yabanci)}")

def _ekstra_etiket(k):
    if k in SISTEM_LISTELERI:
        ad, acik = SISTEM_LISTELERI[k]
        return (k, ad, acik + '  [ekranda düzenlenemez]')
    return (k, k.upper(), 'Veritabanında bulundu — Ayarlar ekranında tanımlı değil')


aktarilacak = list(LISTELER) + [_ekstra_etiket(k) for k in ekstra]

print()
kitap = Workbook()

# ══ ÖZET SAYFASI ══
ozet = kitap.active
ozet.title = 'Özet'
ozet['A1'] = 'MILESTONE ERP — AYARLAR LİSTELERİ'
ozet['A1'].font = BUYUK
ozet['A2'] = f"Dışa aktarma: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
ozet['A2'].font = SOLUK
ozet.merge_cells('A1:D1')
ozet.merge_cells('A2:D2')

basliklar = ['Liste', 'Kayıt Sayısı', 'Sayfa Adı', 'Açıklama']
for s, b in enumerate(basliklar, start=1):
    h = ozet.cell(row=4, column=s, value=b)
    h.font = BASLIK_YAZI
    h.fill = BASLIK_DOLGU
    h.alignment = Alignment(horizontal='center', vertical='center')
    h.border = CERCEVE

satir = 5
toplam_kayit = 0
sayfa_adlari = {}
for kategori, etiket, aciklama in aktarilacak:
    kayitlar = gruplar.get(kategori, [])
    toplam_kayit += len(kayitlar)
    # Excel sayfa adı 31 karakterle sınırlı ve bazı karakterleri kabul etmez
    sayfa_adi = etiket[:31].replace('/', '-').replace('\\', '-')
    sayfa_adlari[kategori] = sayfa_adi

    ozet.cell(row=satir, column=1, value=etiket).font = KALIN
    # Sayı SABİT DEĞİL — ilgili sayfadan sayılır, veri değişirse güncellenir.
    ozet.cell(row=satir, column=2,
              value=f"=COUNTA('{sayfa_adi}'!B2:B10000)").font = NORMAL
    ozet.cell(row=satir, column=3, value=sayfa_adi).font = NORMAL
    ozet.cell(row=satir, column=4, value=aciklama).font = NORMAL
    for s in range(1, 5):
        ozet.cell(row=satir, column=s).border = CERCEVE
    ozet.cell(row=satir, column=2).alignment = Alignment(horizontal='center')
    satir += 1

ozet.cell(row=satir, column=1, value='TOPLAM').font = KALIN
ozet.cell(row=satir, column=2, value=f'=SUM(B5:B{satir - 1})').font = KALIN
ozet.cell(row=satir, column=2).alignment = Alignment(horizontal='center')
for s in range(1, 5):
    ozet.cell(row=satir, column=s).border = CERCEVE

not_satir = satir + 2
ozet.cell(row=not_satir, column=1,
          value='Kaynak: Milestone ERP veritabanı, veriler tablosu '
                '(Ayarlar → Listeler ekranından yönetilir).').font = SOLUK
ozet.cell(row=not_satir + 1, column=1,
          value='Kayıt sayıları formülle hesaplanır; ilgili sayfada satır '
                'eklenip silindiğinde kendiliğinden güncellenir.').font = SOLUK

ozet.column_dimensions['A'].width = 26
ozet.column_dimensions['B'].width = 14
ozet.column_dimensions['C'].width = 24
ozet.column_dimensions['D'].width = 52
ozet.freeze_panes = 'A5'

# ══ HER LİSTE İÇİN SAYFA ══
for kategori, etiket, aciklama in aktarilacak:
    kayitlar = gruplar.get(kategori, [])
    sayfa = kitap.create_sheet(sayfa_adlari[kategori])

    basliklar = ['Sıra', 'Değer', 'Kısaltma', 'Ek Bilgi']
    for s, b in enumerate(basliklar, start=1):
        h = sayfa.cell(row=1, column=s, value=b)
        h.font = BASLIK_YAZI
        h.fill = BASLIK_DOLGU
        h.alignment = Alignment(horizontal='center', vertical='center')
        h.border = CERCEVE

    veriler = [basliklar]
    for i, r in enumerate(kayitlar, start=1):
        sayfa.cell(row=i + 1, column=1, value=i).font = NORMAL
        sayfa.cell(row=i + 1, column=2, value=r[2]).font = NORMAL
        sayfa.cell(row=i + 1, column=3, value=r[3]).font = NORMAL
        sayfa.cell(row=i + 1, column=4, value=r[4]).font = NORMAL
        for s in range(1, 5):
            sayfa.cell(row=i + 1, column=s).border = CERCEVE
        sayfa.cell(row=i + 1, column=1).alignment = Alignment(horizontal='center')
        veriler.append([i, r[2], r[3], r[4]])

    if not kayitlar:
        h = sayfa.cell(row=2, column=2, value='(bu listede kayıt yok)')
        h.font = SOLUK

    sutun_genislet(sayfa, veriler)
    sayfa.column_dimensions['A'].width = 7
    sayfa.freeze_panes = 'A2'
    sayfa.auto_filter.ref = f'A1:D{max(2, len(kayitlar) + 1)}'

    print(f"   ✓ {etiket:<22s} {len(kayitlar):>4} kayıt")

# ── Kaydet ──
if CIKTI:
    hedef = Path(CIKTI).expanduser()
else:
    damga = datetime.now().strftime('%Y%m%d_%H%M')
    hedef = Path.home() / f'milestone_listeler_{damga}.xlsx'
hedef.parent.mkdir(parents=True, exist_ok=True)
kitap.save(hedef)

print()
print("═" * 68)
print(f" ✓ TAMAMLANDI — {toplam_kayit} kayıt, {len(aktarilacak)} liste")
print()
print(f" Dosya: {hedef}")
print(f" Boyut: {hedef.stat().st_size / 1024:.0f} KB")
print()
print(" Özet sayfasındaki kayıt sayıları FORMÜLDÜR. Excel dosyayı ilk")
print(" açtığında hesaplar; bazı görüntüleyicilerde boş görünebilir.")
print("═" * 68)
