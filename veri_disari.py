#!/usr/bin/env python3
# ══════════════════════════════════════════════════════════════════════
#  Milestone ERP — REFERANS VERİ DIŞA AKTARMA  ·  VD1
#
#  Cariler, kasalar, bankalar ve liste tanimlarini (cins, yuzey,
#  ulke...) tek bir Excel dosyasina yazar. Her tablo AYRI SAYFA.
#
#  ── NEDEN "İSTENDİĞİNDE", SÜREKLİ DEĞİL ──
#    Kayit eklendikce Excel'e yazmak, ayni bilginin IKI KOPYASINI
#    olustururdu. Bu projede ayni hata sinifinin bes ornegini
#    duzelttik (ekranda gorunen kayitla uyusmuyor).
#
#    Somut riskler:
#      · Dosya kilitliyken yazma basarisiz olur → kayit var, Excel'de yok
#      · Excel elle duzenlenirse sisteme geri donmez → sessiz ayrisma
#      · Iki kullanici ayni anda eklerse dosya cakisir → biri kaybolur
#
#    Bu betik TEK YONLU: veritabani → Excel. Cikti her zaman o anki
#    gercegi gosterir, cakisma olmaz, hata gizlenmez.
#
#  ── LİSTE KATEGORİLERİ SABİT YAZILMADI ──
#    Veritabaninda GERCEKTEN var olan kategoriler taranir. Sabit
#    liste yazsaydim, yeni bir kategori eklendiginde disarida
#    kalirdi — bu projede o tuzagi birkac kez gorduk.
#
#  ── ÇIKTI GİTHUB'A GİTMEMELİ ──
#    Musteri verisi iceriyor. .gitignore'a eklenmeli:
#        veri_disari/
#        *.xlsx
#
#  KULLANIM (proje klasöründe):
#      venv/bin/python veri_disari.py
#      venv/bin/python veri_disari.py --klasor ~/yedekler
#
#  Zamanlanmis gorev icin (gunde bir, yedeklerin yanina):
#      0 20 * * *  cd /home/mermer/milestoneERP-v2 && \
#                  venv/bin/python veri_disari.py --klasor ~/yedekler
# ══════════════════════════════════════════════════════════════════════
import os
import sys
from datetime import datetime
from pathlib import Path

if not Path('flask_app.py').exists():
    print("HATA: flask_app.py bu klasörde yok. Proje klasöründe çalıştırın.")
    sys.exit(1)

try:
    from dotenv import load_dotenv
    load_dotenv('.env')
except ImportError:
    pass
os.environ['MILESTONE_ACILIS_ATLA'] = '1'
sys.path.insert(0, str(Path('.').resolve()))

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("HATA: openpyxl kurulu değil.")
    print("  venv/bin/pip install openpyxl")
    sys.exit(1)

import flask_app  # noqa: E402
from models import Banka, Cari, Kasa, Veriler, db  # noqa: E402

# --klasor secenegi
KLASOR = Path('veri_disari')
if '--klasor' in sys.argv:
    _i = sys.argv.index('--klasor')
    if _i + 1 < len(sys.argv):
        KLASOR = Path(os.path.expanduser(sys.argv[_i + 1]))

print("═" * 70)
print(" REFERANS VERİ DIŞA AKTARMA")
print("═" * 70)
print()


def sayfa_yaz(kitap, ad, basliklar, satirlar):
    """Bir tabloyu kendi sayfasina yazar."""
    s = kitap.create_sheet(ad[:31])
    s.append(basliklar)
    for h in s[1]:
        h.font = Font(bold=True, color='FFFFFF')
        h.fill = PatternFill('solid', start_color='1F4E3D')
    for r in satirlar:
        s.append(r)
    # Sutun genisligi: en uzun degere gore, makul bir tavanla.
    for i, _ in enumerate(basliklar, start=1):
        en = max([len(str(basliklar[i - 1]))] +
                 [len(str(r[i - 1])) for r in satirlar[:200]
                  if i - 1 < len(r) and r[i - 1] is not None] or [10])
        s.column_dimensions[get_column_letter(i)].width = min(en + 3, 46)
    s.freeze_panes = 'A2'
    return len(satirlar)


with flask_app.app.app_context():
    KLASOR.mkdir(parents=True, exist_ok=True)
    damga = datetime.now().strftime('%Y%m%d_%H%M')
    hedef = KLASOR / f'milestone_referans_{damga}.xlsx'

    kitap = openpyxl.Workbook()
    kitap.remove(kitap.active)          # varsayilan bos sayfa
    ozet = []

    # ── CARİLER ──
    n = sayfa_yaz(kitap, 'Cariler',
                  ['ID', 'Ünvan', 'Tip', 'Ülke', 'Vergi No', 'Vergi Dairesi',
                   'Döviz', 'Telefon', 'E-posta', 'Yetkili', 'Adres',
                   'Risk Limiti', 'Sorumlu', 'Görünürlük'],
                  [[c.id, c.unvan, c.cari_tip, c.ulke, c.vergi_no,
                    c.vergi_dairesi, c.para_birimi, c.telefon, c.email,
                    c.yetkili, c.adres, float(c.risk_limiti or 0),
                    c.sorumlu, c.gorunurluk]
                   for c in Cari.query.order_by(Cari.unvan).all()])
    ozet.append(('Cariler', n))

    # ── KASALAR ──
    n = sayfa_yaz(kitap, 'Kasalar',
                  ['ID', 'Ad', 'Döviz', 'Bakiye', 'Ana Kasa', 'Banka ID',
                   'Varsayılan', 'Aktif', 'Açıklama'],
                  [[k.id, k.ad, k.doviz, float(k.bakiye or 0),
                    'Evet' if k.ana_kasa else '', k.banka_id,
                    'Evet' if k.varsayilan else '',
                    'Evet' if k.aktif else 'Hayır', k.aciklama]
                   for k in Kasa.query.order_by(Kasa.ad).all()])
    ozet.append(('Kasalar', n))

    # ── BANKALAR ──
    n = sayfa_yaz(kitap, 'Bankalar',
                  ['ID', 'Banka Adı', 'Şube', 'Hesap No', 'IBAN', 'SWIFT',
                   'Döviz', 'Varsayılan', 'Aktif', 'Açıklama'],
                  [[b.id, b.banka_adi, b.sube, b.hesap_no, b.iban, b.swift,
                    b.doviz, 'Evet' if b.varsayilan else '',
                    'Evet' if b.aktif else 'Hayır', b.aciklama]
                   for b in Banka.query.order_by(Banka.banka_adi).all()])
    ozet.append(('Bankalar', n))

    # ── LİSTELER (cins, yüzey, ülke, ölçü...) ──
    # Kategoriler SABİT YAZILMADI: veritabanında gerçekten var
    # olanlar taranır. Yeni bir kategori eklenirse kendiliğinden
    # gelir.
    #
    # Ayar kayıtları (SMTP, logo, firma bilgisi) DIŞARIDA:
    # referans listesi değiller ve logo binary veri içerir.
    AYAR_KATEGORI = {'firma', 'firma_logo', 'smtp_ayar', 'kdv_ayar',
                     'muhasebe', 'siparis_ayar'}
    kategoriler = sorted({
        k for (k,) in db.session.query(Veriler.kategori).distinct().all()
        if k and k not in AYAR_KATEGORI})

    liste_satir = []
    for kat in kategoriler:
        for v in Veriler.query.filter_by(kategori=kat).order_by(
                Veriler.deger).all():
            liste_satir.append([kat, v.deger, v.kisaltma, v.ek_bilgi])
    n = sayfa_yaz(kitap, 'Listeler',
                  ['Kategori', 'Değer', 'Kısaltma', 'Ek Bilgi'], liste_satir)
    ozet.append((f'Listeler ({len(kategoriler)} kategori)', n))

    kitap.save(hedef)

for ad, n in ozet:
    print(f"   {ad:<28} {n:>5} kayıt")
print()
print("═" * 70)
print(f" ✓ {hedef}")
print()
print(" ⚠ Bu dosya MÜŞTERİ VERİSİ içerir — GitHub'a göndermeyin.")
print("   .gitignore'a ekleyin:")
print("     veri_disari/")
print("     *.xlsx")
print("═" * 70)
